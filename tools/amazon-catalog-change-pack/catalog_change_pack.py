#!/usr/bin/env python3
"""Build, validate, and review Amazon catalog variation change packs.

This utility writes files only. It never uploads or changes an Amazon catalog.

Usage:
  python3 catalog_change_pack.py build --manifest change.json
  python3 catalog_change_pack.py validate --manifest change.json --file upload.xlsm
  python3 catalog_change_pack.py review-summary --manifest change.json --summary result.xlsm
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import openpyxl


HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent

FULL_UPDATE = "Create or Replace (Full Update)"
PARTIAL_UPDATE = "Edit (Partial Update)"
DELETE = "Delete"
VALID_ACTIONS = {FULL_UPDATE, PARTIAL_UPDATE, DELETE}

OPERATIONS = {
    "create_parent",
    "add_existing_child",
    "create_gtin_exempt_child",
    "delete_parent",
    "rebuild_family",
    "detach_children",
    "reparent_children",
}

ERROR_GUIDANCE = {
    "8032": (
        "Conflicting parentage",
        "The child is still assigned to another parent. Process the old-parent delete file, "
        "confirm detachment, then submit the rebuild or reparent file.",
    ),
    "8066": (
        "Invalid parentage setup",
        "Check Parentage Level, Parent SKU, variation theme, product type, and the distinct "
        "variation value on every child.",
    ),
    "8801": (
        "Duplicate variation attributes",
        "Two children share the same variation value or Amazon is still retaining a conflicting "
        "catalog attribute. Use one unique value per child and verify the selected theme.",
    ),
    "90041": (
        "Missing required data",
        "Inspect the other errors on the same SKU first. This can be secondary to an invalid "
        "listing action; otherwise fill the required fields from Data Definitions.",
    ),
    "90057": (
        "Invalid listing action",
        f"Use one of: {FULL_UPDATE}, {PARTIAL_UPDATE}, or {DELETE}.",
    ),
    "100521": (
        "Amazon review pending",
        "No corrective upload is indicated yet. Wait up to 48 hours, then verify the backend and frontend.",
    ),
}


class CatalogError(Exception):
    """A user-correctable manifest, template, or workbook error."""


def close_workbook(workbook: Any) -> None:
    """Close openpyxl workbooks and their copied VBA archive cleanly."""
    archive = getattr(workbook, "vba_archive", None)
    if archive is not None:
        archive.close()
        workbook.vba_archive = None
    workbook.close()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def normalized(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def load_manifest(path: Path) -> dict[str, Any]:
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise CatalogError(f"Could not read manifest {path}: {exc}") from exc
    manifest["_manifest_path"] = str(path.resolve())
    return manifest


def resolve_source(manifest: dict[str, Any], key: str) -> Path:
    raw = manifest.get("source", {}).get(key)
    if not raw:
        raise CatalogError(f"source.{key} is required")
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = Path(manifest["_manifest_path"]).parent / path
    return path.resolve()


def validate_manifest(manifest: dict[str, Any], *, require_files: bool = True) -> list[str]:
    issues: list[str] = []
    notes: list[str] = []
    for key in ("client", "seller_account", "marketplace", "operation"):
        if not manifest.get(key):
            issues.append(f"{key} is required")
    if manifest.get("schema_version") != 1:
        issues.append("schema_version must be 1")
    operation = manifest.get("operation")
    if operation not in OPERATIONS:
        issues.append(f"operation must be one of: {', '.join(sorted(OPERATIONS))}")

    source = manifest.get("source", {})
    for key in ("category_listings_report", "blank_template"):
        if not source.get(key):
            issues.append(f"source.{key} is required")
        elif require_files:
            try:
                path = resolve_source(manifest, key)
                if not path.exists():
                    issues.append(f"source.{key} does not exist: {path}")
                elif key == "blank_template" and path.suffix.lower() != ".xlsm":
                    issues.append("source.blank_template must be Amazon's macro-enabled .xlsm template")
            except CatalogError as exc:
                issues.append(str(exc))

    family = manifest.get("family") or {}
    theme = family.get("variation_theme")
    variation_attribute = family.get("variation_attribute")
    parent = family.get("parent") or {}
    children = family.get("children") or []
    if operation != "delete_parent":
        if not theme:
            issues.append("family.variation_theme is required")
        if not variation_attribute:
            issues.append("family.variation_attribute is required")
    if not parent.get("sku"):
        issues.append("family.parent.sku is required")
    if not parent.get("product_type"):
        issues.append("family.parent.product_type is required")
    if parent.get("status") and parent.get("status") not in {"new", "existing"}:
        issues.append("family.parent.status must be new or existing")
    if operation == "reparent_children" and parent.get("status") not in {"new", "existing"}:
        issues.append("family.parent.status is required for reparent_children")

    if operation not in {"delete_parent"} and not children:
        issues.append("family.children[] must contain at least one child")

    skus: list[str] = []
    values: list[str] = []
    for index, child in enumerate(children, 1):
        tag = f"family.children[{index}]"
        for key in ("sku", "status"):
            if not child.get(key):
                issues.append(f"{tag}.{key} is required")
        if child.get("status") not in {"existing", "new"}:
            issues.append(f"{tag}.status must be existing or new")
        if child.get("product_type") and child.get("product_type") != parent.get("product_type"):
            issues.append(f"{tag}.product_type must match family.parent.product_type")
        if child.get("status") == "existing":
            asin = str(child.get("asin") or "").strip().upper()
            if not re.fullmatch(r"[A-Z0-9]{10}", asin):
                issues.append(f"{tag}.asin must be a 10-character ASIN for an existing child")
        if not child.get("variation_value"):
            issues.append(f"{tag}.variation_value is required")
        if child.get("sku"):
            skus.append(str(child["sku"]))
        if child.get("variation_value"):
            values.append(str(child["variation_value"]).strip().casefold())
        if child.get("status") == "new":
            if operation == "create_gtin_exempt_child" and not child.get("gtin_exempt"):
                issues.append(f"{tag}.gtin_exempt must be true for this operation")
            if child.get("gtin_exempt") and nonempty(child.get("product_id")):
                issues.append(f"{tag}.product_id must be blank when gtin_exempt is true")
            if not child.get("fields"):
                notes.append(f"{tag}.fields is empty; Amazon may reject missing required attributes")
        elif child.get("fields"):
            issues.append(
                f"{tag}.fields must be empty for an existing child; existing children use minimal partial updates"
            )

    if parent.get("sku"):
        skus.append(str(parent["sku"]))
    duplicate_skus = sorted({sku for sku in skus if skus.count(sku) > 1})
    if duplicate_skus:
        issues.append(f"duplicate SKUs in manifest: {', '.join(duplicate_skus)}")
    duplicate_values = sorted({v for v in values if values.count(v) > 1})
    if duplicate_values:
        issues.append(f"duplicate variation values: {', '.join(duplicate_values)}")

    removed = set(map(str, family.get("removed_child_skus") or []))
    child_skus = {str(c.get("sku")) for c in children if c.get("sku")}
    overlap = sorted(removed & child_skus)
    if overlap:
        issues.append(f"removed_child_skus cannot also be rebuilt as children: {', '.join(overlap)}")
    if operation == "detach_children" and not removed:
        issues.append("family.removed_child_skus is required for detach_children")
    if operation == "reparent_children" and not family.get("old_parent_sku"):
        issues.append("family.old_parent_sku is required for reparent_children")

    if require_files and source.get("category_listings_report"):
        try:
            report_path = resolve_source(manifest, "category_listings_report")
            if report_path.exists():
                report_skus = read_report_skus(report_path)
                allowed_missing = set(map(str, source.get("allow_missing_skus") or []))
                if not report_skus:
                    issues.append("Category Listings Report contains no readable SKU column")
                else:
                    expected_existing = {
                        str(child["sku"])
                        for child in children
                        if child.get("status") == "existing" and child.get("sku")
                    }
                    if operation in {"delete_parent", "rebuild_family", "detach_children"} and parent.get("sku"):
                        expected_existing.add(str(parent["sku"]))
                    if operation == "reparent_children" and family.get("old_parent_sku"):
                        expected_existing.add(str(family["old_parent_sku"]))
                        if parent.get("status") == "existing":
                            expected_existing.add(str(parent["sku"]))
                    if operation == "detach_children":
                        expected_existing.update(map(str, family.get("removed_child_skus") or []))
                    missing = sorted(expected_existing - report_skus - allowed_missing)
                    if missing:
                        issues.append(
                            "existing SKUs missing from Category Listings Report: "
                            + ", ".join(missing)
                            + "; add only known report omissions to source.allow_missing_skus"
                        )
                    new_skus = {
                        str(child["sku"])
                        for child in children
                        if child.get("status") == "new" and child.get("sku")
                    }
                    if operation == "create_parent" and parent.get("sku"):
                        new_skus.add(str(parent["sku"]))
                    if operation == "reparent_children" and parent.get("status") == "new":
                        new_skus.add(str(parent["sku"]))
                    collisions = sorted(new_skus & report_skus)
                    if collisions:
                        issues.append(f"new SKUs already exist in Category Listings Report: {', '.join(collisions)}")
                    if allowed_missing:
                        notes.append(
                            "Category Listings Report omissions explicitly allowed: "
                            + ", ".join(sorted(allowed_missing))
                        )
        except (CatalogError, OSError, ValueError) as exc:
            issues.append(f"Could not inspect Category Listings Report: {exc}")

    if issues:
        raise CatalogError("Manifest validation failed:\n- " + "\n- ".join(issues))
    return notes


def read_report_skus(path: Path) -> set[str]:
    """Read SKU values from modern CLRs or conventional SKU-column reports."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    found: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            settings = str(sheet["A1"].value or "")
            if "settings" in settings.lower():
                try:
                    attribute_row = parse_setting(settings, "attributeRow")
                    data_row = parse_setting(settings, "dataRow")
                    sku_col = next(
                        (
                            col
                            for col in range(1, sheet.max_column + 1)
                            if normalized(sheet.cell(attribute_row, col).value) in {"itemsku", "sku"}
                        ),
                        None,
                    )
                    if sku_col:
                        for row in range(data_row, sheet.max_row + 1):
                            value = sheet.cell(row, sku_col).value
                            if nonempty(value):
                                found.add(str(value).strip())
                        continue
                except CatalogError:
                    pass
            for header_row in range(1, min(sheet.max_row, 25) + 1):
                sku_col = next(
                    (
                        col
                        for col in range(1, sheet.max_column + 1)
                        if normalized(sheet.cell(header_row, col).value) in {"itemsku", "sku", "sellersku"}
                    ),
                    None,
                )
                if not sku_col:
                    continue
                for row in range(header_row + 1, sheet.max_row + 1):
                    value = sheet.cell(row, sku_col).value
                    if nonempty(value):
                        found.add(str(value).strip())
                break
    finally:
        close_workbook(workbook)
    return found


@dataclass
class TemplateInfo:
    workbook: Any
    sheet: Any
    label_row: int
    attribute_row: int
    data_row: int
    attributes: dict[str, int]
    labels: dict[str, int]


def parse_setting(settings: str, key: str) -> int:
    match = re.search(rf"(?:^|[;&,\s]){re.escape(key)}\s*[=:]\s*(\d+)", settings, re.I)
    if not match:
        match = re.search(rf"{re.escape(key)}[^0-9]*(\d+)", settings, re.I)
    if not match:
        raise CatalogError(f"Could not find {key} in template settings cell")
    return int(match.group(1))


def load_template(path: Path) -> TemplateInfo:
    if not path.exists():
        raise CatalogError(f"Template does not exist: {path}")
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = openpyxl.load_workbook(path, keep_vba=keep_vba)
    candidates = []
    for sheet in workbook.worksheets:
        settings = str(sheet["A1"].value or "")
        if "settings" not in settings.lower():
            continue
        try:
            label_row = parse_setting(settings, "labelRow")
            attribute_row = parse_setting(settings, "attributeRow")
            data_row = parse_setting(settings, "dataRow")
        except CatalogError:
            continue
        attrs = {
            str(sheet.cell(attribute_row, col).value).strip(): col
            for col in range(1, sheet.max_column + 1)
            if nonempty(sheet.cell(attribute_row, col).value)
        }
        labels = {
            str(sheet.cell(label_row, col).value).strip(): col
            for col in range(1, sheet.max_column + 1)
            if nonempty(sheet.cell(label_row, col).value)
        }
        if "::record_action" in attrs or any("item_sku" in key for key in attrs):
            candidates.append(TemplateInfo(workbook, sheet, label_row, attribute_row, data_row, attrs, labels))
    if not candidates:
        close_workbook(workbook)
        raise CatalogError("No Amazon template sheet with readable settings and attributes was found")
    return candidates[0]


FIELD_CANDIDATES = {
    "sku": ["item_sku"],
    "product_type": ["product_type"],
    "action": ["::record_action"],
    "parentage": ["parentage_level#1.value", "parentage_level.value"],
    "parent_sku": ["child_parent_sku_relationship#1.parent_sku"],
    "variation_theme": ["variation_theme#1.name"],
    "title": ["item_name#1.value"],
    "external_product_id": ["externally_assigned_product_identifier#1.value"],
    "external_product_id_type": ["externally_assigned_product_identifier#1.type"],
    "price": ["purchasable_offer#1.our_price#1.schedule#1.value_with_tax"],
    "quantity": ["fulfillment_availability#1.quantity"],
    "fulfillment_channel": ["fulfillment_availability#1.fulfillment_channel_code"],
    "lead_time": ["fulfillment_availability#1.lead_time_to_ship_max_days"],
    "condition": ["condition_type#1.value"],
}


def resolve_column(info: TemplateInfo, field: str, *, required: bool = True) -> int | None:
    if field in info.attributes:
        return info.attributes[field]
    candidates = FIELD_CANDIDATES.get(field, [field])
    for candidate in candidates:
        if candidate in info.attributes:
            return info.attributes[candidate]
    for candidate in candidates:
        matches = [col for attr, col in info.attributes.items() if attr.startswith(candidate + "#")]
        if len(matches) == 1:
            return matches[0]
    if required:
        raise CatalogError(f"Template is missing required attribute column: {field}")
    return None


def set_value(info: TemplateInfo, row: int, field: str, value: Any, *, required: bool = True) -> None:
    col = resolve_column(info, field, required=required)
    if col is not None:
        info.sheet.cell(row, col).value = value


def clear_data_rows(info: TemplateInfo) -> None:
    for row in range(info.data_row, info.sheet.max_row + 1):
        for col in range(1, info.sheet.max_column + 1):
            info.sheet.cell(row, col).value = None


def apply_fields(info: TemplateInfo, row: int, fields: dict[str, Any]) -> None:
    for field, value in fields.items():
        set_value(info, row, field, value)


def parent_row(manifest: dict[str, Any]) -> dict[str, Any]:
    family = manifest["family"]
    parent = family["parent"]
    row = dict(parent.get("fields") or {})
    row.update(
        {
            "sku": parent["sku"],
            "product_type": parent["product_type"],
            "action": FULL_UPDATE,
            "parentage": "Parent",
            "parent_sku": None,
            "variation_theme": family["variation_theme"],
        }
    )
    if parent.get("title"):
        row["title"] = parent["title"]
    return row


def child_row(manifest: dict[str, Any], child: dict[str, Any]) -> dict[str, Any]:
    family = manifest["family"]
    parent = family["parent"]
    is_new = child["status"] == "new"
    row = dict(child.get("fields") or {}) if is_new else {}
    row.update(
        {
            "sku": child["sku"],
            "product_type": child.get("product_type") or parent["product_type"],
            "action": FULL_UPDATE if is_new else PARTIAL_UPDATE,
            "parentage": "Child",
            "parent_sku": parent["sku"],
            "variation_theme": family["variation_theme"],
            family["variation_attribute"]: child["variation_value"],
        }
    )
    if is_new:
        if child.get("title"):
            row["title"] = child["title"]
        if child.get("gtin_exempt"):
            row["external_product_id"] = None
            row["external_product_id_type"] = "GTIN Exempt"
        elif child.get("product_id"):
            row["external_product_id"] = child["product_id"]
            row["external_product_id_type"] = child.get("product_id_type")
        for key in ("price", "quantity", "fulfillment_channel", "lead_time", "condition"):
            if key in child:
                row[key] = child[key]
    return row


def delete_row(sku: str, product_type: str) -> dict[str, Any]:
    return {"sku": sku, "product_type": product_type, "action": DELETE}


def build_specs(manifest: dict[str, Any]) -> list[tuple[str, list[dict[str, Any]]]]:
    operation = manifest["operation"]
    family = manifest["family"]
    parent = family["parent"]
    children = family.get("children") or []
    apply_rows: list[dict[str, Any]] = []

    if operation == "create_parent":
        apply_rows = [parent_row(manifest)] + [child_row(manifest, child) for child in children]
        return [("01-create-family", apply_rows)]
    if operation in {"add_existing_child", "create_gtin_exempt_child"}:
        apply_rows = [child_row(manifest, child) for child in children]
        return [("01-add-children", apply_rows)]
    if operation == "delete_parent":
        return [("01-delete-parent", [delete_row(parent["sku"], parent["product_type"])])]

    old_parent = family.get("old_parent_sku") or parent["sku"]
    delete_rows = [delete_row(old_parent, parent["product_type"])]
    if operation == "detach_children":
        removed = set(map(str, family.get("removed_child_skus") or []))
        children = [child for child in children if str(child["sku"]) not in removed]
    apply_rows = [child_row(manifest, child) for child in children]
    if operation != "reparent_children" or parent.get("status") != "existing":
        apply_rows.insert(0, parent_row(manifest))
    return [("01-delete-old-parent", delete_rows), ("02-rebuild-family", apply_rows)]


def output_dir(manifest: dict[str, Any]) -> Path:
    raw = manifest.get("output_dir")
    if raw:
        path = Path(raw).expanduser()
        if not path.is_absolute():
            path = REPO / path
        return path.resolve()
    stamp = manifest.get("date") or date.today().isoformat()
    operation = slugify(manifest["operation"])
    return REPO / "output" / slugify(manifest["client"]) / "catalog" / f"{stamp}-{operation}"


def hash_member(path: Path, member: str) -> str | None:
    try:
        with zipfile.ZipFile(path) as archive:
            if member not in archive.namelist():
                return None
            return hashlib.sha256(archive.read(member)).hexdigest()
    except zipfile.BadZipFile:
        return None


def write_workbook(template: Path, output: Path, rows: list[dict[str, Any]]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    info = load_template(output)
    clear_data_rows(info)
    for offset, values in enumerate(rows):
        row_number = info.data_row + offset
        for field, value in values.items():
            set_value(info, row_number, field, value)
    info.workbook.save(output)
    close_workbook(info.workbook)


def markdown_table(headers: list[str], rows: Iterable[Iterable[Any]]) -> str:
    def clean(value: Any) -> str:
        return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")

    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    lines.extend("| " + " | ".join(clean(v) for v in row) + " |" for row in rows)
    return "\n".join(lines)


def write_pack_docs(manifest: dict[str, Any], pack: Path, specs: list[tuple[str, list[dict[str, Any]]]]) -> None:
    source = manifest["source"]
    source_text = f"""# Catalog Change Source Record

- Client: {manifest['client']}
- Seller account: {manifest['seller_account']}
- Marketplace: {manifest['marketplace']}
- Operation: {manifest['operation']}
- Category Listings Report: `{source['category_listings_report']}`
- Blank template: `{source['blank_template']}`
- Manifest: `{manifest['_manifest_path']}`
- Rule: the Category Listings Report is evidence only and must never be uploaded.
"""
    (pack / "01-source-record.md").write_text(source_text, encoding="utf-8")

    manifest_rows = []
    order = 0
    for file_name, rows in specs:
        order += 1
        for row in rows:
            submitted = ", ".join(sorted(row))
            preserved = "All fields not listed" if row.get("action") == PARTIAL_UPDATE else "Not applicable to full update"
            manifest_rows.append((order, file_name + ".xlsm", row.get("sku"), row.get("action"), submitted, preserved))
    manifest_text = f"""# Catalog Change Manifest

## Approval Summary

- Client: {manifest['client']}
- Marketplace: {manifest['marketplace']}
- Operation: {manifest['operation']}
- Variation theme: {manifest.get('family', {}).get('variation_theme', 'Not applicable')}
- Final submission: manual team action after this manifest and validation report are reviewed.

## Upload Order

{markdown_table(['Order', 'File', 'SKU', 'Action', 'Submitted fields', 'Preserved fields'], manifest_rows)}

## Expected Result

Amazon processes the files in the listed order. For multi-file changes, do not submit the next file until the previous processing summary succeeds and the old parentage is visibly detached.

## Review Checklist

- Only the intended SKUs appear in each file.
- Parent and child product types match.
- Every child has one unique customer-facing variation value.
- Existing children use minimal partial updates.
- New parents and new children use full updates.
- No parent row contains price, quantity, condition, or a child variation value.
- The Category Listings Report is not used as an upload file.
"""
    (pack / "02-change-manifest.md").write_text(manifest_text, encoding="utf-8")


def rows_from_workbook(path: Path) -> tuple[TemplateInfo, list[tuple[int, dict[str, Any]]]]:
    info = load_template(path)
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number in range(info.data_row, info.sheet.max_row + 1):
        values = {
            attr: info.sheet.cell(row_number, col).value
            for attr, col in info.attributes.items()
            if nonempty(info.sheet.cell(row_number, col).value)
        }
        if values:
            rows.append((row_number, values))
    return info, rows


def canonical_value(info: TemplateInfo, values: dict[str, Any], field: str) -> Any:
    col = resolve_column(info, field, required=False)
    if col is None:
        return None
    attr = next((key for key, value in info.attributes.items() if value == col), None)
    return values.get(attr) if attr else None


def expected_for_file(manifest: dict[str, Any], file_path: Path) -> tuple[str, list[dict[str, Any]]]:
    specs = build_specs(manifest)
    if len(specs) == 1:
        return specs[0]
    stem = file_path.stem.lower()
    for name, rows in specs:
        if name.lower() in stem:
            return name, rows
    raise CatalogError(
        "This operation has multiple upload files. The filename must contain one of: "
        + ", ".join(name for name, _ in specs)
    )


def validate_workbook(manifest: dict[str, Any], path: Path) -> list[str]:
    if not path.exists():
        raise CatalogError(f"Workbook does not exist: {path}")
    _, expected_rows = expected_for_file(manifest, path)
    info, actual = rows_from_workbook(path)
    issues: list[str] = []
    notes: list[str] = []
    expected_by_sku = {str(row["sku"]): row for row in expected_rows}
    actual_by_sku: dict[str, tuple[int, dict[str, Any]]] = {}
    for row_number, values in actual:
        sku = canonical_value(info, values, "sku")
        if not sku:
            issues.append(f"row {row_number}: non-empty row has no SKU")
            continue
        sku = str(sku)
        if sku in actual_by_sku:
            issues.append(f"SKU {sku} appears more than once")
        actual_by_sku[sku] = (row_number, values)
    expected_skus = set(expected_by_sku)
    actual_skus = set(actual_by_sku)
    if missing := sorted(expected_skus - actual_skus):
        issues.append(f"missing expected SKUs: {', '.join(missing)}")
    if unexpected := sorted(actual_skus - expected_skus):
        issues.append(f"unexpected SKUs present: {', '.join(unexpected)}")

    variation_values: list[str] = []
    themes: set[str] = set()
    variation_attr = manifest.get("family", {}).get("variation_attribute")
    variation_col = resolve_column(info, variation_attr, required=False) if variation_attr else None
    variation_key = next((key for key, col in info.attributes.items() if col == variation_col), None)
    offer_fields = {key for key in ("price", "quantity", "condition", "fulfillment_channel", "lead_time")}

    for sku in sorted(expected_skus & actual_skus):
        row_number, values = actual_by_sku[sku]
        expected = expected_by_sku[sku]
        action = canonical_value(info, values, "action")
        if action not in VALID_ACTIONS:
            issues.append(f"SKU {sku}: invalid listing action {action!r}")
        if action != expected["action"]:
            issues.append(f"SKU {sku}: action {action!r} should be {expected['action']!r}")
        for field, expected_value in expected.items():
            col = resolve_column(info, field, required=False)
            if col is None:
                issues.append(f"SKU {sku}: expected template field is missing: {field}")
                continue
            attr = next(k for k, v in info.attributes.items() if v == col)
            actual_value = values.get(attr)
            if expected_value is None:
                if nonempty(actual_value):
                    issues.append(f"SKU {sku}: {field} must be blank")
            elif str(actual_value).strip() != str(expected_value).strip():
                issues.append(f"SKU {sku}: {field} is {actual_value!r}; expected {expected_value!r}")

        allowed_columns = set()
        for field in expected:
            col = resolve_column(info, field, required=False)
            if col is not None:
                allowed_columns.add(next(k for k, v in info.attributes.items() if v == col))
        extras = sorted(set(values) - allowed_columns)
        if extras:
            issues.append(f"SKU {sku}: unrelated populated fields: {', '.join(extras)}")

        parentage = str(canonical_value(info, values, "parentage") or "").casefold()
        if parentage == "parent":
            if variation_key and nonempty(values.get(variation_key)):
                issues.append(f"SKU {sku}: parent row contains a child variation value")
            for field in offer_fields:
                if nonempty(canonical_value(info, values, field)):
                    issues.append(f"SKU {sku}: parent row contains prohibited {field}")
        elif parentage == "child":
            theme = canonical_value(info, values, "variation_theme")
            if theme:
                themes.add(str(theme))
            if variation_key and nonempty(values.get(variation_key)):
                variation_values.append(str(values[variation_key]).strip().casefold())
            if action == PARTIAL_UPDATE and extras:
                issues.append(f"SKU {sku}: existing child is not a minimal partial update")
            id_type = canonical_value(info, values, "external_product_id_type")
            product_id = canonical_value(info, values, "external_product_id")
            if str(id_type or "").casefold() == "gtin exempt" and nonempty(product_id):
                issues.append(f"SKU {sku}: GTIN-exempt child must have a blank product ID")

    if len(themes) > 1:
        issues.append(f"mixed variation themes: {', '.join(sorted(themes))}")
    duplicates = sorted({v for v in variation_values if variation_values.count(v) > 1})
    if duplicates:
        issues.append(f"duplicate child variation values: {', '.join(duplicates)}")

    template = resolve_source(manifest, "blank_template")
    template_info = load_template(template)
    for row in range(1, min(info.data_row, template_info.data_row)):
        for col in range(1, min(info.sheet.max_column, template_info.sheet.max_column) + 1):
            if info.sheet.cell(row, col).value != template_info.sheet.cell(row, col).value:
                issues.append(f"template metadata changed at {info.sheet.cell(row, col).coordinate}")
                break
        if issues and issues[-1].startswith("template metadata changed"):
            break
    close_workbook(template_info.workbook)
    close_workbook(info.workbook)
    if hash_member(template, "xl/vbaProject.bin") != hash_member(path, "xl/vbaProject.bin"):
        issues.append("VBA project was not preserved from the blank template")

    if issues:
        raise CatalogError("Workbook validation failed:\n- " + "\n- ".join(issues))
    notes.append(f"PASS: {path.name} contains exactly {len(expected_skus)} intended SKU row(s)")
    return notes


def build(manifest: dict[str, Any]) -> Path:
    notes = validate_manifest(manifest)
    template = resolve_source(manifest, "blank_template")
    pack = output_dir(manifest)
    upload_dir = pack / "03-upload-ready"
    (pack / "04-processing-summary").mkdir(parents=True, exist_ok=True)
    (pack / "05-evidence").mkdir(parents=True, exist_ok=True)
    specs = build_specs(manifest)
    write_pack_docs(manifest, pack, specs)
    for name, rows in specs:
        output = upload_dir / f"{name}.xlsm"
        write_workbook(template, output, rows)
        for message in validate_workbook(manifest, output):
            print(f"  {message}")
    print(f"READY: {pack}")
    for note in notes:
        print(f"NOTE: {note}")
    if len(specs) > 1:
        print("GATE: process and verify file 01 before submitting file 02")
    return pack


def find_summary_records(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    records: list[dict[str, str]] = []
    code_re = re.compile(r"(?<!\d)(8032|8066|8801|90041|90057|100521)(?!\d)")
    any_code_re = re.compile(r"(?<!\d)(\d{4,6})(?!\d)")
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        for row_number, row in enumerate(rows, 1):
            text = " | ".join(str(value) for value in row if nonempty(value))
            if not text:
                continue
            codes = code_re.findall(text)
            if not codes and any(word in text.casefold() for word in ("error", "warning")):
                codes = any_code_re.findall(text)
            for code in dict.fromkeys(codes):
                sku = next(
                    (str(value) for value in row if nonempty(value) and re.match(r"^[A-Za-z0-9][A-Za-z0-9._-]{2,39}$", str(value)) and not str(value).isdigit()),
                    "Unknown",
                )
                records.append({"sheet": sheet.title, "row": str(row_number), "sku": sku, "code": code, "message": text})
    close_workbook(workbook)
    return records


def review_summary(manifest: dict[str, Any], summary: Path, output: Path | None = None) -> Path:
    if not summary.exists():
        raise CatalogError(f"Processing summary does not exist: {summary}")
    pack = output_dir(manifest)
    target = output or pack / "04-processing-summary" / f"{summary.stem}-review.md"
    target.parent.mkdir(parents=True, exist_ok=True)
    records = find_summary_records(summary)
    rows = []
    for record in records:
        title, guidance = ERROR_GUIDANCE.get(
            record["code"],
            ("Unmapped processing code", "Read the complete per-SKU message before preparing another upload."),
        )
        rows.append((record["sku"], record["code"], title, guidance, f"{record['sheet']} row {record['row']}"))
    if rows:
        verdict = "Review required before another upload."
        table = markdown_table(["SKU", "Code", "Meaning", "Next action", "Source"], rows)
    else:
        verdict = "No recognized blocking code was found. Confirm the per-SKU status and verify backend and frontend state."
        table = "No recognized processing codes found."
    content = f"""# Feed Processing Summary Review

- File: `{summary}`
- Client: {manifest['client']}
- Marketplace: {manifest['marketplace']}
- Verdict: {verdict}

## Findings

{table}

## Verification

Check the Feed Processing Summary per SKU, then verify Manage All Inventory, Variation Wizard, a fresh Category Listings Report, and the customer-facing child detail page. Backend and frontend state may update at different times.
"""
    target.write_text(content, encoding="utf-8")
    print(f"REVIEW: {target}")
    return target


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    for command in ("build", "validate", "review-summary"):
        sub = subparsers.add_parser(command)
        sub.add_argument("--manifest", required=True, type=Path)
        if command == "validate":
            sub.add_argument("--file", required=True, type=Path)
        if command == "review-summary":
            sub.add_argument("--summary", required=True, type=Path)
            sub.add_argument("--output", type=Path)
    args = parser.parse_args(argv)
    try:
        manifest = load_manifest(args.manifest.resolve())
        if args.command == "build":
            build(manifest)
        elif args.command == "validate":
            validate_manifest(manifest)
            for note in validate_workbook(manifest, args.file.resolve()):
                print(note)
        else:
            validate_manifest(manifest, require_files=False)
            review_summary(manifest, args.summary.resolve(), args.output.resolve() if args.output else None)
    except CatalogError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
