#!/usr/bin/env python3
"""Synthetic, client-neutral tests for the catalog change-pack utility."""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

import catalog_change_pack as ccp  # noqa: E402


TMP = Path(tempfile.mkdtemp(prefix="catalog-change-pack-selftest-"))
FAILURES: list[str] = []


ATTRIBUTES = [
    "item_sku",
    "product_type",
    "::record_action",
    "parentage_level#1.value",
    "child_parent_sku_relationship#1.parent_sku",
    "variation_theme#1.name",
    "package_size_name#1.value",
    "item_name#1.value",
    "brand#1.value",
    "product_description#1.value",
    "externally_assigned_product_identifier#1.value",
    "externally_assigned_product_identifier#1.type",
    "purchasable_offer#1.our_price#1.schedule#1.value_with_tax",
    "fulfillment_availability#1.quantity",
    "fulfillment_availability#1.fulfillment_channel_code",
    "fulfillment_availability#1.lead_time_to_ship_max_days",
    "condition_type#1.value",
    "bullet_point#1.value",
]


def check(label: str, condition: bool, detail: str = "") -> None:
    status = "ok" if condition else "FAIL"
    print(f"  [{status}] {label}" + (f": {detail}" if detail and not condition else ""))
    if not condition:
        FAILURES.append(f"{label}: {detail}")


def expect_error(label: str, fn, contains: str) -> None:
    try:
        fn()
    except ccp.CatalogError as exc:
        check(label, contains in str(exc), str(exc))
    else:
        check(label, False, "expected CatalogError")


def make_template(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Template"
    sheet["A1"] = "settings=labelRow=4;attributeRow=5;dataRow=8;flavor=full-seller"
    sheet["A2"] = "SYNTHETIC TEMPLATE - DO NOT COMMIT CLIENT DATA"
    for col, attribute in enumerate(ATTRIBUTES, 1):
        sheet.cell(4, col).value = attribute.replace("#1", "").replace("_", " ").title()
        sheet.cell(5, col).value = attribute
    sheet.cell(6, 1).value = "Amazon example row"
    sheet.cell(7, 1).value = "Prefilled attributes notice - do not delete this row"
    workbook.save(path)
    workbook.close()


def add_fake_vba(path: Path) -> None:
    scratch = path.with_suffix(".rewrite")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(scratch, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                text = text.replace(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
                text = text.replace(
                    "</Types>",
                    '<Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
                )
                data = text.encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8")
                text = text.replace(
                    "</Relationships>",
                    '<Relationship Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                    'Target="vbaProject.bin" Id="rId99"/></Relationships>',
                )
                data = text.encode("utf-8")
            target.writestr(item, data)
        target.writestr("xl/vbaProject.bin", b"synthetic-vba-project")
    scratch.replace(path)


def make_clr(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Category Listings Report"
    sheet.append(["SKU", "ASIN", "Parent SKU", "Variation Theme"])
    sheet.append(["CHILD-1", "B000000001", "", ""])
    sheet.append(["CHILD-2", "B000000002", "", ""])
    sheet.append(["CHILD-3", "B000000003", "", ""])
    sheet.append(["OLD-PARENT", "B000000004", "", ""])
    sheet.append(["TARGET-PARENT", "B000000005", "", ""])
    workbook.save(path)
    workbook.close()


def base_manifest(template: Path, clr: Path, output: Path) -> dict:
    return {
        "schema_version": 1,
        "client": "Synthetic Client",
        "seller_account": "Synthetic Seller",
        "marketplace": "US",
        "date": "2026-07-16",
        "operation": "create_parent",
        "output_dir": str(output),
        "source": {"category_listings_report": str(clr), "blank_template": str(template)},
        "family": {
            "variation_theme": "PACKAGE_SIZE_NAME",
            "variation_attribute": "package_size_name#1.value",
            "parent": {
                "sku": "SYNTH-PARENT",
                "product_type": "SKIN_SERUM",
                "status": "new",
                "title": "Synthetic Skin Serum",
                "fields": {
                    "brand#1.value": "Synthetic Brand",
                    "product_description#1.value": "Synthetic parent description",
                },
            },
            "children": [
                {"sku": "CHILD-1", "asin": "B000000001", "status": "existing", "variation_value": "1-Month Supply"},
                {"sku": "CHILD-2", "asin": "B000000002", "status": "existing", "variation_value": "2-Month Supply"},
            ],
        },
    }


def save_manifest(manifest: dict, name: str) -> Path:
    path = TMP / name
    path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return path


def mutate(path: Path, callback) -> None:
    info = ccp.load_template(path)
    callback(info)
    info.workbook.save(path)
    ccp.close_workbook(info.workbook)


def main() -> int:
    print(f"Scratch: {TMP}")
    template = TMP / "blank-template.xlsm"
    clr = TMP / "category-report.xlsx"
    make_template(template)
    add_fake_vba(template)
    make_clr(clr)

    print("\n=== Create parent with two existing children ===")
    manifest = base_manifest(template, clr, TMP / "pack-create")
    manifest_path = save_manifest(manifest, "create.json")
    loaded = ccp.load_manifest(manifest_path)
    pack = ccp.build(loaded)
    upload = pack / "03-upload-ready" / "01-create-family.xlsm"
    check("upload exists", upload.exists())
    check("source record exists", (pack / "01-source-record.md").exists())
    check("change manifest exists", (pack / "02-change-manifest.md").exists())
    check("VBA project preserved", ccp.hash_member(template, "xl/vbaProject.bin") == ccp.hash_member(upload, "xl/vbaProject.bin"))
    check("valid workbook passes", bool(ccp.validate_workbook(loaded, upload)))

    info, rows = ccp.rows_from_workbook(upload)
    check("exactly three rows", len(rows) == 3, str(len(rows)))
    parent_values = rows[0][1]
    child_values = rows[1][1]
    check("parent is full update", ccp.canonical_value(info, parent_values, "action") == ccp.FULL_UPDATE)
    check("existing child is partial", ccp.canonical_value(info, child_values, "action") == ccp.PARTIAL_UPDATE)
    check("existing child has minimal six relationship fields plus variation", len(child_values) == 7, str(child_values))
    ccp.close_workbook(info.workbook)

    print("\n=== Add an existing third child ===")
    add = base_manifest(template, clr, TMP / "pack-add")
    add["operation"] = "add_existing_child"
    add["family"]["children"] = [
        {"sku": "CHILD-3", "asin": "B000000003", "status": "existing", "variation_value": "3-Month Supply"}
    ]
    add_path = save_manifest(add, "add.json")
    add_loaded = ccp.load_manifest(add_path)
    add_pack = ccp.build(add_loaded)
    add_file = add_pack / "03-upload-ready" / "01-add-children.xlsm"
    check("third child file validates", bool(ccp.validate_workbook(add_loaded, add_file)))

    print("\n=== Create a new GTIN-exempt child ===")
    new = base_manifest(template, clr, TMP / "pack-new")
    new["operation"] = "create_gtin_exempt_child"
    new["family"]["children"] = [
        {
            "sku": "CHILD-NEW",
            "status": "new",
            "variation_value": "3-Month Supply",
            "gtin_exempt": True,
            "title": "Synthetic Serum - 3-Month Supply",
            "price": 159,
            "quantity": 100,
            "fulfillment_channel": "DEFAULT",
            "lead_time": 2,
            "condition": "new_new",
            "fields": {
                "brand#1.value": "Synthetic Brand",
                "product_description#1.value": "Synthetic child description",
            },
        }
    ]
    new_path = save_manifest(new, "new.json")
    new_loaded = ccp.load_manifest(new_path)
    new_pack = ccp.build(new_loaded)
    new_file = new_pack / "03-upload-ready" / "01-add-children.xlsm"
    check("new child file validates", bool(ccp.validate_workbook(new_loaded, new_file)))
    info, rows = ccp.rows_from_workbook(new_file)
    values = rows[0][1]
    check("new child is full update", ccp.canonical_value(info, values, "action") == ccp.FULL_UPDATE)
    check("GTIN type is exemption", ccp.canonical_value(info, values, "external_product_id_type") == "GTIN Exempt")
    check("GTIN ID is blank", not ccp.nonempty(ccp.canonical_value(info, values, "external_product_id")))
    ccp.close_workbook(info.workbook)

    print("\n=== Manifest and workbook rejection gates ===")
    duplicate = base_manifest(template, clr, TMP / "duplicate")
    duplicate["family"]["children"][1]["variation_value"] = "1-month supply"
    duplicate_path = save_manifest(duplicate, "duplicate.json")
    duplicate_loaded = ccp.load_manifest(duplicate_path)
    expect_error("duplicate variation values rejected", lambda: ccp.validate_manifest(duplicate_loaded), "duplicate variation values")

    contaminated = TMP / "contaminated.xlsm"
    shutil.copy2(upload, contaminated)
    mutate(contaminated, lambda info: setattr(info.sheet.cell(info.data_row + 1, ccp.resolve_column(info, "bullet_point#1.value")), "value", "Unrelated bullet"))
    expect_error("unrelated child field rejected", lambda: ccp.validate_workbook(loaded, contaminated), "unrelated populated fields")

    unexpected = TMP / "unexpected.xlsm"
    shutil.copy2(upload, unexpected)
    def add_unexpected(info):
        row = info.data_row + 3
        ccp.set_value(info, row, "sku", "OTHER-SKU")
        ccp.set_value(info, row, "product_type", "SKIN_SERUM")
        ccp.set_value(info, row, "action", ccp.PARTIAL_UPDATE)
    mutate(unexpected, add_unexpected)
    expect_error("unexpected SKU rejected", lambda: ccp.validate_workbook(loaded, unexpected), "unexpected SKUs present")

    mixed = TMP / "mixed-theme.xlsm"
    shutil.copy2(upload, mixed)
    mutate(mixed, lambda info: ccp.set_value(info, info.data_row + 2, "variation_theme", "SIZE"))
    expect_error("mixed themes rejected", lambda: ccp.validate_workbook(loaded, mixed), "mixed variation themes")

    parent_offer = TMP / "parent-offer.xlsm"
    shutil.copy2(upload, parent_offer)
    mutate(parent_offer, lambda info: ccp.set_value(info, info.data_row, "price", 99))
    expect_error("parent offer field rejected", lambda: ccp.validate_workbook(loaded, parent_offer), "parent row contains prohibited price")

    invalid_action = TMP / "invalid-action.xlsm"
    shutil.copy2(upload, invalid_action)
    mutate(invalid_action, lambda info: ccp.set_value(info, info.data_row + 1, "action", "Edit (Full Update)"))
    expect_error("invalid action rejected", lambda: ccp.validate_workbook(loaded, invalid_action), "invalid listing action")

    print("\n=== Ordered delete and rebuild ===")
    rebuild = base_manifest(template, clr, TMP / "pack-rebuild")
    rebuild["operation"] = "rebuild_family"
    rebuild["source"]["allow_missing_skus"] = ["SYNTH-PARENT"]
    rebuild_path = save_manifest(rebuild, "rebuild.json")
    rebuild_loaded = ccp.load_manifest(rebuild_path)
    rebuild_pack = ccp.build(rebuild_loaded)
    delete_file = rebuild_pack / "03-upload-ready" / "01-delete-old-parent.xlsm"
    apply_file = rebuild_pack / "03-upload-ready" / "02-rebuild-family.xlsm"
    check("delete file exists", delete_file.exists())
    check("rebuild file exists", apply_file.exists())
    check("delete file validates", bool(ccp.validate_workbook(rebuild_loaded, delete_file)))
    check("rebuild file validates", bool(ccp.validate_workbook(rebuild_loaded, apply_file)))

    print("\n=== Detach and reparent sequencing ===")
    detach = base_manifest(template, clr, TMP / "pack-detach")
    detach["operation"] = "detach_children"
    detach["family"]["parent"]["sku"] = "OLD-PARENT"
    detach["family"]["parent"]["status"] = "existing"
    detach["family"]["removed_child_skus"] = ["CHILD-2"]
    detach["family"]["children"] = [detach["family"]["children"][0]]
    detach_path = save_manifest(detach, "detach.json")
    detach_loaded = ccp.load_manifest(detach_path)
    detach_pack = ccp.build(detach_loaded)
    check("detach creates delete file", (detach_pack / "03-upload-ready" / "01-delete-old-parent.xlsm").exists())
    check("detach creates rebuild file", (detach_pack / "03-upload-ready" / "02-rebuild-family.xlsm").exists())

    reparent = base_manifest(template, clr, TMP / "pack-reparent")
    reparent["operation"] = "reparent_children"
    reparent["family"]["old_parent_sku"] = "OLD-PARENT"
    reparent["family"]["parent"]["sku"] = "TARGET-PARENT"
    reparent["family"]["parent"]["status"] = "existing"
    reparent_path = save_manifest(reparent, "reparent.json")
    reparent_loaded = ccp.load_manifest(reparent_path)
    reparent_pack = ccp.build(reparent_loaded)
    target_file = reparent_pack / "03-upload-ready" / "02-rebuild-family.xlsm"
    info, reparent_rows = ccp.rows_from_workbook(target_file)
    check("existing target parent is not resubmitted", len(reparent_rows) == 2, str(len(reparent_rows)))
    check(
        "reparent file contains children only",
        all(str(ccp.canonical_value(info, values, "parentage")).casefold() == "child" for _, values in reparent_rows),
    )
    ccp.close_workbook(info.workbook)

    print("\n=== Processing-summary review ===")
    summary = TMP / "processing-summary.xlsm"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feed Processing Summary"
    ws.append(["SKU", "Error Code", "Message"])
    ws.append(["CHILD-1", 8032, "Child is already assigned to another parent"])
    ws.append(["CHILD-2", 90057, "Invalid listing action"])
    ws.append(["CHILD-3", 8801, "Duplicate variation attributes"])
    ws.append(["CHILD-4", 8066, "Invalid parentage setup"])
    ws.append(["CHILD-5", 90041, "Missing required fields"])
    wb.save(summary)
    wb.close()
    review = ccp.review_summary(loaded, summary)
    review_text = review.read_text(encoding="utf-8")
    for code in ("8032", "8066", "8801", "90041", "90057"):
        check(f"summary recognizes {code}", code in review_text)

    print("\n=== Result ===")
    if FAILURES:
        print("FAILED")
        for failure in FAILURES:
            print(f"- {failure}")
        return 1
    print("PASS")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    finally:
        if not FAILURES:
            shutil.rmtree(TMP, ignore_errors=True)
