#!/usr/bin/env python3
"""
Amazon Ad/Sales Audit — orchestrator.

  # Preflight: check the config's input contract, emit a Codex download task or READY
  python3 tools/amazon-ad-audit/build_audit.py --config <cfg> --preflight

  # Build: analyze -> audit workbook -> SQP workbook -> master -> narrative scaffold (+docx)
  python3 tools/amazon-ad-audit/build_audit.py --config <cfg>

  # QA gates: spend reconciliation / no >100% ACOS colored green / master tab count
  python3 tools/amazon-ad-audit/build_audit.py --config <cfg> --validate

Everything client-specific lives in the config (see config.TEMPLATE.json / NEW-CLIENT.md).
"""
from __future__ import annotations
import argparse
import csv
import json
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from analyze_audit import REPO, load_config, rp, _slug, as_market_map  # noqa: E402

GREEN_FILLS = {"C6EFCE", "E2EFDA"}  # traffic-light good/ok — never legal on ACOS > 100%
SPEND_TOL = 0.01  # currency units


def outdir_for(cfg) -> Path:
    return REPO / "output" / _slug(cfg["client"]) / "reporting"


# ----------------------------------------------------------------- preflight
def collect_inputs(cfg):
    """Yield (label, path, gatherer) for every contract input."""
    inp = cfg.get("inputs", {})
    default_market = cfg["marketplaces"][0] if cfg.get("marketplaces") else "US"
    for mkt, p in as_market_map(inp.get("ads_bulk_xlsx"), default_market).items():
        yield (f"ads bulk .xlsx [{mkt}]", p, "CODEX")
    for mkt, p in as_market_map(inp.get("business_report_csv"), default_market).items():
        yield (f"Business Report .csv [{mkt}]", p, "CODEX")
    for grp, p in (inp.get("sqp_csvs") or {}).items():
        yield (f"SQP .csv [{grp}]", p, "CODEX")
    if cfg.get("datadive_niche"):
        yield ("DataDive niche keywords JSON", inp.get("datadive_niche_json"), "CLAUDE/MCP")
        if inp.get("datadive_competitors_json"):
            yield ("DataDive competitors JSON", inp.get("datadive_competitors_json"), "CLAUDE/MCP")


def preflight(cfg, cfg_path) -> int:
    missing = []
    print(f"Preflight — {cfg.get('client')} ({', '.join(cfg.get('marketplaces', []))})")
    for label, p, who in collect_inputs(cfg):
        q = rp(p)
        ok = bool(q and q.exists() and q.stat().st_size > 0)
        print(f"  [{'OK' if ok else 'MISSING'}] {label}  ->  {p}")
        if not ok:
            missing.append((label, p, who))
    codex = [m for m in missing if m[2] == "CODEX"]
    claude = [m for m in missing if m[2] == "CLAUDE/MCP"]
    if not missing:
        print("\nREADY — all inputs present. Run the build:")
        print(f"  python3 tools/amazon-ad-audit/build_audit.py --config {cfg_path}")
        return 0
    if codex:
        w = cfg.get("windows", {})
        print("\nWAITING ON CODEX — copy-ready download task:\n")
        print("---8<--- CODEX TASK ---")
        print(f"Objective: download the missing Amazon exports for the {cfg.get('client')} "
              f"{'/'.join(cfg.get('marketplaces', []))} ad/sales audit. Do NOT run the builder, "
              f"write the narrative, or change anything in the ad account.")
        print(f"Account: {cfg.get('amazon_account')} · Window: ads {w.get('ads')} · BR {w.get('business_report')} "
              f"· SQP weeks {', '.join(w.get('sqp_weeks', []))}")
        print("Save each file to the EXACT path below (@Chrome extension for downloads, US VPN):")
        for label, p, _ in codex:
            print(f"  - {label}: {p}")
        print("Sources: ads bulk = Ads console > Bulk Operations (SP required; SB/SD/RAS if running); "
              "Business Report = Seller Central > Business Reports > Detail Page Sales & Traffic by Child ASIN; "
              "SQP = Brand Analytics > Search Query Performance, multi-ASIN weekly export per product group "
              "(caveat: multi-ASIN caps the query grid — SV is a floor).")
        print("Recommended extras (optional — improve depth, do NOT block READY):")
        print("  - SB campaign placement report (Ads console > Reports) — the bulk's SB placement rows are "
              "incomplete; this gives the true Top-of-Search / Rest-of-Search / Product-page split.")
        print("  - SP Search-Term Impression-Share report — Top-of-Search headroom for the placement lever.")
        print("  NOT needed: SB/SD search-term reports — SB is intent-split by TARGET from the bulk itself "
              "(bulk SB search-term coverage is ~half; a dedicated SB ST report adds almost nothing).")
        print("Then: note evidence + caveats and STOP.")
        print("---8<---------------")
    if claude:
        print("\nCLAUDE (MCP) still to pull:")
        for label, p, _ in claude:
            print(f"  - {label}: save get_niche_keywords/get_niche_competitors for "
                  f"niche '{cfg.get('datadive_niche')}' to {p}")
    return 1


# ----------------------------------------------------------------- build
def build(cfg, cfg_path) -> int:
    import analyze_audit
    import build_audit_workbook
    import build_sqp_workbook
    import build_master_workbook
    import narrative_scaffold

    outdir, _ = analyze_audit.run(cfg_path)
    audit_path = build_audit_workbook.build(cfg_path, outdir)
    sqp_path = build_sqp_workbook.build(cfg_path, outdir)
    master_path = build_master_workbook.build(cfg_path, outdir, audit_path, sqp_path)
    scaffold_md = narrative_scaffold.build(cfg_path, outdir)
    docx_path = None
    try:
        import subprocess
        docx_path = Path(str(scaffold_md)).with_suffix(".docx")
        subprocess.run([sys.executable, str(HERE / "md_to_docx.py"), str(scaffold_md), str(docx_path)],
                       check=True)
    except Exception as e:  # docx is a convenience; the .md scaffold is the contract
        print(f"[build] md_to_docx skipped: {e}")
        docx_path = None
    print("\nArtifacts:")
    for p in (audit_path, sqp_path, master_path, scaffold_md, docx_path):
        if p:
            print(f"  {p}")
    _completeness_panel(outdir)
    print(f"\nNext: QA gates ->  python3 tools/amazon-ad-audit/build_audit.py --config {cfg_path} --validate")
    return 0


def _completeness_panel(outdir):
    """Print the data-completeness summary from metrics.json after a build."""
    mp = Path(outdir) / "metrics.json"
    if not mp.exists():
        return
    dc = json.loads(mp.read_text()).get("data_completeness")
    if not dc:
        return
    print("\nDATA COMPLETENESS")
    print(f"  intent-split coverage : {dc['intent_coverage']:.1%}"
          + ("  ⚠ below 90% — the split does not represent all spend" if dc["intent_coverage_low"] else ""))
    if dc["sqp_revenue_gap"] > 0.001:
        print(f"  SQP revenue gap       : {dc['sqp_revenue_gap']:.1%} of sales in groups with NO SQP file"
              + (f"  ⚠  ({', '.join(dc['sqp_uncovered_groups'])})" if dc["sqp_revenue_gap"] > 0.20 else ""))
    else:
        print("  SQP revenue gap       : 0% (every selling group has SQP)")
    print(f"  channels missing      : {', '.join(dc['channels_missing']) or 'none'}")
    print(f"  multi-parent ad groups: {dc['multi_parent_ad_groups']}"
          + ("  ⚠ ad groups advertise several product families" if dc["multi_parent_ad_groups"] else ""))


# ----------------------------------------------------------------- validate
def _gate(name, ok, detail):
    print(f"  [{'PASS' if ok else 'FAIL'}] {name} — {detail}")
    return ok


def validate(cfg, cfg_path) -> int:
    from openpyxl import load_workbook

    outdir = outdir_for(cfg)
    metrics_p = outdir / "metrics.json"
    if not metrics_p.exists():
        print(f"FAIL — no metrics.json in {outdir}; run the build first.")
        return 1
    M = json.loads(metrics_p.read_text())
    ok_all = True
    print(f"Validate — {cfg.get('client')} ({outdir})")

    # (a) Branded+Generic+Competitor spend reconciles to the ST-classified total
    stb_spend = sum(v.get("spend", 0) for v in M.get("searchterm_bucket", {}).values())
    st_csv = outdir / "clean" / "search_terms_classified.csv"
    st_spend = 0.0
    if st_csv.exists():
        with open(st_csv) as f:
            st_spend = sum(float(r["spend"] or 0) for r in csv.DictReader(f))
    ok = abs(stb_spend - st_spend) <= max(SPEND_TOL, 0.0001 * max(stb_spend, st_spend))
    ok_all &= _gate("spend reconciliation",
                    ok, f"buckets {stb_spend:,.2f} vs search-term rows {st_spend:,.2f} "
                        f"(coverage {stb_spend / M['totals']['spend']:.1%} of total SP+ spend)")

    # (b) no ACOS ratio > 1.0 carrying a green fill, in any built workbook
    markets = "-".join(M.get("marketplaces", []) or ["x"])
    books = sorted(outdir.glob("*.xlsx"))
    bad = []
    for wb_path in books:
        wb = load_workbook(wb_path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, (int, float)) or v <= 1.0:
                        continue
                    if "%" not in (cell.number_format or ""):
                        continue
                    fill = cell.fill
                    rgb = getattr(getattr(fill, "fgColor", None), "rgb", None) if fill and fill.patternType else None
                    if rgb and str(rgb)[-6:].upper() in GREEN_FILLS:
                        bad.append(f"{wb_path.name}!{ws.title}!{cell.coordinate}={v:.2f}")
    ok_all &= _gate("no >100% ACOS colored green",
                    not bad, f"scanned {len(books)} workbooks" + (f"; offenders: {bad[:5]}" if bad else ""))

    # (c) master tab count = overview + (audit - sources) + (sqp - sources) + master sources
    master = next((p for p in books if "MASTER" in p.name), None)
    audit = next((p for p in books if "Ad_Audit" in p.name), None)
    sqp = next((p for p in books if "SQP_Intelligence" in p.name), None)
    if master and audit and sqp:
        n_master = len(load_workbook(master, read_only=True).sheetnames)
        n_expect = (len(load_workbook(audit, read_only=True).sheetnames) - 1
                    + len(load_workbook(sqp, read_only=True).sheetnames) - 1 + 2)
        ok_all &= _gate("master tab count", n_master == n_expect, f"{n_master} tabs vs expected {n_expect}")
    else:
        ok_all &= _gate("master tab count", False,
                        f"missing workbook(s): master={bool(master)} audit={bool(audit)} sqp={bool(sqp)}")

    # soft data-completeness warnings (not gate failures — imperfect data ≠ a bug)
    dc = M.get("data_completeness", {})
    warns = []
    if dc.get("intent_coverage_low"):
        warns.append(f"intent-split coverage {dc['intent_coverage']:.1%} < 90% — state it explicitly; the split is not all spend")
    if dc.get("sqp_revenue_gap", 0) > 0.20:
        warns.append(f"SQP revenue gap {dc['sqp_revenue_gap']:.1%} — groups with no SQP: {', '.join(dc.get('sqp_uncovered_groups', []))}. Confirm the data truly doesn't exist, else download it; disclose in Method Notes.")
    if dc.get("channels_missing"):
        warns.append(f"no {', '.join(dc['channels_missing'])} campaigns (no brand-defense / retargeting motion)")
    if dc.get("multi_parent_ad_groups"):
        warns.append(f"{dc['multi_parent_ad_groups']} ad groups advertise several parent families — keyword→product fit uncontrolled")
    if warns:
        print("\nWARNINGS (data completeness — resolve or disclose before delivery):")
        for w in warns:
            print(f"  ⚠ {w}")

    print("\nALL GATES PASS" if ok_all else "\nGATES FAILED")
    return 0 if ok_all else 1


# ----------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Amazon Ad/Sales Audit orchestrator")
    ap.add_argument("--config", required=True)
    ap.add_argument("--preflight", action="store_true")
    ap.add_argument("--validate", action="store_true")
    args = ap.parse_args()
    cfg = load_config(args.config)
    if args.preflight:
        sys.exit(preflight(cfg, args.config))
    if args.validate:
        sys.exit(validate(cfg, args.config))
    sys.exit(build(cfg, args.config))


if __name__ == "__main__":
    main()
