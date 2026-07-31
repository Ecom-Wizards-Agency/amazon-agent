#!/usr/bin/env python3
"""
Amazon Ad/Sales Audit — client-agnostic parser.
Reads a per-client config and produces the intermediate contract the builders consume:
  <outdir>/metrics.json  +  <outdir>/clean/*.csv  +  <outdir>/clean/sqp_summary.json

Generalizes the original per-client analyze script:
 - paths, brand/competitor tokens, break-even, asin_groups, marketplaces come from config
 - Sponsored Products is parsed in full (totals, branded/generic/competitor split from the
   Search Term Report, placement, bid hygiene, structure diagnosis)
 - SB / SB-Multi / SD / RAS are added to channel_totals when their sheets carry data rows
 - single OR multiple marketplaces (inputs.ads_bulk_xlsx / business_report_csv may be a
   string or a {marketplace: path} map); SQP is per product-group.
"""
from __future__ import annotations
import csv, json, re, sys, warnings
from collections import defaultdict
from pathlib import Path
warnings.filterwarnings("ignore")
import openpyxl

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO / "tools"))
import ew_paths  # noqa: E402

def rp(p):
    """Resolve a config path: {drive}/{pcloud}/{vault} token, ~ expansion, then
    repo-relative fallback. Configs should use the token form — an absolute
    Drive path only ever resolves on the machine it was written on."""
    if not p:
        return None
    p = ew_paths.expand_tokens(str(p))
    q = Path(p).expanduser()
    if not q.is_absolute() and not q.exists():
        alt = REPO / str(p)
        if alt.exists():
            return alt
    return q

def nf(x):
    if x is None:
        return 0.0
    s = str(x).replace(",", "").replace("%", "").replace("$", "").replace("€", "").strip()
    if s in ("", "-", "—", "nan", "None"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def acos(s, sa):
    return (s / sa) if sa else None

# ----------------------------------------------------------------- config
def load_config(path):
    cfg = json.loads(Path(path).read_text())
    cfg["_brand_re"] = re.compile("|".join(re.escape(t.lower()) for t in cfg.get("brand_tokens", []) if t), re.I) if cfg.get("brand_tokens") else None
    cfg["_comp_re"] = re.compile("|".join(re.escape(t.lower()) for t in cfg.get("competitor_tokens", []) if t), re.I) if cfg.get("competitor_tokens") else None
    cfg["_own_asins"] = {a.lower() for asins in (cfg.get("asin_groups") or {}).values() for a in asins}
    cfg["_core_re"] = re.compile("|".join(re.escape(t.lower()) for t in cfg.get("core_tokens", []) if t), re.I) if cfg.get("core_tokens") else None
    return cfg

_ASIN_RE = re.compile(r"^b0[a-z0-9]{8}$")
_ASIN_ANY = re.compile(r"b0[a-z0-9]{8}")

def classify(cfg, term):
    t = (term or "").lower().strip()
    # Product-targeting rows surface the targeted ASIN as the "search term":
    # own ASIN = defense (Branded); a foreign detail page = conquesting (Competitor).
    if _ASIN_RE.fullmatch(t):
        return "Branded" if t in cfg.get("_own_asins", set()) else "Competitor"
    if cfg["_brand_re"] and cfg["_brand_re"].search(t):
        return "Branded"
    if cfg["_comp_re"] and cfg["_comp_re"].search(t):
        return "Competitor"
    return "Generic"

DEMAND_SEGMENTS = ["Branded", "Competitor", "Core", "Head"]

def classify_demand(cfg, term):
    """Four-way EXCLUSIVE demand segmentation for the SQP demand charts.

    Branded and Competitor follow classify(). Generic then splits in two, because
    lumping them together is what makes a category look 20x bigger than it is:
      Core = matches cfg['core_tokens'], the category language this product actually
             competes on ("under eye brightener", "eye corrector"). Winnable.
      Head = the undifferentiated remainder ("makeup", "beauty") that no single
             listing wins. Real demand, but not addressable, so it must never sit
             inside the number you measure your category share against.
    With no core_tokens configured every Generic term falls to Core, which keeps the
    old two-way behaviour rather than silently reclassifying an existing client."""
    it = classify(cfg, term)
    if it != "Generic":
        return it
    rx = cfg.get("_core_re")
    if not rx:
        return "Core"
    return "Core" if rx.search((term or "").lower()) else "Head"


def classify_target(cfg, text):
    """Classify an SB keyword or product-target EXPRESSION (not a search term).
    ASIN targets (asin="B0…") → own = Branded defense, foreign = Competitor
    conquesting; else fall back to brand/competitor token match on the text."""
    t = (text or "").lower()
    m = _ASIN_ANY.search(t)
    if m and "asin" in t:
        return "Branded" if m.group(0) in cfg.get("_own_asins", set()) else "Competitor"
    if cfg["_brand_re"] and cfg["_brand_re"].search(t):
        return "Branded"
    if cfg["_comp_re"] and cfg["_comp_re"].search(t):
        return "Competitor"
    return "Generic"

def as_market_map(v, default_market):
    """Accept a string (single market) or a {market: path} dict."""
    if v is None:
        return {}
    if isinstance(v, dict):
        return v
    return {default_market: v}

# ----------------------------------------------------------------- ads bulk (per marketplace)
SP_SHEET = "Sponsored Products Campaigns"
CHANNEL_SHEETS = {
    "SB": "Sponsored Brands Campaigns",
    "SB-Multi": "SB Multi Ad Group Campaigns",
    "SD": "Sponsored Display Campaigns",
    "RAS": "RAS Campaigns",
}

SB_SHEETS = ["SB Multi Ad Group Campaigns", "Sponsored Brands Campaigns"]  # superset first

# Bounds for streaming the ads bulk (see sheet_io): Amazon writes a bogus "A1:A1" dimension
# that openpyxl read_only clips to, so iter_rows is forced past it with explicit maxima.
# Generous headroom: bulk sheets have <~80 columns; big accounts run into 6-figure row counts.
BULK_MAX_COLS = 256
BULK_MAX_ROWS = 5_000_000

def parse_bulk(cfg, market, path, agg):
    # Ads bulk sheets can be huge (a real account hit 288k rows on "Sponsored Products
    # Campaigns"). read_only=False + per-cell .cell() access turned that into a ~40-min,
    # multi-GB parse. Stream with read_only=True + iter_rows instead; reset_dimensions
    # ignores the (often absent/wrong) stored <dimension> so every row is read. Results are
    # memoized because the SB sheets get read up to 3x.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    _sheet_cache = {}

    def sheet_io(sheet):
        if sheet in _sheet_cache:
            return _sheet_cache[sheet]
        ws = wb[sheet]
        ws.reset_dimensions = True
        # Amazon's bulk export writes a bogus "A1:A1" <dimension>, and openpyxl 3.1.5
        # read_only clips iter_rows to it on BOTH axes (verified: 288k real <row> elements,
        # dimension says 1x1) even with reset_dimensions set. Explicit min/max bounds force
        # the full stream: read the header with a generous column bound, trim trailing empty
        # header cells to the true width, then stream the data at exactly that width. The row
        # stream still stops at the real last row (no empty padding).
        hdr = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=BULK_MAX_COLS,
                                values_only=True), ())
        H = list(hdr)
        while H and H[-1] is None:              # drop the None cells beyond the real header
            H.pop()
        ncol = len(H)
        I = {h: i for i, h in enumerate(H)}
        rows = []
        if ncol:
            for r in ws.iter_rows(min_row=2, max_row=BULK_MAX_ROWS, min_col=1, max_col=ncol,
                                  values_only=True):
                rows.append(list(r))
        _sheet_cache[sheet] = (H, I, rows)
        return _sheet_cache[sheet]

    # Amazon lists every Sponsored Brands campaign in BOTH the legacy "Sponsored Brands
    # Campaigns" sheet and the "SB Multi Ad Group Campaigns" sheet — same Campaign IDs,
    # same spend/sales. Summing both double-counts SB. Assign each SB campaign to ONE
    # sheet (superset wins) and only count/parse it there.
    sb_owner = {}  # campaign_id -> sheet that owns it
    for sheet in SB_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        _, I, rows = sheet_io(sheet)
        cid = I.get("Campaign ID")
        if cid is None or "Entity" not in I:
            continue
        for r in rows:
            if r[I["Entity"]] == "Campaign" and r[cid] is not None:
                sb_owner.setdefault(r[cid], sheet)
    agg["_sb_owner"] = sb_owner

    # ---- channel presence + additive totals ----
    # SB (both sheets, deduped by owner) collapse into one "SB" channel.
    for sheet in SB_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        _, I, rows = sheet_io(sheet)
        if not rows or "Entity" not in I:
            continue
        cid = I.get("Campaign ID")
        camp = [r for r in rows if r[I["Entity"]] == "Campaign"
                and (cid is None or sb_owner.get(r[cid]) == sheet)]
        sp = sum(nf(r[I["Spend"]]) for r in camp if "Spend" in I)
        sa = sum(nf(r[I["Sales"]]) for r in camp if "Sales" in I)
        if sp or sa or camp:
            agg["channels"].setdefault("SB", {"spend": 0.0, "sales": 0.0, "campaigns": 0})
            agg["channels"]["SB"]["spend"] += sp
            agg["channels"]["SB"]["sales"] += sa
            agg["channels"]["SB"]["campaigns"] += len(camp)
    # SD / RAS — independent sheets, no dedup needed
    for ch, sheet in (("SD", "Sponsored Display Campaigns"), ("RAS", "RAS Campaigns")):
        if sheet not in wb.sheetnames:
            continue
        _, I, rows = sheet_io(sheet)
        if not rows or "Entity" not in I:
            continue
        camp = [r for r in rows if r[I["Entity"]] == "Campaign"]
        sp = sum(nf(r[I["Spend"]]) for r in camp if "Spend" in I)
        sa = sum(nf(r[I["Sales"]]) for r in camp if "Sales" in I)
        if sp or sa or camp:
            agg["channels"].setdefault(ch, {"spend": 0.0, "sales": 0.0, "campaigns": 0})
            agg["channels"][ch]["spend"] += sp
            agg["channels"][ch]["sales"] += sa
            agg["channels"][ch]["campaigns"] += len(camp)

    # ---- Sponsored Products (full detail) ----
    H, I, R = sheet_io(SP_SHEET)
    def g(row, col): return row[I[col]] if col in I else None
    ENT = lambda e: [r for r in R if g(r, "Entity") == e]
    camp_rows = ENT("Campaign"); ag_rows = ENT("Ad Group")
    kw_rows = ENT("Keyword"); pt_rows = ENT("Product Targeting")
    ba_rows = ENT("Bidding Adjustment")
    neg_kw = ENT("Negative Keyword"); camp_neg = ENT("Campaign Negative Keyword")

    camp_meta = {g(r, "Campaign ID"): dict(name=g(r, "Campaign Name"), state=g(r, "State"),
                                           targeting=g(r, "Targeting Type"), budget=nf(g(r, "Daily Budget")))
                 for r in camp_rows}
    enabled = {c for c, m in camp_meta.items() if m["state"] == "enabled"}

    # SP totals from leaf targets (Keyword + Product Targeting)
    for r in kw_rows + pt_rows:
        agg["sp"]["spend"] += nf(g(r, "Spend")); agg["sp"]["sales"] += nf(g(r, "Sales"))
        agg["sp"]["orders"] += nf(g(r, "Orders")); agg["sp"]["clicks"] += nf(g(r, "Clicks"))
        agg["sp"]["impr"] += nf(g(r, "Impressions"))

    # placement (Bidding Adjustment rows). Audience-cohort adjustment rows slice the
    # SAME traffic by audience, not by placement — including them double-counts.
    for r in ba_rows:
        if g(r, "Shopper Cohort Type") and not g(r, "Placement"):
            continue
        p = (g(r, "Placement") or "(none)").replace("Placement ", "")
        d = agg["placement"].setdefault(p, {"spend": 0.0, "sales": 0.0, "clicks": 0.0, "impr": 0.0})
        d["spend"] += nf(g(r, "Spend")); d["sales"] += nf(g(r, "Sales"))
        d["clicks"] += nf(g(r, "Clicks")); d["impr"] += nf(g(r, "Impressions"))

    # ad-group ASIN sets (Product Ad rows) for the multi-parent-family diagnosis
    for r in ENT("Product Ad"):
        a = (g(r, "ASIN (Informational only)") or "").strip()
        if a:
            agg.setdefault("ag_asins", {}).setdefault(
                (market, g(r, "Campaign Name (Informational only)"), g(r, "Ad Group Name (Informational only)")), set()).add(a)

    # bid hygiene
    def cents(v): return round((round(float(v), 2) * 100)) % 100
    for r in kw_rows + pt_rows:
        bid = g(r, "Bid")
        if bid in (None, ""):
            continue
        bv = nf(bid); agg["bids"]["total"] += 1
        if cents(bv) in (0, 25, 50, 75):
            agg["bids"]["round"] += 1

    # structure
    kw_camp = defaultdict(set)
    for r in kw_rows:
        cid = g(r, "Campaign ID")
        if cid in enabled and g(r, "Keyword Text") and g(r, "Match Type"):
            kw_camp[(str(g(r, "Keyword Text")).lower().strip(), g(r, "Match Type"))].add(cid)
    dups = {k: v for k, v in kw_camp.items() if len(v) > 1}
    camps_with_neg = set(g(r, "Campaign ID") for r in camp_neg) | set(g(r, "Campaign ID") for r in neg_kw)
    ag_kwcount = defaultdict(int)
    for r in kw_rows:
        if g(r, "Campaign ID") in enabled:
            ag_kwcount[g(r, "Ad Group ID")] += 1
    camp_buckets = defaultdict(set)
    for r in kw_rows:
        if g(r, "Campaign ID") in enabled:
            camp_buckets[g(r, "Campaign ID")].add(classify(cfg, g(r, "Keyword Text")))
    s = agg["structure"]
    s["total_campaigns"] += len(camp_rows); s["enabled"] += len(enabled)
    s["paused"] += len(camp_rows) - len(enabled); s["ad_groups"] += len(ag_rows)
    s["dup_kw_pairs"] += len(dups); s["dup_placements"] += sum(len(v) for v in dups.values())
    s["enabled_no_negatives"] += len([c for c in enabled if c not in camps_with_neg])
    s["total_neg_kw"] += len(neg_kw) + len(camp_neg)
    s["mixed_brand_generic_campaigns"] += len([c for c, bs in camp_buckets.items() if len(bs) > 1])
    if ag_kwcount:
        s["kw_per_ag_min"] = min(s.get("kw_per_ag_min", 10**9), min(ag_kwcount.values()))
        s["kw_per_ag_max"] = max(s.get("kw_per_ag_max", 0), max(ag_kwcount.values()))

    # keyword targets (for clean CSV + organic overlay)
    for r in kw_rows:
        sp, sa = nf(g(r, "Spend")), nf(g(r, "Sales"))
        agg["kw_targets"].append(dict(market=market, keyword=g(r, "Keyword Text"), match=g(r, "Match Type"),
                                      bucket=classify(cfg, g(r, "Keyword Text")),
                                      campaign=camp_meta.get(g(r, "Campaign ID"), {}).get("name"),
                                      state=g(r, "State"), bid=g(r, "Bid"), spend=sp, sales=sa,
                                      orders=nf(g(r, "Orders")), acos=acos(sp, sa)))

    # ---- SP intent split from the Search Term Report (what customers typed) ----
    # A customer search term is the gold standard: it is what the shopper asked for,
    # which is what an intent split is meant to measure. SB falls back to its TARGETS
    # only when its search-term report does not cover the channel (see below).
    for sheet in ("SP Search Term Report",):
        if sheet not in wb.sheetnames:
            continue
        SH, SI, SR = sheet_io(sheet)
        if "Customer Search Term" not in SI:
            continue
        for row in SR:
            def sg(col, _row=row): return _row[SI[col]] if col in SI else None
            term = sg("Customer Search Term")
            if term is None:
                continue
            b = classify(cfg, term)
            spd, sad, o, cl, im = nf(sg("Spend")), nf(sg("Sales")), nf(sg("Orders")), nf(sg("Clicks")), nf(sg("Impressions"))
            d = agg["st_bucket"].setdefault(b, {"spend": 0.0, "sales": 0.0, "orders": 0.0, "clicks": 0.0, "impr": 0.0})
            d["spend"] += spd; d["sales"] += sad; d["orders"] += o; d["clicks"] += cl; d["impr"] += im
            agg["st_rows"].append(dict(market=market, term=term, kw=sg("Keyword Text"), match=sg("Match Type"),
                                       bucket=b, spend=spd, sales=sad, orders=o, clicks=cl, impr=im,
                                       acos=acos(spd, sad), cvr=(o / cl if cl else 0), source="SP·search-term"))

    # ---- SB intent split: search term when it covers the channel, else TARGETS ----
    # Classifying SB by target silently mislabels conquesting. An SB campaign targeting
    # a rival ASIN still serves against searches, and on the Heusom 2026-07 run 53% of
    # that spend reached generic queries and 45% reached the brand's OWN name — only 2%
    # reached someone typing a competitor. Read by target it looked like $16.6k of
    # efficient conquesting at 24.6% ACOS; read by search term it was $6.9k at 45.4%,
    # the weakest bucket rather than the second best. So prefer the SB Search Term
    # Report whenever it accounts for the channel, and fall back to targets only when
    # it does not (older exports, or accounts whose SB spend is mostly video).
    sb_owner = agg.get("_sb_owner", {})
    sb_channel_spend = agg.get("channels", {}).get("SB", {}).get("spend", 0.0)
    sb_st_sheet = "SB Search Term Report"
    sb_st_spend = 0.0
    if sb_st_sheet in wb.sheetnames:
        _, _SI, _SR = sheet_io(sb_st_sheet)
        if "Customer Search Term" in _SI and "Spend" in _SI:
            sb_st_spend = sum(nf(r[_SI["Spend"]]) for r in _SR)
    # 90% is the bar: below it the report is not describing the channel and the target
    # fallback, imperfect as it is, at least covers all of the spend.
    sb_by_search_term = bool(sb_channel_spend) and sb_st_spend >= 0.90 * sb_channel_spend
    agg["sb_intent_method"] = "search-term" if sb_by_search_term else "target"
    agg["sb_st_coverage"] = (sb_st_spend / sb_channel_spend) if sb_channel_spend else 0.0

    if sb_by_search_term:
        SBH, SBI, SBR = sheet_io(sb_st_sheet)
        for row in SBR:
            def sg2(col, _row=row): return _row[SBI[col]] if col in SBI else None
            term = (str(sg2("Customer Search Term") or "")).strip()
            if term:
                b = classify(cfg, term); label = term
            else:
                # No typed query on the row (product targeting served on a detail page):
                # fall back to the target expression, which is the honest read there.
                label = (str(sg2("Product Targeting Expression") or "") + " " +
                         str(sg2("Keyword Text") or "")).strip()
                if not label:
                    continue
                b = classify_target(cfg, label)
            spd, sad = nf(sg2("Spend")), nf(sg2("Sales"))
            o, cl, im = nf(sg2("Orders")), nf(sg2("Clicks")), nf(sg2("Impressions"))
            d = agg["st_bucket"].setdefault(b, {"spend": 0.0, "sales": 0.0, "orders": 0.0, "clicks": 0.0, "impr": 0.0})
            d["spend"] += spd; d["sales"] += sad; d["orders"] += o; d["clicks"] += cl; d["impr"] += im
            agg["st_rows"].append(dict(market=market, term=label[:120], kw=sg2("Keyword Text"),
                                       match=sg2("Match Type"), bucket=b, spend=spd, sales=sad, orders=o,
                                       clicks=cl, impr=im, acos=acos(spd, sad), cvr=(o / cl if cl else 0),
                                       source="SB·search-term"))
    # Targets are only walked when the search-term read did not happen. Assigning to
    # SB_SHEETS itself would make the module-level name local to this whole function
    # and break its two earlier uses, so route through a local instead.
    sb_target_sheets = () if sb_by_search_term else SB_SHEETS

    # Dedupe across the two SB sheets by the owner map so overlapping campaigns'
    # targets aren't counted twice.
    for sheet in sb_target_sheets:
        if sheet not in wb.sheetnames:
            continue
        SBH, SBI, SBR = sheet_io(sheet)
        if "Entity" not in SBI:
            continue
        cidx = SBI.get("Campaign ID")
        for row in SBR:
            def bg(col, _row=row): return _row[SBI[col]] if col in SBI else None
            cid_val = row[cidx] if cidx is not None else None
            if cid_val is not None and sb_owner.get(cid_val, sheet) != sheet:
                continue  # this campaign is owned by the other SB sheet — skip (dedupe)
            ent = bg("Entity")
            if ent == "Keyword":
                label = bg("Keyword Text"); b = classify_target(cfg, label)
            elif ent == "Product Targeting":
                label = str(bg("Product Targeting Expression") or "") + " " + \
                        str(bg("Resolved Product Targeting Expression (Informational only)") or "")
                b = classify_target(cfg, label.strip())
            else:
                continue
            spd, sad = nf(bg("Spend")), nf(bg("Sales"))
            o, cl, im = nf(bg("Orders")), nf(bg("Clicks")), nf(bg("Impressions"))
            if not (spd or sad or cl or im):
                continue
            d = agg["st_bucket"].setdefault(b, {"spend": 0.0, "sales": 0.0, "orders": 0.0, "clicks": 0.0, "impr": 0.0})
            d["spend"] += spd; d["sales"] += sad; d["orders"] += o; d["clicks"] += cl; d["impr"] += im
            agg["st_rows"].append(dict(market=market, term=(label or "").strip()[:120], kw=bg("Keyword Text"),
                                       match=(bg("Match Type") or ent), bucket=b, spend=spd, sales=sad, orders=o,
                                       clicks=cl, impr=im, acos=acos(spd, sad), cvr=(o / cl if cl else 0),
                                       source="SB·target"))

# ----------------------------------------------------------------- business report
def parse_business_report(cfg, market, path, agg):
    asin_group = {}
    for grp, asins in (cfg.get("asin_groups") or {}).items():
        for a in asins:
            asin_group[a] = grp
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            asin = row.get("(Child) ASIN") or row.get("Child ASIN") or row.get("ASIN")
            if asin:
                agg.setdefault("asin_parent", {})[asin] = row.get("(Parent) ASIN") or row.get("Parent ASIN") or asin
            sales = nf(row.get("Ordered Product Sales")); units = int(nf(row.get("Units Ordered")))
            sess = int(nf(row.get("Sessions - Total") or row.get("Sessions – Total") or 0))
            usp = nf(row.get("Unit Session Percentage")); bb = nf(row.get("Featured Offer (Buy Box) Percentage"))
            agg["br_rows"].append(dict(market=market, asin=asin, title=row.get("Title", ""),
                                       group=asin_group.get(asin, "—"), sessions=sess, units=units,
                                       sales=sales, cvr=usp / 100.0, buybox=bb / 100.0))
            agg["br_total_sales"] += sales; agg["br_total_units"] += units; agg["br_total_sessions"] += sess

# ----------------------------------------------------------------- SQP (week-aware)
def parse_sqp(cfg, agg):
    def cls(q): return classify(cfg, q)
    mkt_wk = defaultdict(dict)                 # query -> week -> market metrics (deduped once)
    grp_wk = defaultdict(lambda: defaultdict(lambda: dict(imp=0, clk=0, cart=0, pur=0)))  # (group,query)->week->ASIN sums
    grp_intent = {}; grp_qtext = {}
    coverage = defaultdict(lambda: defaultdict(set))
    weeks = set()
    for group, path in (cfg["inputs"].get("sqp_csvs") or {}).items():
        p = rp(path)
        if not p or not p.exists():
            continue
        for x in csv.DictReader(open(p, encoding="utf-8-sig")):
            wk = x["Reporting Date"].strip(); q = x["Search Query"].strip(); ql = q.lower(); a = x["ASIN"]
            weeks.add(wk); coverage[(group, a)][wk].add(ql)
            if wk not in mkt_wk[ql]:
                mkt_wk[ql][wk] = dict(sv=nf(x["Search Query Volume"]), imp=nf(x["Impressions: Total Count"]),
                                      clk=nf(x["Clicks: Total Count"]), cart=nf(x["Cart Adds: Total Count"]),
                                      pur=nf(x["Purchases: Total Count"]))
            gg = grp_wk[(group, ql)][wk]
            gg["imp"] += nf(x["Impressions: ASIN Count"]); gg["clk"] += nf(x["Clicks: ASIN Count"])
            gg["cart"] += nf(x["Cart Adds: ASIN Count"]); gg["pur"] += nf(x["Purchases: ASIN Count"])
            grp_intent[(group, ql)] = cls(q); grp_qtext[(group, ql)] = q
    agg["sqp"] = dict(mkt_wk=mkt_wk, grp_wk=grp_wk, grp_intent=grp_intent, grp_qtext=grp_qtext,
                      coverage=coverage, weeks=sorted(weeks))

# ----------------------------------------------------------------- DataDive
def parse_datadive(cfg, agg):
    p = rp(cfg["inputs"].get("datadive_niche_json"))
    if not p or not p.exists():
        agg["dd"] = None
        return
    niche = json.loads(p.read_text())
    brand_asins = [a for asins in (cfg.get("asin_groups") or {}).values() for a in asins]
    rows = []
    for k in niche.get("keywords", []):
        ranks = k.get("asinRanks", {}) or {}
        best = None
        for a in brand_asins:
            if ranks.get(a) is not None:
                best = ranks[a] if best is None else min(best, ranks[a])
        rows.append(dict(kw=k.get("keyword"), sv=k.get("searchVolume"), rel=k.get("relevancy"), rank=best))
    agg["dd"] = dict(keywords=rows)

# ----------------------------------------------------------------- assemble
def build_metrics(cfg, agg, outdir):
    clean = outdir / "clean"; clean.mkdir(parents=True, exist_ok=True)
    default_market = cfg["marketplaces"][0] if cfg.get("marketplaces") else "US"

    sp = agg["sp"]
    tot_sp, tot_sa = sp["spend"], sp["sales"]
    # channel totals include SP + any additive channels
    channel_totals = {"SP": {"spend": tot_sp, "sales": tot_sa}}
    channel_totals.update({k: {"spend": v["spend"], "sales": v["sales"]} for k, v in agg["channels"].items()})
    all_spend = sum(v["spend"] for v in channel_totals.values())
    all_sales = sum(v["sales"] for v in channel_totals.values())

    STB = {k: dict(spend=v["spend"], sales=v["sales"], orders=v["orders"], clicks=v["clicks"], impr=v["impr"],
                   acos=acos(v["spend"], v["sales"]), cvr=(v["orders"] / v["clicks"] if v["clicks"] else 0))
           for k, v in agg["st_bucket"].items()}
    placement = {p: dict(spend=d["spend"], sales=d["sales"], clicks=d["clicks"], impr=d["impr"],
                         acos=acos(d["spend"], d["sales"])) for p, d in agg["placement"].items()}

    if agg["structure"]["kw_per_ag_min"] >= 10**9:
        agg["structure"]["kw_per_ag_min"] = 0
    # ad groups whose Product Ads span >1 parent family (keyword→product mismatch risk)
    pmap = agg.get("asin_parent", {})
    agg["structure"]["multi_parent_ad_groups"] = sum(
        1 for asins in agg.get("ag_asins", {}).values()
        if len({pmap.get(a, a) for a in asins}) > 1)
    br_total = agg["br_total_sales"]
    metrics = dict(
        client=cfg.get("client"), date=cfg.get("date"), marketplaces=cfg.get("marketplaces"),
        currency=cfg.get("currency", "USD"), breakeven=cfg.get("breakeven_acos", 0.50),
        windows=cfg.get("windows", {}),
        channels_present=sorted(channel_totals.keys()),
        channel_totals=channel_totals,
        totals=dict(spend=all_spend, sales=all_sales, orders=sp["orders"], clicks=sp["clicks"], impr=sp["impr"],
                    acos=acos(all_spend, all_sales), roas=(all_sales / all_spend if all_spend else 0),
                    br_total_sales=br_total, tacos=(all_spend / br_total if br_total else 0),
                    ad_dependency=(all_sales / br_total if br_total else 0)),
        searchterm_bucket=STB,
        st_coverage=(sum(v["spend"] for v in STB.values()) / all_spend if all_spend else 0),
        st_method=("SP classified by customer search term; SB by customer search term "
                   f"({agg.get('sb_st_coverage', 0):.0%} of SB spend covered by its search-term report). "
                   "Rows with no typed query fall back to the target expression. Remainder = SD, which "
                   "targets audiences and product pages and carries no search term."
                   if agg.get("sb_intent_method") == "search-term" else
                   "SP classified by customer search term; SB/SB-Multi by keyword + product-target "
                   f"expression, because SB's search-term report covers only {agg.get('sb_st_coverage', 0):.0%} "
                   "of SB spend. Target-based SB intent OVERSTATES conquesting: an ASIN-targeted SB ad "
                   "still serves against searches, most of which are not for that rival's name."),
        placement=placement,
        bid_hygiene=dict(total_bids=agg["bids"]["total"], round_bids=agg["bids"]["round"],
                         round_pct=(agg["bids"]["round"] / agg["bids"]["total"] if agg["bids"]["total"] else 0)),
        structure=agg["structure"],
        business_report=dict(rows=agg["br_rows"], total_sales=br_total,
                             total_units=agg["br_total_units"], total_sessions=agg["br_total_sessions"]),
    )

    # ---- data-completeness audit (surfaces thin/missing inputs before delivery) ----
    intent_cov = metrics["st_coverage"]
    sqp_groups = set((cfg["inputs"].get("sqp_csvs") or {}).keys())
    grp_rev = defaultdict(float)
    for r in agg["br_rows"]:
        grp_rev[r.get("group") or "—"] += r.get("sales") or 0.0
    uncovered = {g: rev for g, rev in grp_rev.items()
                 if g not in sqp_groups and g not in ("—", "Bundles") and rev > 0}
    sqp_gap = (sum(uncovered.values()) / br_total) if br_total else 0.0
    metrics["data_completeness"] = dict(
        intent_coverage=intent_cov,
        intent_coverage_low=bool(intent_cov < 0.90),
        sqp_revenue_gap=sqp_gap,
        sqp_uncovered_groups=sorted(uncovered, key=lambda g: -uncovered[g]),
        # RAS (Retail Ad Service, ads on non-Amazon retailer sites) is a deliberate
        # opt-in, not a gap. Its absence is never a finding; when it IS running it is
        # reported like any other channel from channel_totals.
        channels_missing=[c for c in ("SB", "SD") if c not in channel_totals],
        multi_parent_ad_groups=agg["structure"].get("multi_parent_ad_groups", 0),
    )

    # SQP summary (branded/generic/competitor capture) for the Overview + narrative
    sqp = agg.get("sqp")
    if sqp:
        latest = sqp["weeks"][-1] if sqp["weeks"] else None
        bi = defaultdict(lambda: dict(queries=0, avg_wk_sv=0.0, mkt_purch=0.0, brand_purch=0.0))
        # market SV / purchases by intent (deduped per query)
        from statistics import mean
        for q, weekmap in sqp["mkt_wk"].items():
            intent = classify(cfg, q)
            b = bi[intent]; b["queries"] += 1
            b["avg_wk_sv"] += mean([weekmap[w]["sv"] for w in weekmap]) if weekmap else 0
            b["mkt_purch"] += sum(weekmap[w]["pur"] for w in weekmap)
        for (group, ql), weekmap in sqp["grp_wk"].items():
            bi[sqp["grp_intent"][(group, ql)]]["brand_purch"] += sum(weekmap[w]["pur"] for w in weekmap)
        tot_avg = sum(bi[i]["avg_wk_sv"] for i in bi) or 1
        sqp_summary = {i: dict(queries=bi[i]["queries"], avg_wk_sv=bi[i]["avg_wk_sv"],
                               sv_share=bi[i]["avg_wk_sv"] / tot_avg,
                               brand_purch=int(bi[i]["brand_purch"]), mkt_purch=int(bi[i]["mkt_purch"]),
                               capture=(bi[i]["brand_purch"] / bi[i]["mkt_purch"] if bi[i]["mkt_purch"] else 0))
                       for i in bi}
        (clean / "sqp_summary.json").write_text(json.dumps(sqp_summary, indent=2))
        metrics["sqp_summary"] = sqp_summary

    # DataDive summary
    if agg.get("dd"):
        dd = agg["dd"]["keywords"]
        advertised = set(str(r["keyword"]).lower().strip() for r in agg["kw_targets"] if r["bucket"] == "Generic")
        metrics["datadive"] = dict(n_kw=len(dd),
                                   # Named for its actual threshold, not "p1". It was called
                                   # ranked_p1 while counting rank <=15, which invites anyone
                                   # reading metrics.json to quote it as a page-1 figure next to
                                   # a chart that bands 1-4/5-10/11-20/21-50. Page 1 is rank <=48.
                                   ranked_top15=len([d for d in dd if d["rank"] and d["rank"] <= 15]),
                                   advertised_and_ranked_top10=len([d for d in dd if d["kw"] and d["kw"].lower().strip() in advertised and d["rank"] and d["rank"] <= 10]))
        _wcsv(clean / "datadive_keywords.csv", sorted(dd, key=lambda x: -(x["sv"] or 0)), ["kw", "sv", "rel", "rank"])

    (outdir / "metrics.json").write_text(json.dumps(metrics, indent=2, default=str))

    # clean CSVs
    agg["st_rows"].sort(key=lambda x: -x["spend"])
    _wcsv(clean / "search_terms_classified.csv", agg["st_rows"],
          ["market", "term", "kw", "match", "bucket", "spend", "sales", "orders", "clicks", "impr", "acos", "cvr", "source"])
    agg["kw_targets"].sort(key=lambda x: -x["spend"])
    _wcsv(clean / "keyword_targets.csv", agg["kw_targets"],
          ["market", "keyword", "match", "bucket", "campaign", "state", "bid", "spend", "sales", "orders", "acos"])
    _wcsv(clean / "business_report.csv", agg["br_rows"],
          ["market", "asin", "group", "title", "sessions", "units", "sales", "cvr", "buybox"])
    return metrics

def _wcsv(path, rows, cols):
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c) for c in cols})

# ----------------------------------------------------------------- main
def run(config_path, outdir=None):
    cfg = load_config(config_path)
    outdir = Path(outdir) if outdir else Path(rp(cfg["delivery"]["drive_folder"]) if False else (REPO / "output" / _slug(cfg["client"]) / "reporting"))
    outdir.mkdir(parents=True, exist_ok=True)
    default_market = cfg["marketplaces"][0] if cfg.get("marketplaces") else "US"

    agg = dict(sp=dict(spend=0.0, sales=0.0, orders=0.0, clicks=0.0, impr=0.0),
               channels={}, placement={}, bids={"total": 0, "round": 0},
               structure=dict(total_campaigns=0, enabled=0, paused=0, ad_groups=0, dup_kw_pairs=0,
                              dup_placements=0, enabled_no_negatives=0, total_neg_kw=0,
                              mixed_brand_generic_campaigns=0, kw_per_ag_min=10**9, kw_per_ag_max=0),
               st_bucket={}, st_rows=[], kw_targets=[],
               br_rows=[], br_total_sales=0.0, br_total_units=0, br_total_sessions=0)

    for market, path in as_market_map(cfg["inputs"].get("ads_bulk_xlsx"), default_market).items():
        p = rp(path)
        if p and p.exists():
            parse_bulk(cfg, market, p, agg)
    for market, path in as_market_map(cfg["inputs"].get("business_report_csv"), default_market).items():
        p = rp(path)
        if p and p.exists():
            parse_business_report(cfg, market, p, agg)
    parse_sqp(cfg, agg)
    parse_datadive(cfg, agg)
    m = build_metrics(cfg, agg, outdir)
    print(f"[analyze] {cfg['client']}: spend {m['totals']['spend']:,.0f} sales {m['totals']['sales']:,.0f} "
          f"ACOS {m['totals']['acos']:.1%} | channels {m['channels_present']} | outdir {outdir}")
    return outdir, m

def _slug(s):
    return re.sub(r"[^a-z0-9]+", "-", (s or "client").lower()).strip("-")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: analyze_audit.py <config.json> [outdir]"); sys.exit(1)
    run(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
