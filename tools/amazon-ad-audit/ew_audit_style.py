#!/usr/bin/env python3
"""
Ecom Wizards — Amazon Ad/Sales Audit shared styling system.
Single source of truth for the EW CI palette, fonts, and workbook helpers used
by build_audit_workbook / build_sqp_workbook / build_master_workbook.

CRITICAL RULE (the bug this module exists to prevent):
  ACOS is ALWAYS stored as a ratio (0.23 = 23%, 1.09 = 109%). Colour helpers must
  NEVER divide by 100. Any `v = v if v<=1 else v/100` normalization is a bug — it
  turns every >100% ACOS green. Do not reintroduce it.
"""
from __future__ import annotations
from copy import copy
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- Ecom Wizards CI ----------
C = {"obsidian": "0F1318", "carbon": "171C24", "slate": "1E242C",
     "coral": "FD4807", "violet": "3322E0", "deep": "0E01A2",
     "mist": "9AA5B4", "cloud": "F5F6F8", "hairline": "E3E7ED", "white": "FFFFFF", "ink": "1E242C"}
# soft traffic-lights (used on decision columns only)
TL = {"good": "C6EFCE", "ok": "E2EFDA", "warn": "FFEB9C", "bad": "FFC7CE", "grey": "E7EAEF"}
DISPLAY = "Aptos Display"; BODY = "Aptos"

def F(sz=11, bold=False, color="1E242C", name=BODY):
    return Font(name=name, size=sz, bold=bold, color=color)

HDR_FILL = PatternFill("solid", fgColor=C["obsidian"])
BAND_FILL = PatternFill("solid", fgColor=C["slate"])
CORAL_FILL = PatternFill("solid", fgColor=C["coral"])
SUB_FILL = PatternFill("solid", fgColor=C["hairline"])
_thin = Side(style="thin", color=C["hairline"])
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

USD = '$#,##0'; USD2 = '$#,##0.00'; PCT = '0.0%'; PCT2 = '0.00%'; RO = '0.00'; INT = '#,##0'
EUR = '€#,##0'; EUR2 = '€#,##0.00'

def numf(x):
    """Parse a possibly-formatted number (strips $, €, %, commas) to float."""
    if x is None:
        return 0.0
    s = str(x).replace(",", "").replace("%", "").replace("$", "").replace("€", "").strip()
    if s in ("", "-", "—", "nan", "None"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def money_fmt(currency="USD"):
    return EUR if currency in ("EUR", "€") else USD

# ---------- traffic-light fills ----------
def acos_fill(v, breakeven=0.50):
    """ACOS colour. v is a RATIO (never divide by 100).
    green <30% · light-green <break-even · amber ≤60% · red >60%."""
    if v is None or v == 0:
        return None
    return PatternFill("solid", fgColor=(
        TL["good"] if v < .30 else
        TL["ok"] if v < breakeven else
        TL["warn"] if v <= .60 else
        TL["bad"]))

def roas_fill(v):
    if v is None or v == 0:
        return None
    return PatternFill("solid", fgColor=(
        TL["good"] if v >= 3 else TL["ok"] if v >= 2 else TL["warn"] if v >= 1.5 else TL["bad"]))

def gap_fill(v):
    """CTR/CVR gap in percentage points vs market: green ≥0, amber small-neg, red well below."""
    return PatternFill("solid", fgColor=(TL["good"] if v >= 0 else TL["warn"] if v > -1 else TL["bad"]))

def tl_fill(value, metric, breakeven=0.50):
    if metric == "acos":
        return acos_fill(value, breakeven)
    if metric == "roas":
        return roas_fill(value)
    return None

# ---------- workbook frame helpers ----------
def title_block(ws, title, subtitle, width_cols, banner="ECOM WIZARDS  ·  Amazon Advertising Audit"):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width_cols)
    c = ws.cell(1, 1, title); c.font = F(16, True, C["white"], DISPLAY); c.fill = HDR_FILL; c.alignment = LEFT
    for col in range(1, width_cols + 1):
        ws.cell(1, col).fill = HDR_FILL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width_cols)
    a = ws.cell(2, 1, banner); a.font = F(9, True, C["white"]); a.fill = CORAL_FILL; a.alignment = LEFT
    for col in range(1, width_cols + 1):
        ws.cell(2, col).fill = CORAL_FILL
    if subtitle:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=width_cols)
        s = ws.cell(3, 1, subtitle); s.font = F(9, False, C["mist"]); s.alignment = LEFT
    ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 16

def header_row(ws, r, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h); cell.fill = HDR_FILL; cell.font = F(10, True, C["white"])
        cell.alignment = CENTER; cell.border = BORDER
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def datarow(ws, rr, vals, fmts, acos_cols=(), roas_cols=(), left_cols=(1,), breakeven=0.50):
    for c, (v, fm) in enumerate(zip(vals, fmts), 1):
        cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(10)
        if fm:
            cell.number_format = fm; cell.alignment = RIGHT
        if c in left_cols:
            cell.alignment = LEFT
        if c in acos_cols:
            f = acos_fill(v, breakeven)
            if f:
                cell.fill = f
        if c in roas_cols:
            f = roas_fill(v)
            if f:
                cell.fill = f

def note(ws, r, text, width, color=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
    cc = ws.cell(r, 1, text); cc.font = F(9, False, color or C["mist"]); cc.alignment = WRAP
    ws.row_dimensions[r].height = max(14, 14 * (1 + len(text) // (width * 11)))

def band(ws, r, text, width, font_color=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
    cc = ws.cell(r, 1, text); cc.font = F(11, True, font_color or C["white"]); cc.fill = BAND_FILL; cc.alignment = LEFT
    for col in range(1, width + 1):
        ws.cell(r, col).fill = BAND_FILL

def copy_sheet(src_ws, dst_wb, new_title):
    """Cell-by-cell copy (value + style + merges + widths + freeze) for the master merge.
    Safe because audit/SQP workbooks contain tables only (no charts/images)."""
    nw = dst_wb.create_sheet(new_title[:31])
    nw.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    for col, dim in src_ws.column_dimensions.items():
        if dim.width:
            nw.column_dimensions[col].width = dim.width
    for rr, dim in src_ws.row_dimensions.items():
        if dim.height:
            nw.row_dimensions[rr].height = dim.height
    for mr in src_ws.merged_cells.ranges:
        nw.merge_cells(str(mr))
    for row in src_ws.iter_rows():
        for cell in row:
            nc = nw.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                nc.font = copy(cell.font); nc.fill = copy(cell.fill); nc.border = copy(cell.border)
                nc.alignment = copy(cell.alignment); nc.number_format = cell.number_format
    if src_ws.freeze_panes:
        nw.freeze_panes = src_ws.freeze_panes
    return nw
