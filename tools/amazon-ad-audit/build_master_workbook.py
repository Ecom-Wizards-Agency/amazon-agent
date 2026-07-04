#!/usr/bin/env python3
"""
Amazon Ad/Sales Audit — MASTER workbook (client-agnostic).
One file: a built ① Overview one-pager + all audit tabs + all SQP tabs (cell-copied)
+ one combined Sources tab. Ports the per-client master-workbook builder.
"""
from __future__ import annotations
import json, sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ew_audit_style as ew
from ew_audit_style import C, TL, F, BORDER, RIGHT, LEFT, WRAP, copy_sheet
from analyze_audit import load_config

INT = ew.INT; PCT = ew.PCT; RO = ew.RO


def build(config_path, outdir, audit_path, sqp_path):
    cfg = load_config(config_path)
    outdir = Path(outdir)
    M = json.loads((outdir / "metrics.json").read_text())
    SS = json.loads((outdir / "clean" / "sqp_summary.json").read_text()) if (outdir / "clean" / "sqp_summary.json").exists() else {}
    T = M["totals"]; STB = M["searchterm_bucket"]; P = M["placement"]
    BE = M.get("breakeven", 0.50); MONEY = ew.money_fmt(M.get("currency", "USD"))
    CLIENT = M.get("client", cfg.get("client", "Client"))
    markets = ", ".join(M.get("marketplaces", []) or [])
    channels = M.get("channels_present", ["SP"])
    win = M.get("windows", {})

    def acos_fill(v):
        return ew.acos_fill(v, BE)

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("① Overview")
    ws.sheet_view.showGridLines = False
    W = 6
    for c, wd in enumerate([30, 16, 14, 16, 16, 20], 1):
        ws.column_dimensions[get_column_letter(c)].width = wd
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=W)
    c = ws.cell(1, 1, f"{CLIENT} — {markets} Amazon Audit  ·  Master Summary"); c.font = F(16, True, C["white"], ew.DISPLAY); c.fill = ew.HDR_FILL; c.alignment = LEFT
    for col in range(1, W + 1):
        ws.cell(1, col).fill = ew.HDR_FILL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=W)
    a = ws.cell(2, 1, ew.brand_banner(f"{win.get('ads','')}  ·  {'/'.join(channels)}  ·  Break-even ACOS assumption: {BE:.0%}")); a.font = F(9, True, C["white"]); a.fill = ew.CORAL_FILL; a.alignment = LEFT
    for col in range(1, W + 1):
        ws.cell(2, col).fill = ew.CORAL_FILL
    ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 16
    r = [4]

    gen_acos = STB.get("Generic", {}).get("acos") or 0
    gen_cap = (SS.get("Generic", {}) or {}).get("capture")
    br_cap = (SS.get("Branded", {}) or {}).get("capture")
    OV = cfg.get("overview", {}) or {}
    oneliner = OV.get("oneliner") or (
        f"Profitable in aggregate — but branded carries it. Generic prospecting runs at {gen_acos:.0%} ACOS"
        + (f", and the SQP data shows why: the brand captures {gen_cap:.1%} of category purchases vs {br_cap:.0%} on its own name." if gen_cap is not None else "."))
    ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=W)
    ws.cell(r[0], 1, oneliner).font = F(10, True, C["coral"]); ws.cell(r[0], 1).alignment = WRAP; ws.row_dimensions[r[0]].height = 40; r[0] += 2

    def band(txt):
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=W)
        cc = ws.cell(r[0], 1, txt); cc.font = F(11, True, C["white"]); cc.fill = ew.BAND_FILL; cc.alignment = LEFT
        for col in range(1, W + 1):
            ws.cell(r[0], col).fill = ew.BAND_FILL
        r[0] += 1

    organic = T["br_total_sales"] - T["sales"]
    band(f"Account KPIs ({'/'.join(channels)}, {win.get('ads','30 days')})")
    kpis = [("Ad spend", T["spend"], MONEY, None), ("Ad sales", T["sales"], MONEY, None),
            ("Ad ACOS", T["acos"], PCT, "acos"), ("Ad ROAS", T["roas"], RO, None),
            ("Total sales (all traffic)", T["br_total_sales"], MONEY, None), ("TACOS", T["tacos"], PCT, None),
            ("Organic (implied)", organic, MONEY, None),
            ("Ad : organic", f"{T['ad_dependency']*100:.0f} : {(1-T['ad_dependency'])*100:.0f}", None, None)]
    ci = 1
    for lbl, val, fmt, metric in kpis:
        ws.cell(r[0], ci, lbl).font = F(9, False, C["mist"]); ws.cell(r[0], ci).border = BORDER
        cell = ws.cell(r[0] + 1, ci, val); cell.font = F(12, True); cell.border = BORDER
        if fmt:
            cell.number_format = fmt
        if metric == "acos":
            f = acos_fill(val)
            if f:
                cell.fill = f
        ci += 1
        if ci > W:
            ci = 1; r[0] += 2
    r[0] += 2

    st_cov = sum(v.get("spend", 0) for v in STB.values()) / T["spend"] if T["spend"] else 0
    band(f"Traffic mix — spend efficiency (ads) × purchase capture (SQP) — intent split covers {st_cov:.0%} of spend (SP by search term, SB by target)")
    hdr = ["Bucket", "Ad spend", "% spend", "Ad ACOS", "SQP SV share", "SQP purchase capture"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(r[0], c, h); cell.fill = ew.HDR_FILL; cell.font = F(9, True, C["white"]); cell.border = BORDER; cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r[0] += 1
    for b in ("Branded", "Generic", "Competitor"):
        d = STB.get(b)
        if not d:
            continue
        s = SS.get(b, {})
        vals = [b, d["spend"], d["spend"]/T["spend"], d["acos"], s.get("sv_share"), s.get("capture")]
        fmts = [None, MONEY, PCT, PCT, PCT, PCT]
        for c, (v, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r[0], c, v); cell.border = BORDER; cell.font = F(10)
            if fm:
                cell.number_format = fm; cell.alignment = RIGHT
            if c == 1:
                cell.alignment = LEFT
            if c == 4:
                f = acos_fill(v)
                if f:
                    cell.fill = f
            if c == 6 and v is not None:
                cell.fill = PatternFill("solid", fgColor=(TL["good"] if v > 0.3 else TL["bad"] if v < 0.05 else TL["warn"]))
        r[0] += 1
    # remainder without search-term detail so % spend sums to 100%
    un_sp = T["spend"] - sum(v.get("spend", 0) for v in STB.values())
    un_sa = T["sales"] - sum(v.get("sales", 0) for v in STB.values())
    if un_sp > 0.005 * T["spend"]:
        vals = ["Not classified (SB video/reach + SD)", un_sp, un_sp/T["spend"], (un_sp/un_sa if un_sa else None), None, None]
        fmts = [None, MONEY, PCT, PCT, None, None]
        for c, (v, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r[0], c, v); cell.border = BORDER; cell.font = F(10)
            if fm:
                cell.number_format = fm; cell.alignment = RIGHT
            if c == 1:
                cell.alignment = LEFT
            if c == 4:
                f = acos_fill(v)
                if f:
                    cell.fill = f
        r[0] += 1
    r[0] += 1

    band("Placement — same keywords, different ROI")
    for c, h in enumerate(["Placement", "Spend", "ACOS"], 1):
        cell = ws.cell(r[0], c, h); cell.fill = ew.HDR_FILL; cell.font = F(9, True, C["white"]); cell.border = BORDER
    r[0] += 1
    for p, d in sorted(P.items(), key=lambda x: -x[1]["spend"]):
        ws.cell(r[0], 1, p).border = BORDER; ws.cell(r[0], 1).font = F(10)
        cc = ws.cell(r[0], 2, d["spend"]); cc.number_format = MONEY; cc.border = BORDER; cc.alignment = RIGHT; cc.font = F(10)
        cell = ws.cell(r[0], 3, d["acos"]); cell.number_format = PCT; cell.border = BORDER; cell.alignment = RIGHT; cell.font = F(10)
        f = acos_fill(d["acos"])
        if f:
            cell.fill = f
        r[0] += 1
    r[0] += 1

    band("Top findings")
    for t in _findings(M, SS, MONEY, channels):
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=W)
        ws.cell(r[0], 1, t).font = F(10); ws.cell(r[0], 1).alignment = WRAP; ws.row_dimensions[r[0]].height = 26; r[0] += 1
    band(OV.get("recommendations_title") or "Recommendations (short / medium term)")
    for t in OV.get("recommendations") or [
        "Lever 1 — Earn 5-star reviews by resetting buyer expectations (accurate imagery/sizing/claims + Vine/inserts). The rating gap is the true cap on category demand.",
        "Lever 2 — Focus positioning on the winning use case across listing, brand store & brand videos (stop the 'works everywhere' dilution).",
        "Ads track (parallel) — Restructure PPC (waste falls out of a clean setup); rebalance toward Top of Search; add the missing channel motions.",
    ]:
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=W)
        ws.cell(r[0], 1, t).font = F(10); ws.cell(r[0], 1).alignment = WRAP; ws.row_dimensions[r[0]].height = 28; r[0] += 1
    r[0] += 1
    ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=W)
    ws.cell(r[0], 1, f"Break-even ACOS {BE:.0%} is an ASSUMPTION pending confirmed margin. All ACOS colour verdicts update on the real number.").font = F(9, True, C["violet"])
    ws.cell(r[0], 1).alignment = WRAP; ws.row_dimensions[r[0]].height = 30
    ws.freeze_panes = "A3"

    # ---- copy audit + sqp sheets ----
    aud = load_workbook(audit_path)
    for sn in aud.sheetnames:
        if sn == "Sources & Notes":
            continue
        title = "② Executive Summary" if sn == "Executive Summary" else sn
        copy_sheet(aud[sn], wb, title)
    sqp = load_workbook(sqp_path)
    for sn in sqp.sheetnames:
        if sn == "Sources & Notes":
            continue
        copy_sheet(sqp[sn], wb, ("SQP · " + sn))
    copy_sheet(aud["Sources & Notes"], wb, "ⓩ Sources & Notes")

    out = outdir / f"{_slug(CLIENT)}_{_slug(markets)}_Amazon_Audit_MASTER.xlsx"
    wb.save(out)
    print("[master] wrote", out.name, "| tabs:", len(wb.sheetnames))
    return out


def _findings(M, SS, MONEY, channels):
    T = M["totals"]; STB = M["searchterm_bucket"]; P = M["placement"]
    br = STB.get("Branded", {}); gen = STB.get("Generic", {})
    out = [f"Branded carries the account ({br.get('spend',0)/T['spend']:.0%} spend @ {(br.get('acos') or 0):.0%} ACOS)."]
    out.append(f"Generic bleeds — {gen.get('spend',0)/T['spend']:.0%} spend @ {(gen.get('acos') or 0):.0%} ACOS.")
    if P:
        best = min(P.items(), key=lambda x: (x[1]['acos'] if x[1]['acos'] else 9))
        worst = max(P.items(), key=lambda x: (x[1]['acos'] if x[1]['acos'] else 0) * (x[1]['spend'] > 0))
        if best[0] != worst[0]:
            out.append(f"Placement gap — {worst[0]} {worst[1]['acos']:.0%} vs {best[0]} {best[1]['acos']:.0%}.")
    mp = (M.get("structure") or {}).get("multi_parent_ad_groups")
    if mp:
        out.append(f"{mp} of {(M.get('structure') or {}).get('ad_groups', '?')} ad groups advertise SEVERAL parent families in one ad group — Amazon, not you, picks which product serves each query; keyword→product fit and per-product stats are uncontrolled.")
    miss = [c for c in ("SB", "SD", "RAS") if c not in channels]
    if miss:
        out.append(f"Channel gaps — no {', '.join(miss)} (no brand-defense / retargeting).")
    gc = (SS.get("Generic", {}) or {}).get("capture")
    if gc is not None:
        out.append(f"SQP capture — {gc:.1%} of category purchases; the category demand is largely unconverted.")
    return [f"{i}. {s}" for i, s in enumerate(out, 1)]


def _slug(s):
    import re
    return re.sub(r"[^A-Za-z0-9]+", "-", (s or "x")).strip("-")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
