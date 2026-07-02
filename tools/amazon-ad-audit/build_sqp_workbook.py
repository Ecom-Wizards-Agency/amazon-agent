#!/usr/bin/env python3
"""
Amazon Ad/Sales Audit — SQP Intelligence workbook (client-agnostic).
Week-aware, market-SV deduped per query+week, ASIN-vs-market CTR/CVR gaps, PPC overlay
(from clean/search_terms_classified.csv aggregated by term), DataDive organic-rank overlay.
Ports the per-client SQP-workbook builder; styling from ew_audit_style.
"""
from __future__ import annotations
import csv, json, sys
from collections import defaultdict
from pathlib import Path
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ew_audit_style as ew
from ew_audit_style import C, TL, F, header_row, numf, RIGHT, LEFT, WRAP, CENTER, BORDER, gap_fill, title_block as _tb
from analyze_audit import rp, load_config, classify

INT = ew.INT; PCT = ew.PCT; PCT2 = ew.PCT2; RO = ew.RO


def _num(x):
    return numf(x)


def build(config_path, outdir):
    cfg = load_config(config_path)
    outdir = Path(outdir)
    M = json.loads((outdir / "metrics.json").read_text())
    CLEAN = outdir / "clean"
    CLIENT = M.get("client", cfg.get("client", "Client"))
    MONEY = ew.money_fmt(M.get("currency", "USD"))
    markets = ", ".join(M.get("marketplaces", []) or [])
    w = M.get("windows", {})

    def avg(xs):
        xs = list(xs); return mean(xs) if xs else 0.0

    def safe(nn, dd):
        return nn / dd if dd else 0.0

    # ---- load SQP (week-aware) ----
    mkt_wk = defaultdict(dict); mkt_intent = {}
    grp_wk = defaultdict(lambda: defaultdict(lambda: dict(imp=0, clk=0, cart=0, pur=0)))
    grp_intent = {}; grp_qtext = {}
    coverage = defaultdict(lambda: defaultdict(set)); weeks = set()
    for group, path in (cfg["inputs"].get("sqp_csvs") or {}).items():
        p = rp(path)
        if not p or not p.exists():
            continue
        for row in csv.DictReader(open(p, encoding="utf-8-sig")):
            wk = row["Reporting Date"].strip(); q = row["Search Query"].strip(); ql = q.lower(); a = row["ASIN"]
            weeks.add(wk); coverage[(group, a)][wk].add(ql)
            if wk not in mkt_wk[ql]:
                mkt_wk[ql][wk] = dict(sv=_num(row["Search Query Volume"]), imp=_num(row["Impressions: Total Count"]),
                                      clk=_num(row["Clicks: Total Count"]), cart=_num(row["Cart Adds: Total Count"]),
                                      pur=_num(row["Purchases: Total Count"]))
                mkt_intent[ql] = classify(cfg, q)
            g = grp_wk[(group, ql)][wk]
            g["imp"] += _num(row["Impressions: ASIN Count"]); g["clk"] += _num(row["Clicks: ASIN Count"])
            g["cart"] += _num(row["Cart Adds: ASIN Count"]); g["pur"] += _num(row["Purchases: ASIN Count"])
            grp_intent[(group, ql)] = classify(cfg, q); grp_qtext[(group, ql)] = q
    KEPT = sorted(weeks)

    # ---- query intelligence ----
    qi = []
    for (group, ql), weekmap in grp_wk.items():
        mweeks = mkt_wk.get(ql, {})
        sv_series = [mweeks[x]["sv"] for x in mweeks]
        mctr = [safe(mweeks[x]["clk"], mweeks[x]["imp"]) for x in mweeks]
        mcvr = [safe(mweeks[x]["pur"], mweeks[x]["clk"]) for x in mweeks]
        mcart = [safe(mweeks[x]["cart"], mweeks[x]["clk"]) for x in mweeks]
        actr = [safe(weekmap[x]["clk"], weekmap[x]["imp"]) for x in weekmap]
        acvr = [safe(weekmap[x]["pur"], weekmap[x]["clk"]) for x in weekmap]
        acart = [safe(weekmap[x]["cart"], weekmap[x]["clk"]) for x in weekmap]
        impshare = [safe(weekmap[x]["imp"], mweeks[x]["imp"]) for x in weekmap if x in mweeks]
        pur_asin = sum(weekmap[x]["pur"] for x in weekmap); pur_mkt = sum(mweeks[x]["pur"] for x in mweeks)
        qi.append([group, grp_qtext[(group, ql)], grp_intent[(group, ql)], len(weekmap),
                   round(avg(sv_series)), round(sum(sv_series)), avg(impshare),
                   avg(actr), avg(mctr), (avg(actr) - avg(mctr)) * 100,
                   avg(acvr), avg(mcvr), (avg(acvr) - avg(mcvr)) * 100,
                   avg(acart), avg(mcart), safe(pur_asin, pur_mkt)])
    qi.sort(key=lambda x: -x[4])

    # ---- PPC overlay: aggregate clean/search_terms_classified.csv by term ----
    ppc = defaultdict(lambda: dict(sp=0.0, sa=0.0))
    stp = CLEAN / "search_terms_classified.csv"
    if stp.exists():
        for row in csv.DictReader(open(stp)):
            d = ppc[str(row["term"]).lower().strip()]
            d["sp"] += _num(row["spend"]); d["sa"] += _num(row["sales"])

    # ---- DataDive organic rank ----
    ddrank = {}
    ddp = CLEAN / "datadive_keywords.csv"
    if ddp.exists():
        for row in csv.DictReader(open(ddp)):
            rk = row.get("rank")
            ddrank[row["kw"].lower().strip()] = (int(_num(rk)) if rk not in (None, "", "None") else None)
    has_dd = bool(ddrank)

    wb = Workbook(); wb.remove(wb.active)
    banner = f"ECOM WIZARDS  ·  {CLIENT} SQP Intelligence ({markets})  ·  {len(KEPT)} weeks"

    def title_block(ws, t, sub, wd):
        _tb(ws, t, sub, wd, banner=banner)

    def notes_block(ws, rr, lines, wd, h=42):
        for nt in lines:
            ws.cell(rr, 1, nt).font = F(10); ws.cell(rr, 1).alignment = WRAP
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=wd); ws.row_dimensions[rr].height = h; rr += 1
        return rr

    # TAB 1: Data Completeness
    ws = wb.create_sheet("Data Completeness")
    title_block(ws, "Data Completeness — read this first", "Multi-ASIN SQP exports cap the query grid per ASIN per week. Long-tail is missing; SV totals are a FLOOR.", 6)
    rr = notes_block(ws, 4, [
        f"Built from {len(cfg['inputs'].get('sqp_csvs',{}))} multi-ASIN SQP exports. Weeks used: {', '.join(KEPT)}.",
        "CAP: the multi-ASIN SQP tool returns only the top queries per ASIN per week. Below the cap is missing every week, so SV/non-branded totals are a FLOOR.",
        "FIX for full data: export Brand Analytics > Search Query Performance per single ASIN in Seller Central (uncapped).",
        "Market metrics deduped once per query+week. ASIN counts summed across the ASINs in each parent group.",
    ], 6)
    rr += 1; ws.cell(rr, 1, "Per-ASIN coverage").font = F(11, True, C["deep"]); rr += 1
    hr = rr; header_row(ws, hr, ["Group", "ASIN", "Weeks", "Max queries / week", "Total unique queries", "Status"], [14, 16, 8, 18, 18, 12]); rr += 1
    for (group, a), wkmap in sorted(coverage.items()):
        perwk = [len(wkmap[x]) for x in wkmap]; maxq = max(perwk) if perwk else 0
        uq = len(set().union(*wkmap.values())) if wkmap else 0
        vals = [group, a, len(wkmap), maxq, uq, "CAPPED" if maxq >= 100 else ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(10)
            if c in (3, 4, 5):
                cell.alignment = RIGHT; cell.number_format = INT
            if c == 6 and v == "CAPPED":
                cell.fill = PatternFill("solid", fgColor=TL["warn"])
        rr += 1
    ws.freeze_panes = f"A{hr+1}"

    # TAB 2: Weekly Trend
    ws = wb.create_sheet("Weekly Trend")
    title_block(ws, "Weekly Trend — deduped market demand + brand funnel", "Per week: deduped market SV / CTR / CVR, and the brand's pooled impression→click→cart→purchase funnel.", 8)
    hr = 5; header_row(ws, hr, ["Week", "Market SV", "Mkt CTR", "Mkt CVR", "Brand impr", "Brand clicks", "Brand cart", "Brand purch"], [14, 14, 10, 10, 12, 12, 12, 12])
    rr = hr + 1
    for x in KEPT:
        sv = sum(mkt_wk[q][x]["sv"] for q in mkt_wk if x in mkt_wk[q])
        imp = sum(mkt_wk[q][x]["imp"] for q in mkt_wk if x in mkt_wk[q]); clk = sum(mkt_wk[q][x]["clk"] for q in mkt_wk if x in mkt_wk[q]); pur = sum(mkt_wk[q][x]["pur"] for q in mkt_wk if x in mkt_wk[q])
        limp = sum(grp_wk[k][x]["imp"] for k in grp_wk if x in grp_wk[k]); lclk = sum(grp_wk[k][x]["clk"] for k in grp_wk if x in grp_wk[k])
        lcart = sum(grp_wk[k][x]["cart"] for k in grp_wk if x in grp_wk[k]); lpur = sum(grp_wk[k][x]["pur"] for k in grp_wk if x in grp_wk[k])
        vals = [x, sv, safe(clk, imp), safe(pur, clk), limp, lclk, lcart, lpur]; fmts = [None, INT, PCT2, PCT, INT, INT, INT, INT]
        for c, (v, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(10)
            if fm:
                cell.number_format = fm; cell.alignment = RIGHT
        rr += 1
    ws.freeze_panes = f"A{hr+1}"

    # TAB 3: Branded vs Non-Branded
    ws = wb.create_sheet("Branded vs Non-Branded")
    title_block(ws, "Branded vs Non-Branded — SQP demand & purchase capture", "SV deduped per query+week. Capture = brand purchases ÷ market purchases on those queries.", 7)
    hr = 5; header_row(ws, hr, ["Intent", "Unique queries", "Avg weekly SV", "Total SV", "SV share", "Brand purch", "Mkt purch"], [14, 14, 14, 14, 10, 12, 12])
    bi = defaultdict(lambda: dict(q=0, avg=0.0, tot=0.0, ap=0.0, mp=0.0))
    for q, weekmap in mkt_wk.items():
        it = mkt_intent[q]; b = bi[it]; b["q"] += 1
        b["avg"] += avg([weekmap[x]["sv"] for x in weekmap]); b["tot"] += sum(weekmap[x]["sv"] for x in weekmap)
        b["mp"] += sum(weekmap[x]["pur"] for x in weekmap)
    for (group, ql), weekmap in grp_wk.items():
        bi[grp_intent[(group, ql)]]["ap"] += sum(weekmap[x]["pur"] for x in weekmap)
    tot_avg = sum(bi[i]["avg"] for i in bi) or 1
    rr = hr + 1
    for it in ("Branded", "Generic", "Competitor"):
        b = bi.get(it)
        if not b:
            continue
        vals = [it, b["q"], round(b["avg"]), round(b["tot"]), safe(b["avg"], tot_avg), int(b["ap"]), int(b["mp"])]; fmts = [None, INT, INT, INT, PCT, INT, INT]
        for c, (v, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(10)
            if fm:
                cell.number_format = fm; cell.alignment = RIGHT
            if c == 1 and it == "Branded":
                cell.fill = PatternFill("solid", fgColor=TL["good"])
            if c == 1 and it == "Generic":
                cell.fill = PatternFill("solid", fgColor=TL["warn"])
        rr += 1
    notes_block(ws, rr + 1, ["Read: the category (non-branded) queries are where the search volume is — but the brand captures a far smaller share of purchases there than on its own name. Category is the gap, gated by price + reviews."], 7, 44)
    ws.freeze_panes = f"A{hr+1}"

    # TAB 4: Query Intelligence
    ws = wb.create_sheet("Query Intelligence")
    title_block(ws, "Query Intelligence — brand vs market (multi-week avg)", "CTR/CVR gap vs market = where you under-index. Green = at/above market, red = below. Sorted by avg weekly SV.", 16)
    hr = 5; header_row(ws, hr, ["Group", "Search query", "Intent", "Wks", "Avg SV", "Total SV", "Imp Share", "Brand CTR", "Mkt CTR", "CTR gap pp", "Brand CVR", "Mkt CVR", "CVR gap pp", "Cart rate", "Mkt cart", "Purch share"],
                       [12, 30, 9, 5, 9, 10, 9, 9, 9, 9, 9, 9, 9, 9, 9, 10])
    rr = hr + 1
    for row in qi[:400]:
        fmts = [None, None, None, INT, INT, INT, PCT, PCT2, PCT2, '0.00', PCT, PCT, '0.00', PCT, PCT, PCT]
        for c, (v, fm) in enumerate(zip(row, fmts), 1):
            cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(9)
            if fm:
                cell.number_format = fm; cell.alignment = RIGHT
            if c == 2:
                cell.alignment = LEFT
            if c == 10:
                cell.fill = gap_fill(row[9])
            if c == 13:
                cell.fill = gap_fill(row[12])
        rr += 1
    ws.freeze_panes = f"A{hr+1}"

    # TAB 5: Top Opportunities
    ws = wb.create_sheet("Top Opportunities")
    title_block(ws, "Top Opportunities — non-branded, high demand, brand invisible", "Non-branded queries, avg weekly SV ≥ 500, ranked by lowest impression share.", 8)
    hr = 5; header_row(ws, hr, ["Group", "Search query", "Avg SV", "Total SV", "Imp Share", "Brand CTR", "Mkt CTR", "CTR gap pp", "Purch share"], [12, 30, 10, 11, 9, 9, 9, 10, 10])
    opp = [r for r in qi if r[2] != "Branded" and r[4] >= 500]
    best = {}
    for r in opp:
        if r[1] not in best or r[6] < best[r[1]][6]:
            best[r[1]] = r
    opp = sorted(best.values(), key=lambda x: x[6])[:60]
    rr = hr + 1
    for r in opp:
        vals = [r[0], r[1], r[4], r[5], r[6], r[7], r[8], r[9], r[15]]; fmts = [None, None, INT, INT, PCT, PCT2, PCT2, '0.00', PCT]
        for c, (v, fm) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(9)
            if fm:
                cell.number_format = fm; cell.alignment = RIGHT
            if c == 2:
                cell.alignment = LEFT
            if c == 5:
                cell.fill = PatternFill("solid", fgColor=(TL["bad"] if r[6] < 0.02 else TL["warn"] if r[6] < 0.05 else TL["ok"]))
        rr += 1
    notes_block(ws, rr + 1, ["Big category queries where the brand barely shows up. Winning them needs rank + conversion (price/reviews), not just ad bids."], 8, 44)
    ws.freeze_panes = f"A{hr+1}"

    # TAB 6: PPC Overlay
    ws = wb.create_sheet("PPC Overlay")
    title_block(ws, "PPC Overlay — buying visibility vs earning it", "Top SQP queries matched to ad spend/ACOS. High organic imp-share + high ad spend = paying for traffic you may already get.", 9)
    hr = 5; header_row(ws, hr, ["Group", "Search query", "Intent", "Avg SV", "Imp Share", "Ad spend", "Ad sales", "Ad ACOS", "Read"], [12, 28, 9, 9, 9, 11, 11, 9, 26])
    rr = hr + 1
    seen_q = {}
    for r in qi:
        key = r[1].lower().strip()
        if key not in seen_q or r[6] > seen_q[key][6]:
            seen_q[key] = r
    rows_ov = []
    for r in seen_q.values():
        p = ppc.get(r[1].lower().strip())
        if not p or p["sp"] < 20:
            continue
        ac = safe(p["sp"], p["sa"])
        read = "Branded — defensive" if r[2] == "Branded" else ("High spend, low visibility" if r[6] < 0.05 else "Category spend")
        rows_ov.append([r[0], r[1], r[2], r[4], r[6], p["sp"], p["sa"], ac if p["sa"] else None, read])
    for r in sorted(rows_ov, key=lambda x: -x[5])[:40]:
        fmts = [None, None, None, INT, PCT, MONEY, MONEY, PCT, None]
        for c, (v, fm) in enumerate(zip(r, fmts), 1):
            cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(9)
            if fm:
                cell.number_format = fm; cell.alignment = RIGHT
            if c in (2, 9):
                cell.alignment = LEFT
            if c == 8 and v is not None:
                f = ew.acos_fill(v, M.get("breakeven", 0.50))
                if f:
                    cell.fill = f
        rr += 1
    ws.freeze_panes = f"A{hr+1}"

    # TAB 7: Organic Rank (DataDive) — only if data present
    if has_dd:
        ws = wb.create_sheet("Organic Rank (DataDive)")
        title_block(ws, "Organic Rank overlay (DataDive)", "Brand organic rank on the SQP queries. Ranked top-10 + high SV + high ad spend = candidate to reduce PPC dependence.", 7)
        hr = 5; header_row(ws, hr, ["Group", "Search query", "Intent", "Avg SV", "Imp Share", "Organic rank", "Ad spend"], [12, 30, 9, 9, 10, 16, 11])
        rr = hr + 1
        seen_or = {}
        for r in qi:
            key = r[1].lower().strip()
            if key not in seen_or or r[6] > seen_or[key][6]:
                seen_or[key] = r
        for r in sorted(seen_or.values(), key=lambda x: -x[4]):
            if r[4] < 300:
                continue
            rk = ddrank.get(r[1].lower().strip())
            p = ppc.get(r[1].lower().strip())
            sp = p["sp"] if p else 0
            vals = [r[0], r[1], r[2], r[4], r[6], (rk if rk else "not ranked"), sp]; fmts = [None, None, None, INT, PCT, None, MONEY]
            for c, (v, fm) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(rr, c, v); cell.border = BORDER; cell.font = F(9)
                if fm:
                    cell.number_format = fm; cell.alignment = RIGHT
                if c == 2:
                    cell.alignment = LEFT
                if c == 6 and isinstance(rk, int):
                    cell.fill = PatternFill("solid", fgColor=(TL["good"] if rk <= 10 else TL["ok"] if rk <= 20 else TL["warn"]))
            rr += 1
            if rr > hr + 120:
                break
        ws.freeze_panes = f"A{hr+1}"

    # TAB 8: Sources & Notes
    ws = wb.create_sheet("Sources & Notes")
    title_block(ws, "Sources, method & caveats", None, 2)
    lines = [
        ("Data", f"{len(cfg['inputs'].get('sqp_csvs',{}))} multi-ASIN SQP exports, {len(KEPT)} weekly snapshots ({', '.join(KEPT)}). {markets}."),
        ("SV dedup", "Search Query Volume + market impressions/clicks/purchases deduped once per query+week. ASIN counts summed across the ASINs in each parent group."),
        ("Averages", "CTR/CVR/cart shown as mean of weekly values; SV shown as avg weekly and total."),
        ("Branded", f"Queries containing {', '.join(cfg.get('brand_tokens',[]))} = Branded. Competitor brand tokens = Competitor. Else Generic."),
        ("Cap caveat", "Multi-ASIN SQP returns only top queries per ASIN per week; long-tail missing; SV totals are a FLOOR."),
        ("PPC overlay", "Search Term Report (full window) matched to SQP queries by exact term. Ad window vs SQP weeks differ — directional alignment."),
        ("Legend", f"CTR/CVR gap: green ≥ market. ACOS: green <30% / <{M.get('breakeven',0.5):.0%} / amber ≤60% / red >60%."),
    ]
    rr = 4
    for k, v in lines:
        ws.cell(rr, 1, k).font = F(10, True); ws.cell(rr, 1).alignment = WRAP
        ws.cell(rr, 2, v).font = F(10); ws.cell(rr, 2).alignment = WRAP
        ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 120
        ws.row_dimensions[rr].height = 30; rr += 1

    for s in wb.worksheets:
        s.sheet_view.showGridLines = False
    out = outdir / f"{_slug(CLIENT)}_{_slug(markets)}_SQP_Intelligence.xlsx"
    wb.save(out)
    print("[sqp] wrote", out.name, "| tabs:", wb.sheetnames, "| weeks:", KEPT)
    return out


def _slug(s):
    import re
    return re.sub(r"[^A-Za-z0-9]+", "-", (s or "x")).strip("-")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2])
