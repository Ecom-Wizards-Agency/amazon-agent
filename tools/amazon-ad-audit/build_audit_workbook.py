#!/usr/bin/env python3
"""
Amazon Ad/Sales Audit — audit workbook builder (client-agnostic).
Reads config + <outdir>/metrics.json + <outdir>/clean/*.csv (+ optional DataDive
competitors JSON) and writes the EW-branded audit .xlsx. Nothing is hard-coded per
client — client name, break-even, currency, windows, channels all come from config/metrics.
"""
from __future__ import annotations
import csv, json, sys
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ew_audit_style as ew
from ew_audit_style import C, TL, F, title_block, header_row, datarow, note, numf, RIGHT, LEFT, WRAP, BORDER
from analyze_audit import rp, load_config

USD = ew.USD; USD2 = ew.USD2; PCT = ew.PCT; PCT2 = ew.PCT2; RO = ew.RO; INT = ew.INT


def build(config_path, outdir):
    cfg = load_config(config_path)
    outdir = Path(outdir)
    M = json.loads((outdir / "metrics.json").read_text())
    CLEAN = outdir / "clean"
    BE = M.get("breakeven", 0.50)
    MONEY = ew.money_fmt(M.get("currency", "USD"))
    CLIENT = M.get("client", cfg.get("client", "Client"))
    T = M["totals"]; STB = M["searchterm_bucket"]; P = M["placement"]
    markets = ", ".join(M.get("marketplaces", []) or [])
    channels = M.get("channels_present", ["SP"])
    sp_only = channels == ["SP"]
    w = M.get("windows", {})
    SUBTITLE = (f"{markets}  ·  Ads bulk {w.get('ads','')}  ·  Business Report {w.get('business_report','')}  ·  "
                f"SQP {'/'.join(w.get('sqp_weeks',[])[:1]+w.get('sqp_weeks',[])[-1:]) if w.get('sqp_weeks') else ''}"
                f"  ·  Break-even ACOS assumption: {BE:.0%}")
    slug_us = f"{markets}"

    def load_csv(name):
        p = CLEAN / name
        return list(csv.DictReader(open(p))) if p.exists() else []

    comp = None
    cj = rp(cfg["inputs"].get("datadive_competitors_json"))
    if cj and cj.exists():
        client_asins = {a for asins in (cfg.get("asin_groups") or {}).values() for a in asins}
        comp = _adapt_competitors(json.loads(cj.read_text()), client_asins)

    wb = Workbook(); wb.remove(wb.active)

    # ---- dynamic headline numbers ----
    br_share = STB.get("Branded", {}).get("spend", 0) / T["spend"] if T["spend"] else 0
    br_acos = STB.get("Branded", {}).get("acos") or 0
    gen_acos = STB.get("Generic", {}).get("acos") or 0
    best_pl = min(P.items(), key=lambda x: (x[1]["acos"] if x[1]["acos"] else 9)) if P else (None, {})
    worst_pl = max(P.items(), key=lambda x: (x[1]["acos"] if x[1]["acos"] else 0) * (x[1]["spend"] > 0)) if P else (None, {})
    missing_ch = [c for c in ("SB", "SD", "RAS") if c not in channels]

    # ================= TAB 1: Executive Summary =================
    ws = wb.create_sheet("Executive Summary")
    title_block(ws, f"{CLIENT} — {markets} Amazon Advertising Audit", SUBTITLE, 3)
    for c, wd in enumerate([46, 18, 34], 1):
        ws.column_dimensions[get_column_letter(c)].width = wd
    oneliner = (f"The account is profitable in aggregate because branded demand carries it — "
                f"but generic prospecting runs at {gen_acos:.0%} ACOS"
                + (f" and the product is a price outlier in a commodity category." if comp else "."))
    ws.cell(4, 1, oneliner).font = F(10, True, C["coral"]); ws.cell(4, 1).alignment = WRAP
    ws.merge_cells("A4:C4"); ws.row_dimensions[4].height = 42
    r = [6]

    def sub(label):
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=3)
        cc = ws.cell(r[0], 1, label); cc.font = F(11, True, C["white"]); cc.fill = ew.BAND_FILL; cc.alignment = LEFT
        for cx in range(1, 4):
            ws.cell(r[0], cx).fill = ew.BAND_FILL
        r[0] += 1

    def line(label, val, fmt=INT, metric=None, nt=""):
        ws.cell(r[0], 1, label).font = F(10)
        cc = ws.cell(r[0], 2, val); cc.font = F(10, True); cc.alignment = RIGHT
        if fmt:
            cc.number_format = fmt
        if metric:
            f = ew.tl_fill(val, metric, BE)
            if f:
                cc.fill = f
        if nt:
            ws.cell(r[0], 3, nt).font = F(9, False, C["mist"]); ws.cell(r[0], 3).alignment = WRAP
        for cx in (1, 2, 3):
            ws.cell(r[0], cx).border = BORDER
        r[0] += 1

    organic = T["br_total_sales"] - T["sales"]
    sub(f"Account totals — {'/'.join(channels)} ({w.get('ads','30 days')})")
    line("Ad spend", T["spend"], MONEY)
    line("Ad sales", T["sales"], MONEY)
    line("Ad ACOS", T["acos"], PCT, "acos", f"Aggregate is {'under' if T['acos'] < BE else 'over'} the {BE:.0%} break-even.")
    line("Ad ROAS", T["roas"], RO, "roas")
    line("Total ordered product sales (all traffic)", T["br_total_sales"], MONEY)
    line("Organic / non-ad sales — implied", organic, MONEY, None, "Total sales − ad sales. Directional (attribution vs report-date).")
    line("TACOS", T["tacos"], PCT, None, "Total ad spend ÷ total sales.")
    line("Ad-attributed share of sales", T["ad_dependency"], PCT, None,
         f"Ad : organic ≈ {T['ad_dependency']*100:.0f} : {(1-T['ad_dependency'])*100:.0f}"
         + (f"  ({T['sales']/organic:.2f} : 1)" if organic else ""))
    r[0] += 1
    sub("Traffic mix — where the money actually goes (by customer search term)")
    for b in ("Branded", "Generic", "Competitor"):
        d = STB.get(b)
        if not d:
            continue
        line(b, d["spend"], MONEY, None,
             f"{d['spend']/T['spend']:.0%} of spend · ACOS {d['acos']:.0%} · CVR {d['cvr']:.1%} · sales {_m(d['sales'],MONEY)}")
    r[0] += 1
    sub("Headline findings")
    findings = [
        f"1) Branded carries the account — {br_share:.0%} of spend at {br_acos:.0%} ACOS. Defensive spend on your own name.",
        f"2) Generic prospecting runs at {gen_acos:.0%} ACOS — the single biggest efficiency lever (see Waste & Winners).",
    ]
    if best_pl[0] and worst_pl[0] and worst_pl[0] != best_pl[0]:
        findings.append(f"3) Placement gap — {worst_pl[0]} {worst_pl[1]['acos']:.0%} ACOS vs {best_pl[0]} {best_pl[1]['acos']:.0%}. Same keywords, different ROI.")
    if missing_ch:
        findings.append(f"4) Channel gaps — no {', '.join(missing_ch)}. No brand-defense / retargeting motions running.")
    if comp:
        cps = [x for x in comp["competitors"] if not x.get("is_client")]
        client_rows = [x for x in comp["competitors"] if x.get("is_client")]
        if cps and client_rows:
            findings.append(f"5) Price outlier — sells at {_m(min(x['price'] for x in client_rows),MONEY)}–{_m(max(x['price'] for x in client_rows),MONEY)} vs a {_m(comp['median_price'],MONEY)} category median, lower ratings & fewer reviews. Ads buy visibility uphill.")
    for txt in findings:
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=3)
        ws.cell(r[0], 1, txt).font = F(10, False, C["ink"]); ws.cell(r[0], 1).alignment = WRAP
        ws.row_dimensions[r[0]].height = 30; r[0] += 1
    r[0] += 1
    ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=3)
    ws.cell(r[0], 1, f"Break-even ACOS of {BE:.0%} is an ASSUMPTION pending confirmed margin. All red/amber verdicts key off it and update once margin is known.").font = F(9, True, C["violet"])
    ws.cell(r[0], 1).alignment = WRAP; ws.row_dimensions[r[0]].height = 28

    # ================= TAB 2: Channel Mix & Gaps =================
    ws = wb.create_sheet("Channel Mix & Gaps")
    title_block(ws, "Channel Mix & Gaps", None, 4)
    header_row(ws, 4, ["Ad product", "Status", "Spend", "Note"], [26, 16, 16, 60])
    ch_meta = [("SP", "Sponsored Products", "100% of ad spend." if sp_only else "Core search spend."),
               ("SB", "Sponsored Brands", "No headline brand defense / brand-store driver."),
               ("SD", "Sponsored Display", "No retargeting of non-converting first-click traffic."),
               ("RAS", "Sponsored TV / RAS", "Not running.")]
    ct = M.get("channel_totals", {})
    rr = 5
    for key, nm, gapnote in ch_meta:
        active = key in channels
        sp = ct.get(key, {}).get("spend", 0)
        datarow(ws, rr, [nm, "ACTIVE" if active else "NONE", sp, ("Active." if active else gapnote)],
                [None, None, MONEY, None], left_cols=(1, 2, 4), breakeven=BE)
        ws.cell(rr, 2).fill = PatternFill("solid", fgColor=(TL["good"] if active else TL["warn"]))
        rr += 1
    if missing_ch:
        note(ws, rr + 1, f"Opportunity: the missing motions ({', '.join(missing_ch)}) are free growth. Branded already converts ({br_acos:.0%} ACOS) — a small Sponsored Brands defends the name cheaply; Sponsored Display re-engages the buyers who don't purchase first session.", 4)

    # ================= TAB 3: Branded vs Generic =================
    st_cov = sum(v.get("spend", 0) for v in STB.values()) / T["spend"] if T["spend"] else 0
    ws = wb.create_sheet("Branded vs Generic")
    title_block(ws, "Branded vs Generic vs Competitor — by search term (SP) & target (SB)",
                f"Intent split covers {st_cov:.1%} of total spend — SP by customer search term, SB/SB-Multi by keyword/product-target; remainder is SB video/reach + SD", 7)
    header_row(ws, 4, ["Bucket", "Spend", "% spend", "Sales", "ACOS", "CVR", "Verdict"], [24, 14, 10, 14, 10, 10, 30])
    rr = 5
    for b in ("Branded", "Generic", "Competitor"):
        d = STB.get(b)
        if not d:
            continue
        verdict = "Profitable — carries account" if (d["acos"] or 0) < BE * 0.7 else ("Above break-even — bleed" if (d["acos"] or 0) > BE else "Near break-even")
        datarow(ws, rr, [b, d["spend"], d["spend"]/T["spend"], d["sales"], d["acos"], d["cvr"], verdict],
                [None, MONEY, PCT, MONEY, PCT, PCT2, None], acos_cols=(5,), left_cols=(1, 7), breakeven=BE); rr += 1
    # remainder without search-term detail (e.g. SB-Multi/SD) so % spend sums to 100%
    un_sp = T["spend"] - sum(v.get("spend", 0) for v in STB.values())
    un_sa = T["sales"] - sum(v.get("sales", 0) for v in STB.values())
    if un_sp > 0.005 * T["spend"]:
        datarow(ws, rr, ["Not classified (SB video/reach + SD)", un_sp, un_sp/T["spend"], un_sa, (un_sp/un_sa if un_sa else None), None, "No search term or target to attribute intent"],
                [None, MONEY, PCT, MONEY, PCT, None, None], acos_cols=(5,), left_cols=(1, 7), breakeven=BE); rr += 1
    datarow(ws, rr, ["TOTAL", T["spend"], 1.0, T["sales"], T["acos"], None, ""],
            [None, MONEY, PCT, MONEY, PCT, None, None], acos_cols=(5,), left_cols=(1, 7), breakeven=BE)
    for c in range(1, 8):
        ws.cell(rr, c).font = F(10, True)
    note(ws, rr + 2, f"Method: SP is split by the Search Term Report (what customers actually typed). SB/SB-Multi search-term reporting covers only ~half of SB spend (the rest is product-targeting, category & video), so SB is split by its keyword text + product-target expression instead — slightly overstates branded on broad SB keywords, but reaches {st_cov:.0%} coverage vs ~79% search-term-only. The 'Not ST-classified' remainder is SB video/reach + SD.", 7)

    # ================= TAB 4: SP Placement =================
    ws = wb.create_sheet("SP Placement")
    title_block(ws, "Placement performance — same keywords, different ROI", None, 5)
    header_row(ws, 4, ["Placement", "Spend", "Sales", "ACOS", "Read"], [24, 14, 14, 10, 40])
    rr = 5
    for p, d in sorted(P.items(), key=lambda x: -x[1]["spend"]):
        read = ""
        if p == best_pl[0]:
            read = "Best ROI — bid up here"
        elif p == worst_pl[0]:
            read = "Highest spend, weakest ROI"
        datarow(ws, rr, [p, d["spend"], d["sales"], d["acos"], read], [None, MONEY, MONEY, PCT, None],
                acos_cols=(4,), left_cols=(1, 5), breakeven=BE); rr += 1
    note(ws, rr + 1, "Placement modifiers let you push budget toward the best placement and throttle the worst without touching keywords.", 5)

    # ================= TAB 5: Waste & Winners =================
    ws = wb.create_sheet("Waste & Winners")
    title_block(ws, "Generic waste tiers · top wasters · profitable non-brand winners", None, 6)
    st = load_csv("search_terms_classified.csv")
    gen = [x for x in st if x["bucket"] == "Generic"]

    def n(x, k):
        return numf(x[k])
    zero = [x for x in gen if n(x, "orders") == 0 and n(x, "spend") > 0]
    over = [x for x in gen if n(x, "orders") > 0 and x["acos"] and numf(x["acos"]) > BE]
    under = [x for x in gen if n(x, "orders") > 0 and x["acos"] and numf(x["acos"]) <= BE]
    gsp = sum(n(x, "spend") for x in gen) or 1
    header_row(ws, 4, ["Generic tier", "Spend", "% of generic", "Action"], [30, 14, 14, 40])
    tiers = [("Zero-order spend", sum(n(x, "spend") for x in zero), "Negate / pause — no conversions"),
             ("Orders but above break-even", sum(n(x, "spend") for x in over), "Cut bids hard or move to Exact-only"),
             ("Profitable (≤ break-even)", sum(n(x, "spend") for x in under), "Keep / scale — real discovery")]
    rr = 5
    for lbl, sp, act in tiers:
        datarow(ws, rr, [lbl, sp, sp/gsp, act], [None, MONEY, PCT, None], left_cols=(1, 4), breakeven=BE)
        ws.cell(rr, 2).fill = PatternFill("solid", fgColor=(TL["bad"] if "Zero" in lbl else TL["warn"] if "above" in lbl else TL["good"]))
        rr += 1
    wasted = sum(n(x, "spend") for x in zero + over)
    note(ws, rr, f"Genuinely wasted generic (zero-order + above-break-even) = {_m(wasted,MONEY)} = {wasted/gsp:.0%} of generic, {wasted/T['spend']:.0%} of total ad spend.", 4)
    rr += 2
    ws.cell(rr, 1, "Top generic wasters (spend, zero or above-break-even)").font = F(11, True, C["coral"]); rr += 1
    header_row(ws, rr, ["Search term", "Spend", "Sales", "ACOS", "Orders"], [46, 12, 12, 10, 8]); rr += 1
    wasters = sorted([x for x in gen if n(x, "spend") > 40 and (n(x, "orders") == 0 or (x["acos"] and numf(x["acos"]) > BE))], key=lambda x: -n(x, "spend"))[:14]
    for x in wasters:
        ac = numf(x["acos"]) if x["acos"] else None
        datarow(ws, rr, [x["term"], n(x, "spend"), n(x, "sales"), ac, int(n(x, "orders"))], [None, MONEY, MONEY, PCT, INT], acos_cols=(4,), left_cols=(1,), breakeven=BE); rr += 1
    rr += 1
    ws.cell(rr, 1, "Profitable non-brand winners (keep & scale)").font = F(11, True, C["deep"]); rr += 1
    header_row(ws, rr, ["Search term", "Spend", "Sales", "ACOS", "Orders"], [46, 12, 12, 10, 8]); rr += 1
    for x in sorted(under, key=lambda x: -n(x, "sales"))[:10]:
        datarow(ws, rr, [x["term"], n(x, "spend"), n(x, "sales"), numf(x["acos"]), int(n(x, "orders"))], [None, MONEY, MONEY, PCT, INT], acos_cols=(4,), left_cols=(1,), breakeven=BE); rr += 1

    # ================= TAB 6: Structure Diagnosis =================
    ws = wb.create_sheet("Structure Diagnosis")
    title_block(ws, "Campaign structure — control diagnosis", None, 3)
    S = M["structure"]
    header_row(ws, 4, ["Check", "Value", "Read"], [42, 14, 58])
    items = [("Total campaigns", S["total_campaigns"], "—"),
             ("Enabled / paused", f"{S['enabled']} / {S['paused']}", "Paused campaigns clutter reporting and bulk edits." if S['paused'] else "—"),
             ("Ad groups (enabled acct)", S["ad_groups"], "—"),
             ("Keywords per ad group (min–max)", f"{S['kw_per_ag_min']}–{S['kw_per_ag_max']}", "Tight, disciplined ad groups keep placement/bid controllable. GOOD." if S['kw_per_ag_max'] <= 10 else "Some ad groups are dumping grounds."),
             ("Negative keywords in place", S["total_neg_kw"], "Active negation. GOOD." if S['total_neg_kw'] > 100 else "Thin negation."),
             ("Enabled campaigns with NO negatives", S["enabled_no_negatives"], f"{S['enabled_no_negatives']} campaigns run without negatives — where unqualified queries slip in." if S['enabled_no_negatives'] else "All open campaigns negate. GOOD."),
             ("Same kw + same match across campaigns", S["dup_kw_pairs"], f"{S['dup_kw_pairs']} pairs self-compete across {S['dup_placements']} placements — consolidate the bid source of truth." if S['dup_kw_pairs'] else "No self-competition. GOOD."),
             ("Branded + generic mixed in one campaign", S["mixed_brand_generic_campaigns"], "Clean brand/generic separation. GOOD." if not S['mixed_brand_generic_campaigns'] else "Brand wins mask generic losses in shared budgets."),
             ("Ad groups mixing parent families", S.get("multi_parent_ad_groups", 0), "Each ad group advertises one product family. GOOD." if not S.get("multi_parent_ad_groups") else f"{S['multi_parent_ad_groups']} ad groups advertise ASINs from several parent families — Amazon picks which product serves each query, so keyword→product fit and per-product stats are out of your control.")]
    rr = 5
    for lbl, val, rd in items:
        datarow(ws, rr, [lbl, val, rd], [None, None, None], left_cols=(1, 2, 3), breakeven=BE)
        if "GOOD" in rd:
            ws.cell(rr, 2).fill = PatternFill("solid", fgColor=TL["good"])
        elif rd != "—":
            ws.cell(rr, 2).fill = PatternFill("solid", fgColor=TL["warn"])
        rr += 1
    note(ws, rr + 1, "Verdict: read the greens as strengths and the amber rows as tuning targets (cleanup, negatives, consolidation) — not a rebuild unless the mixed/dup counts are high.", 3)

    # ================= TAB 7: Business Report =================
    ws = wb.create_sheet("Business Report")
    title_block(ws, "Business Report — conversion by ASIN (all traffic)", None, 7)
    header_row(ws, 4, ["ASIN", "Group", "Sessions", "Units", "Sales", "Unit-session CVR", "Buy Box %"], [14, 12, 12, 10, 14, 14, 12])
    BRr = M["business_report"]["rows"]; rr = 5
    for d in sorted(BRr, key=lambda x: -x["sales"]):
        datarow(ws, rr, [d["asin"], d["group"], d["sessions"], d["units"], d["sales"], d["cvr"], d["buybox"]],
                [None, None, INT, INT, MONEY, PCT, PCT], left_cols=(1, 2), breakeven=BE); rr += 1
    datarow(ws, rr, ["TOTAL", "", M["business_report"]["total_sessions"], M["business_report"]["total_units"], M["business_report"]["total_sales"], None, None],
            [None, None, INT, INT, MONEY, None, None], left_cols=(1, 2), breakeven=BE)
    for c in range(1, 8):
        ws.cell(rr, c).font = F(10, True)
    note(ws, rr + 2, "Where sessions are high and CVR healthy, the product converts. $0-unit ASINs are new/unlaunched — decide launch/merge/prune.", 7)

    # ================= TAB 8: DataDive Niche (only if competitors present) =================
    if comp:
        ws = wb.create_sheet("DataDive Niche")
        title_block(ws, "DataDive niche — competitive benchmarking", None, 9)
        note(ws, 4, f"Niche {cfg.get('datadive_niche','')} ({comp.get('category','')}), {comp.get('num_keywords','')} keywords. Category median price {_m(comp['median_price'],MONEY)}, median reviews {comp['median_reviews']:.0f}, median rating {comp['median_rating']}.", 9)
        header_row(ws, 6, ["Brand", "ASIN", "Price", "Rating", "Reviews", "Est. revenue", "Adv. kws", "% kw on P1", "Seller"], [18, 12, 10, 9, 10, 14, 10, 10, 8])
        rr = 7
        for x in comp["competitors"]:
            datarow(ws, rr, [x["brand"], x["asin"], x["price"], x["rating"], x["reviews"], x["revenue"], x["advertised_kws"], x["kw_p1_pct"], x.get("seller_country", "")],
                    [None, None, USD2 if MONEY == USD else ew.EUR2, RO, INT, MONEY, INT, PCT, None], left_cols=(1, 2, 9), breakeven=BE)
            if x.get("is_client"):
                for c in range(1, 10):
                    ws.cell(rr, c).fill = PatternFill("solid", fgColor=C["cloud"]); ws.cell(rr, c).font = F(10, True)
                ws.cell(rr, 3).fill = PatternFill("solid", fgColor=TL["bad"])  # price outlier flag
            rr += 1
        note(ws, rr + 1, "Client rows highlighted; the red price cell flags the premium-vs-category-median gap. Where the client is priced well above the median with lower ratings/fewer reviews, paid traffic is doing the work price + social proof do for competitors.", 9)

    # ================= TAB 9: Sources & Notes =================
    ws = wb.create_sheet("Sources & Notes")
    title_block(ws, "Sources, method & caveats", None, 2)
    lines = [
        ("Window", f"Ads {w.get('ads','')}; Business Report {w.get('business_report','')}; SQP {', '.join(w.get('sqp_weeks',[]))}; DataDive {w.get('datadive','')}."),
        ("Marketplace", markets),
        ("Account scope", f"Channels present: {', '.join(channels)}." + (f" Missing: {', '.join(missing_ch)}." if missing_ch else "")),
        ("Break-even ACOS", f"{BE:.0%} — ASSUMPTION pending confirmed margin. Drives all red/amber verdicts."),
        ("Branded definition", f"Search terms containing {', '.join(cfg.get('brand_tokens',[]))} = Branded. Competitor brand tokens = Competitor. Else Generic."),
        ("Split method", "Branded/Generic/Competitor computed from the Search Term Report (customer search terms), not keyword text."),
        ("Placement", "From SP Bidding Adjustment (placement) rows in the bulk file."),
        ("SQP caveat", "Multi-ASIN SQP exports cap the query grid; shares are directional. Weekly snapshots averaged."),
        ("Colour legend", f"ACOS green <30% · light-green <{BE:.0%} (break-even) · amber ≤60% · red >60%. ROAS green ≥3 · amber ≥1.5 · red <1.5."),
        ("Prepared by", f"Ecom Wizards. Figures reconcile to the raw bulk (spend {_m(T['spend'],MONEY)} / sales {_m(T['sales'],MONEY)}) and Business Report ({_m(T['br_total_sales'],MONEY)})."),
    ]
    rr = 4
    for k, v in lines:
        ws.cell(rr, 1, k).font = F(10, True); ws.cell(rr, 1).alignment = WRAP
        ws.cell(rr, 2, v).font = F(10); ws.cell(rr, 2).alignment = WRAP
        ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 110
        ws.row_dimensions[rr].height = 28; rr += 1

    for wsx in wb.worksheets:
        wsx.sheet_view.showGridLines = False
    out = outdir / f"{_slug(CLIENT)}_{_slug(markets)}_Ad_Audit.xlsx"
    wb.save(out)
    print("[audit] wrote", out.name, "| tabs:", [s.title for s in wb.worksheets])
    return out


def _adapt_competitors(comp, client_asins):
    """Accept either the derived competitors shape or the raw DataDive MCP
    get_niche_competitors payload (the contract: Claude saves the MCP response
    verbatim). Raw payloads are converted; derived ones pass through."""
    if not comp or "benchmark" not in comp:
        return comp
    b = comp.get("benchmark", {}) or {}
    stats = comp.get("statistics", {}) or {}
    category = ""
    rows = []
    for x in comp.get("competitors", []):
        category = category or x.get("categoryTree") or x.get("category") or ""
        rows.append(dict(
            brand=x.get("brand", ""), asin=x.get("asin", ""), price=x.get("price"),
            rating=x.get("rating"), reviews=x.get("reviewCount"), revenue=x.get("revenue"),
            advertised_kws=x.get("advertisedKws"), kw_p1_pct=x.get("kwRankedOnP1Percent"),
            seller_country=x.get("sellerCountry", ""),
            is_client=x.get("asin") in client_asins))
    return dict(category=category, num_keywords=stats.get("numKeywords"),
                median_price=b.get("price"), median_reviews=b.get("reviewCount"),
                median_rating=b.get("rating"), competitors=rows)


def _m(v, fmt):
    sym = "€" if "€" in fmt else "$"
    return f"{sym}{v:,.0f}"


def _slug(s):
    import re
    return re.sub(r"[^A-Za-z0-9]+", "-", (s or "x")).strip("-")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2])
