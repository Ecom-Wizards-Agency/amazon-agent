import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


TOOL_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(TOOL_DIR))

from analyze_audit import (  # noqa: E402
    assess_ads_business_window_alignment,
    blended_totals,
)
from build_audit_workbook import build as build_workbook  # noqa: E402
from narrative_scaffold import build as build_narrative  # noqa: E402


def fixture_metrics(status):
    windows = {
        "ads": "2026-01-01..2026-01-28",
        "business_report": "2026-01-01..2026-01-28",
        "sqp_weeks": [],
    }
    if status == "mismatched":
        windows["ads"] = "Snapshot 2026-01-20..2026-01-21 only"
    elif status == "unknown":
        windows["ads"] = "latest available snapshot"
    alignment = assess_ads_business_window_alignment(windows)
    blended = blended_totals(100.0, 250.0, 500.0, alignment)
    return {
        "client": "Fixture",
        "currency": "USD",
        "marketplaces": ["US"],
        "breakeven": 0.5,
        "ads_snapshot_directional": status != "matched",
        "channels_present": ["SP"],
        "channel_totals": {"SP": {"spend": 100.0, "sales": 250.0}},
        "windows": windows,
        "window_alignment": alignment,
        "totals": {
            "spend": 100.0,
            "sales": 250.0,
            "br_total_sales": 500.0,
            "acos": 0.4,
            "roas": 2.5,
            **blended,
        },
        "searchterm_bucket": {
            "Branded": {"spend": 40.0, "sales": 120.0, "acos": 1 / 3, "cvr": 0.2},
            "Generic": {"spend": 60.0, "sales": 130.0, "acos": 60 / 130, "cvr": 0.1},
        },
        "placement": {
            "Top of Search": {"spend": 60.0, "sales": 200.0, "acos": 0.3},
            "Product Pages": {"spend": 40.0, "sales": 50.0, "acos": 0.8},
        },
        "structure": {
            "total_campaigns": 2,
            "enabled": 2,
            "paused": 0,
            "ad_groups": 2,
            "kw_per_ag_min": 2,
            "kw_per_ag_max": 4,
            "total_neg_kw": 120,
            "enabled_no_negatives": 0,
            "dup_kw_pairs": 0,
            "dup_placements": 0,
            "mixed_brand_generic_campaigns": 0,
            "multi_parent_ad_groups": 0,
        },
        "business_report": {
            "rows": [{
                "asin": "B0FIXTURE0",
                "group": "Core",
                "sessions": 100,
                "units": 20,
                "sales": 500.0,
                "cvr": 0.2,
                "buybox": 0.95,
            }],
            "total_sessions": 100,
            "total_units": 20,
            "total_sales": 500.0,
        },
    }


class WindowAlignmentAnalysisTest(unittest.TestCase):
    def test_matched_mismatched_unknown_and_disrupted_control(self):
        cases = [
            ({"ads": "2026-01-01..2026-01-28", "business_report": "2026-01-01..2026-01-28"}, "matched"),
            ({"ads": "2026-01-20..2026-01-21", "business_report": "2026-01-01..2026-01-28"}, "mismatched"),
            ({"ads": "latest available", "business_report": "2026-01-01..2026-01-28"}, "unknown"),
            ({
                "ads": "Amazon Ads snapshot 2026-06-18..2026-06-19 only",
                "business_report": "Online control 2026-05-31..2026-06-27",
            }, "mismatched"),
        ]
        for windows, expected in cases:
            with self.subTest(expected=expected, windows=windows):
                result = assess_ads_business_window_alignment(windows)
                self.assertEqual(expected, result["ads_vs_business_report"])

    def test_blended_values_are_nullable_unless_windows_match(self):
        for status in ("mismatched", "unknown"):
            totals = fixture_metrics(status)["totals"]
            for key in ("tacos", "organic_implied", "ad_dependency", "ad_attributed_share"):
                self.assertIsNone(totals[key])
        matched = fixture_metrics("matched")["totals"]
        self.assertEqual(0.2, matched["tacos"])
        self.assertEqual(250.0, matched["organic_implied"])
        self.assertEqual(0.5, matched["ad_attributed_share"])


class WindowAlignmentOutputTest(unittest.TestCase):
    def _write_fixture(self, root, status):
        cfg = {
            "client": "Fixture",
            "marketplaces": ["US"],
            "brand_tokens": ["fixture"],
            "competitor_tokens": ["rival"],
            "inputs": {},
            "ads_snapshot_directional": status != "matched",
            "narrative": {"include_levers": False},
        }
        config_path = root / "config.json"
        config_path.write_text(json.dumps(cfg))
        (root / "metrics.json").write_text(json.dumps(fixture_metrics(status)))
        (root / "clean").mkdir()
        return config_path

    def test_narrative_renders_n_a_and_reason_for_nonmatching_windows(self):
        for status in ("mismatched", "unknown"):
            with self.subTest(status=status), tempfile.TemporaryDirectory() as tmp:
                root = Path(tmp)
                config_path = self._write_fixture(root, status)
                rendered = build_narrative(config_path, root, force=True).read_text()
                self.assertIn("| TACOS | N/A |", rendered)
                self.assertIn("| Ad-attributed share | N/A |", rendered)
                self.assertIn("Blended Ads and Business Report KPIs are not reported", rendered)
                self.assertNotIn("| **TACOS** | **20.0%** |", rendered)

    def test_narrative_keeps_blended_values_for_matching_windows(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config_path = self._write_fixture(root, "matched")
            rendered = build_narrative(config_path, root, force=True).read_text()
            self.assertIn("| **TACOS** | **20.0%** |", rendered)
            self.assertIn("| Ad-attributed share | 50.0% |", rendered)
            self.assertNotIn("| TACOS | N/A |", rendered)

    def test_workbook_renders_n_a_without_dropping_source_sales(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            config_path = self._write_fixture(root, "mismatched")
            path = build_workbook(config_path, root)
            wb = load_workbook(path, data_only=False)
            ws = wb["Executive Summary"]
            rows = {ws.cell(row, 1).value: (ws.cell(row, 2).value, ws.cell(row, 3).value)
                    for row in range(1, 20)}
            self.assertEqual(500.0, rows["Business Report sales"][0])
            self.assertEqual("N/A", rows["Organic / non-ad sales (implied)"][0])
            self.assertEqual("N/A", rows["TACOS"][0])
            self.assertEqual("N/A", rows["Ad-attributed share of sales"][0])
            self.assertIn("2026-01-20..2026-01-21", rows["TACOS"][1])
            source_rows = {ws2.cell(row, 1).value: ws2.cell(row, 2).value
                           for ws2 in [wb["Sources & Notes"]] for row in range(1, 20)}
            self.assertIn("Do not calculate TACOS", source_rows["Window alignment"])


if __name__ == "__main__":
    unittest.main()
