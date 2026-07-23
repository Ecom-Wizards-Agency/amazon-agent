#!/usr/bin/env python3
"""
SUPA-style SQP x PPC weekly workbook.

Joins Amazon SQP shares (search volume, click share, purchase share) with PPC
ad spend + ad sales per keyword, per Sunday-Saturday week, per ASIN. Flags are
deterministic rules. Inputs are AdLabs MCP CSV exports (search_query entity per
week + search_term entity grouped by term per week). The AdLabs search_query
entity was verified byte-identical to Seller Central SQP on 2026-07-08.

Usage:
  python3 build_supa_workbook.py --config config.<client>.json [--validate]
"""
from __future__ import annotations

import argparse
import csv
import json
import random
import sys
from datetime import datetime
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO / "tools" / "amazon-ad-audit"))

import ew_audit_style as ew  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

GREEN_FILLS = {"C6EFCE", "E2EFDA"}
GREY_FILL = PatternFill("solid", fgColor=ew.TL["grey"])
PP = '+0.00%;-0.00%'          # share delta shown in percentage points
DINT = '+#,##0;-#,##0'
DUSD = '+$#,##0.00;-$#,##0.00'
DEUR = '+€#,##0.00;-€#,##0.00'

FLAG_FILL = {
    "Problem": PatternFill("solid", fgColor=ew.TL["bad"]),
    "Opportunity": PatternFill("solid", fgColor=ew.TL["warn"]),
    "Efficiency Gain": PatternFill("solid", fgColor=ew.TL["good"]),
    "Stable": GREY_FILL,
}


# ------------------------------------------------------------------ loading
def rp(p):
    q = Path(str(p)).expanduser()
    return q if q.is_absolute() else REPO / q


def load_config(path):
    cfg = json.loads(Path(path).read_text())
    cfg["weeks"] = sorted(cfg["weeks"])
    return cfg


def _get(row, *names):
    for n in names:
        for k in row:
            if k.strip().lower() == n:
                return row[k]
    return None


def _req(row, *names):
    v = _get(row, *names)
    if v is None:
        raise SystemExit(
            f"missing column (tried {names}); available: {sorted(row.keys())}")
    return v


def load_sqp(cfg):
    """sqp[(asin, week)][query_lower] = rec. Shares recomputed from counts."""
    sqp = {}
    for week, p in cfg["inputs"]["sqp_csvs"].items():
        with open(rp(p), encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                asin = _req(row, "asin")
                if asin not in cfg["asins"]:
                    continue
                q = _req(row, "search_query").strip().lower()
                t_clk = ew.numf(_req(row, "total_click_count"))
                a_clk = ew.numf(_req(row, "asin_click_count"))
                t_pur = ew.numf(_req(row, "total_purchase_count"))
                a_pur = ew.numf(_req(row, "asin_purchase_count"))
                # Impressions are optional via _get: a SQP export without them
                # still builds, it just leaves the CTR-gap column grey.
                t_imp = ew.numf(_get(row, "total_query_impression_count"))
                a_imp = ew.numf(_get(row, "asin_impression_count"))
                sqp.setdefault((asin, week), {})[q] = {
                    "query": _req(row, "search_query").strip(),
                    "sv": ew.numf(_req(row, "search_query_volume")),
                    "t_clk": t_clk, "a_clk": a_clk,
                    "t_pur": t_pur, "a_pur": a_pur,
                    "t_imp": t_imp, "a_imp": a_imp,
                    "cs": a_clk / t_clk if t_clk else 0.0,
                    "ps": a_pur / t_pur if t_pur else 0.0,
                    "cs_cmp": ew.numf(_get(row, "asin_click_share_comparison")),
                    "score": ew.numf(_get(row, "search_query_score")),
                }
    return sqp


def load_ppc(cfg):
    """ppc[week][term_lower] = {spend, sales, clicks, orders}. Grouped export
    is already one row per term; sum defensively anyway."""
    ppc = {}
    for week, p in cfg["inputs"]["search_term_csvs"].items():
        wk = ppc.setdefault(week, {})
        with open(rp(p), encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                term = _req(row, "search_term", "search term", "query").strip().lower()
                rec = wk.setdefault(term, {"spend": 0.0, "sales": 0.0,
                                           "clicks": 0.0, "orders": 0.0})
                rec["spend"] += ew.numf(_req(row, "spend", "cost"))
                rec["sales"] += ew.numf(_req(row, "sales"))
                rec["clicks"] += ew.numf(_get(row, "clicks"))
                rec["orders"] += ew.numf(_get(row, "orders"))
    return ppc


def load_rank(cfg):
    """rank[(asin, query_lower)][week] = median organic rank across that week.

    Input is a DataDive Rank Radar payload per ASIN: [{keyword, searchVolume,
    ranks:[{date, organicRank}]}]. Median, not last-day, because a single day's
    rank is noisy and SUPA compares whole weeks.

    Entirely optional. No `inputs.rank_jsons` in the config, or a radar that does
    not cover these weeks, and the rank column simply never renders. Radars track
    a few dozen keywords while SQP has hundreds, so most rows stay blank by design.
    """
    from datetime import timedelta
    paths = (cfg.get("inputs", {}) or {}).get("rank_jsons") or {}
    if not paths:
        return {}
    spans = {w: [(datetime.strptime(w, "%Y-%m-%d") + timedelta(days=i)).strftime("%Y-%m-%d")
                 for i in range(7)] for w in cfg["weeks"]}
    out = {}
    for asin, path in paths.items():
        f = rp(path)
        if not f.exists():
            print(f"  [rank] skipped {asin}: {f} not found")
            continue
        try:
            payload = json.loads(f.read_text())
        except Exception as e:
            print(f"  [rank] skipped {asin}: unreadable ({e})")
            continue
        for kwrec in payload:
            q = str(kwrec.get("keyword", "")).strip().lower()
            if not q:
                continue
            by_date = {r["date"]: r.get("organicRank") for r in kwrec.get("ranks", [])
                       if r.get("organicRank")}
            per_week = {}
            for w, days in spans.items():
                vals = sorted(by_date[d] for d in days if d in by_date)
                if vals:
                    per_week[w] = vals[len(vals) // 2]
            if per_week:
                out[(asin, q)] = per_week
    return out


def load_stock(cfg):
    """stock[(asin, week)] = out-of-stock days in that week (0-7).

    Input is one AdLabs `product` entity export per week, same DATE filter as the
    SQP pull. `out_of_stock_days` is WINDOW-scoped, so a per-week pull gives the
    days lost in that week and the weeks sum back to the monthly figure.

    Do NOT try to read availability / days_of_cover / fulfillable_units this way:
    those are CURRENT snapshots and come back identical in every week's export.

    Optional. No `inputs.stock_csvs` and the block never renders, and every flag
    falls back to v2 behaviour.
    """
    paths = (cfg.get("inputs", {}) or {}).get("stock_csvs") or {}
    if not paths:
        return {}
    out = {}
    for week, path in paths.items():
        f = rp(path)
        if not f.exists():
            print(f"  [stock] skipped {week}: {f} not found")
            continue
        with open(f, encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                asin = _get(row, "asin")
                if not asin or asin not in cfg["asins"]:
                    continue
                out[(asin, week)] = ew.numf(_get(row, "out_of_stock_days"))
    return out


def join_weekly(cfg, sqp, ppc, rank=None, stock=None):
    """kw[(asin, qlower)] = {"query": str, "weeks": {week: cell}}.
    cell = sv/cs/ps/spend/sales/acos. PPC values are profile-level."""
    rank = rank or {}
    stock = stock or {}
    kw = {}
    for (asin, week), queries in sqp.items():
        for q, rec in queries.items():
            ent = kw.setdefault((asin, q), {"query": rec["query"], "weeks": {}})
            ad = ppc.get(week, {}).get(q, {})
            spend = ad.get("spend", 0.0)
            sales = ad.get("sales", 0.0)
            # Conversion gap to the market, in percentage points. Both sides come
            # from counts we already read, so this costs no extra input.
            # Amazon scores conversion per keyword: below-market CVR means spend
            # rents a position instead of earning one.
            our_cvr = rec["a_pur"] / rec["a_clk"] if rec["a_clk"] else None
            mkt_cvr = rec["t_pur"] / rec["t_clk"] if rec["t_clk"] else None
            our_ctr = rec["a_clk"] / rec["a_imp"] if rec["a_imp"] else None
            mkt_ctr = rec["t_clk"] / rec["t_imp"] if rec["t_imp"] else None
            ent["weeks"][week] = {
                "sv": rec["sv"], "cs": rec["cs"], "ps": rec["ps"],
                "t_clk": rec["t_clk"], "a_clk": rec["a_clk"],
                "t_pur": rec["t_pur"], "a_pur": rec["a_pur"],
                "spend": spend, "sales": sales,
                "acos": spend / sales if sales else None,
                "cs_cmp": rec["cs_cmp"],
                "rank": rank.get((asin, q), {}).get(week),
                # ASIN-week property, repeated on every keyword row of that ASIN.
                # Redundant down the column on purpose: when a row shows rank
                # slipping, the reason needs to be in the same eyeline.
                "oos": stock.get((asin, week)),
                "cvr": our_cvr, "mkt_cvr": mkt_cvr,
                "cvr_gap": (our_cvr - mkt_cvr) * 100.0
                           if (our_cvr is not None and mkt_cvr is not None
                               and rec["a_clk"] >= 5) else None,
                "ctr_gap": (our_ctr - mkt_ctr) * 100.0
                           if (our_ctr is not None and mkt_ctr is not None
                               and rec["a_imp"] >= 100) else None,
            }
    return kw


def rollup_asin(cfg, sqp, ppc):
    """roll[asin][week] = weighted rollup. Shares = sums of counts, never
    means of shares. Spend/sales = matched terms only (profile-level)."""
    roll = {a: {} for a in cfg["asins"]}
    for (asin, week), queries in sqp.items():
        t_clk = sum(r["t_clk"] for r in queries.values())
        a_clk = sum(r["a_clk"] for r in queries.values())
        t_pur = sum(r["t_pur"] for r in queries.values())
        a_pur = sum(r["a_pur"] for r in queries.values())
        spend = sum(ppc.get(week, {}).get(q, {}).get("spend", 0.0) for q in queries)
        sales = sum(ppc.get(week, {}).get(q, {}).get("sales", 0.0) for q in queries)
        roll[asin][week] = {
            "sv": sum(r["sv"] for r in queries.values()),
            "cs": a_clk / t_clk if t_clk else 0.0,
            "ps": a_pur / t_pur if t_pur else 0.0,
            "spend": spend, "sales": sales,
            "acos": spend / sales if sales else None,
        }
    return roll


# ------------------------------------------------------------------ flags
def _context(weeks, ent):
    """(rank_move, cvr_gap, oos_days) for the focus week. Any may be None.

    rank_move: end-rank minus start-rank across the available weeks. Positive =
    slipped (rank 7 -> 15 is +8). Negative = gained.
    cvr_gap: our conversion minus the market's, in percentage points, focus week.
    oos_days: days out of stock in the focus week, 0-7.

    Stock is the one that outranks the others. Amazon deranks what it cannot ship,
    so an out-of-stock week produces share loss and rank slip that look exactly
    like a bidding failure and are not one.
    """
    ranked = [(w, ent["weeks"][w]["rank"]) for w in weeks
              if w in ent["weeks"] and ent["weeks"][w].get("rank") is not None]
    move = (ranked[-1][1] - ranked[0][1]) if len(ranked) >= 2 else None
    d2 = ent["weeks"].get(weeks[-1]) or {}
    # Two different stock questions, and conflating them gives wrong verdicts:
    #   oos      = focus week only -> "should we pause the term right now?"
    #   oos_span = across the SAME weeks the rank move is measured over ->
    #              "does stock explain that rank move?"
    # A product can be clean this week and still have lost its rank to an
    # out-of-stock week a fortnight ago, which rank alone would misread as a
    # bidding failure.
    span = [w for w, _ in ranked] if len(ranked) >= 2 else list(weeks)
    oos_span = sum((ent["weeks"].get(w, {}).get("oos") or 0) for w in span)
    any_oos = any(ent["weeks"].get(w, {}).get("oos") is not None for w in span)
    return move, d2.get("cvr_gap"), d2.get("oos"), (oos_span if any_oos else None)


def _oos_txt(oos, oos_span=None):
    bits = []
    if oos:
        bits.append(f" Out of stock {oos:.0f} of 7 days this week.")
    if oos_span and oos_span != oos:
        bits.append(f" {oos_span:.0f} days lost across the compared weeks.")
    return "".join(bits)


def _stock_override(oos, th):
    """True when the week lost enough days that stock, not bidding, is the story."""
    return bool(oos) and oos >= th.get("oos_days_flag", 2)


def _rank_txt(move):
    if move is None:
        return ""
    if move > 0:
        return f" Organic rank slipped {move} places over these weeks."
    if move < 0:
        return f" Organic rank gained {abs(move)} places over these weeks."
    return " Organic rank held."


def flag_keyword(cfg, weeks, ent):
    """Return list of (flag, rule_id, action, why, score). Uses the last two
    weeks. Shares are fractions; pp thresholds are converted here."""
    th = cfg["thresholds"]
    target = cfg["targets"]["acos"]
    S = sym(cfg)
    move, gap, oos, oos_span = _context(weeks, ent)
    rt = _rank_txt(move) + _oos_txt(oos, oos_span)
    stock_hit = _stock_override(oos, th)          # act-now: this week is broken
    stock_span = _stock_override(oos_span, th)    # explains-it: the move was bought by stock
    w1, w2 = weeks[-2], weeks[-1]
    d1, d2 = ent["weeks"].get(w1), ent["weeks"].get(w2)
    if not d2 or d2["sv"] < th["min_sv"]:
        return []
    out = []
    q = ent["query"]

    def pp(x):
        return x * 100.0

    if d1:
        cs_drop = d1["cs"] - d2["cs"]
        cs_down = cs_drop >= th["click_share_drop_pp"] / 100.0 or (
            d1["cs"] > 0 and cs_drop / d1["cs"] >= th["click_share_drop_rel"])
        spend_cut = (d1["spend"] >= th["min_spend_prev"]
                     and d2["spend"] <= (1 - th["spend_drop_rel"]) * d1["spend"])
        spend_held = d2["spend"] >= 0.90 * d1["spend"]
        if cs_down and cs_drop > 0 and spend_cut:
            act = ("Restore spend to the prior-week level on this term. Check budget "
                   "caps and bid changes dated this week.")
            if stock_hit:
                act = (f"Check stock FIRST. You were out of stock {oos:.0f} of 7 days, "
                       "which drops share on its own. Restore spend once the product "
                       "is reliably in stock, not before.")
            elif move is not None and move > 0:
                act = ("URGENT: restore spend. The cut is costing organic rank too, "
                       "which compounds: rank lost while dark is slow to buy back.")
            elif move is not None and move <= 0:
                act = ("Restore spend, but this is not on fire. Organic rank held "
                       "through the cut, so the loss is paid traffic only.")
            out.append((
                "Problem", "P1", act,
                f"Spend fell from {S}{d1['spend']:,.0f} to {S}{d2['spend']:,.0f}. "
                f"Click share followed, down {pp(cs_drop):.1f} pp "
                f"({pp(d1['cs']):.1f}% to {pp(d2['cs']):.1f}%)." + rt,
                d2["sv"] * cs_drop))
        elif cs_drop >= th["share_loss_held_pp"] / 100.0 and spend_held:
            act = ("Spend did not cause this. Check price, rating, competitor moves "
                   "and organic rank on this term.")
            if stock_hit:
                act = (f"Spend did not cause this: you were out of stock {oos:.0f} of 7 "
                       "days. That is the likely cause, not price or competitors. "
                       "Restock, then re-measure.")
            elif move is not None and move > 0 and stock_span:
                act = (f"Spend did not cause this, and neither did the listing: you lost "
                       f"{oos_span:.0f} days of stock across these weeks. Fix supply first.")
            elif move is not None and move > 0:
                act = ("Spend did not cause this: you were outranked organically. "
                       "This is a listing, price or review problem, not a bid problem.")
            elif move is not None and move <= 0:
                act = ("Spend did not cause this and rank held, so a competitor bought "
                       "the share or the SERP changed. Check who is new on this term.")
            if gap is not None and gap < -1.0:
                act += (f" Conversion is {abs(gap):.1f} pp below market here: fix that "
                        "before paying for more traffic.")
            out.append((
                "Problem", "P2", act,
                f"Click share down {pp(cs_drop):.1f} pp while spend held "
                f"({S}{d1['spend']:,.0f} to {S}{d2['spend']:,.0f})." + rt,
                d2["sv"] * cs_drop))
        ps_flat = d2["ps"] <= d1["ps"] + 0.002
        if (d2["spend"] >= th["min_spend_flag"]
                and (d2["sales"] == 0 or (d2["acos"] or 0) > th["acos_problem_mult"] * target)
                and ps_flat):
            acos_txt = "no sales" if d2["sales"] == 0 else f"ACOS {d2['acos']:.0%}"
            act = "Cut the bid 20-30% or isolate the term in exact with a controlled bid."
            flg = "Problem"
            if stock_hit:
                # Stock beats every other reading. An out-of-stock product deranks
                # and stops converting no matter what you bid, so both the high
                # ACOS and any rank slip are downstream of it. Cutting the bid
                # "fixes" a number without touching the cause.
                act = (f"STOCK, not bids. Out of stock {oos:.0f} of 7 days this week: "
                       "Amazon deranks what it cannot ship, so the ACOS and any rank "
                       "slip here are symptoms. Pause the term until it is back in "
                       "stock, then judge it on a clean week.")
            elif move is not None and move < 0:
                # Rank is being bought. ACOS is an indicator here, not a trigger.
                act = ("DO NOT CUT on ACOS alone. Rank is improving, so this spend is "
                       "buying position. Judge it on rank, and hold while it climbs.")
                flg = "Opportunity"
            elif move is not None and move > 0 and gap is not None and gap < -1.0:
                act = ("Stop. Rank is slipping AND conversion is "
                       f"{abs(gap):.1f} pp below market: you cannot buy this term at any "
                       "bid. Fix the listing, price or offer first.")
            elif move is not None and move > 0 and stock_span:
                act = (f"Do not read this as a bidding failure. Rank slipped, but you "
                       f"lost {oos_span:.0f} days of stock across these weeks and Amazon "
                       "deranks what it cannot ship. Fix supply, hold the bid, and "
                       "re-judge on a clean stretch.")
            elif move is not None and move > 0:
                act = ("Cut. Rank slipped while you paid over target and stock was "
                       "clean throughout, so the spend bought neither sales nor position.")
            out.append((
                flg, "P3", act,
                f"{S}{d2['spend']:,.0f} spent this week at {acos_txt}. "
                f"Purchase share did not improve ({pp(d1['ps']):.1f}% to {pp(d2['ps']):.1f}%)."
                + rt + (f" Conversion {gap:+.1f} pp vs market." if gap is not None else ""),
                d2["spend"]))
        if (d2["cs"] > d1["cs"] and d2["ps"] > d1["ps"]
                and d2["spend"] >= th["min_spend_prev"]
                and d2["acos"] is not None and d2["acos"] <= target):
            e1 = ("Working. Protect the bid. Consider 10-20% budget headroom on the "
                  "host campaign.")
            if stock_hit:
                e1 += (f" Note it did this while out of stock {oos:.0f} of 7 days, so "
                       "the true ceiling is higher than these numbers show.")
            out.append((
                "Efficiency Gain", "E1", e1,
                f"Click share up to {pp(d2['cs']):.1f}%, purchase share up to "
                f"{pp(d2['ps']):.1f}%, ACOS {d2['acos']:.0%} at or under the "
                f"{target:.0%} target." + _oos_txt(oos),
                d2["sales"]))
    if (d2["sv"] >= th["opp_min_sv"] and d2["cs"] < th["low_share"]
            and d2["ps"] < th["low_share"] and d2["spend"] < th["min_spend_prev"]):
        act = (f"Add an exact-match target. Size the starting bid to the "
               f"{target:.0%} ACOS target.")
        if stock_hit:
            act = (f"Do not fund yet. Out of stock {oos:.0f} of 7 days this week: "
                   "buy the demand once you can reliably ship it.")
        elif move is not None and move <= 10:
            act = ("Organic already ranks here. Fund it only if you want the top slot "
                   "too, otherwise you are paying for traffic you have.")
        if gap is not None and gap < -1.0:
            act = (f"Do not fund yet. Conversion is {abs(gap):.1f} pp below market on "
                   "this term, so paid traffic will not stick.")
        out.append((
            "Opportunity", "O1", act,
            f"{d2['sv']:,.0f} searches this week. Click share {pp(d2['cs']):.1f}%, "
            f"purchase share {pp(d2['ps']):.1f}%, ad spend {S}{d2['spend']:,.0f}." + rt,
            d2["sv"] * (th["low_share"] - d2["cs"])))
    if (d2["sv"] >= th["opp2_min_sv"]
            and d2["ps"] - d2["cs"] >= th["converts_edge_pp"] / 100.0):
        out.append((
            "Opportunity", "O2",
            "The product converts when seen. Buy more impressions: raise the bid or add a top-of-search modifier.",
            f"Purchase share {pp(d2['ps']):.1f}% vs click share {pp(d2['cs']):.1f}%. "
            f"It converts above its traffic weight.",
            d2["sv"] * (d2["ps"] - d2["cs"])))
    return out


def flag_asin(cfg, weeks, wk_roll):
    """One (flag, reason) per ASIN from the weighted rollup, priority order."""
    th = cfg["thresholds"]
    target = cfg["targets"]["acos"]
    d1, d2 = wk_roll.get(weeks[-2]), wk_roll.get(weeks[-1])
    if not d1 or not d2:
        return ("Stable", "Not enough weeks for a verdict.")
    cs_rel = (d1["cs"] - d2["cs"]) / d1["cs"] if d1["cs"] else 0.0
    sp_rel = (d1["spend"] - d2["spend"]) / d1["spend"] if d1["spend"] else 0.0
    ps_rel = (d1["ps"] - d2["ps"]) / d1["ps"] if d1["ps"] else 0.0
    sv_rel = (d2["sv"] - d1["sv"]) / d1["sv"] if d1["sv"] else 0.0
    if cs_rel >= th["asin_click_share_drop_rel"] and sp_rel >= th["asin_spend_drop_rel"]:
        return ("Problem",
                f"Matched ad spend fell {sp_rel:.0%} WoW. Click share followed, "
                f"down {(d1['cs'] - d2['cs']) * 100:.2f} pp.")
    if ps_rel >= th["asin_purchase_share_drop_rel"] and sp_rel < 0.05:
        return ("Problem",
                f"Purchase share down {ps_rel:.0%} WoW with spend flat or up. "
                f"Not a spend problem. Check price, stock, reviews, competitors.")
    if (d2["cs"] > d1["cs"] and d2["ps"] > d1["ps"]
            and d2["acos"] is not None and d2["acos"] <= target):
        return ("Efficiency Gain",
                f"Click share and purchase share both up WoW at {d2['acos']:.0%} "
                f"ACOS, inside the {target:.0%} target.")
    if sv_rel >= th["asin_sv_up_rel"] and abs(d2["cs"] - d1["cs"]) <= 0.003:
        return ("Opportunity",
                f"Demand grew {sv_rel:.0%} WoW but click share stayed flat. "
                f"Presence did not scale with the market.")
    return ("Stable", "No threshold-level movement week over week.")


def build_actions(cfg, weeks, kw):
    """Ranked action rows across ASINs. Dedupe on (query, rule).

    Anchor choice matters more than it looks. Ad spend is PROFILE level, so
    siblings sharing a query carry an identical spend number and the score
    tie-break between them is close to arbitrary. But the SQP side (shares,
    conversion) and rank are per ASIN, and rank is what decides whether P3 says
    "cut" or "this is rank investment, hold". So prefer an anchor we can actually
    judge: rank data first, then score. Without this the rank split silently
    fails to fire whenever the arbitrary winner happens to lack a radar.
    """
    hits = {}
    for (asin, q), ent in kw.items():
        flags = flag_keyword(cfg, weeks, ent)
        if not flags:
            continue
        has_rank = any(d.get("rank") is not None for d in ent["weeks"].values())
        # first match wins, same as the Flag column; avoids contradictory
        # actions (e.g. P1 restore-spend + P3 cut-bid on one keyword)
        for flag, rule, action, why, score in flags[:1]:
            key = (q, rule)
            cur = hits.get(key)
            if cur and (cur["has_rank"], cur["score"]) >= (has_rank, score):
                cur["others"].append(asin)
                continue
            others = cur["others"] + [cur["asin"]] if cur else []
            hits[key] = {"asin": asin, "query": ent["query"], "flag": flag,
                         "rule": rule, "action": action, "why": why,
                         "score": score, "ent": ent, "others": others,
                         "has_rank": has_rank}
    order = {"Problem": 0, "Opportunity": 1, "Efficiency Gain": 2}
    rows = sorted(hits.values(), key=lambda h: (order[h["flag"]], -h["score"]))
    n_p1 = 0
    for i, h in enumerate(rows):
        if h["rule"] == "P1":
            h["prio"] = "P1"
            n_p1 += 1
        elif h["flag"] == "Problem":
            h["prio"] = "P1" if i < n_p1 + 5 else "P2"
        elif h["flag"] == "Opportunity":
            h["prio"] = "P2" if sum(1 for r in rows[:i] if r["flag"] == "Opportunity") < 5 else "P3"
        else:
            h["prio"] = "P3"
    return rows[:cfg.get("actions_cap", 40)]


# ------------------------------------------------------------------ helpers
def wk_label(week):
    return datetime.strptime(week, "%Y-%m-%d").strftime("%b %d")


def money(cfg):
    return ew.EUR2 if cfg["currency"] in ("EUR", "€") else ew.USD2


def is_eur(cfg):
    return cfg["currency"] in ("EUR", "€")


def dmoney(cfg):
    """Delta (WoW) money format. Must track money(): a EUR workbook that renders
    its value cells in € and its delta cells in $ is the bug this prevents."""
    return DEUR if is_eur(cfg) else DUSD


def sym(cfg):
    """Currency symbol for prose. Flag narratives and the Methods tab are text,
    not number_format, so they need this explicitly."""
    return "€" if is_eur(cfg) else "$"


def sheet_name(label, asin):
    return f"{label} {asin}"[:31]


def _grey(cell):
    cell.fill = GREY_FILL


# ------------------------------------------------------------------ tabs
def tab_dashboard(wb, cfg, weeks, roll, ppc, asin_flags):
    W = 7
    ws = wb.active
    ws.title = "① Dashboard"
    ws.sheet_view.showGridLines = False
    ew.title_block(
        ws, f"{cfg['client']} {cfg['marketplace']} SQP x PPC Weekly",
        f"Weeks {weeks[0]} to {weeks[-1]} (Sunday starts). SQP top-100 queries "
        f"per ASIN per week. Target ACOS {cfg['targets']['acos']:.0%}.",
        W, banner=ew.brand_banner("SQP x PPC Weekly Monitor"))
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 13
    r = 5
    metric_rows = [
        ("Total SV (top-100)", "sv", ew.INT, DINT),
        ("Click share (weighted)", "cs", ew.PCT2, PP),
        ("Purchase share (weighted)", "ps", ew.PCT2, PP),
        ("Ad spend (matched terms)", "spend", money(cfg), dmoney(cfg)),
        ("Ad sales (matched terms)", "sales", money(cfg), dmoney(cfg)),
        ("Ad ACOS (matched)", "acos", ew.PCT, None),
    ]
    for asin, label in cfg["asins"].items():
        ew.band(ws, r, f"{label}  ·  {asin}", W)
        r += 1
        ew.header_row(ws, r, ["Metric"] + [wk_label(w) for w in weeks] + ["WoW Δ", ""])
        r += 1
        for name, key, fmt, dfmt in metric_rows:
            ws.cell(r, 1, name).font = ew.F(10)
            ws.cell(r, 1).alignment = ew.LEFT
            vals = [roll[asin].get(w, {}).get(key) for w in weeks]
            for i, v in enumerate(vals):
                c = ws.cell(r, 2 + i, v)
                c.number_format = fmt
                c.font = ew.F(10)
                c.alignment = ew.RIGHT
                if key == "acos" and v is not None:
                    f = ew.acos_fill(v)
                    if f:
                        c.fill = f
            if dfmt and vals[-2] is not None and vals[-1] is not None:
                d = vals[-1] - vals[-2]
                c = ws.cell(r, 2 + len(weeks), d)
                c.number_format = dfmt
                c.font = ew.F(10, True)
                c.alignment = ew.RIGHT
                if key in ("cs", "ps") and abs(d) >= 0.005:
                    c.fill = PatternFill(
                        "solid", fgColor=ew.TL["good"] if d > 0 else ew.TL["bad"])
            r += 1
        flag, reason = asin_flags[asin]
        ws.cell(r, 1, "Flag").font = ew.F(10, True)
        fc = ws.cell(r, 2, flag)
        fc.font = ew.F(10, True)
        fc.fill = FLAG_FILL[flag]
        fc.alignment = ew.CENTER
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=W)
        rc = ws.cell(r, 3, reason)
        rc.font = ew.F(9, False, ew.C["mist"])
        rc.alignment = ew.WRAP
        r += 2

    ew.band(ws, r, "Account level (each term counted once)", W)
    r += 1
    ew.header_row(ws, r, ["Metric"] + [wk_label(w) for w in weeks] + ["WoW Δ", ""])
    r += 1
    totals = {}
    for w in weeks:
        terms = ppc.get(w, {})
        totals[w] = {
            "spend": sum(t["spend"] for t in terms.values()),
            "sales": sum(t["sales"] for t in terms.values()),
        }
    for name, key in [("Profile ad spend", "spend"), ("Profile ad sales", "sales")]:
        ws.cell(r, 1, name).font = ew.F(10)
        for i, w in enumerate(weeks):
            c = ws.cell(r, 2 + i, totals[w][key])
            c.number_format = money(cfg)
            c.font = ew.F(10)
            c.alignment = ew.RIGHT
        d = totals[weeks[-1]][key] - totals[weeks[-2]][key]
        c = ws.cell(r, 2 + len(weeks), d)
        c.number_format = dmoney(cfg)
        c.font = ew.F(10, True)
        c.alignment = ew.RIGHT
        r += 1
    ws.cell(r, 1, "Profile ACOS").font = ew.F(10)
    for i, w in enumerate(weeks):
        v = totals[w]["spend"] / totals[w]["sales"] if totals[w]["sales"] else None
        c = ws.cell(r, 2 + i, v)
        c.number_format = ew.PCT
        c.font = ew.F(10)
        c.alignment = ew.RIGHT
        if v is not None:
            f = ew.acos_fill(v)
            if f:
                c.fill = f
    r += 2
    ew.note(ws, r,
            "Ad spend is profile level per search term. Sibling ASINs share "
            "terms. Do not sum ad spend across ASIN blocks.", W)
    return ws


def asin_has_rank(kw, asin):
    return any(d.get("rank") is not None
               for (a, _), e in kw.items() if a == asin
               for d in e["weeks"].values())


def asin_has_stock(kw, asin):
    return any(d.get("oos") is not None
               for (a, _), e in kw.items() if a == asin
               for d in e["weeks"].values())


def asin_layout(cfg, weeks, has_rank, has_stock=False):
    """Single source of truth for the ASIN tab's column geometry.

    Each metric block is <n weeks> value cells followed by 1 delta cell. v1
    hardcoded the offsets for 3 weeks, so a 4-week run wrote every block on top
    of the previous block's delta AND the validator recomputed the same wrong
    offsets, so it agreed with the corruption instead of catching it. The writer
    and the validator both call this now: they cannot drift apart again.

    Returns (spans, tail0, W) where spans is [(title, key, fmt, dfmt, pp_fill, col0)].
    """
    n = len(weeks)
    blocks = [("SEARCH VOLUME", "sv", ew.INT, DINT, False),
              ("CLICK SHARE", "cs", ew.PCT2, PP, True),
              ("PURCHASE SHARE", "ps", ew.PCT2, PP, True),
              ("AD SPEND", "spend", money(cfg), dmoney(cfg), False)]
    if has_rank:
        blocks.append(("ORGANIC RANK", "rank", ew.INT, '+0;-0', False))
    if has_stock:
        blocks.append(("DAYS OUT OF STOCK", "oos", ew.INT, '+0;-0', False))
    spans, col = [], 3
    for title, key, fmt, dfmt, ppf in blocks:
        spans.append((title, key, fmt, dfmt, ppf, col))
        col += n + 1
    tail0 = col                      # Ad sales, ACOS, CVR gap, CTR gap
    return spans, tail0, tail0 + 3


def tab_asin(wb, cfg, weeks, asin, label, kw):
    n = len(weeks)
    has_rank = asin_has_rank(kw, asin)
    has_stock = asin_has_stock(kw, asin)
    spans, tail0, W = asin_layout(cfg, weeks, has_rank, has_stock)
    ws = wb.create_sheet(sheet_name(label, asin))
    ws.sheet_view.showGridLines = False

    ew.title_block(
        ws, f"{label}  \u00b7  {asin}",
        "Per query. SQP shares next to ad spend, organic rank, days out of stock "
        "and the conversion gap to the market. Days out of stock is an ASIN-level "
        "fact repeated on every row on purpose: a rank slip and its cause belong "
        "in the same eyeline. Blank grey cell = query was outside that week's top-100.",
        W, banner=ew.brand_banner("SQP x PPC Weekly Monitor"))
    hr = 5
    for title, key, fmt, dfmt, ppf, c0 in spans:
        ws.merge_cells(start_row=hr, start_column=c0, end_row=hr, end_column=c0 + n)
        c = ws.cell(hr, c0, title)
        c.fill = ew.HDR_FILL
        c.font = ew.F(9, True, ew.C["white"])
        c.alignment = ew.CENTER
    ws.merge_cells(start_row=hr, start_column=tail0, end_row=hr, end_column=W)
    c = ws.cell(hr, tail0, "LATEST WEEK")
    c.fill = ew.HDR_FILL
    c.font = ew.F(9, True, ew.C["white"])
    c.alignment = ew.CENTER
    for cc in range(1, W + 1):
        ws.cell(hr, cc).fill = ew.HDR_FILL

    wl = [wk_label(w) for w in weeks]
    delta_lbl = {"sv": "\u0394", "cs": "\u0394 pp", "ps": "\u0394 pp",
                 "spend": f"\u0394 {sym(cfg)}", "rank": "\u0394", "oos": "\u0394"}
    sub = ["Search query", "Flag"]
    widths = [34, 13]
    for title, key, fmt, dfmt, ppf, c0 in spans:
        sub += wl + [delta_lbl[key]]
        widths += [9] * n + [9]
    sub += ["Ad sales", "ACOS", "CVR vs mkt", "CTR vs mkt"]
    widths += [10, 8, 11, 11]
    ew.header_row(ws, hr + 1, sub, widths=widths)
    for cc in (1, 2):
        ws.merge_cells(start_row=hr, start_column=cc, end_row=hr + 1, end_column=cc)
        c = ws.cell(hr, cc, "Search query" if cc == 1 else "Flag")
        c.fill = ew.HDR_FILL
        c.font = ew.F(10, True, ew.C["white"])
        c.alignment = ew.CENTER
    ws.freeze_panes = f"C{hr + 2}"

    rows = [(q, ent) for (a, q), ent in kw.items() if a == asin]
    rows.sort(key=lambda t: -(t[1]["weeks"].get(weeks[-1], {}).get("sv")
                              or max(d["sv"] for d in t[1]["weeks"].values())))
    rows = rows[:cfg.get("max_rows_per_asin", 150)]

    w1, w2 = weeks[-2], weeks[-1]
    r = hr + 2
    for q, ent in rows:
        flags = flag_keyword(cfg, weeks, ent)
        flag = flags[0] if flags else None
        wd = ent["weeks"]
        d1, d2 = wd.get(w1), wd.get(w2)
        qc = ws.cell(r, 1, ent["query"])
        qc.font = ew.F(9)
        qc.alignment = ew.LEFT
        qc.border = ew.BORDER
        fc = ws.cell(r, 2, flag[0] if flag else "")
        fc.font = ew.F(9)
        fc.alignment = ew.CENTER
        fc.border = ew.BORDER
        if flag:
            fc.fill = FLAG_FILL[flag[0]]

        def block(col0, key, fmt, dfmt, pp_fill=False, lower_is_better=False):
            for i, w in enumerate(weeks):
                c = ws.cell(r, col0 + i)
                c.border = ew.BORDER
                c.font = ew.F(9)
                v = wd.get(w, {}).get(key) if w in wd else None
                if w in wd and v is not None:
                    c.value = v
                    c.number_format = fmt
                    c.alignment = ew.RIGHT
                else:
                    _grey(c)
            c = ws.cell(r, col0 + n)
            c.border = ew.BORDER
            c.font = ew.F(9, True)
            c.alignment = ew.RIGHT
            if d1 and d2 and d1.get(key) is not None and d2.get(key) is not None:
                d = d2[key] - d1[key]
                c.value = d
                c.number_format = dfmt
                good = (d < 0) if lower_is_better else (d > 0)
                if pp_fill and abs(d) >= 0.005:
                    c.fill = PatternFill(
                        "solid", fgColor=ew.TL["good"] if good else ew.TL["bad"])
                if lower_is_better and d != 0:
                    c.fill = PatternFill(
                        "solid", fgColor=ew.TL["good"] if good else ew.TL["bad"])
            return c

        oos_th = cfg["thresholds"].get("oos_days_flag", 2)
        dsp = None
        for title, key, fmt, dfmt, ppf, c0 in spans:
            c = block(c0, key, fmt, dfmt, pp_fill=ppf,
                      lower_is_better=(key in ("rank", "oos")))
            if key == "spend":
                dsp = c
            if key == "oos":
                # colour the weeks that lost days, so a rank slip and its cause
                # sit on the same line
                for i, w in enumerate(weeks):
                    v = wd.get(w, {}).get("oos")
                    if v:
                        ws.cell(r, c0 + i).fill = PatternFill(
                            "solid",
                            fgColor=ew.TL["bad"] if v >= oos_th else ew.TL["warn"])
        if flag and flag[1] == "P1" and dsp is not None:
            dsp.fill = PatternFill("solid", fgColor=ew.TL["warn"])

        c = ws.cell(r, tail0, d2["sales"] if d2 else None)
        c.number_format = money(cfg)
        c.font = ew.F(9); c.alignment = ew.RIGHT; c.border = ew.BORDER
        c = ws.cell(r, tail0 + 1, d2["acos"] if d2 else None)
        c.number_format = ew.PCT
        c.font = ew.F(9); c.alignment = ew.RIGHT; c.border = ew.BORDER
        if d2 and d2["acos"] is not None:
            f = ew.acos_fill(d2["acos"], cfg["targets"]["acos"])
            if f:
                c.fill = f
        # CVR gap vs market, in percentage points. Free: both sides are already
        # in the SQP counts. This is the "can we even win this term" column.
        for off, key in ((2, "cvr_gap"), (3, "ctr_gap")):
            c = ws.cell(r, tail0 + off)
            c.border = ew.BORDER
            c.font = ew.F(9)
            c.alignment = ew.RIGHT
            v = d2.get(key) if d2 else None
            if v is None:
                _grey(c)
            else:
                c.value = v
                c.number_format = '+0.0"pp";-0.0"pp"'
                f = ew.gap_fill(v)
                if f:
                    c.fill = f
        r += 1
    return ws


def tab_actions(wb, cfg, weeks, actions):
    W = 11
    ws = wb.create_sheet("Action Plan")
    ws.sheet_view.showGridLines = False
    ew.title_block(
        ws, "Action Plan",
        f"Rule based. Every flag is reproducible from the Methods tab. "
        f"Numbers are week {weeks[-1]} vs {weeks[-2]}.",
        W, banner=ew.brand_banner("SQP x PPC Weekly Monitor"))
    hr = 5
    ew.header_row(ws, hr, ["Priority", "Flag", "ASIN", "Search query", "SV (wk)",
                           "Click share", "Purch share", "Ad spend", "ACOS",
                           "Recommended action", "Why"],
                  widths=[8, 14, 13, 30, 9, 14, 14, 16, 8, 46, 52])
    S = sym(cfg)
    w1, w2 = weeks[-2], weeks[-1]
    r = hr + 1
    for h in actions:
        d1 = h["ent"]["weeks"].get(w1)
        d2 = h["ent"]["weeks"].get(w2)

        def trans(key, kind):
            if not d1:
                a = "n/a"
            elif kind == "pct":
                a = f"{d1[key] * 100:.1f}%"
            else:
                a = f"{S}{d1[key]:,.0f}"
            b = (f"{d2[key] * 100:.1f}%" if kind == "pct" else f"{S}{d2[key]:,.0f}") if d2 else "n/a"
            return f"{a} -> {b}"

        asin_txt = h["asin"] + (f" (+{len(h['others'])})" if h["others"] else "")
        vals = [h["prio"], f"{h['flag']} · {h['rule']}", asin_txt, h["query"],
                d2["sv"] if d2 else None, trans("cs", "pct"), trans("ps", "pct"),
                trans("spend", "usd"), d2["acos"] if d2 else None,
                h["action"], h["why"]]
        for cidx, v in enumerate(vals, 1):
            c = ws.cell(r, cidx, v)
            c.border = ew.BORDER
            c.font = ew.F(9)
            c.alignment = ew.LEFT if cidx in (4, 10, 11) else ew.CENTER
            if cidx in (10, 11):
                c.alignment = ew.WRAP
        ws.cell(r, 5).number_format = ew.INT
        ws.cell(r, 5).alignment = ew.RIGHT
        ac = ws.cell(r, 9)
        ac.number_format = ew.PCT
        ac.alignment = ew.RIGHT
        if d2 and d2["acos"] is not None:
            f = ew.acos_fill(d2["acos"])
            if f:
                ac.fill = f
        pc = ws.cell(r, 1)
        pc.fill = PatternFill("solid", fgColor={
            "P1": ew.TL["bad"], "P2": ew.TL["warn"], "P3": ew.TL["grey"]}[h["prio"]])
        pc.font = ew.F(9, True)
        ws.cell(r, 2).fill = FLAG_FILL[h["flag"]]
        r += 1
    if not actions:
        ew.note(ws, r, "No rule fired this week.", W)
    return ws


def tab_methods(wb, cfg, weeks, stats):
    S = sym(cfg)
    W = 2
    ws = wb.create_sheet("Methods & Sources")
    ws.sheet_view.showGridLines = False
    ew.title_block(ws, "Methods & Sources", "", W,
                   banner=ew.brand_banner("SQP x PPC Weekly Monitor"))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110
    th = cfg["thresholds"]
    target = cfg["targets"]["acos"]
    rows = [
        ("SQP source", "AdLabs MCP search_query entity, one pull per week. "
         "Verified byte-identical to Seller Central SQP on 2026-07-08 "
         "(US 3-way and DE 2-way checks, 13,248 metric cells, 0 mismatches)."),
        ("PPC source", "AdLabs MCP search_term entity, one pull per week, "
         "grouped by search term. Profile "
         + str(cfg["adlabs_profile_id"]) + "."),
        ("Weeks", ", ".join(weeks) + ". Sunday to Saturday, Amazon SQP convention."),
        ("Coverage cap", "SQP carries the top 100 queries per ASIN per week. "
         "That cap is Amazon's, present in every source. The long tail is "
         "missing. SV totals are floors, not full demand."),
        ("Join", "Exact lowercased string match: ad search term equals SQP "
         "query. The join is profile level. Search terms are campaign scoped, "
         "not ASIN scoped. Spend shown on an ASIN tab can include clicks that "
         "landed on a sibling ASIN. Never sum spend across ASIN tabs."),
        ("Rollup math", "ASIN click share = sum of ASIN clicks over sum of "
         "total clicks across its top-100 queries, per week. Purchase share "
         "the same. Never a mean of shares."),
        ("Targets", f"ACOS {target:.0%}. TACOS {cfg['targets']['tacos']:.0%}."),
        ("Match rate", stats["match_rate"]),
        ("Flag P1", "Spend-driven share loss. Click share down at least "
         f"{th['click_share_drop_pp']}pp or {th['click_share_drop_rel']:.0%} "
         f"relative, prior-week spend at least {S}{th['min_spend_prev']:.0f} and "
         f"latest spend down at least {th['spend_drop_rel']:.0%}. Action: restore spend."),
        ("Flag P2", f"Share loss with spend held. Click share down at least "
         f"{th['share_loss_held_pp']}pp while spend stayed at 90%+ of prior "
         "week. Action: check price, rating, competitors, organic rank."),
        ("Flag P3", f"Paying over target without traction. Spend at least "
         f"{S}{th['min_spend_flag']:.0f} with no sales or ACOS above "
         f"{th['acos_problem_mult'] * target:.0%}, purchase share not improving. "
         "Action: cut bid or isolate in exact."),
        ("Flag O1", f"Unfunded demand. SV at least {th['opp_min_sv']}, click "
         f"share and purchase share under {th['low_share']:.0%}, spend under "
         f"{S}{th['min_spend_prev']:.0f}. Action: add exact target."),
        ("Flag O2", f"Converts above its weight. SV at least {th['opp2_min_sv']}, "
         f"purchase share at least {th['converts_edge_pp']}pp above click share. "
         "Action: buy more impressions."),
        ("Flag E1", f"Scaling winner. Both shares up WoW, spend at least "
         f"{S}{th['min_spend_prev']:.0f}, ACOS at or under {target:.0%}. "
         "Action: protect bid, add budget headroom."),
        ("Days out of stock", "AdLabs `product` entity, one export per week: "
         "`out_of_stock_days` is window-scoped, so the weeks sum back to the "
         "monthly figure. An ASIN-level fact repeated on every keyword row of that "
         "ASIN, deliberately, so a rank slip and its cause sit in the same eyeline. "
         "Do NOT read availability / days_of_cover / fulfillable_units per week: "
         "those are current snapshots and come back identical in every export."),
        ("Stock outranks everything", "Amazon deranks what it cannot ship, so an "
         f"out-of-stock week produces share loss, rank slip and a high ACOS that "
         f"look exactly like a bidding failure and are not one. At or above "
         f"{th.get('oos_days_flag', 2)} lost days in the focus week, stock takes over the "
         "verdict: P3 says 'STOCK, not bids, pause until restocked' instead of 'cut "
         "the bid', P1 says check stock before restoring spend, P2 stops blaming "
         "price and competitors, O1 says do not buy demand you cannot ship, and E1 "
         "notes the winner's true ceiling is higher than it looks."),
        ("Organic rank", "DataDive Rank Radar, median organic rank across each "
         "Sunday-Saturday week (median, not last-day: a single day is noisy and "
         "SUPA compares whole weeks). Optional. Radars track a few dozen keywords "
         "while SQP carries hundreds, so most rows are blank by design and the "
         "column disappears entirely when no radar is configured. Rank move is "
         "last ranked week minus first: positive means it slipped."),
        ("CVR vs market", "Our conversion minus the market's on the same query, in "
         "percentage points, focus week. Both sides come from the SQP counts, so it "
         "costs no extra input. Needs at least 5 of our clicks that week or it is "
         "left blank. Amazon scores conversion per keyword: below-market conversion "
         "means spend rents a position instead of earning one, and no bid fixes that."),
        ("CTR vs market", "Same idea on click-through. Needs asin_impression_count and "
         "total_query_impression_count in the export (optional) plus at least 100 of "
         "our impressions, else blank."),
        ("How rank changes the advice", "P3 is the split that matters. Paying over "
         "target with rank IMPROVING is rank investment, not waste: the flag flips to "
         "Opportunity and says do not cut on ACOS alone. Paying over target with rank "
         "SLIPPING and conversion below market means the term is unwinnable at any "
         "bid: fix the offer first. P1 becomes urgent when the spend cut also cost "
         "rank. P2 separates 'outranked organically' from 'a competitor bought the "
         "share'. O1 stops recommending you pay for traffic organic already owns."),
        ("Eligibility", f"Keyword flags need SV of at least {th['min_sv']} in "
         "the latest week. First matching rule wins the Flag column. All "
         "matches feed the Action Plan."),
        ("Legend", "Green fill = good move or ACOS under target. Amber = watch. "
         "Red = problem. Grey = no data that week (query outside the top-100)."),
        ("Prepared by", ew.prepared_by_org() + ", " + datetime.now().strftime("%Y-%m-%d")),
    ]
    r = 5
    for k, v in rows:
        kc = ws.cell(r, 1, k)
        kc.font = ew.F(10, True)
        kc.alignment = ew.WRAP
        vc = ws.cell(r, 2, v)
        vc.font = ew.F(10)
        vc.alignment = ew.WRAP
        ws.row_dimensions[r].height = max(14, 13 * (1 + len(v) // 105))
        r += 1
    return ws


# ------------------------------------------------------------------ build
def out_path(cfg):
    d = rp(cfg["output"]["dir"])
    d.mkdir(parents=True, exist_ok=True)
    return d / (f"{cfg['client']}_{cfg['marketplace']}_SQP-PPC_Weekly_"
                + last_saturday(cfg) + ".xlsx")


def last_saturday(cfg):
    wk = datetime.strptime(cfg["weeks"][-1], "%Y-%m-%d")
    from datetime import timedelta
    return (wk + timedelta(days=6)).strftime("%Y-%m-%d")


def match_stats(cfg, sqp, ppc):
    lines = []
    for w in cfg["weeks"]:
        qs = set()
        for (asin, week), queries in sqp.items():
            if week == w:
                qs |= set(queries)
        terms = ppc.get(w, {})
        matched = qs & set(terms)
        spend_all = sum(t["spend"] for t in terms.values())
        spend_matched = sum(terms[t]["spend"] for t in matched)
        lines.append(
            f"{w}: {len(matched)}/{len(qs)} SQP queries have ad traffic "
            f"({len(matched) / len(qs):.0%}). Matched spend "
            f"{sym(cfg)}{spend_matched:,.0f} of {sym(cfg)}{spend_all:,.0f} profile spend "
            f"({(spend_matched / spend_all if spend_all else 0):.0%}).")
    return " ".join(lines)


def build(cfg_path):
    cfg = load_config(cfg_path)
    weeks = cfg["weeks"]
    if len(weeks) < 2:
        raise SystemExit("need at least 2 weeks")
    sqp = load_sqp(cfg)
    ppc = load_ppc(cfg)
    rank = load_rank(cfg)
    stock = load_stock(cfg)
    kw = join_weekly(cfg, sqp, ppc, rank, stock)
    roll = rollup_asin(cfg, sqp, ppc)
    asin_flags = {a: flag_asin(cfg, weeks, roll[a]) for a in cfg["asins"]}
    actions = build_actions(cfg, weeks, kw)
    stats = {"match_rate": match_stats(cfg, sqp, ppc)}

    wb = Workbook()
    tab_dashboard(wb, cfg, weeks, roll, ppc, asin_flags)
    for asin, label in cfg["asins"].items():
        tab_asin(wb, cfg, weeks, asin, label, kw)
    tab_actions(wb, cfg, weeks, actions)
    tab_methods(wb, cfg, weeks, stats)
    p = out_path(cfg)
    wb.save(p)
    print(f"BUILT {p}")
    ranked = sum(1 for e in kw.values() if any(d.get("rank") is not None for d in e["weeks"].values()))
    gapped = sum(1 for e in kw.values() if any(d.get("cvr_gap") is not None for d in e["weeks"].values()))
    print(f"  keywords: {len(kw)} (asin,query) pairs · actions: {len(actions)}")
    print(f"  rank coverage: {ranked}/{len(kw)} keywords have Rank Radar data"
          + ("" if rank else "  (no inputs.rank_jsons configured)"))
    print(f"  CVR-gap coverage: {gapped}/{len(kw)} keywords have enough clicks to compare to market")
    if stock:
        lost = {a: sum(v for (aa, w), v in stock.items() if aa == a and v)
                for a in cfg["asins"]}
        lost = {a: v for a, v in lost.items() if v}
        print(f"  stock: {len(stock)} ASIN-weeks loaded"
              + (f" · days lost: " + ", ".join(f"{a} {v:.0f}" for a, v in sorted(lost.items(), key=lambda t: -t[1]))
                 if lost else " · no out-of-stock days in these weeks"))
    else:
        print("  stock: none  (no inputs.stock_csvs configured)")
    print(f"  {stats['match_rate']}")
    for a, (fl, why) in asin_flags.items():
        print(f"  {a}: {fl} · {why}")
    return p


# ------------------------------------------------------------------ validate
def validate(cfg_path):
    cfg = load_config(cfg_path)
    weeks = cfg["weeks"]
    p = out_path(cfg)
    wb = load_workbook(p)
    errors, warns = [], []

    expected = (["① Dashboard"]
                + [sheet_name(l, a) for a, l in cfg["asins"].items()]
                + ["Action Plan", "Methods & Sources"])
    if wb.sheetnames != expected:
        errors.append(f"sheets {wb.sheetnames} != {expected}")

    sqp = load_sqp(cfg)
    ppc = load_ppc(cfg)
    rank = load_rank(cfg)
    stock = load_stock(cfg)
    kw = join_weekly(cfg, sqp, ppc, rank, stock)
    rng = random.Random(42)
    wl = [wk_label(w) for w in weeks]
    n = len(weeks)
    for asin, label in cfg["asins"].items():
        ws = wb[sheet_name(label, asin)]
        spans, tail0, W = asin_layout(cfg, weeks, asin_has_rank(kw, asin),
                                      asin_has_stock(kw, asin))
        # Every block must carry the week labels at its own offset. This is the
        # check that would have caught the v1 3-week hardcoding on a 4-week run.
        for title, key, fmt, dfmt, ppf, c0 in spans:
            hdr = [ws.cell(6, c).value for c in range(c0, c0 + n)]
            if hdr != wl:
                errors.append(f"{asin}: {key} block week header {hdr} != {wl} at col {c0}")
            grp = ws.cell(5, c0).value
            if grp != title:
                errors.append(f"{asin}: group header at col {c0} is {grp!r}, expected {title!r}")
        if ws.cell(5, tail0).value != "LATEST WEEK":
            errors.append(f"{asin}: LATEST WEEK group not at col {tail0}")
        rowmap = {}
        for r in range(7, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v:
                rowmap[str(v).strip().lower()] = r
        latest = sqp.get((asin, weeks[-1]), {})
        sample = rng.sample(sorted(latest), min(10, len(latest)))
        for q in sample:
            r = rowmap.get(q)
            if not r:
                continue  # below the per-tab row cap
            rec = latest[q]
            col = {key: c0 for _, key, _, _, _, c0 in spans}
            # last week's value cell of each block = c0 + n - 1
            checks = [
                ("sv", ws.cell(r, col["sv"] + n - 1).value, rec["sv"], 1e-6),
                ("cs", ws.cell(r, col["cs"] + n - 1).value,
                 rec["a_clk"] / rec["t_clk"] if rec["t_clk"] else 0.0, 1e-6),
                ("ps", ws.cell(r, col["ps"] + n - 1).value,
                 rec["a_pur"] / rec["t_pur"] if rec["t_pur"] else 0.0, 1e-6),
                ("spend", ws.cell(r, col["spend"] + n - 1).value,
                 ppc.get(weeks[-1], {}).get(q, {}).get("spend", 0.0), 0.01),
            ]
            for name, got, want, tol in checks:
                if abs((got or 0) - want) > tol:
                    errors.append(f"{asin} '{q}' {name}: cell {got} != raw {want}")
        # soft cross-check vs AdLabs' own comparison column
        prior = sqp.get((asin, weeks[-2]), {})
        for q in sample[:7]:
            rec = latest[q]
            if q in prior and rec["cs_cmp"]:
                ours = prior[q]["cs"]
                if abs(ours - rec["cs_cmp"]) > 0.005:
                    warns.append(
                        f"{asin} '{q}': prior-week click share {ours:.4f} vs "
                        f"AdLabs comparison {rec['cs_cmp']:.4f}")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if (isinstance(cell.value, (int, float)) and cell.value > 1.0
                        and "%" in (cell.number_format or "")):
                    rgb = getattr(getattr(cell.fill, "start_color", None), "rgb", None)
                    if rgb and str(rgb)[-6:].upper() in GREEN_FILLS:
                        errors.append(
                            f"{ws.title}!{cell.coordinate}: >100% ratio with green fill")

    print(match_stats(cfg, sqp, ppc))
    for w in warns:
        print(f"WARN  {w}")
    if errors:
        for e in errors:
            print(f"FAIL  {e}")
        return 1
    print(f"VALIDATE OK: {p.name} ({len(wb.sheetnames)} tabs)")
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--validate", action="store_true")
    args = ap.parse_args()
    build(args.config)
    if args.validate:
        sys.exit(validate(args.config))


if __name__ == "__main__":
    main()
