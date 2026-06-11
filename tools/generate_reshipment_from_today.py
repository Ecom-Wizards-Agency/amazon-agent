#!/usr/bin/env python3
import csv
import json
import math
import re
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook = None


RUN_DATE = "2026-05-25"
REPORT_DAYS = 30
TARGET_DAYS = 45 + 7 + 14
MULTIPLIER = 1.2

DOWNLOADS = Path("/Users/victoruhl/Downloads")
ROOT = Path("/Users/victoruhl/Codex Projects/Amazon Agent")

CLIENTS = [
    {
        "key": "Swissker US",
        "brand": "Swissker",
        "market": "US",
        "country": "United States",
        "fba": DOWNLOADS / "636074020598.csv",
        "business": DOWNLOADS / "BusinessReport-5-24-26 (3).csv",
        "inventory": None,
        "restock": None,
        "notes": "Same-day FBA Inventory + same-day Business Report used. Restock/Inventory Report unavailable in this partial batch.",
    },
    {
        "key": "AlphaInfuse US",
        "brand": "AlphaInfuse",
        "market": "US",
        "country": "United States",
        "fba": DOWNLOADS / "99704020598.csv",
        "business": DOWNLOADS / "BusinessReport-5-24-26.csv",
        "inventory": DOWNLOADS / "Inventory+Report_05-25-2026 (2).txt",
        "restock": None,
        "notes": "Same-day FBA Inventory, Business Report, and Inventory Report used. Restock Report unavailable in this partial batch.",
    },
    {
        "key": "Goda US",
        "brand": "Goda",
        "market": "US",
        "country": "United States",
        "fba": DOWNLOADS / "273773020598.csv",
        "business": DOWNLOADS / "BusinessReport-5-24-26 (1).csv",
        "inventory": DOWNLOADS / "Inventory+Report_05-25-2026 (3).txt",
        "restock": DOWNLOADS / "273774020598.csv",
        "notes": "Same-day FBA Inventory, Business Report, Inventory Report, and Restock Report used.",
    },
    {
        "key": "Seranova US",
        "brand": "Seranova",
        "market": "US",
        "country": "United States",
        "fba": DOWNLOADS / "253947020598.csv",
        "business": DOWNLOADS / "BusinessReport-5-24-26 (2).csv",
        "inventory": DOWNLOADS / "Inventory+Report_05-25-2026 (4).txt",
        "restock": DOWNLOADS / "253948020598.csv",
        "notes": "Same-day FBA Inventory, Business Report, Inventory Report, and Restock Report used.",
    },
    {
        "key": "Piercing XXL DE",
        "brand": "Piercing XXL",
        "market": "DE",
        "country": "Germany",
        "fba": DOWNLOADS / "240124020598.csv",
        "business": None,
        "inventory": DOWNLOADS / "Inventory+report_05-25-2026.txt",
        "restock": DOWNLOADS / "240125020598.csv",
        "notes": "Same-day FBA Inventory, Inventory Report, and Restock Report used. Business Report is not in the partial batch, so demand falls back to FBA 30-day shipped units and Restock 30-day sold units.",
    },
    {
        "key": "JBS DE",
        "brand": "JBS",
        "market": "DE",
        "country": "Germany",
        "fba": DOWNLOADS / "380905020598.txt",
        "business": DOWNLOADS / "BusinessReport-25-05-2026.csv",
        "inventory": None,
        "restock": DOWNLOADS / "380943020598.csv",
        "restock_country": "DE",
        "notes": "Same-day FBA Inventory, Business Report, and Restock Report used. Business Report date range is 24/04/2026-24/05/2026 as shown in Seller Central on 25/05/2026.",
    },
]


def n(value):
    if value is None:
        return 0
    text = str(value).strip().replace(",", "").replace("$", "").replace("€", "").replace("%", "")
    if not text or text == "--":
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def i(value):
    return int(round(n(value)))


def short(text, limit=66):
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def read_rows(path):
    if not path or not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        delimiter = "\t" if "\t" in sample and sample.count("\t") > sample.count(",") else ","
        return list(csv.DictReader(fh, delimiter=delimiter))


def add_item(items, asin, **updates):
    if not asin:
        return
    item = items[asin]
    item["asin"] = asin
    for key, value in updates.items():
        if value in ("", None):
            continue
        if key in {"title", "sku"}:
            item.setdefault(key, value)
        elif key == "sources":
            item.setdefault("sources", set()).update(value)
        else:
            item[key] = item.get(key, 0) + value


def calculate(client):
    items = defaultdict(dict)
    source_files = []

    for row in read_rows(client["business"]):
        asin = row.get("(Child) ASIN") or row.get("ASIN")
        add_item(
            items,
            asin,
            title=row.get("Title"),
            business_units=i(row.get("Units Ordered")),
            sources={"business"},
        )
    if client["business"] and client["business"].exists():
        source_files.append(str(client["business"]))

    for row in read_rows(client["fba"]):
        asin = row.get("asin")
        add_item(
            items,
            asin,
            sku=row.get("sku"),
            title=row.get("product-name"),
            available=i(row.get("available")),
            inbound=i(row.get("inbound-quantity")),
            reserved=i(row.get("Total Reserved Quantity")),
            unfulfillable=i(row.get("unfulfillable-quantity")),
            fba_units_30=i(row.get("units-shipped-t30")),
            fba_units_7=i(row.get("units-shipped-t7")),
            estimated_excess=i(row.get("estimated-excess-quantity")),
            days_of_supply=i(row.get("days-of-supply")),
            sources={"fba"},
        )
    if client["fba"] and client["fba"].exists():
        source_files.append(str(client["fba"]))

    for row in read_rows(client["restock"]):
        if client.get("restock_country") and row.get("Country") != client["restock_country"]:
            continue
        asin = row.get("ASIN")
        add_item(
            items,
            asin,
            sku=row.get("Merchant SKU"),
            title=row.get("Product Name"),
            restock_units_30=i(row.get("Units Sold Last 30 Days")),
            restock_total=i(row.get("Total Units")),
            restock_inbound=i(row.get("Inbound")),
            restock_available=i(row.get("Available")),
            restock_fc_transfer=i(row.get("FC transfer")),
            restock_fc_processing=i(row.get("FC Processing")),
            restock_customer_order=i(row.get("Customer Order")),
            unfulfillable=i(row.get("Unfulfillable")),
            sources={"restock"},
        )
    if client["restock"] and client["restock"].exists():
        source_files.append(str(client["restock"]))

    for row in read_rows(client["inventory"]):
        asin = row.get("asin")
        add_item(items, asin, sku=row.get("sku"), price=n(row.get("price")), sources={"inventory"})
    if client["inventory"] and client["inventory"].exists():
        source_files.append(str(client["inventory"]))

    rows = []
    for asin, item in sorted(items.items()):
        demand = item.get("business_units") or item.get("fba_units_30") or item.get("restock_units_30") or 0
        demand_source = "Business Report" if item.get("business_units") else ("FBA Inventory" if item.get("fba_units_30") else "Restock Report")
        available = max(item.get("available", 0), item.get("restock_available", 0))
        inbound = max(item.get("inbound", 0), item.get("restock_inbound", 0))
        reserved = max(
            item.get("reserved", 0),
            item.get("restock_fc_transfer", 0) + item.get("restock_fc_processing", 0) + item.get("restock_customer_order", 0),
        )
        required = demand / REPORT_DAYS * MULTIPLIER * TARGET_DAYS
        reship = int(math.ceil(max(0, required - available - inbound - reserved)))
        excess = item.get("estimated_excess", 0)
        if not excess and demand > 0 and available > (demand / REPORT_DAYS * 120):
            excess = int(max(0, available - demand / REPORT_DAYS * 90))
        rows.append(
            {
                "ASIN": asin,
                "SKU": item.get("sku", ""),
                "Product Name": item.get("title", ""),
                "Demand 30d": demand,
                "Demand Source": demand_source,
                "Available": available,
                "Inbound": inbound,
                "Reserved": reserved,
                "Unfulfillable": item.get("unfulfillable", 0),
                "Required Units (66d x 1.2)": round(required, 1),
                "Reshipment Units": reship,
                "Estimated Excess Units": excess,
                "Days of Supply": item.get("days_of_supply", ""),
                "FBA Units 7d": item.get("fba_units_7", 0),
                "FBA Units 30d": item.get("fba_units_30", 0),
                "Restock Units 30d": item.get("restock_units_30", 0),
            }
        )

    rows.sort(key=lambda r: r["Reshipment Units"], reverse=True)
    out_dir = ROOT / "output" / client["key"] / "inventory"
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{RUN_DATE}_Inventory Overview_{client['brand']}_{client['market']}"
    csv_path = out_dir / f"{stem}.csv"
    xlsx_path = out_dir / f"{stem}.xlsx"
    slack_path = out_dir / f"{stem}_slack.txt"
    json_path = out_dir / f"{stem}_manifest.json"

    fields = list(rows[0].keys()) if rows else ["ASIN"]
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    if Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Overview"
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row in rows:
            ws.append([row[k] for k in fields])
        send = wb.create_sheet("Reshipment")
        send.append(fields)
        for row in [r for r in rows if r["Reshipment Units"] > 0]:
            send.append([row[k] for k in fields])
        excess = wb.create_sheet("Excess Inventory")
        excess.append(fields)
        for row in [r for r in rows if r["Estimated Excess Units"] > 0]:
            excess.append([row[k] for k in fields])
        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            for col in sheet.columns:
                letter = col[0].column_letter
                sheet.column_dimensions[letter].width = min(48, max(12, max(len(str(c.value or "")) for c in col[:80]) + 2))
        wb.save(xlsx_path)

    send_rows = [r for r in rows if r["Reshipment Units"] > 0]
    excess_rows = [r for r in rows if r["Estimated Excess Units"] > 0]
    parts = [
        f"{client['brand']} Inventory Overview - {client['country']}",
        "",
        f"Source: {client['notes']} Prime Day demand multiplier: 1.2x. Output saved: `{csv_path.name}` / `{xlsx_path.name}`.",
    ]
    if send_rows:
        parts.extend(["", "**Reshipment**"])
        for row in send_rows[:30]:
            parts.append(
                f"`{row['ASIN']}` {short(row['Product Name'])} - {row['Reshipment Units']:,} units needed | Available: {row['Available']:,} | Inbound: {row['Inbound']:,} | Reserved: {row['Reserved']:,}"
            )
        if len(send_rows) > 30:
            parts.append(f"Plus {len(send_rows) - 30} more low-volume rows in the workbook.")
        parts.append(f"Total: {sum(r['Reshipment Units'] for r in send_rows):,} units")
    else:
        parts.extend(["", "**Reshipment**", "No positive Prime Day reshipment quantities from today’s source data."])
    if excess_rows:
        parts.extend(["", "**Excess Inventory / Plan Sales**"])
        for row in excess_rows[:20]:
            parts.append(
                f"`{row['ASIN']}` {short(row['Product Name'])} - {row['Estimated Excess Units']:,} excess units | Available: {row['Available']:,} | 30d demand: {row['Demand 30d']:,}"
            )
        if len(excess_rows) > 20:
            parts.append(f"Plus {len(excess_rows) - 20} more excess rows in the workbook.")
        parts.append(f"Total: {sum(r['Estimated Excess Units'] for r in excess_rows):,} excess units")
    slack_path.write_text("\n".join(parts) + "\n", encoding="utf-8")

    manifest = {
        "account": client["key"],
        "date": RUN_DATE,
        "sources": source_files,
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
        "slack": str(slack_path),
        "sendRows": len(send_rows),
        "sendUnits": sum(r["Reshipment Units"] for r in send_rows),
        "excessRows": len(excess_rows),
        "excessUnits": sum(r["Estimated Excess Units"] for r in excess_rows),
        "notes": client["notes"],
    }
    json_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest


def main():
    manifests = [calculate(client) for client in CLIENTS]
    bundle = ROOT / "output" / "reshipment-plans-2026-05-25" / "summary_partial_ready.json"
    bundle.parent.mkdir(parents=True, exist_ok=True)
    bundle.write_text(json.dumps(manifests, indent=2), encoding="utf-8")
    print(json.dumps(manifests, indent=2))


if __name__ == "__main__":
    main()
