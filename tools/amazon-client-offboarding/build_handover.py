#!/usr/bin/env python3
"""Build and validate a branded Amazon client handover.

Inputs are a run config, an evidence manifest, and operator-written Markdown.
The builder is read-only with respect to Amazon systems. It creates local DOCX/XLSX
intermediaries and can optionally pass validated files to the existing native Google
Drive conversion tool when the operator explicitly adds ``--deliver``.
"""
from __future__ import annotations

import argparse
import importlib
import json
import re
import subprocess
import sys
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
AUDIT_TOOLS = REPO / "tools" / "amazon-ad-audit"
sys.path.insert(0, str(AUDIT_TOOLS))

WORKBOOK_TABS = [
    "Read Me & Data Watermark",
    "Market Scoreboard",
    "Non-Brand RPC Diagnostics",
    "Rank & Query Tracker",
    "Action Register",
]
MARKET_NAMES = {
    "DE": "Germany", "IT": "Italy", "ES": "Spain", "FR": "France",
    "SE": "Sweden", "UK": "United Kingdom", "US": "United States",
    "NL": "Netherlands", "PL": "Poland", "BE": "Belgium",
}
FORBIDDEN_LINK_HOSTS = {
    "slack.com", "app.slack.com", "notion.so", "www.notion.so",
    "localhost", "127.0.0.1",
}
ECON_STATUSES = {"verified", "historical-lead", "not-verified"}


class PreflightError(Exception):
    """Raised when a handover fails its evidence or delivery contract."""


def _read_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise PreflightError(f"missing file: {path}") from exc
    except json.JSONDecodeError as exc:
        raise PreflightError(f"invalid JSON in {path}: {exc}") from exc


def _resolve(base: Path, raw: str | None) -> Path:
    if not raw:
        return Path()
    path = Path(raw).expanduser()
    return path if path.is_absolute() else (base / path).resolve()


def _parse_date(raw: str, label: str) -> date:
    try:
        return date.fromisoformat(raw)
    except (TypeError, ValueError) as exc:
        raise PreflightError(f"{label} must be YYYY-MM-DD, got {raw!r}") from exc


def _parse_timestamp(raw: str, label: str) -> datetime:
    if not raw:
        raise PreflightError(f"{label} is missing")
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except (TypeError, ValueError) as exc:
        raise PreflightError(f"{label} must be an ISO timestamp, got {raw!r}") from exc
    if parsed.tzinfo is None:
        raise PreflightError(f"{label} must include a timezone")
    return parsed


def _client_name(cfg: dict) -> str:
    client = cfg.get("client")
    if isinstance(client, dict):
        return str(client.get("name") or "").strip()
    return str(client or "").strip()


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def _window_days(window: dict, prefix: str) -> int:
    start = _parse_date(window.get(f"{prefix}_start"), f"{prefix}_start")
    end = _parse_date(window.get(f"{prefix}_end"), f"{prefix}_end")
    if end < start:
        raise PreflightError(f"{prefix} window ends before it starts")
    return (end - start).days + 1


def _market_aliases(code: str) -> list[str]:
    aliases = [code]
    if code in MARKET_NAMES:
        aliases.append(MARKET_NAMES[code])
    return aliases


def _narrative_has_market(md: str, market: str) -> bool:
    aliases = "|".join(re.escape(a) for a in _market_aliases(market))
    return bool(re.search(rf"^##+\s+.*(?:\b(?:{aliases})\b)", md, re.M))


def _forbidden_market_in_narrative(md: str, market: str) -> bool:
    aliases = _market_aliases(market)
    for alias in aliases:
        flags = 0 if alias == market else re.I
        if re.search(rf"\b{re.escape(alias)}\b", md, flags):
            return True
    return False


def _iter_records(value, path="evidence"):
    if isinstance(value, dict):
        yield path, value
        for key, child in value.items():
            yield from _iter_records(child, f"{path}.{key}")
    elif isinstance(value, list):
        for idx, child in enumerate(value):
            yield from _iter_records(child, f"{path}[{idx}]")


def _economics_by_market(evidence: dict) -> dict[str, dict]:
    rows = evidence.get("market_economics") or []
    return {str(row.get("market", "")).upper(): row for row in rows}


def _source_ids(evidence: dict) -> set[str]:
    return {str(row.get("id")) for row in evidence.get("sources", []) if row.get("id")}


def _validate_links(cfg: dict, evidence: dict, md: str, errors: list[str]) -> None:
    links = list(evidence.get("assets") or [])
    appendix = cfg.get("existing_appendix")
    if appendix:
        links.append(appendix)
    for idx, item in enumerate(links):
        label = item.get("name") or item.get("label") or f"link {idx + 1}"
        url = str(item.get("url") or "")
        parsed = urlparse(url)
        host = parsed.netloc.lower()
        if parsed.scheme not in {"http", "https"} or not host:
            errors.append(f"{label}: URL is missing or invalid")
        if host in FORBIDDEN_LINK_HOSTS or host.endswith(".slack.com"):
            errors.append(f"{label}: internal-only host is forbidden")
        if item.get("client_accessible") is not True:
            errors.append(f"{label}: client_accessible must be true")
        if item.get("status") != "verified":
            errors.append(f"{label}: status must be verified")
        try:
            _parse_timestamp(item.get("verified_at"), f"{label}.verified_at")
        except PreflightError as exc:
            errors.append(str(exc))
        if url and url not in md:
            errors.append(f"{label}: verified client link is absent from the narrative")


def _validate_branding(cfg: dict, base: Path, errors: list[str]) -> tuple[Path, Path, dict]:
    branding = cfg.get("branding") or {}
    branding_path = _resolve(base, branding.get("branding_json"))
    brand_dir = _resolve(base, branding.get("brand_dir"))
    if branding.get("approved") is not True:
        errors.append("branding.approved must be true")
    if not branding_path.is_file():
        errors.append(f"approved branding file is missing: {branding_path or '<unset>'}")
        return branding_path, brand_dir, {}
    try:
        data = _read_json(branding_path)
    except PreflightError as exc:
        errors.append(str(exc))
        return branding_path, brand_dir, {}
    if not data.get("agency_name"):
        errors.append("approved branding file has no agency_name")
    if str(data.get("agency_name", "")).strip().lower() != "ecom wizards":
        errors.append("v0.1 requires the approved Ecom Wizards branding identity")
    for palette_name in ("palette_doc", "palette_xlsx"):
        palette = data.get(palette_name) or {}
        for key, value in palette.items():
            if key.startswith("_"):
                continue
            if not re.fullmatch(r"[0-9A-Fa-f]{6}", str(value or "")):
                errors.append(f"approved branding has invalid {palette_name}.{key}")
    assets = data.get("assets") or {}
    fonts = data.get("fonts") or {}
    for key in ("doc_font_name", "doc_font_file", "xlsx_font_display", "xlsx_font_body"):
        if not fonts.get(key) or str(fonts[key]).startswith("<"):
            errors.append(f"approved branding is missing fonts.{key}")
    required = [assets.get("logo_black"), fonts.get("doc_font_file")]
    if not brand_dir.is_dir():
        errors.append(f"branding asset directory is missing: {brand_dir or '<unset>'}")
    for asset in required:
        if not asset or not (brand_dir / asset).is_file():
            errors.append(f"required branding asset is missing: {asset or '<unset>'}")
    return branding_path, brand_dir, data


def preflight(config_path: Path) -> dict:
    cfg = _read_json(config_path)
    base = config_path.parent.resolve()
    errors: list[str] = []
    client = _client_name(cfg)
    if not client:
        errors.append("client name is required")
    if isinstance(cfg.get("client"), dict) and not cfg["client"].get("slug"):
        errors.append("client.slug is required")
    for field in ("successor", "audience", "currency"):
        if not str(cfg.get(field) or "").strip():
            errors.append(f"{field} is required")
    included = [str(v).upper() for v in cfg.get("included_markets") or []]
    excluded = [str(v).upper() for v in cfg.get("excluded_markets") or []]
    if not included:
        errors.append("included_markets must not be empty")
    overlap = sorted(set(included) & set(excluded))
    if overlap:
        errors.append(f"included and excluded markets overlap: {', '.join(overlap)}")
    if len(set(included)) != len(included):
        errors.append("included_markets contains duplicates")

    narrative_path = _resolve(base, cfg.get("narrative_path"))
    evidence_path = _resolve(base, cfg.get("evidence_path"))
    if not narrative_path.is_file():
        errors.append(f"narrative is missing: {narrative_path or '<unset>'}")
        md = ""
    else:
        md = narrative_path.read_text(encoding="utf-8")
    md_lower = md.lower()
    if not evidence_path.is_file():
        errors.append(f"evidence manifest is missing: {evidence_path or '<unset>'}")
        evidence = {}
    else:
        try:
            evidence = _read_json(evidence_path)
        except PreflightError as exc:
            errors.append(str(exc))
            evidence = {}

    try:
        cutoff = _parse_timestamp(cfg.get("cutoff_timestamp"), "cutoff_timestamp")
    except PreflightError as exc:
        errors.append(str(exc))
        cutoff = datetime.now(timezone.utc)

    windows = cfg.get("comparison_windows") or []
    if not windows:
        errors.append("comparison_windows must not be empty")
    if evidence.get("comparison_windows") != windows:
        errors.append("evidence comparison_windows must exactly match the run config")
    for idx, window in enumerate(windows):
        try:
            current_days = _window_days(window, "current")
            prior_days = _window_days(window, "prior")
            if current_days != prior_days:
                errors.append(f"comparison window {idx + 1} has unequal day counts")
            current_end = _parse_date(window.get("current_end"), "current_end")
            prior_end = _parse_date(window.get("prior_end"), "prior_end")
            if current_end >= cutoff.date() or prior_end >= cutoff.date():
                errors.append(f"comparison window {idx + 1} includes current-day or future data")
            complete_through = window.get("attribution_complete_through")
            if not complete_through:
                errors.append(f"comparison window {idx + 1} is missing attribution_complete_through")
            elif current_end > _parse_date(complete_through, "attribution_complete_through"):
                errors.append(f"comparison window {idx + 1} extends beyond attribution-complete data")
        except PreflightError as exc:
            errors.append(f"comparison window {idx + 1}: {exc}")

    source_ids = _source_ids(evidence)
    if not source_ids:
        errors.append("evidence.sources must contain at least one timestamped source")
    for idx, source in enumerate(evidence.get("sources") or []):
        if not source.get("id") or not source.get("name"):
            errors.append(f"source {idx + 1} requires id and name")
        try:
            _parse_timestamp(source.get("extracted_at"), f"source {source.get('id') or idx + 1}.extracted_at")
        except PreflightError as exc:
            errors.append(str(exc))

    scoreboard_markets = {str(r.get("market", "")).upper() for r in evidence.get("market_scoreboard") or []}
    economics = _economics_by_market(evidence)
    for market in included:
        if market not in scoreboard_markets:
            errors.append(f"market_scoreboard has no row for included market {market}")
        if market not in economics:
            errors.append(f"market_economics has no row for included market {market}")
        if md and not _narrative_has_market(md, market):
            errors.append(f"narrative has no market chapter for {market}")
    for market in excluded:
        if market in scoreboard_markets:
            errors.append(f"excluded market {market} appears in market_scoreboard")
        if md and _forbidden_market_in_narrative(md, market):
            errors.append(f"excluded market {market} appears in the narrative")
        evidence_text = json.dumps(evidence, ensure_ascii=False)
        full_name = MARKET_NAMES.get(market)
        if re.search(rf'"{re.escape(market)}"', evidence_text) or (
            full_name and re.search(rf"\b{re.escape(full_name)}\b", evidence_text, re.I)
        ):
            errors.append(f"excluded market {market} appears in the evidence manifest")
    for path, record in _iter_records(evidence):
        market = str(record.get("market", "")).upper()
        if market in excluded:
            errors.append(f"excluded market {market} appears at {path}")
        refs = record.get("source_refs")
        if refs is not None:
            for ref in refs:
                if str(ref) not in source_ids:
                    errors.append(f"unknown source ref {ref!r} at {path}")

    for section_name in ("market_scoreboard", "non_brand_rpc", "rank_queries", "engagement_delivery", "advertising_changes", "open_gaps", "actions"):
        for idx, row in enumerate(evidence.get(section_name) or []):
            if not row.get("source_refs"):
                errors.append(f"{section_name}[{idx}] requires source_refs")

    for market, econ in economics.items():
        status = econ.get("status")
        if status not in ECON_STATUSES:
            errors.append(f"{market} economics status must be one of {sorted(ECON_STATUSES)}")
        if status == "verified":
            if not isinstance(econ.get("break_even_acos"), (int, float)) or not 0 < econ["break_even_acos"] <= 1:
                errors.append(f"{market} verified economics requires break_even_acos as a ratio")
            if not econ.get("source_refs"):
                errors.append(f"{market} verified economics requires source_refs")

    for idx, record in enumerate(evidence.get("engagement_delivery") or []):
        for field in ("date", "area", "market", "built", "changed", "learned", "source_refs"):
            if not record.get(field):
                errors.append(f"engagement_delivery[{idx}] is missing {field}")

    for section_name in ("market_scoreboard", "non_brand_rpc"):
        for idx, row in enumerate(evidence.get(section_name) or []):
            market = str(row.get("market", "")).upper()
            verified = (economics.get(market) or {}).get("status") == "verified"
            if not verified:
                if row.get("required_rpc") not in (None, ""):
                    errors.append(f"{section_name}[{idx}] exposes Required RPC with unverified economics")
                if row.get("profitability_label") not in (None, ""):
                    errors.append(f"{section_name}[{idx}] exposes profitability with unverified economics")

    lag = int(cfg.get("attribution_lag_days", 14))
    for idx, change in enumerate(evidence.get("advertising_changes") or []):
        try:
            effective = _parse_date(change.get("effective_date"), f"advertising_changes[{idx}].effective_date")
            review = _parse_date(change.get("review_date"), f"advertising_changes[{idx}].review_date")
            recent = (cutoff.date() - effective).days < lag or review > cutoff.date()
            if recent and change.get("attribution_status") != "provisional":
                errors.append(f"advertising_changes[{idx}] must remain attribution-provisional until review")
            if review < effective:
                errors.append(f"advertising_changes[{idx}] review date precedes the change")
        except PreflightError as exc:
            errors.append(str(exc))
        for field in ("market", "summary", "details", "source_refs"):
            if not change.get(field):
                errors.append(f"advertising_changes[{idx}] is missing {field}")
        summary = str(change.get("summary") or "").strip()
        if summary and summary.lower() not in md_lower:
            errors.append(f"advertising_changes[{idx}] summary is absent from the narrative")

    required_sections = [
        r"executive summary", r"evidence watermark", r"engagement delivery record",
        r"account operating model", r"advertising change history",
        r"open items", r"client asset (?:and|&) link index",
        r"supersession", r"read-only",
    ]
    for pattern in required_sections:
        if not re.search(pattern, md_lower):
            errors.append(f"narrative is missing required section marker: {pattern}")
    if evidence.get("non_brand_rpc") and "non-brand rpc" not in md_lower:
        errors.append("advertising evidence exists but narrative has no Non-Brand RPC playbook")
    if any(evidence.get(k) for k in ("listing_assets", "creative_assets", "poe_findings")):
        if not all(term in md_lower for term in ("listing", "creative", "poe")):
            errors.append("listing/creative/PoE evidence exists but its handover section is incomplete")
        for term in ("reuse", "reorder", "localiz", "redesign"):
            if term not in md_lower:
                errors.append(f"listing/creative/PoE handover is missing the {term} decision rule")

    if not evidence.get("engagement_delivery"):
        errors.append("engagement_delivery must record what was built, changed, and learned")
    if not evidence.get("advertising_changes"):
        errors.append("advertising_changes must contain the material change history")
    if not evidence.get("open_gaps"):
        errors.append("open_gaps must disclose unresolved evidence or operating items")
    if not evidence.get("actions"):
        errors.append("actions must contain the operator queue")
    for idx, action in enumerate(evidence.get("actions") or []):
        for field in ("market", "area", "action", "owner", "timing", "trigger", "expected_outcome", "stop_condition", "review_date", "source_refs"):
            if not action.get(field):
                errors.append(f"actions[{idx}] is missing {field}")

    _validate_links(cfg, evidence, md, errors)
    branding_path, brand_dir, branding_data = _validate_branding(cfg, base, errors)

    destination = cfg.get("destination") or {}
    dest = str(destination.get("folder") or "")
    if destination.get("client_visible") is not True or destination.get("existing") is not True:
        errors.append("destination must be an existing client-visible folder")
    if not dest:
        errors.append("destination.folder is required")
    elif not Path(dest).expanduser().is_dir() and not (
        re.fullmatch(r"[A-Za-z0-9_-]{10,}", dest)
        or re.search(r"/folders/[A-Za-z0-9_-]{10,}", dest)
    ):
        errors.append("destination.folder must be an existing local folder, Drive folder ID, or Drive folder URL")
    try:
        _parse_timestamp(destination.get("verified_at"), "destination.verified_at")
    except PreflightError as exc:
        errors.append(str(exc))

    if errors:
        raise PreflightError("Preflight failed:\n- " + "\n- ".join(errors))
    return {
        "config": cfg, "evidence": evidence, "narrative": md,
        "config_path": config_path, "base": base,
        "narrative_path": narrative_path, "evidence_path": evidence_path,
        "branding_path": branding_path, "brand_dir": brand_dir,
        "branding": branding_data, "included": included, "excluded": excluded,
        "client": client, "cutoff": cutoff,
    }


def _style_module(cfg: dict):
    branding = importlib.import_module("branding")
    branding.activate_branding(cfg)
    style = importlib.import_module("ew_audit_style")
    style.configure_branding(cfg)
    return style


def _cell_value(row: dict, key: str, default=""):
    value = row.get(key, default)
    return default if value is None else value


def _setup_sheet(ws, title, subtitle, headers, widths, style):
    style.title_block(ws, title, subtitle, len(headers), banner=style.brand_banner("Amazon Account Handover"))
    style.header_row(ws, 5, headers, widths)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}5"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{max(5, ws.max_row)}"


def _write_rows(ws, rows, headers, keys, formats, style, left_cols=None, breakeven=0.5):
    left_cols = left_cols or tuple(range(1, len(headers) + 1))
    for ridx, row in enumerate(rows, 6):
        vals = [_cell_value(row, key) for key in keys]
        style.datarow(ws, ridx, vals, formats, left_cols=left_cols, breakeven=breakeven)
        for cidx in left_cols:
            ws.cell(ridx, cidx).alignment = style.WRAP
    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{max(5, ws.max_row)}"


def build_workbook(ctx: dict, out_path: Path) -> None:
    cfg, evidence = ctx["config"], ctx["evidence"]
    style = _style_module(cfg)
    money2 = style.EUR2 if str(cfg.get("currency", "EUR")).upper() in {"EUR", "€"} else style.USD2
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(WORKBOOK_TABS[0])
    headers = ["Field", "Value", "Source / Rule"]
    _setup_sheet(ws, WORKBOOK_TABS[0], f"Evidence cutoff: {cfg['cutoff_timestamp']}", headers, [30, 70, 70], style)
    rows = [
        {"field": "Client", "value": ctx["client"], "rule": "Run config"},
        {"field": "Included markets", "value": ", ".join(ctx["included"]), "rule": "Excluded markets never enter deliverables"},
        {"field": "Successor", "value": cfg.get("successor", ""), "rule": "Primary operator after handover"},
        {"field": "Audience", "value": cfg.get("audience", ""), "rule": "Client-facing"},
        {"field": "Currency", "value": cfg.get("currency", "EUR"), "rule": "Marketplace-native values remain native when needed"},
        {"field": "Evidence watermark", "value": cfg["cutoff_timestamp"], "rule": "Recent advertising changes remain provisional"},
        {"field": "Read-only scope", "value": "No Amazon, AdLabs, listing, inventory, or communication changes", "rule": "Builder and workflow contract"},
        {"field": "Economics rule", "value": "Profitability and Required RPC are withheld unless break-even ACOS is verified", "rule": "Evidence manifest"},
    ]
    for source in evidence.get("sources", []):
        rows.append({"field": f"Source: {source['name']}", "value": source["extracted_at"], "rule": source.get("coverage", "")})
    _write_rows(ws, rows, headers, ["field", "value", "rule"], [None, None, None], style)

    ws = wb.create_sheet(WORKBOOK_TABS[1])
    headers = ["Market", "Area", "Period", "Spend", "Ad Sales", "Clicks", "Orders", "CPC", "CVR", "RPC", "ACOS", "Economics Status", "Profitability", "Source Refs", "Diagnosis", "Next Move"]
    keys = ["market", "area", "period", "spend", "ad_sales", "clicks", "orders", "cpc", "cvr", "rpc", "acos", "economics_status", "profitability_label", "source_refs_text", "diagnosis", "next_move"]
    widths = [10, 18, 20, 13, 13, 11, 11, 11, 11, 11, 11, 18, 16, 18, 42, 42]
    _setup_sheet(ws, WORKBOOK_TABS[1], "Performance and operating status by in-scope market", headers, widths, style)
    market_rows = []
    economics = _economics_by_market(evidence)
    for row in evidence.get("market_scoreboard", []):
        r = dict(row)
        market = str(r.get("market", "")).upper()
        r["economics_status"] = (economics.get(market) or {}).get("status", "not-verified")
        r["source_refs_text"] = ", ".join(r.get("source_refs") or [])
        market_rows.append(r)
    _write_rows(ws, market_rows, headers, keys,
                [None, None, None, money2, money2, style.INT, style.INT, money2, style.PCT, money2, style.PCT, None, None, None, None, None],
                style, left_cols=(1, 2, 3, 12, 13, 14, 15, 16))

    ws = wb.create_sheet(WORKBOOK_TABS[2])
    headers = ["Market", "ASIN", "Pack", "Campaign", "Query / Target", "Brand Class", "Spend", "Clicks", "Orders", "Sales", "CTR", "CPC", "CVR", "AOV", "RPC", "ACOS", "Top of Search Share", "Organic Rank", "Required RPC", "Diagnosis", "Recommendation", "Owner", "Review Date"]
    widths = [9, 15, 12, 30, 30, 15, 12, 10, 10, 12, 10, 10, 10, 11, 11, 11, 17, 13, 14, 38, 42, 18, 14]
    _setup_sheet(ws, WORKBOOK_TABS[2], "RPC = Ad Sales / Clicks = CVR x AOV. Bid changes affect ACOS, not RPC.", headers, widths, style)
    econ = _economics_by_market(evidence)
    for ridx, row in enumerate(evidence.get("non_brand_rpc") or [], 6):
        for cidx, key in enumerate(["market", "asin", "pack", "campaign", "query_target", "brand_class", "spend", "clicks", "orders", "sales", "ctr", "cpc", "cvr", "aov", "rpc", "acos", "top_of_search_share", "organic_rank", "required_rpc", "diagnosis", "recommendation", "owner", "review_date"], 1):
            ws.cell(ridx, cidx, _cell_value(row, key))
        ws.cell(ridx, 12, f'=IFERROR(G{ridx}/H{ridx},0)')
        ws.cell(ridx, 13, f'=IFERROR(I{ridx}/H{ridx},0)')
        ws.cell(ridx, 14, f'=IFERROR(J{ridx}/I{ridx},0)')
        ws.cell(ridx, 15, f'=IFERROR(J{ridx}/H{ridx},0)')
        ws.cell(ridx, 16, f'=IFERROR(G{ridx}/J{ridx},0)')
        market = str(row.get("market", "")).upper()
        market_econ = econ.get(market) or {}
        if market_econ.get("status") == "verified":
            ws.cell(ridx, 19, f'=IFERROR(L{ridx}/{float(market_econ["break_even_acos"])},0)')
        else:
            ws.cell(ridx, 19, "")
        for cidx in range(1, len(headers) + 1):
            cell = ws.cell(ridx, cidx)
            cell.border = style.BORDER
            cell.font = style.F(10)
            cell.alignment = style.WRAP if cidx in (1, 2, 3, 4, 5, 6, 18, 20, 21, 22, 23) else style.RIGHT
        for cidx in (7, 10, 12, 14, 15, 19):
            ws.cell(ridx, cidx).number_format = money2
        for cidx in (11, 13, 16, 17):
            ws.cell(ridx, cidx).number_format = style.PCT
        for cidx in (8, 9, 18):
            ws.cell(ridx, cidx).number_format = style.INT
    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{max(5, ws.max_row)}"

    ws = wb.create_sheet(WORKBOOK_TABS[3])
    headers = ["Market", "ASIN", "Query", "Query Type", "Organic Rank", "Prior Rank", "Ad Rank / Share", "Status", "Opportunity", "Recommendation", "Owner", "Review Date", "Source Refs"]
    keys = ["market", "asin", "query", "query_type", "organic_rank", "prior_rank", "ad_rank_share", "status", "opportunity", "recommendation", "owner", "review_date", "source_refs_text"]
    widths = [9, 15, 30, 15, 13, 13, 16, 18, 36, 42, 18, 14, 20]
    _setup_sheet(ws, WORKBOOK_TABS[3], "Rank graduation, query opportunity, and verification queue", headers, widths, style)
    rank_rows = []
    for row in evidence.get("rank_queries") or []:
        r = dict(row); r["source_refs_text"] = ", ".join(r.get("source_refs") or []); rank_rows.append(r)
    _write_rows(ws, rank_rows, headers, keys, [None, None, None, None, style.INT, style.INT, None, None, None, None, None, None, None], style)

    ws = wb.create_sheet(WORKBOOK_TABS[4])
    headers = ["Market", "Area", "Action", "Owner", "Timing", "Trigger", "Expected Outcome", "Stop Condition", "Review Date", "Status", "Source Refs"]
    keys = ["market", "area", "action", "owner", "timing", "trigger", "expected_outcome", "stop_condition", "review_date", "status", "source_refs_text"]
    widths = [9, 18, 46, 18, 16, 34, 36, 36, 14, 16, 20]
    _setup_sheet(ws, WORKBOOK_TABS[4], "Owner-ready queue. Timing is evidence-led, not forced into arbitrary phases.", headers, widths, style)
    action_rows = []
    for row in evidence.get("actions") or []:
        r = dict(row); r["source_refs_text"] = ", ".join(r.get("source_refs") or []); action_rows.append(r)
    _write_rows(ws, action_rows, headers, keys, [None] * len(headers), style)

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = style.C["coral"]
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(out_path)


def build_doc(ctx: dict, outdir: Path, out_path: Path) -> None:
    cfg = dict(ctx["config"])
    cfg["client"] = ctx["client"]
    cfg["marketplaces"] = ctx["included"]
    cfg["date"] = ctx["cutoff"].date().isoformat()
    cfg["branding"] = dict(cfg.get("branding") or {})
    cfg["branding"]["branding_json"] = str(ctx["branding_path"])
    cfg["branding"]["brand_dir"] = str(ctx["brand_dir"])
    cfg["branding"]["doc_label"] = "Amazon Account Handover"
    cfg["branding"]["first_time"] = False
    metrics = {
        "currency": cfg.get("currency", "EUR"),
        "windows": {"business_report": cfg["comparison_windows"][0]["current_start"] + " to " + cfg["comparison_windows"][0]["current_end"]},
        "custom_kpis": [
            [str(len(ctx["included"])), "Markets in scope", ", ".join(ctx["included"])],
            [str(len(ctx["evidence"].get("engagement_delivery") or [])), "Delivery records", "built, changed, learned"],
            [str(len(ctx["evidence"].get("actions") or [])), "Open actions", "owner and stop condition assigned"],
            [ctx["cutoff"].date().isoformat(), "Evidence cutoff", "recent changes remain provisional"],
        ],
    }
    (outdir / "metrics.json").write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    branding = importlib.import_module("branding")
    branding.activate_branding(cfg)
    renderer = importlib.import_module("render_branded")
    rendered = renderer.render(cfg, outdir, ctx["narrative_path"], cover=False, brand_dir=ctx["brand_dir"])
    generated = Path(rendered["docx"])
    if generated != out_path:
        generated.replace(out_path)


def validate_outputs(docx_path: Path, xlsx_path: Path, branding: dict) -> None:
    errors = []
    if not docx_path.is_file() or docx_path.stat().st_size < 1000:
        errors.append("DOCX was not created or is implausibly small")
    else:
        with zipfile.ZipFile(docx_path) as archive:
            names = set(archive.namelist())
            doc_xml = archive.read("word/document.xml").decode("utf-8")
            styles_xml = archive.read("word/styles.xml").decode("utf-8")
            if "word/header1.xml" not in names or "word/footer1.xml" not in names:
                errors.append("DOCX is missing the branded running header or footer")
            if not any(name.startswith("word/media/") for name in names):
                errors.append("DOCX header contains no logo asset")
            font_name = (branding.get("fonts") or {}).get("doc_font_name", "Inter")
            if font_name not in styles_xml:
                errors.append(f"DOCX typography does not contain {font_name}")
            if "w:type=\"page\"" in doc_xml[-2000:]:
                errors.append("DOCX ends with an explicit page break that may strand a final page")
            if "w:titlePg" in doc_xml:
                errors.append("DOCX unexpectedly contains first-page cover treatment")
    if not xlsx_path.is_file() or xlsx_path.stat().st_size < 1000:
        errors.append("XLSX was not created or is implausibly small")
    else:
        wb = load_workbook(xlsx_path, data_only=False)
        if wb.sheetnames != WORKBOOK_TABS:
            errors.append(f"workbook tabs differ from required topology: {wb.sheetnames}")
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and "#REF!" in value:
                        errors.append(f"#REF! found in {ws.title}!{cell.coordinate}")
    if errors:
        raise PreflightError("Artifact validation failed:\n- " + "\n- ".join(errors))


def build(ctx: dict) -> dict:
    cfg = ctx["config"]
    output_dir = _resolve(ctx["base"], cfg.get("output_dir") or f"output/{_slug(ctx['client'])}/offboarding")
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = ctx["cutoff"].date().isoformat()
    markets = "-".join(ctx["included"])
    safe_client = re.sub(r"[^A-Za-z0-9._-]+", "_", ctx["client"]).strip("._")
    stem = f"{stamp}_{safe_client}_{markets}"
    docx_path = output_dir / f"{stem}_Amazon_Account_Handover_v1.docx"
    xlsx_path = output_dir / f"{stem}_Handover_Evidence_Workbook_v1.xlsx"
    build_doc(ctx, output_dir, docx_path)
    build_workbook(ctx, xlsx_path)
    validate_outputs(docx_path, xlsx_path, ctx["branding"])
    manifest = {
        "built_at": datetime.now(timezone.utc).isoformat(),
        "client": ctx["client"], "markets": ctx["included"],
        "evidence_cutoff": cfg["cutoff_timestamp"],
        "docx": str(docx_path), "xlsx": str(xlsx_path),
        "workbook_tabs": WORKBOOK_TABS,
        "branding_source": str(ctx["branding_path"]),
        "validated": True, "delivered": False,
    }
    manifest_path = output_dir / "handover-build-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return {"docx": docx_path, "xlsx": xlsx_path, "manifest": manifest_path}


def deliver(ctx: dict, outputs: dict) -> None:
    destination = str(ctx["config"]["destination"]["folder"])
    deliverer = REPO / "tools" / "gdrive-deliver" / "deliver.py"
    for key in ("docx", "xlsx"):
        local = outputs[key]
        subprocess.run([
            sys.executable, str(deliverer), str(local), destination,
            "--name", local.stem,
        ], cwd=REPO, check=True)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--preflight", action="store_true", help="Validate inputs without building")
    parser.add_argument("--deliver", action="store_true", help="Convert validated outputs to native Google files")
    args = parser.parse_args()
    try:
        ctx = preflight(args.config.resolve())
        print(f"[preflight] PASS: {ctx['client']} ({', '.join(ctx['included'])})")
        if args.preflight:
            return 0
        outputs = build(ctx)
        print(f"[build] DOCX: {outputs['docx']}")
        print(f"[build] XLSX: {outputs['xlsx']}")
        print(f"[build] manifest: {outputs['manifest']}")
        if args.deliver:
            deliver(ctx, outputs)
    except (PreflightError, subprocess.CalledProcessError) as exc:
        print(str(exc), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
