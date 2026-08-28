#!/usr/bin/env python3
"""Synthetic contract tests for the Amazon client offboarding builder."""
from __future__ import annotations

import copy
import importlib.util
import json
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw
from pypdf import PdfReader


HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
SPEC = importlib.util.spec_from_file_location("offboarding_builder", HERE / "build_handover.py")
builder = importlib.util.module_from_spec(SPEC)
assert SPEC.loader
SPEC.loader.exec_module(builder)


def write_json(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def make_branding(root: Path) -> tuple[Path, Path]:
    brand_dir = root / "brand"
    brand_dir.mkdir()
    logo = Image.new("RGBA", (800, 160), (255, 255, 255, 0))
    draw = ImageDraw.Draw(logo)
    draw.rectangle((10, 25, 790, 135), fill=(18, 24, 31, 255))
    draw.text((45, 65), "ECOM WIZARDS", fill=(255, 255, 255, 255))
    logo.save(brand_dir / "logo_black.png")
    font_candidates = [
        Path("/System/Library/Fonts/Supplemental/Arial.ttf"),
        Path("/System/Library/Fonts/SFNS.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    font_source = next((p for p in font_candidates if p.is_file()), None)
    if not font_source:
        raise RuntimeError("self-test needs one installed TrueType font")
    shutil.copyfile(font_source, brand_dir / "Inter-Variable.ttf")
    template = json.loads((REPO / "tools" / "amazon-ad-audit" / "branding.EXAMPLE-neutral.json").read_text())
    template["agency_name"] = "Ecom Wizards"
    template["agency_url"] = "https://example.com"
    template["fonts"]["doc_font_name"] = "Inter"
    template["fonts"]["doc_font_file"] = "Inter-Variable.ttf"
    template["assets"]["logo_black"] = "logo_black.png"
    template["assets"]["brand_dir"] = str(brand_dir)
    branding_path = root / "branding.json"
    write_json(branding_path, template)
    return branding_path, brand_dir


def fixture(root: Path) -> tuple[Path, dict, dict]:
    branding_path, brand_dir = make_branding(root)
    destination = root / "existing-client-folder"
    destination.mkdir()
    narrative = root / "handover.md"
    asset_url = "https://docs.google.com/document/d/SYNTHETICASSET123/edit"
    appendix_url = "https://drive.google.com/file/d/SYNTHETICAPPENDIX123/view"
    narrative.write_text(f"""# Synthetic Brand Amazon Account Handover

## Executive Summary and Evidence Watermark

Evidence is current through 2026-08-07T12:00:00+02:00. Two markets transfer with an evidence-led operating queue.

## Engagement Delivery Record

The engagement built the campaign system, changed routing, and learned where conversion constrains scale.

## Account Operating Model

The successor owns weekly rank, stock, query, negative, and economics checks.

## Advertising Change History

On 2026-08-05, the Latest bid set changed in DE. Its result remains provisional until 2026-08-19.

## DE - Germany

Advertising and rank evidence support controlled graduation.

## FR - France

The market needs a clean baseline and verified economics before profitability decisions.

## Non-Brand RPC Playbook

Improve CVR, AOV, query relevance, pack routing, fulfillment, and landing ASIN before calling a CPC reduction an RPC improvement.

## Listing, Creative and POE Handover

Reuse the image system. Reorder images and localize their text before commissioning a complete redesign.

## Open Items

FR break-even economics remain unverified.

## Client Asset and Link Index

- [Synthetic client asset]({asset_url})
- [Tactical appendix]({appendix_url})

## Supersession and Read-Only Statement

This handover supersedes the tactical appendix for ongoing decisions. It made no Amazon or communication changes.
""", encoding="utf-8")
    evidence = {
        "comparison_windows": [{"name": "Primary", "current_start": "2026-07-06", "current_end": "2026-08-04", "prior_start": "2026-06-06", "prior_end": "2026-07-05", "attribution_complete_through": "2026-08-04"}],
        "sources": [
            {"id": "ads", "name": "Synthetic ads export", "extracted_at": "2026-08-07T10:00:00+02:00", "coverage": "DE and FR"},
            {"id": "rank", "name": "Synthetic rank export", "extracted_at": "2026-08-07T10:05:00+02:00", "coverage": "DE and FR"},
        ],
        "market_economics": [
            {"market": "DE", "status": "verified", "break_even_acos": 0.35, "source_refs": ["ads"]},
            {"market": "FR", "status": "not-verified", "break_even_acos": None, "source_refs": []},
        ],
        "market_scoreboard": [
            {"market": "DE", "area": "Advertising", "period": "2026-07-06 to 2026-08-04", "spend": 20, "ad_sales": 60, "clicks": 10, "orders": 2, "cpc": 2, "cvr": 0.2, "rpc": 6, "acos": 1 / 3, "profitability_label": "", "diagnosis": "Controlled", "next_move": "Verify", "source_refs": ["ads"]},
            {"market": "FR", "area": "Readiness", "period": "At cutoff", "spend": 0, "ad_sales": 0, "clicks": 0, "orders": 0, "cpc": 0, "cvr": 0, "rpc": 0, "acos": 0, "profitability_label": "", "required_rpc": None, "diagnosis": "Baseline missing", "next_move": "Establish baseline", "source_refs": ["ads"]},
        ],
        "engagement_delivery": [
            {"date": "2026-07-01", "area": "Advertising", "market": "DE", "built": "Campaign system", "changed": "Routing", "learned": "Conversion constraint", "source_refs": ["ads"]}
        ],
        "advertising_changes": [
            {"effective_date": "2026-08-05", "review_date": "2026-08-19", "market": "DE", "summary": "Latest bid set", "details": "Complete synthetic entity record", "attribution_status": "provisional", "source_refs": ["ads"]}
        ],
        "non_brand_rpc": [
            {"market": "DE", "asin": "B000000001", "pack": "1-pack", "campaign": "Synthetic exact", "query_target": "generic protein drink", "brand_class": "Non-brand", "spend": 20, "clicks": 10, "orders": 2, "sales": 60, "ctr": 0.01, "top_of_search_share": 0.2, "organic_rank": 12, "required_rpc": 2 / 0.35, "diagnosis": "Healthy RPC / limited traffic", "recommendation": "Controlled exact test", "owner": "Successor", "review_date": "2026-08-19", "source_refs": ["ads"]},
            {"market": "FR", "asin": "B000000002", "pack": "1-pack", "campaign": "Synthetic exact", "query_target": "generic meal drink", "brand_class": "Non-brand", "spend": 12, "clicks": 8, "orders": 1, "sales": 25, "ctr": 0.008, "top_of_search_share": 0.1, "organic_rank": 31, "required_rpc": None, "diagnosis": "Weak conversion", "recommendation": "Fix offer first", "owner": "Successor", "review_date": "2026-08-19", "source_refs": ["ads"]},
        ],
        "rank_queries": [
            {"market": "DE", "asin": "B000000001", "query": "generic protein drink", "query_type": "Non-brand", "organic_rank": 12, "prior_rank": 20, "ad_rank_share": "Measured", "status": "Graduate", "opportunity": "Secondary query", "recommendation": "Reduce in stages", "owner": "Successor", "review_date": "2026-08-19", "source_refs": ["rank"]}
        ],
        "listing_assets": [{"market": "DE", "finding": "Reuse images", "source_refs": ["rank"]}],
        "creative_assets": [{"market": "DE", "finding": "Reorder and localize", "source_refs": ["rank"]}],
        "poe_findings": [{"market": "DE", "finding": "Country-specific language", "source_refs": ["rank"]}],
        "assets": [{"name": "Synthetic client asset", "url": asset_url, "client_accessible": True, "status": "verified", "verified_at": "2026-08-07T11:00:00+02:00"}],
        "open_gaps": [{"market": "FR", "area": "Economics", "gap": "Break-even unverified", "impact": "Withhold profitability", "owner": "Client", "review_date": "2026-08-19", "source_refs": ["ads"]}],
        "actions": [
            {"market": "DE", "area": "Advertising", "action": "Review test", "owner": "Successor", "timing": "At maturity", "trigger": "Attribution mature", "expected_outcome": "Decision evidence", "stop_condition": "Spend threshold without order", "review_date": "2026-08-19", "status": "Open", "source_refs": ["ads"]},
            {"market": "FR", "area": "Economics", "action": "Verify break-even", "owner": "Client", "timing": "Before scale", "trigger": "Approved finance source", "expected_outcome": "Valid guardrail", "stop_condition": "No approved source", "review_date": "2026-08-19", "status": "Open", "source_refs": ["ads"]},
        ],
    }
    evidence_path = root / "evidence.json"
    write_json(evidence_path, evidence)
    config = {
        "client": {"name": "Synthetic Brand", "slug": "synthetic-brand"},
        "successor": "Successor", "audience": "Client leadership", "included_markets": ["DE", "FR"],
        "excluded_markets": ["SE"], "currency": "EUR", "cutoff_timestamp": "2026-08-07T12:00:00+02:00",
        "attribution_lag_days": 14,
        "comparison_windows": [{"name": "Primary", "current_start": "2026-07-06", "current_end": "2026-08-04", "prior_start": "2026-06-06", "prior_end": "2026-07-05", "attribution_complete_through": "2026-08-04"}],
        "narrative_path": str(narrative), "evidence_path": str(evidence_path),
        "existing_appendix": {"name": "Tactical appendix", "url": appendix_url, "client_accessible": True, "status": "verified", "verified_at": "2026-08-07T11:00:00+02:00"},
        "branding": {"approved": True, "branding_json": str(branding_path), "brand_dir": str(brand_dir), "doc_label": "Amazon Account Handover"},
        "destination": {"folder": str(destination), "client_visible": True, "existing": True, "verified_at": "2026-08-07T11:00:00+02:00"},
        "output_dir": str(root / "output"),
    }
    config_path = root / "config.json"
    write_json(config_path, config)
    return config_path, config, evidence


def expect_failure(config_path: Path, config: dict, evidence: dict, mutate, label: str) -> None:
    cfg = copy.deepcopy(config)
    ev = copy.deepcopy(evidence)
    mutate(cfg, ev)
    write_json(Path(cfg["evidence_path"]), ev)
    write_json(config_path, cfg)
    try:
        builder.preflight(config_path)
    except builder.PreflightError:
        return
    raise AssertionError(f"expected preflight failure: {label}")


def run() -> None:
    with tempfile.TemporaryDirectory(prefix="offboarding-selftest-") as tmp:
        root = Path(tmp)
        config_path, config, evidence = fixture(root)
        ctx = builder.preflight(config_path)
        outputs = builder.build(ctx)
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            raise RuntimeError("self-test needs LibreOffice for visual-layout preflight")
        converted = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(root), str(outputs["docx"])],
            capture_output=True, text=True, check=True,
        )
        pdf_path = root / (Path(outputs["docx"]).stem + ".pdf")
        assert pdf_path.is_file(), converted.stdout + converted.stderr
        pdf = PdfReader(pdf_path)
        assert len(pdf.pages) >= 2
        first_text = pdf.pages[0].extract_text() or ""
        last_text = pdf.pages[-1].extract_text() or ""
        assert "Executive Summary" in first_text
        assert "AMAZON ACCOUNT HANDOVER" in first_text
        assert "example.com" in first_text
        assert len(last_text.strip()) > 100, "stranded or blank final page"
        wb = load_workbook(outputs["xlsx"], data_only=False)
        assert wb.sheetnames == builder.WORKBOOK_TABS
        readme = wb["Read Me & Data Watermark"]
        assert str(readme["A2"].value).startswith("ECOM WIZARDS")
        assert readme["A1"].fill.fgColor.rgb.endswith("141821")
        ws = wb["Non-Brand RPC Diagnostics"]
        assert ws["L6"].value == "=IFERROR(G6/H6,0)"
        assert ws["M6"].value == "=IFERROR(I6/H6,0)"
        assert ws["N6"].value == "=IFERROR(J6/I6,0)"
        assert ws["O6"].value == "=IFERROR(J6/H6,0)"
        assert ws["P6"].value == "=IFERROR(G6/J6,0)"
        assert ws["S6"].value == "=IFERROR(L6/0.35,0)"
        assert ws["S7"].value in (None, "")
        assert abs((20 / 10) / 0.35 - evidence["non_brand_rpc"][0]["required_rpc"]) < 1e-9

        cases = [
            (lambda c, e: c["excluded_markets"].append("DE"), "overlapping markets"),
            (lambda c, e: e["market_scoreboard"].pop(), "missing in-scope market coverage"),
            (lambda c, e: e["open_gaps"][0].update({"market": "SE"}), "forbidden market appearance"),
            (lambda c, e: c["comparison_windows"][0].update({"prior_start": "2026-06-07"}), "unequal windows"),
            (lambda c, e: c["comparison_windows"][0].update({"current_end": "2026-08-07", "attribution_complete_through": "2026-08-07"}), "current day"),
            (lambda c, e: e["sources"][0].pop("extracted_at"), "missing timestamp"),
            (lambda c, e: (e["assets"][0].update({"url": "https://app.slack.com/client/internal"}), Path(c["narrative_path"]).write_text(Path(c["narrative_path"]).read_text() + "\nhttps://app.slack.com/client/internal\n")), "internal link"),
            (lambda c, e: c["branding"].update({"branding_json": str(root / "missing-branding.json")}), "missing branding"),
            (lambda c, e: e["non_brand_rpc"][1].update({"required_rpc": 4.5}), "unverified Required RPC"),
            (lambda c, e: e["market_scoreboard"][1].update({"profitability_label": "Profitable"}), "unverified profitability"),
            (lambda c, e: e["advertising_changes"][0].update({"attribution_status": "settled"}), "recent change attribution"),
        ]
        for mutate, label in cases:
            Path(config["narrative_path"]).write_text(Path(config["narrative_path"]).read_text().replace("\nhttps://app.slack.com/client/internal\n", "\n"), encoding="utf-8")
            write_json(Path(config["evidence_path"]), evidence)
            write_json(config_path, config)
            expect_failure(config_path, config, evidence, mutate, label)
        print(f"[selftest] PASS: synthetic build and {len(cases)} negative preflight cases")


if __name__ == "__main__":
    run()
