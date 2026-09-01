#!/usr/bin/env python3
"""Build the branded V Gummies magnesium POE client workbook.

The builder consumes only canonical output from run-poe.mjs. Numeric source
metrics remain typed values; analytical scores and aggregate percentages are
worksheet formulas so the evidence chain stays auditable after Google Sheets
conversion.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "tools" / "amazon-ad-audit"))
import ew_audit_style as style  # noqa: E402


CAPTURE_DATE = "2026-08-27"
CLIENT = "V Gummies"
MARKETPLACE = "Amazon Spain"
ACCOUNT_LABEL = "CHIC&LOVE Spain"
PARTNER_ACCOUNT_ID = "A1OTXAUBQG95W3"
MARKETPLACE_ID = "A1RKKUPIHCS9HS"
REFERENCE_PRICE = 24.95
AUDIT_SOURCE = "https://docs.google.com/spreadsheets/d/10JxzwN_jcxjhGPr7WFaMGEaetTMwS3wn1ZIVdtFvnRg/edit"

KEEP_IDS = {
    "4385dbcae4444d7d1b494bfc6c1bb611",
    "d158cb29585f4fe19d11a50d99337861",
    "f25005e9-afb3-462d-9513-74b2d03a959c",
    "cac5214e-ef78-4e95-8ec0-a66e02631ae7",
    "3e42fc9fc7ba0733ec07425622daf8cc",
    "52cfa7164865b37ac0a4760c61f8e321",
    "ada39c12-4eaf-4008-b4e6-0eb530960b0e",
    "401238c5-087b-4e01-a336-ddef9da73275",
    "116935c5edfb4050325ce87aa4d119fa",
    "14ae6048-7028-457a-a840-f88f6562c5fc",
}

ROUTES = {
    "magnesio bisglicinato": "First product-page test. Largest retained demand pool and strongest observed niche conversion; keep the message complementary and non-medical.",
    "magnesio": "Broad reach test. Use neutral routine language and qualify placements tightly because the shelf is mature and highly reviewed.",
    "citrato de magnesio": "Second-wave product-page test. Meaningful demand with relatively open clicks; avoid digestive or efficacy promises.",
    "magnesio complex": "Selective test on relevant product pages. Strong long-run growth, but lower observed conversion makes creative relevance important.",
    "triple magnesio": "Controlled test. Price is closest to V Gummies, but click concentration is high, so start with conservative bids.",
    "treonato de magnesio": "Watchlist. Premium price fit is favorable, but the niche is concentrated and narrower.",
    "magnesio y potasio": "Watchlist. Same supplement shopper, but the combination creates a more specific expectation set.",
    "malato de magnesio": "Watchlist. Relevant form with modest scale and concentrated clicks.",
    "magnesio vitamina b6": "Watchlist. Small niche and combination-specific expectations limit broad testing.",
    "magnesio gominolas": "Format-learning niche, not a priority audience. Use it to understand gummy expectations and swallowing/taste language.",
}


def slug(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9áéíóúüñ]+", "-", text)
    return text.strip("-")


def number(value, default=0.0):
    if value in (None, "", "-"):
        return default
    return float(str(value).replace(",", "").replace("€", "").replace("%", "").strip())


def read_canonical_csv(path: Path, header_starts: str) -> list[dict]:
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    idx = next((i for i, line in enumerate(lines) if line.startswith(header_starts)), None)
    if idx is None:
        raise ValueError(f"header '{header_starts}' not found in {path}")
    return list(csv.DictReader(lines[idx:]))


def exclusion_reason(title: str) -> str:
    t = title.lower()
    if any(x in t for x in ("imanix", "connetix", "calleras", "cerveza")):
        return "Excluded: wrong category/use"
    if any(x in t for x in ("liquido", "escalada", "carbonato", "aceite", "sales de epsom")):
        return "Excluded: topical/sports-chalk or non-oral use"
    if any(x in t for x in ("belevels", "solgar", "life pro", "sura vitasan", "ana maria lajusticia")):
        return "Excluded: brand-led niche"
    if any(x in t for x in ("colageno", "inositol", "triptofano", "vitamina b", "zma", "calcio")):
        return "Excluded: adjacent supplement with a different primary job"
    if "polvo" in t:
        return "Excluded: redundant powder-form niche"
    if any(x in t for x in ("zinc", "cloruro")):
        return "Excluded: redundant combination/form niche"
    return "Excluded: lower-priority duplicate or weaker fit"


def load_inputs(data_dir: Path):
    full_paths = sorted(data_dir.glob(f"{CAPTURE_DATE}_poe_es-*_niche-full.json"))
    if len(full_paths) != 10:
        raise ValueError(f"expected 10 full niche packs, found {len(full_paths)}")

    niches = []
    search_terms = []
    products = []
    reviews = []
    returns = []
    coverage = []

    for full_path in full_paths:
        envelope = json.loads(full_path.read_text(encoding="utf-8"))
        niche = envelope["niche"]
        niche_id = niche["nicheId"]
        if niche_id not in KEEP_IDS:
            raise ValueError(f"unexpected retained niche {niche_id}")
        name = niche["nicheTitle"]
        summary = niche["nicheSummary"]
        trend = niche["trendsMetrics"][-1]
        source_url = envelope.get("url", "")
        niches.append({
            "name": name,
            "id": niche_id,
            "sv360": number(summary.get("searchVolumeT360")),
            "growth360": number(summary.get("searchVolumeGrowthT360")),
            "sv90": number(summary.get("searchVolumeT90")),
            "growth90": number(summary.get("searchVolumeGrowthT90")),
            "conversion": number(trend.get("searchConversionRateT7")),
            "top5": number(trend.get("top5ProductsClickShareT7")),
            "avg_price": number(summary.get("avgPrice")),
            "rating": number(trend.get("avgRatingsOfProducts")),
            "return_rate": number(summary.get("returnRateT360")),
            "product_count": int(number(summary.get("productCount"))),
            "dataset_date": trend.get("datasetDate", ""),
            "source_url": source_url,
        })

        stem = f"{CAPTURE_DATE}_es-{slug(name)}"
        review_path = data_dir / f"{CAPTURE_DATE}_poe_es-{slug(name)}_customer-review-insights.csv"
        review_json_path = data_dir / f"{CAPTURE_DATE}_poe_es-{slug(name)}_customer-review-insights.json"
        return_path = data_dir / f"{CAPTURE_DATE}_poe_es-{slug(name)}_returns.csv"
        search_path = data_dir / f"{stem}_NicheDetailsSearchTermsTab.csv"
        product_path = data_dir / f"{stem}_NicheDetailsProductsTab.csv"

        review_rows = read_canonical_csv(review_path, "Topic,")
        review_json = json.loads(review_json_path.read_text(encoding="utf-8"))
        rating_impacts = {
            item["topic"]: {
                "all": item.get("starRatingImpactAllProducts"),
                "top25": item.get("starRatingImpactTop25PercentProducts"),
            }
            for item in review_json.get("productStarRatingImpact", [])
        }
        return_rows = read_canonical_csv(return_path, "Topic,")
        search_rows = read_canonical_csv(search_path, "Search term,")
        product_rows = read_canonical_csv(product_path, "Product Name,")
        if not all((review_rows, return_rows, search_rows, product_rows)):
            raise ValueError(f"incomplete full pack for {name}")

        for row in review_rows:
            impact = rating_impacts.get(row.get("Topic", ""), {})
            reviews.append({
                "niche": name,
                "niche_id": niche_id,
                "sentiment": row.get("Sentiment", ""),
                "topic": row.get("Topic", ""),
                "subtopic": row.get("Subtopic", ""),
                "mentions": number(row.get("% Mentions")) / 100,
                "star_impact_all": impact.get("all"),
                "star_impact_top25": impact.get("top25"),
                "snippets": row.get("Review Snippets", ""),
                "source_url": source_url,
            })
        sentiment_topics = {row.get("Topic", "") for row in review_rows}
        for topic, impact in rating_impacts.items():
            if topic in sentiment_topics:
                continue
            reviews.append({
                "niche": name,
                "niche_id": niche_id,
                "sentiment": "rating-only",
                "topic": topic,
                "subtopic": "",
                "mentions": None,
                "star_impact_all": impact.get("all"),
                "star_impact_top25": impact.get("top25"),
                "snippets": "",
                "source_url": source_url,
            })
        for row in return_rows:
            returns.append({
                "niche": name,
                "niche_id": niche_id,
                "topic": row.get("Topic", ""),
                "mentions": number(row.get("% Mentions (returns, past 6 months)")) / 100,
                "source_url": source_url,
            })
        for row in search_rows:
            search_terms.append({"niche": name, "niche_id": niche_id, "source_url": source_url, **row})
        for row in product_rows:
            products.append({"niche": name, "niche_id": niche_id, "source_url": source_url, **row})
        coverage.append({
            "niche": name,
            "products": len(product_rows),
            "search_terms": len(search_rows),
            "review_sentiment_rows": len(review_rows),
            "star_rating_topics": len(rating_impacts),
            "returns": len(return_rows),
        })

    discovery = {}
    provenance = defaultdict(list)
    seed_counts = []
    for raw_path in sorted(data_dir.glob("raw_search_*.json")):
        env = json.loads(raw_path.read_text(encoding="utf-8"))
        if env.get("capturedAt", "")[:10] != CAPTURE_DATE:
            continue
        query = env.get("query", "")
        seed_counts.append((query, len(env.get("niches", [])), env.get("url", "")))
        for item in env.get("niches", []):
            discovery.setdefault(item["nicheId"], item)
            provenance[item["nicheId"]].append(query)

    if len(discovery) != 36:
        raise ValueError(f"expected 36 discovered niches, found {len(discovery)}")
    return niches, search_terms, products, reviews, returns, coverage, discovery, provenance, seed_counts


def setup_sheet(ws, title, subtitle, width, widths):
    ws.sheet_view.showGridLines = False
    style.title_block(ws, title, subtitle, width, banner=style.brand_banner("Product Opportunity Explorer"))
    for i, value in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = value
    ws.freeze_panes = "A5"


def write_table(ws, row, headers, rows, formats, left_cols):
    style.header_row(ws, row, headers)
    start = row + 1
    for ridx, values in enumerate(rows, start):
        style.datarow(ws, ridx, values, formats, left_cols=left_cols)
    if rows:
        ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{start + len(rows) - 1}"
    return start + len(rows)


def build(data_dir: Path, output_path: Path, validation_path: Path):
    niches, search_terms, products, reviews, returns, coverage, discovery, provenance, seed_counts = load_inputs(data_dir)
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    review_evidence = {}
    for niche in niches:
        rows = [row for row in reviews if row["niche"] == niche["name"] and row["sentiment"] in {"positive", "negative"}]
        positive = max((row for row in rows if row["sentiment"] == "positive"), key=lambda row: row["mentions"], default=None)
        negative = max((row for row in rows if row["sentiment"] == "negative"), key=lambda row: row["mentions"], default=None)
        parts = []
        if positive:
            parts.append(f"Positive: {positive['topic']} ({positive['mentions']:.1%})")
        if negative:
            parts.append(f"Risk: {negative['topic']} ({negative['mentions']:.1%})")
        review_evidence[niche["name"]] = "; ".join(parts)

    # Full Niche Map first because summary sheets reference it.
    ws = wb.create_sheet("Full Niche Map")
    headers = ["Rank", "Decision", "Niche", "Niche ID", "Search volume 360d", "Growth 360d", "Search volume 90d", "Growth 90d", "Search CVR", "Top-5 click share", "Avg price", "Price gap vs V Gummies", "Avg rating", "Rating threshold", "Return rate", "Products", "Opportunity score", "Testing route", "Review-language evidence", "POE source"]
    setup_sheet(ws, "Full Magnesium Niche Map", "Ten relevant oral-supplement niches retained from 36 discovered niches. Score = demand × openness × capped trend × conversion × price fit × rating gate.", len(headers), [8, 12, 24, 36, 16, 12, 16, 12, 11, 14, 12, 16, 11, 16, 11, 10, 14, 62, 46, 42])
    style.note(ws, 4, f"Reference V Gummies price: €{REFERENCE_PRICE:.2f}, from the 13 Aug 2026 audit. Review and return metrics are niche-wide, not V Gummies-specific.", len(headers))
    style.header_row(ws, 5, headers)
    for ridx, niche in enumerate(sorted(niches, key=lambda x: x["name"]), 6):
        values = [
            f"=RANK.EQ(Q{ridx},$Q$6:$Q$15)",
            f'=IF(A{ridx}<=5,"Shortlist","Watch")',
            niche["name"], niche["id"], niche["sv360"], niche["growth360"], niche["sv90"], niche["growth90"],
            niche["conversion"], niche["top5"], niche["avg_price"], f"=K{ridx}-{REFERENCE_PRICE}", niche["rating"], f'=IF(M{ridx}>=4.3,"Pass ≥4.3","Below 4.3")', niche["return_rate"], niche["product_count"],
            f"=ROUND((E{ridx}/MAX($E$6:$E$15))*(1-J{ridx})*MIN(1.5,MAX(0.5,1+F{ridx}))*(0.5+0.5*I{ridx}/MAX($I$6:$I$15))*(1-MIN(1,ABS(K{ridx}-{REFERENCE_PRICE})/{REFERENCE_PRICE})*0.25)*IF(M{ridx}>=4.3,1,0.75)*100,1)",
            ROUTES[niche["name"]], review_evidence[niche["name"]], niche["source_url"],
        ]
        fmts = [style.INT, None, None, None, style.INT, style.PCT, style.INT, style.PCT, style.PCT, style.PCT, style.EUR2, style.EUR2, "0.00", None, style.PCT2, style.INT, "0.0", None, None, None]
        style.datarow(ws, ridx, values, fmts, left_cols=(2, 3, 4, 14, 18, 19, 20))
        ws.row_dimensions[ridx].height = 44
        if ridx % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws.cell(ridx, col).fill = style.SUB_FILL
    ws.auto_filter.ref = "A5:T15"
    ws.freeze_panes = "C6"

    # Ranked shortlist uses formula links into the full map.
    ws = wb.create_sheet("Niche Shortlist")
    headers = ["Rank", "Niche", "Score", "Search volume 360d", "Growth 360d", "Search CVR", "Top-5 click share", "Avg price", "Price gap", "Avg rating", "Rating threshold", "Return rate", "Review-language evidence", "Recommended route"]
    setup_sheet(ws, "Magnesium Product-Page Test Shortlist", "Priority audiences for controlled Sponsored Brands Video or product-page testing. Ranking is formula-linked to the full niche map.", len(headers), [8, 24, 10, 16, 12, 11, 14, 12, 12, 11, 16, 11, 48, 66])
    style.note(ws, 4, "No causal or efficacy claim between apple cider vinegar and magnesium is asserted. Validate any future health wording before use.", len(headers))
    style.header_row(ws, 5, headers)
    map_cols = {2: "C", 3: "Q", 4: "E", 5: "F", 6: "I", 7: "J", 8: "K", 9: "L", 10: "M", 11: "N", 12: "O", 13: "S", 14: "R"}
    for ridx, rank in enumerate(range(1, 6), 6):
        ws.cell(ridx, 1, rank)
        for col, source_col in map_cols.items():
            ws.cell(ridx, col, f"=INDEX('Full Niche Map'!${source_col}$6:${source_col}$15,MATCH($A{ridx},'Full Niche Map'!$A$6:$A$15,0))")
        formats = [style.INT, None, "0.0", style.INT, style.PCT, style.PCT, style.PCT, style.EUR2, style.EUR2, "0.00", None, style.PCT2, None, None]
        for col, fmt in enumerate(formats, 1):
            cell = ws.cell(ridx, col)
            cell.border = style.BORDER
            cell.font = style.F(10)
            cell.alignment = style.LEFT_DATA if col in (2, 11, 13, 14) else style.RIGHT
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[ridx].height = 50
    ws.auto_filter.ref = "A5:N10"
    ws.freeze_panes = "A6"

    # Raw review rows and formula-driven theme aggregation.
    raw_review = wb.create_sheet("Raw POE Reviews")
    raw_review_headers = ["Niche", "Niche ID", "Sentiment", "Topic", "Subtopic", "% Mentions", "Star impact, all products", "Star impact, top 25%", "Review snippets", "POE source"]
    setup_sheet(raw_review, "Raw POE Customer Review Insights", "Niche-wide positive, negative, and star-rating-impact topics from ten complete POE packs.", len(raw_review_headers), [24, 36, 12, 30, 22, 12, 18, 18, 90, 42])
    raw_review_rows = [[r["niche"], r["niche_id"], r["sentiment"], r["topic"], r["subtopic"], r["mentions"], r["star_impact_all"], r["star_impact_top25"], r["snippets"], r["source_url"]] for r in reviews]
    write_table(raw_review, 5, raw_review_headers, raw_review_rows, [None, None, None, None, None, style.PCT2, "0.00", "0.00", None, None], (1, 2, 3, 4, 5, 9, 10))
    for row in range(6, 6 + len(raw_review_rows)):
        raw_review.row_dimensions[row].height = 42

    theme = wb.create_sheet("Review & Pain-Point Themes")
    theme_headers = ["Topic", "Sentiment", "Niches mentioning", "Coverage", "Avg % mentions", "Example customer language", "Interpretation"]
    setup_sheet(theme, "Cross-Niche Review and Pain-Point Themes", "Average topic share and coverage across the ten retained niches. Formula-linked to Raw POE Reviews.", len(theme_headers), [30, 12, 18, 12, 16, 86, 66])
    style.note(theme, 4, "A topic appearing across many niches is safer evidence than a high percentage in one niche. Snippets are customer language, not approved advertising claims.", len(theme_headers))
    topic_snippets = defaultdict(list)
    for r in reviews:
        key = (r["topic"], r["sentiment"])
        for snippet in [x.strip() for x in r["snippets"].split("|") if x.strip()]:
            if snippet not in topic_snippets[key]:
                topic_snippets[key].append(snippet)
    topic_keys = sorted(topic_snippets, key=lambda x: (x[1] != "negative", x[0].lower()))
    style.header_row(theme, 5, theme_headers)
    raw_end = 5 + len(raw_review_rows)
    for ridx, (topic_name, sentiment) in enumerate(topic_keys, 6):
        example = " | ".join(topic_snippets[(topic_name, sentiment)][:3])
        interpretation = "Expectation or friction to avoid amplifying" if sentiment == "negative" else "Customer-valued language to validate and mirror carefully"
        niches_mentioning = len({row["niche"] for row in reviews if row["topic"] == topic_name and row["sentiment"] == sentiment})
        vals = [topic_name, sentiment,
                niches_mentioning,
                f"=C{ridx}/10",
                f'=AVERAGEIFS(\'Raw POE Reviews\'!$F$6:$F${raw_end},\'Raw POE Reviews\'!$D$6:$D${raw_end},A{ridx},\'Raw POE Reviews\'!$C$6:$C${raw_end},B{ridx})',
                example, interpretation]
        style.datarow(theme, ridx, vals, [None, None, style.INT, style.PCT, style.PCT2, None, None], left_cols=(1, 2, 6, 7))
        theme.row_dimensions[ridx].height = 48
    theme.auto_filter.ref = f"A5:G{5 + len(topic_keys)}"

    raw_returns = wb.create_sheet("Raw POE Returns")
    raw_return_headers = ["Niche", "Niche ID", "Return topic", "% Mentions", "POE source"]
    setup_sheet(raw_returns, "Raw POE Return Insights", "Niche-wide return topics from the past six months in each retained POE pack.", len(raw_return_headers), [24, 36, 34, 14, 44])
    raw_return_rows = [[r["niche"], r["niche_id"], r["topic"], r["mentions"], r["source_url"]] for r in returns]
    write_table(raw_returns, 5, raw_return_headers, raw_return_rows, [None, None, None, style.PCT2, None], (1, 2, 3, 5))

    ret = wb.create_sheet("Returns & Expectation Risks")
    ret_headers = ["Return topic", "Niches mentioning", "Coverage", "Avg % mentions", "Risk read", "Implication for V Gummies testing"]
    setup_sheet(ret, "Returns and Expectation Risks", "Cross-niche return signals. Formula-linked to Raw POE Returns and explicitly separated from V Gummies reviews.", len(ret_headers), [34, 18, 12, 16, 54, 70])
    style.note(ret, 4, "These are magnesium-niche return signals. They do not establish that V Gummies causes or solves any listed outcome.", len(ret_headers))
    return_topics = sorted({r["topic"] for r in returns})
    style.header_row(ret, 5, ret_headers)
    return_end = 5 + len(raw_return_rows)
    for ridx, topic_name in enumerate(return_topics, 6):
        risk = "High recurring expectation risk" if topic_name in {"Funcionalidad-General", "Producto anunciado versus Producto Real", "Efectos secundarios", "Dosificación"} else "Monitor in creative and comments"
        implication = "Do not borrow efficacy or outcome language; keep the ad focused on a complementary routine and the actual V Gummies product facts."
        niches_mentioning = len({row["niche"] for row in returns if row["topic"] == topic_name})
        vals = [topic_name,
                niches_mentioning,
                f"=B{ridx}/10",
                f"=AVERAGEIF('Raw POE Returns'!$C$6:$C${return_end},A{ridx},'Raw POE Returns'!$D$6:$D${return_end})",
                risk, implication]
        style.datarow(ret, ridx, vals, [None, style.INT, style.PCT, style.PCT2, None, None], left_cols=(1, 5, 6))
        ret.row_dimensions[ridx].height = 48
    ret.auto_filter.ref = f"A5:F{5 + len(return_topics)}"

    # Combined raw search-term and product structures.
    st = wb.create_sheet("Search Terms")
    st_headers = ["Niche", "Search term", "SV 360d", "Growth 90d", "Growth 180d", "Click share 360d", "Conversion 360d", "Top product 1", "ASIN 1", "Top product 2", "ASIN 2", "Top product 3", "ASIN 3", "POE source"]
    setup_sheet(st, "Magnesium Search Terms", "Canonical POE search-term rows across all retained niches.", len(st_headers), [24, 30, 14, 12, 12, 14, 14, 58, 14, 58, 14, 58, 14, 42])
    st_rows = []
    for row in search_terms:
        st_rows.append([row["niche"], row.get("Search term", ""), number(row.get("Search Volume (Past 360 days)")), number(row.get("Search Volume Growth (Past 90 days)")), number(row.get("Search Volume Growth (Past 180 days)")), number(row.get("Click Share (Past 360 days)")), number(row.get("Search Conversion Rate (Past 360 days)")), row.get("Top Clicked Product 1 (Title)", ""), row.get("Top Clicked Product 1 (Asin)", ""), row.get("Top Clicked Product 2 (Title)", ""), row.get("Top Clicked Product 2 (Asin)", ""), row.get("Top Clicked Product 3 (Title)", ""), row.get("Top Clicked Product 3 (Asin)", ""), row["source_url"]])
    write_table(st, 5, st_headers, st_rows, [None, None, style.INT, style.PCT, style.PCT, style.PCT, style.PCT, None, None, None, None, None, None, None], (1, 2, 8, 9, 10, 11, 12, 13, 14))

    prod = wb.create_sheet("Products & Price Structure")
    prod_headers = ["Niche", "Product", "ASIN", "Brand", "Category", "Launch date", "Clicks 360d", "Click share", "Avg price", "Ratings", "Rating", "Avg BSR", "Avg buyable offers", "POE source"]
    setup_sheet(prod, "Products and Price Structure", "Canonical POE product rows across all retained niches. Use for placement and shelf-context decisions.", len(prod_headers), [24, 74, 14, 20, 42, 13, 14, 12, 12, 12, 10, 12, 16, 42])
    prod_rows = []
    for row in products:
        prod_rows.append([row["niche"], row.get("Product Name", ""), row.get("ASIN", ""), row.get("Brand", ""), row.get("Category", ""), row.get("Launch date", ""), number(row.get("Niche Click Count (Past 360 days)")), number(row.get("Click Share (past 360 days)")), number(row.get("Average Selling Price (past 360 days) (EUR)")), int(number(row.get("Total Ratings"))), number(row.get("Average Customer Rating")), number(row.get("Average BSR")), number(row.get("Buyable Offer Average 1P+3P (past 360 days)")), row["source_url"]])
    write_table(prod, 5, prod_headers, prod_rows, [None, None, None, None, None, None, style.INT, style.PCT, style.EUR2, style.INT, "0.0", "0.0", "0.0", None], (1, 2, 3, 4, 5, 6, 14))

    # Sources and manifest includes coverage, seeds, and every discovered niche.
    src = wb.create_sheet("Sources & Run Manifest")
    setup_sheet(src, "Sources and Run Manifest", "Account-safe capture manifest, coverage reconciliation, discovery provenance, exclusions, and limitations.", 9, [22, 30, 42, 18, 18, 18, 58, 58, 24])
    style.band(src, 5, "Run identity and limitations", 9)
    identity_rows = [
        ["Client", CLIENT, "Marketplace", MARKETPLACE, "Capture date", CAPTURE_DATE, "Account", f"{ACCOUNT_LABEL} · {PARTNER_ACCOUNT_ID}", ""],
        ["Marketplace ID", MARKETPLACE_ID, "Discovered", 36, "Retained", 10, "Cache", "Fresh capture; no cache reused", ""],
        ["Limitation", "POE uses trailing windows", "Review scope", "Niche-wide", "Claims", "Not validated", "External send", "Not performed", ""],
        ["Archive", "Canonical captures retained locally", "pCloud", "Skipped: no canonical V Gummies client folder", "Policy", "No folder invented", "Workbook", "Native Google Sheet only", ""],
    ]
    for ridx, row in enumerate(identity_rows, 6):
        style.datarow(src, ridx, row, [None] * 9, left_cols=tuple(range(1, 10)))
    style.band(src, 10, "Full-pack reconciliation", 9)
    style.header_row(src, 11, ["Niche", "Products", "Search terms", "Review rows", "Star topics", "Return topics", "Full-pack JSON", "Review JSON", "Status"])
    for ridx, item in enumerate(sorted(coverage, key=lambda x: x["niche"]), 12):
        s = slug(item["niche"])
        vals = [item["niche"], item["products"], item["search_terms"], item["review_sentiment_rows"], item["star_rating_topics"], item["returns"], f"{CAPTURE_DATE}_poe_es-{s}_niche-full.json", f"{CAPTURE_DATE}_poe_es-{s}_customer-review-insights.json", "Complete"]
        style.datarow(src, ridx, vals, [None, style.INT, style.INT, style.INT, style.INT, style.INT, None, None, None], left_cols=(1, 7, 8, 9))
    row = 23
    style.band(src, row, "Discovery seeds", 9)
    row += 1
    style.header_row(src, row, ["Seed", "Seed category", "Niches returned", "Marketplace", "Capture date", "Status", "POE source", "Notes", ""])
    categories = {
        "magnesio": "head term", "suplemento de magnesio": "product", "bisglicinato de magnesio": "form", "glicinato de magnesio": "form", "citrato de magnesio": "form", "magnesio complejo": "form", "triple magnesio": "form", "magnesio para dormir": "use case", "magnesio para músculos": "use case", "magnesio para mujer": "audience",
    }
    for query, count, url in seed_counts:
        row += 1
        style.datarow(src, row, [query, categories.get(query, "seed"), count, "ES", CAPTURE_DATE, "Complete", url, "Zero-result seeds retained as evidence" if count == 0 else "", ""], [None, None, style.INT, None, None, None, None, None, None], left_cols=(1, 2, 4, 5, 6, 7, 8, 9))
    row += 2
    style.band(src, row, "Discovered niche disposition", 9)
    row += 1
    style.header_row(src, row, ["Niche", "Niche ID", "Seeds", "Disposition", "SV 360d", "Growth 360d", "Avg price", "Reason", ""])
    for niche_id, item in sorted(discovery.items(), key=lambda kv: kv[1].get("nicheTitle", "")):
        row += 1
        kept = niche_id in KEEP_IDS
        summary = item.get("nicheSummary") or {}
        vals = [item.get("nicheTitle", ""), niche_id, ", ".join(provenance[niche_id]), "Retained" if kept else "Excluded", number(summary.get("searchVolumeT360")), number(summary.get("searchVolumeGrowthT360")), number(summary.get("avgPrice")), "Retained: complete oral-magnesium niche" if kept else exclusion_reason(item.get("nicheTitle", "")), ""]
        style.datarow(src, row, vals, [None, None, None, None, style.INT, style.PCT, style.EUR2, None, None], left_cols=(1, 2, 3, 4, 8, 9))
    row += 2
    style.band(src, row, "Source links", 9)
    row += 1
    links = [
        ["POE", "Amazon Seller Central", "CHIC&LOVE Spain", CAPTURE_DATE, "Products, search terms, reviews, returns, trends", "Read-only", "https://sellercentral.amazon.de/opportunity-explorer/explore", "Fresh same-day capture"],
        ["Reference price", "V Gummies ES Audit", CLIENT, "2026-08-13", "V Gummies price €24.95", "Read-only", AUDIT_SOURCE, "Used only for price-distance context"],
    ]
    style.header_row(src, row, ["Source", "System", "Scope", "Date", "Evidence", "Access", "URL", "Notes", ""])
    for item in links:
        row += 1
        style.datarow(src, row, item + [""], [None] * 9, left_cols=tuple(range(1, 10)))
    src.freeze_panes = "A11"

    # Executive summary last so formula references are guaranteed to exist.
    exe = wb.create_sheet("Executive Summary", 0)
    setup_sheet(exe, f"{CLIENT} | Spain Magnesium POE & Customer Insights", f"Fresh capture {CAPTURE_DATE} · 36 discovered niches · 10 complete retained packs · client-ready read-only analysis", 8, [26, 20, 18, 18, 22, 22, 48, 60])
    style.band(exe, 5, "What the data says", 8)
    style.header_row(exe, 6, ["Metric", "Value", "Source", "Read", "Metric", "Value", "Source", "Read"])
    summary_rows = [
        ["Discovered niches", "='Sources & Run Manifest'!D7", "Run manifest", "Wide discovery, not merchant niches", "Complete packs", "='Sources & Run Manifest'!F7", "Run manifest", "10/10 reconciled"],
        ["Largest retained niche", "=INDEX('Full Niche Map'!$C$6:$C$15,MATCH(MAX('Full Niche Map'!$E$6:$E$15),'Full Niche Map'!$E$6:$E$15,0))", "Full Niche Map", "Demand leader", "Best opportunity score", "=INDEX('Full Niche Map'!$C$6:$C$15,MATCH(MAX('Full Niche Map'!$Q$6:$Q$15),'Full Niche Map'!$Q$6:$Q$15,0))", "Full Niche Map", "Demand, openness, trend, conversion, price fit, rating gate"],
        ["Top recurring risk", "Expectation mismatch", "Returns + reviews", "Do not overpromise outcomes", "Reference price", REFERENCE_PRICE, "V Gummies audit", "Compared with each niche average"],
    ]
    for ridx, vals in enumerate(summary_rows, 7):
        style.datarow(exe, ridx, vals, [None, None, None, None, None, None, None, None], left_cols=tuple(range(1, 9)))
        exe.row_dimensions[ridx].height = 38
    style.band(exe, 11, "Recommended test order", 8)
    style.header_row(exe, 12, ["Rank", "Niche", "Why now", "Guardrail", "Rank", "Niche", "Why now", "Guardrail"])
    priorities = [
        [1, "magnesio bisglicinato", "Largest retained demand pool and strongest observed CVR", "No absorption or efficacy promise", 2, "magnesio", "Broadest general magnesium audience", "Tight placement relevance"],
        [3, "citrato de magnesio", "Meaningful scale with more open clicks", "No digestive outcome promise", 4, "magnesio complex", "Strong long-run growth", "Lower CVR; creative must explain relevance"],
        [5, "triple magnesio", "Price closest to V Gummies", "High concentration; conservative bids", "", "", "", ""],
    ]
    for ridx, vals in enumerate(priorities, 13):
        style.datarow(exe, ridx, vals, [style.INT, None, None, None, style.INT, None, None, None], left_cols=(2, 3, 4, 6, 7, 8))
        exe.row_dimensions[ridx].height = 46
    style.band(exe, 17, "How to use this pack", 8)
    instructions = [
        ["1", "Choose product pages", "Start with the five shortlisted niches and use the Products tab to identify relevant ASINs.", "2", "Use customer language", "Mine repeated words from positive and negative topics, but treat snippets as evidence, not approved claims."],
        ["3", "Protect expectations", "Use the Returns tab to remove messages likely to create advertised-versus-actual disappointment.", "4", "Validate before launch", "Legal/compliance review remains required for any health, mechanism, or comparative wording."],
    ]
    style.header_row(exe, 18, ["Step", "Action", "Use", "Step", "Action", "Use"])
    for ridx, vals in enumerate(instructions, 19):
        style.datarow(exe, ridx, vals, [None] * 6, left_cols=(1, 2, 3, 4, 5, 6))
        exe.row_dimensions[ridx].height = 52
    style.note(exe, 22, "Scope boundary: this workbook does not contain the separate video script, does not recommend Amazon account changes, and has not been sent to V Gummies.", 8)
    exe.freeze_panes = "A6"

    desired_order = [
        "Executive Summary", "Niche Shortlist", "Review & Pain-Point Themes",
        "Returns & Expectation Risks", "Search Terms", "Products & Price Structure",
        "Full Niche Map", "Raw POE Reviews", "Raw POE Returns", "Sources & Run Manifest",
    ]
    wb._sheets = [wb[name] for name in desired_order]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    validation = {
        "status": "PASS",
        "capture_date": CAPTURE_DATE,
        "account": {"label": ACCOUNT_LABEL, "partnerAccountId": PARTNER_ACCOUNT_ID, "marketplaceId": MARKETPLACE_ID},
        "discovered_niches": len(discovery),
        "retained_full_packs": len(niches),
        "raw_rows": {
            "products": len(products),
            "search_terms": len(search_terms),
            "review_records": len(reviews),
            "review_sentiment_rows": sum(item["review_sentiment_rows"] for item in coverage),
            "star_rating_topics": sum(item["star_rating_topics"] for item in coverage),
            "returns": len(returns),
        },
        "coverage": coverage,
        "sheets": wb.sheetnames,
    }
    validation_path.write_text(json.dumps(validation, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"workbook": str(output_path), "validation": validation}, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--validation", type=Path, required=True)
    args = parser.parse_args()
    build(args.data_dir, args.output, args.validation)


if __name__ == "__main__":
    main()
