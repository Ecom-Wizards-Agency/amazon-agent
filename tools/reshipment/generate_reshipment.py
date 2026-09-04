#!/usr/bin/env python3
"""Generate FBA reshipment / inventory-overview workbooks from same-day Seller Central reports.

Client list, run date, and input paths are read from a JSON config (default:
`tools/reshipment/config.json`, which is gitignored). Copy `config.TEMPLATE.json`
to `config.json` and fill in the current run before running:

    python3 tools/reshipment/generate_reshipment.py --config tools/reshipment/config.json

Outputs land under `<output_root>/output/<client key>/inventory/`.
"""
import argparse
import csv
import json
import math
import re
import subprocess
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook = None


# Repo root = two levels up from this file (tools/reshipment/generate_reshipment.py).
REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG = Path(__file__).resolve().parent / "config.json"
PROFILE_LOOKUP = REPO_ROOT / "tools" / "client-profiles" / "find-client-profile.mjs"


def load_shared_profile(profile_key):
    result = subprocess.run(
        ["node", str(PROFILE_LOOKUP), profile_key],
        cwd=REPO_ROOT,
        check=False,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        detail = result.stderr.strip() or result.stdout.strip() or "unknown lookup error"
        raise ValueError(f"Could not load shared profile {profile_key!r}: {detail}")
    payload = json.loads(result.stdout)
    exact = [
        profile
        for profile in payload.get("matches", [])
        if profile.get("profile_key") == profile_key
    ]
    if len(exact) != 1:
        raise ValueError(f"Expected one exact shared profile for {profile_key!r}, found {len(exact)}")
    return exact[0]


def load_config(config_path):
    cfg = json.loads(Path(config_path).read_text(encoding="utf-8"))
    cfg.setdefault("report_days", 30)
    cfg.setdefault("downloads_dir", "~/Downloads")
    cfg.setdefault("output_root", str(REPO_ROOT))
    cfg["downloads"] = Path(cfg["downloads_dir"]).expanduser()
    cfg["output_root"] = Path(cfg["output_root"]).expanduser()

    def resolve(value):
        if not value:
            return None
        p = Path(value).expanduser()
        return p if p.is_absolute() else cfg["downloads"] / value

    forbidden_overrides = {"target_days", "multiplier", "fba_exclude_patterns", "fba_exclude_asins"}
    if forbidden_overrides.intersection(cfg):
        raise ValueError("Stable reshipment settings must come from the shared Amazon Ops profile, not top-level config")

    for client in cfg.get("clients", []):
        present_overrides = sorted(forbidden_overrides.intersection(client))
        if present_overrides:
            raise ValueError(
                f"{client.get('key', 'Client')} duplicates shared profile fields in config: {', '.join(present_overrides)}"
            )
        profile_key = client.get("profile_key")
        if not profile_key:
            raise ValueError(f"{client.get('key', 'Client')} requires profile_key")
        profile = load_shared_profile(profile_key)
        reshipment = profile.get("reshipment") or {}
        if reshipment.get("enabled") is not True:
            raise ValueError(f"Shared profile {profile_key!r} is not enabled for reshipment planning")
        target_days = reshipment.get("effective_coverage_days")
        multiplier = reshipment.get("scaling_multiplier")
        if not isinstance(target_days, (int, float)) or not isinstance(multiplier, (int, float)):
            raise ValueError(f"Shared profile {profile_key!r} lacks complete numeric reshipment settings")
        client["target_days"] = target_days
        client["multiplier"] = multiplier
        client["minimum_monthly_sales_for_fba"] = reshipment.get("minimum_monthly_sales_for_fba")
        client["fba_exclude_patterns"] = reshipment.get("fba_exclude_patterns", [])
        client["fba_exclude_asins"] = reshipment.get("fba_exclude_asins", [])
        # Optional per-ASIN display and packing data. Both live in the shared
        # profile like every other stable planning input: carton_units maps
        # ASIN -> units per shipping carton (from Send to Amazon case-pack
        # templates or the client), short_names maps ASIN -> the name humans
        # use, so Slack never shows a truncated Amazon title.
        carton_specs = reshipment.get("carton_specs") or {}
        client["carton_units"] = {
            **{
                asin: spec.get("units_per_box")
                for asin, spec in carton_specs.items()
                if isinstance(spec, dict) and spec.get("units_per_box")
            },
            **(reshipment.get("carton_units") or {}),
        }
        client["carton_specs"] = carton_specs
        client["short_names"] = reshipment.get("short_names") or {}
        client["profile_source"] = profile.get("source_file")
        client["profile_components"] = {
            "target_stock_days": reshipment.get("target_stock_days"),
            "lead_time_days": reshipment.get("lead_time_days"),
            "amazon_booking_buffer_days": reshipment.get("amazon_booking_buffer_days"),
        }
        for field in ("fba", "business", "inventory", "restock"):
            client[field] = resolve(client.get(field))
    return cfg


def n(value):
    if value is None:
        return 0
    text = str(value).strip().replace(",", "").replace("$", "").replace("€", "").replace("%", "")
    if not text or text == "--":
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def i(value):
    return int(round(n(value)))


def read_rows(path):
    if not path or not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        delimiter = "\t" if "\t" in sample and sample.count("\t") > sample.count(",") else ","
        return list(csv.DictReader(fh, delimiter=delimiter))


def add_item(items, asin, **updates):
    if not asin:
        return
    item = items[asin]
    item["asin"] = asin
    for key, value in updates.items():
        if value in ("", None):
            continue
        if key in {"title", "sku"}:
            item.setdefault(key, value)
        elif key == "sources":
            item.setdefault("sources", set()).update(value)
        else:
            item[key] = item.get(key, 0) + value


def calculate(client, cfg):
    report_days = cfg["report_days"]
    target_days = client["target_days"]
    multiplier = client["multiplier"]
    minimum_monthly_sales = client.get("minimum_monthly_sales_for_fba")
    fba_exclude_patterns = [
        re.compile(pattern, re.IGNORECASE)
        for pattern in client.get("fba_exclude_patterns", [])
    ]
    fba_exclude_asins = {
        str(asin).strip().upper()
        for asin in client.get("fba_exclude_asins", [])
        if str(asin).strip()
    }
    run_date = cfg["run_date"]
    output_root = cfg["output_root"]
    planning_blocker = str(client.get("planning_blocker") or "").strip()
    # AWD stock replenishes FBA on its own, so it counts against the send
    # quantity. awd_by_key maps SKU or ASIN (case-insensitive) to units; when
    # only an account-level total exists it must NOT be subtracted blindly
    # (there is no per-product attribution), it becomes a loud warning instead.
    awd_by_key = {
        str(key).strip().upper(): int(value or 0)
        for key, value in (client.get("awd_by_key") or {}).items()
        if str(key).strip()
    }
    awd_stored = int(client.get("awd_stored") or 0)
    awd_consumed = set()
    committed_by_key = {
        str(key).strip().upper(): int(value or 0)
        for key, value in (client.get("committed_shipments_by_key") or {}).items()
        if str(key).strip()
    }
    committed_consumed = set()

    items = defaultdict(dict)
    source_files = []

    for row in read_rows(client["business"]):
        asin = row.get("(Child) ASIN") or row.get("ASIN")
        add_item(
            items,
            asin,
            title=row.get("Title"),
            business_units=i(row.get("Units Ordered")),
            sources={"business"},
        )
    if client["business"] and client["business"].exists():
        source_files.append(str(client["business"]))

    for row in read_rows(client["fba"]):
        asin = row.get("asin")
        add_item(
            items,
            asin,
            sku=row.get("sku"),
            title=row.get("product-name"),
            available=i(row.get("available")),
            inbound=i(row.get("inbound-quantity")),
            reserved=i(row.get("Total Reserved Quantity")),
            fba_fc_transfer=i(row.get("FC Transfer")),
            fba_customer_order=i(row.get("Customer Order Reserved")),
            fba_fc_processing=i(row.get("FC Processing")),
            unfulfillable=i(row.get("unfulfillable-quantity")),
            fba_units_30=i(row.get("units-shipped-t30")),
            fba_units_7=i(row.get("units-shipped-t7")),
            estimated_excess=i(row.get("estimated-excess-quantity")),
            days_of_supply=i(row.get("days-of-supply")),
            sources={"fba"},
        )
    if client["fba"] and client["fba"].exists():
        source_files.append(str(client["fba"]))

    for row in read_rows(client["restock"]):
        if client.get("restock_country") and row.get("Country") != client["restock_country"]:
            continue
        asin = row.get("ASIN")
        add_item(
            items,
            asin,
            sku=row.get("Merchant SKU"),
            title=row.get("Product Name"),
            restock_units_30=i(row.get("Units Sold Last 30 Days")),
            restock_total=i(row.get("Total Units")),
            restock_inbound=i(row.get("Inbound")),
            restock_available=i(row.get("Available")),
            restock_fc_transfer=i(row.get("FC transfer")),
            restock_fc_processing=i(row.get("FC Processing")),
            restock_customer_order=i(row.get("Customer Order")),
            unfulfillable=i(row.get("Unfulfillable")),
            sources={"restock"},
        )
    if client["restock"] and client["restock"].exists():
        source_files.append(str(client["restock"]))

    for row in read_rows(client["inventory"]):
        asin = row.get("asin")
        add_item(items, asin, sku=row.get("sku"), price=n(row.get("price")), sources={"inventory"})
    if client["inventory"] and client["inventory"].exists():
        source_files.append(str(client["inventory"]))
    for evidence in client.get("evidence_files") or []:
        evidence_path = Path(evidence).expanduser()
        if evidence_path.exists():
            source_files.append(str(evidence_path))

    rows = []
    for asin, item in sorted(items.items()):
        demand = item.get("business_units") or item.get("fba_units_30") or item.get("restock_units_30") or 0
        demand_source = "Business Report" if item.get("business_units") else ("FBA Inventory" if item.get("fba_units_30") else "Restock Report")
        available = max(item.get("available", 0), item.get("restock_available", 0))
        inbound = max(item.get("inbound", 0), item.get("restock_inbound", 0))
        fc_transfer = max(item.get("fba_fc_transfer", 0), item.get("restock_fc_transfer", 0))
        fc_processing = max(item.get("fba_fc_processing", 0), item.get("restock_fc_processing", 0))
        customer_order = max(item.get("fba_customer_order", 0), item.get("restock_customer_order", 0))
        reserved = max(
            item.get("reserved", 0),
            fc_transfer + fc_processing + customer_order,
        )
        required = demand / report_days * multiplier * target_days
        product_identity = f"{item.get('title', '')} {item.get('sku', '')}"
        fba_excluded = (
            asin.strip().upper() in fba_exclude_asins
            or any(pattern.search(product_identity) for pattern in fba_exclude_patterns)
        )
        below_fba_threshold = (
            isinstance(minimum_monthly_sales, (int, float))
            and demand < minimum_monthly_sales
        )
        sku_key = str(item.get("sku", "")).strip().upper()
        asin_key = asin.strip().upper()
        awd_key = sku_key if sku_key in awd_by_key else (asin_key if asin_key in awd_by_key else None)
        awd_units = awd_by_key[awd_key] if awd_key else 0
        if awd_key:
            awd_consumed.add(awd_key)
        committed_key = (sku_key if sku_key in committed_by_key
                         else asin_key if asin_key in committed_by_key else None)
        committed_units = committed_by_key[committed_key] if committed_key else 0
        if committed_key:
            committed_consumed.add(committed_key)
        raw_reship = (
            0
            if (fba_excluded or below_fba_threshold)
            else int(math.ceil(max(0, required - available - inbound
                                   - fc_transfer - fc_processing - awd_units
                                   - committed_units)))
        )
        per_carton = client["carton_units"].get(asin)
        per_carton = int(per_carton) if isinstance(per_carton, (int, float)) and per_carton > 0 else None
        reship_boxes = math.ceil(raw_reship / per_carton) if raw_reship and per_carton else ""
        reship = reship_boxes * per_carton if reship_boxes else raw_reship
        excess = item.get("estimated_excess", 0)
        if not excess and demand > 0 and available > (demand / report_days * 120):
            excess = int(max(0, available - demand / report_days * 90))
        rows.append(
            {
                "ASIN": asin,
                "SKU": item.get("sku", ""),
                "Product Name": item.get("title", ""),
                "Demand 30d": demand,
                "Demand Source": demand_source,
                "Available": available,
                "Inbound": inbound,
                "Committed Shipment Units": committed_units,
                "Reserved": reserved,
                "FC Transfer": fc_transfer,
                "FC Processing": fc_processing,
                "Customer Order Reserved": customer_order,
                "AWD Units": awd_units,
                "Unfulfillable": item.get("unfulfillable", 0),
                f"Required Units ({target_days}d x {multiplier})": round(required, 1),
                "Raw Reshipment Units": raw_reship,
                "Units per Box": per_carton or "",
                "Reshipment Boxes": reship_boxes,
                "Reshipment Units": reship,
                "Estimated Excess Units": excess,
                "Days of Supply": item.get("days_of_supply", ""),
                "FBA Units 7d": item.get("fba_units_7", 0),
                "FBA Units 30d": item.get("fba_units_30", 0),
                "Restock Units 30d": item.get("restock_units_30", 0),
                "FBA Eligibility": (
                    "FBM only - bundle/multipack"
                    if fba_excluded
                    else "Below monthly FBA threshold"
                    if below_fba_threshold
                    else "Eligible"
                ),
            }
        )

    rows.sort(key=lambda r: r["Reshipment Units"], reverse=True)
    out_dir = output_root / "output" / client["key"] / "inventory"
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{run_date}_Inventory Overview_{client['brand']}_{client['market']}"
    csv_path = out_dir / f"{stem}.csv"
    xlsx_path = out_dir / f"{stem}.xlsx"
    slack_path = out_dir / f"{stem}_slack.txt"
    json_path = out_dir / f"{stem}_manifest.json"

    fields = list(rows[0].keys()) if rows else ["ASIN"]
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    if Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Overview"
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row in rows:
            ws.append([row[k] for k in fields])
        send = wb.create_sheet("Reshipment")
        send.append(fields)
        for row in [r for r in rows if r["Reshipment Units"] > 0]:
            send.append([row[k] for k in fields])
        excess = wb.create_sheet("Excess Inventory")
        excess.append(fields)
        for row in [r for r in rows if r["Estimated Excess Units"] > 0]:
            excess.append([row[k] for k in fields])
        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            for col in sheet.columns:
                letter = col[0].column_letter
                sheet.column_dimensions[letter].width = min(48, max(12, max(len(str(c.value or "")) for c in col[:80]) + 2))
        wb.save(xlsx_path)

    send_rows = [r for r in rows if r["Reshipment Units"] > 0]
    excess_rows = [r for r in rows if r["Estimated Excess Units"] > 0]
    missing_carton_asins = sorted(
        r["ASIN"] for r in send_rows if not r["Units per Box"]
    )
    awd_attributed = sum(awd_by_key[key] for key in awd_consumed)
    awd_unattributed = max(0, awd_stored - awd_attributed)
    committed_attributed = sum(committed_by_key[key] for key in committed_consumed)
    committed_unattributed = sum(
        units for key, units in committed_by_key.items() if key not in committed_consumed)
    parts = [
        f"Source: {client['notes']} Shared profile: {client['profile_key']} ({target_days} effective coverage days, {multiplier}x demand multiplier). Output saved: `{csv_path.name}` / `{xlsx_path.name}`.",
    ]
    if planning_blocker:
        parts.insert(0, f"⚠️ PLAN BLOCKED - {planning_blocker} Calculated quantities are diagnostic only and must not be shipped.")
    # Slack lines lead with the actionable quantity in bold: the units needed
    # were mid-line and easy to miss, and the truncated Amazon title with its
    # "..." buried them further (feedback 31.08.2026). A profile short_name
    # replaces the title entirely; without one the title is clipped at a word
    # boundary, never with an ellipsis. When carton_units knows the pack size,
    # the recommendation is whole boxes - nobody ships 1,234 units in 80-packs.
    def display_name(row):
        name = client["short_names"].get(row["ASIN"])
        if name:
            return name
        title = re.sub(r"\s+", " ", str(row["Product Name"] or "")).strip()
        return title if len(title) <= 44 else title[:44].rsplit(" ", 1)[0]

    def quantity_lead(row):
        if row["Units per Box"]:
            lead = (f"*Send {row['Reshipment Boxes']} boxes × {row['Units per Box']} = "
                    f"{row['Reshipment Units']:,} units*")
            if row["Reshipment Units"] != row["Raw Reshipment Units"]:
                lead += f" (raw need {row['Raw Reshipment Units']:,})"
            return lead
        return f"*Send {row['Reshipment Units']:,} units* (carton size missing)"

    if send_rows:
        parts.extend(["", "*Reshipment*"])
        for row in send_rows[:10]:
            lead = quantity_lead(row)
            awd_note = f" · awd {row['AWD Units']:,}" if row["AWD Units"] else ""
            committed_note = (f" · committed shipment {row['Committed Shipment Units']:,}"
                              if row["Committed Shipment Units"] else "")
            parts.append(
                f"{lead} - {display_name(row)} (`{row['ASIN']}`) · avail {row['Available']:,} · inbound {row['Inbound']:,}{committed_note} · reserved {row['Reserved']:,}{awd_note}"
            )
        if len(send_rows) > 10:
            parts.append(f"Plus {len(send_rows) - 10} more low-volume rows in the workbook.")
        rounded_total = sum(r["Reshipment Units"] for r in send_rows)
        raw_total = sum(r["Raw Reshipment Units"] for r in send_rows)
        total_note = f"Total: {rounded_total:,} shippable units"
        if rounded_total != raw_total:
            total_note += f" (raw need {raw_total:,})"
        parts.append(total_note)
        if missing_carton_asins:
            parts.append(
                f"⚠️ Carton size is missing for {len(missing_carton_asins)} positive-send ASIN(s); "
                "those rows remain provisional unit quantities."
            )
    else:
        parts.extend(["", "*Reshipment*", "No positive reshipment quantities from today’s source data."])
    if awd_unattributed:
        parts.append(
            f"⚠️ AWD holds {awd_unattributed:,} unit(s) with no per-SKU attribution; send quantities may overshoot by up to that amount."
        )
    if committed_unattributed:
        parts.append(
            f"⚠️ {committed_unattributed:,} committed shipment unit(s) could not be mapped to a current FBA row."
        )
    if excess_rows:
        parts.extend(["", "*Excess Inventory / Plan Sales*"])
        for row in excess_rows[:6]:
            parts.append(
                f"*{row['Estimated Excess Units']:,} excess* - {display_name(row)} (`{row['ASIN']}`) · avail {row['Available']:,} · 30d demand {row['Demand 30d']:,}"
            )
        if len(excess_rows) > 6:
            parts.append(f"Plus {len(excess_rows) - 6} more excess rows in the workbook.")
        parts.append(f"Total: {sum(r['Estimated Excess Units'] for r in excess_rows):,} excess units")
    slack_path.write_text("\n".join(parts) + "\n", encoding="utf-8")

    manifest = {
        "account": client["key"],
        "date": run_date,
        "actionable": not bool(planning_blocker),
        "planningBlocker": planning_blocker or None,
        "profileKey": client["profile_key"],
        "profileSource": client["profile_source"],
        "planningComponents": client["profile_components"],
        "effectiveCoverageDays": target_days,
        "scalingMultiplier": multiplier,
        "minimumMonthlySalesForFba": minimum_monthly_sales,
        "sources": source_files,
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
        "slack": str(slack_path),
        "sendRows": len(send_rows),
        "sendUnits": sum(r["Reshipment Units"] for r in send_rows),
        "rawSendUnits": sum(r["Raw Reshipment Units"] for r in send_rows),
        "sendBoxes": sum(r["Reshipment Boxes"] for r in send_rows if r["Reshipment Boxes"]),
        "missingCartonAsins": missing_carton_asins,
        "awdUnitsAttributed": awd_attributed,
        "awdUnitsUnattributed": awd_unattributed,
        "committedShipmentUnitsAttributed": committed_attributed,
        "committedShipmentUnitsUnattributed": committed_unattributed,
        "excessRows": len(excess_rows),
        "excessUnits": sum(r["Estimated Excess Units"] for r in excess_rows),
        "notes": client["notes"],
    }
    json_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest


def main():
    parser = argparse.ArgumentParser(description="Generate FBA reshipment / inventory-overview workbooks.")
    parser.add_argument(
        "--config",
        default=str(DEFAULT_CONFIG),
        help="Path to the run config JSON (default: tools/reshipment/config.json).",
    )
    parser.add_argument(
        "--check-config",
        action="store_true",
        help="Resolve and validate shared profiles without reading reports or writing outputs.",
    )
    args = parser.parse_args()

    cfg = load_config(args.config)
    if args.check_config:
        print(
            json.dumps(
                [
                    {
                        "key": client["key"],
                        "profileKey": client["profile_key"],
                        "profileSource": client["profile_source"],
                        "planningComponents": client["profile_components"],
                        "effectiveCoverageDays": client["target_days"],
                        "scalingMultiplier": client["multiplier"],
                        "minimumMonthlySalesForFba": client[
                            "minimum_monthly_sales_for_fba"
                        ],
                    }
                    for client in cfg.get("clients", [])
                ],
                indent=2,
            )
        )
        return
    manifests = [calculate(client, cfg) for client in cfg.get("clients", [])]
    bundle = cfg["output_root"] / "output" / f"reshipment-plans-{cfg['run_date']}" / "summary.json"
    bundle.parent.mkdir(parents=True, exist_ok=True)
    bundle.write_text(json.dumps(manifests, indent=2), encoding="utf-8")
    print(json.dumps(manifests, indent=2))


if __name__ == "__main__":
    main()
