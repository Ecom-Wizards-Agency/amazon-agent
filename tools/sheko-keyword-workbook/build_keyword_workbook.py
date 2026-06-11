#!/usr/bin/env python3
"""
Sheko DE Kollagen — reusable keyword-research workbook builder.

Deterministically rebuilds the Sheko keyword workbook from raw DataDive +
Amazon Product Opportunity Explorer (POE) exports, while preserving the style
and structure of an existing template workbook.

Design
------
* Template-base: an existing .xlsx is loaded as the styled base, so column
  widths, freeze panes, tab colors, and curated tabs are preserved verbatim.
* Exact-paste tabs (DataDive root/master, POE products/search-terms) are
  cleared and re-pasted 1:1 from their source CSVs. They are NEVER transformed.
* Rebuilt tabs:
    - "POE Raw - Related Niches": rebuilt from the related-niches JSON with a
      STRICT keep-list relevance filter (drops unrelated `glow`/skincare drift).
    - "Outlier - Opportunity KWs": rule-based triage over the master keyword
      list, plus a Final Action column.
* Curated tabs are carried forward from the template unchanged.
* All rules live in config.json (keep-list, brand/form/claim tokens, action
  map) so the workflow is auditable and reusable next cycle.
* Validation + an evidence manifest (JSON) are emitted every run.

Usage
-----
    .venv/bin/python tools/sheko-keyword-workbook/build_keyword_workbook.py \
        [--config tools/sheko-keyword-workbook/config.json] \
        [--template "<v2 .xlsx>"] \
        [--out "output/sheko/seo/...v3.xlsx"] \
        [--manifest "output/sheko/seo/..._summary.json"] \
        [--roots-csv ...] [--master-csv ...] [--competitors-csv ...] \
        [--poe-products-csv ...] [--poe-search-terms-csv ...] \
        [--related-niches-json ...]

Defaults target the 2026-06-10 Sheko cycle. The script does NOT write to Google
Drive; copy the reviewed local output there separately.
"""
from __future__ import annotations

import argparse
import copy
import csv
import datetime as _dt
import json
import os
import re
import shutil
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------- #
# Repo + default paths
# --------------------------------------------------------------------------- #
REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
HOME = os.path.expanduser("~")

DEFAULTS = {
    "config": os.path.join(os.path.dirname(__file__), "config.json"),
    "template": os.path.join(
        REPO, "output/sheko/seo/Sheko DE Kollagen Powder Keyword Research 2026-06-10 v2.xlsx"
    ),
    "out": os.path.join(
        REPO, "output/sheko/seo/Sheko DE Kollagen Powder Keyword Research 2026-06-10 v3.xlsx"
    ),
    "manifest": os.path.join(
        REPO, "output/sheko/seo/2026-06-10_sheko_keyword_research_v3_summary.json"
    ),
    "roots_csv": os.path.join(HOME, "Downloads/niche-m6202AaAgV-niche-analysis-roots.csv"),
    "master_csv": os.path.join(HOME, "Downloads/master list-niche-m6202AaAgV-keywords.csv"),
    "competitors_csv": os.path.join(HOME, "Downloads/niche-m6202AaAgV-competitors.csv"),
    "poe_products_csv": os.path.join(HOME, "Downloads/NicheDetailsProductsTab_10_06_2026.csv"),
    "poe_search_terms_csv": os.path.join(
        HOME, "Downloads/NicheDetailsSearchTermsTab_10_06_2026.csv"
    ),
    "related_niches_json": os.path.join(
        REPO,
        "output/sheko/opportunity-data/2026-06-10_poe_collagen-pulver_related-niches_chrome.json",
    ),
    "poe_reviews_json": "",
    "poe_returns_json": "",
    "poe_structured_json": "",
    "seo_content": os.path.join(os.path.dirname(__file__), "seo_content.json"),
    "curated_tabs": "",  # optional: JSON of curated tab content (e.g. POE Raw - Reviews) to write deterministically.
    "drive_dir": "",  # optional: folder to copy the finished .xlsx into (e.g. the synced Drive folder). Off by default.
}


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def read_csv(path: str) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def norm(s: Any) -> str:
    """Normalize text for matching: lowercase, collapse whitespace."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def to_int(x: Any):
    try:
        return int(float(str(x).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def to_float(x: Any):
    try:
        return float(str(x).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# --------------------------------------------------------------------------- #
# Style preservation
# --------------------------------------------------------------------------- #
def snapshot_styles(ws: Worksheet, upto_row: int = 8) -> list[list[dict]]:
    """Capture per-cell style for the first `upto_row` rows so they can be
    re-applied after the tab is cleared and re-pasted."""
    snap: list[list[dict]] = []
    rows = min(ws.max_row, upto_row) if ws.max_row else 0
    cols = ws.max_column or 1
    for r in range(1, rows + 1):
        row_styles = []
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            row_styles.append(
                {
                    "font": copy.copy(cell.font),
                    "fill": copy.copy(cell.fill),
                    "border": copy.copy(cell.border),
                    "alignment": copy.copy(cell.alignment),
                    "number_format": cell.number_format,
                }
            )
        snap.append(row_styles)
    return snap


def apply_style(cell, snap: list[list[dict]], row_idx: int, col_idx: int) -> None:
    """Apply a captured style. Rows past the snapshot reuse the last captured
    row (uniform data styling); columns past it reuse the last captured col."""
    if not snap:
        return
    r = min(row_idx, len(snap)) - 1
    row_styles = snap[r]
    if not row_styles:
        return
    c = min(col_idx, len(row_styles)) - 1
    st = row_styles[c]
    cell.font = copy.copy(st["font"])
    cell.fill = copy.copy(st["fill"])
    cell.border = copy.copy(st["border"])
    cell.alignment = copy.copy(st["alignment"])
    cell.number_format = st["number_format"]


def clear_rows(ws: Worksheet) -> None:
    """Remove all row content but keep sheet-level props (widths via
    column_dimensions, freeze_panes, tabColor all persist)."""
    if ws.max_row and ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)


def write_matrix(ws: Worksheet, rows: list[list[Any]], snap: list[list[dict]]) -> int:
    """Write a 2D matrix starting at A1, applying captured styles. Returns the
    number of rows written."""
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(i, j, value=val)
            apply_style(cell, snap, i, j)
    return len(rows)


# --------------------------------------------------------------------------- #
# Tab builders
# --------------------------------------------------------------------------- #
def build_exact_paste(ws: Worksheet, csv_rows: list[list[str]], warnings: list[str]) -> int:
    """Clear a tab and paste CSV rows 1:1 with preserved styling."""
    snap = snapshot_styles(ws, upto_row=8)
    clear_rows(ws)
    n = write_matrix(ws, csv_rows, snap)
    return n


def build_related_niches(
    ws: Worksheet, related_json_path: str, cfg: dict, result: dict, warnings: list[str]
) -> dict:
    """Rebuild the related-niches tab from the POE related-niches JSON with the
    strict keep-list filter. Preserves the template's metadata + column-label
    header rows (everything above the niche data block)."""
    keep = [norm(x) for x in cfg["related_niche_filter"]["keep"]]
    exclude = [norm(x) for x in cfg["related_niche_filter"]["exclude_examples"]]

    data = json.load(open(related_json_path, encoding="utf-8"))
    # Two capture shapes are supported:
    #  (a) tables[0] = {headers, rows} (chrome table scrape; the scraper
    #      mis-labels the first niche row as `headers`, so real rows = headers + rows)
    #  (b) relatedNiches = [{cells, niche, href}, ...] (visible-structured capture)
    niche_rows: list = []
    tables = data.get("tables") or []
    if tables and isinstance(tables[0], dict) and (tables[0].get("rows") or tables[0].get("headers")):
        table = tables[0]
        if isinstance(table.get("headers"), list):
            niche_rows.append(table["headers"])
        niche_rows.extend(table.get("rows", []))
    elif isinstance(data.get("relatedNiches"), list):
        for item in data["relatedNiches"]:
            cells = item.get("cells") if isinstance(item, dict) else None
            if isinstance(cells, list):
                niche_rows.append(cells)

    if not niche_rows:
        warnings.append(f"Related niches: no usable rows in {related_json_path}; tab left as-is.")
        result.update({"kept": [], "dropped": [], "rows_written": 0, "skipped": True})
        return result

    # The template's tab carries the column-label header in the rows above the
    # niche data block (freeze_panes marks the data start). Snapshot the whole
    # tab style, capture the header block, then rewrite niches under it.
    snap = snapshot_styles(ws, upto_row=8)
    freeze = ws.freeze_panes  # e.g. "A4" -> niche data starts at row 4
    header_rows_count = 3
    if freeze:
        m = re.match(r"[A-Z]+(\d+)", freeze)
        if m:
            header_rows_count = int(m.group(1)) - 1

    header_block = []
    for r in range(1, header_rows_count + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, (ws.max_column or 1) + 1)]
        header_block.append(row_vals)

    # Label rows that can repeat inside a scraped table; never niches.
    label_skip = {"niche details", "niche name", ""}
    kept, dropped = [], []
    kept_rows = []
    for nrow in niche_rows:
        name = norm(nrow[0]) if nrow else ""
        if name in label_skip:
            continue
        if name in keep:
            kept.append(nrow[0])
            kept_rows.append(nrow)
        else:
            dropped.append(nrow[0])

    # Validation safety: ensure no known drift example survived.
    survived_excluded = [k for k in kept if norm(k) in exclude]

    clear_rows(ws)
    out_rows = header_block + kept_rows
    write_matrix(ws, out_rows, snap)
    if freeze:
        ws.freeze_panes = freeze

    result.update(
        {
            "kept": kept,
            "dropped": dropped,
            "rows_written": len(out_rows),
            "niche_rows_written": len(kept_rows),
            "survived_excluded": survived_excluded,
            "skipped": False,
        }
    )
    return result


# --- Outlier triage ------------------------------------------------------- #
REASONS = {
    "Competitor/brand term": "Competitor or third-party brand token in the term.",
    "Wrong product form": "Term implies a non-powder form (capsule/drink/serum/etc.).",
    "Unsupported claim/health-risk": "Term implies a health/disease claim not supported for copy.",
    "Negative candidate": "Term is off-product/irrelevant to a collagen powder.",
    "Semantic opportunity": "Relevant term where Sheko under-ranks vs competitors.",
    "Ignore": "No special signal; normal mainstream term in the master list.",
}
RECOMMENDED_USE = {
    "Competitor/brand term": "Keep for competitor/PPC research; do not use in copy.",
    "Wrong product form": "Add as negative; not relevant to powder form.",
    "Unsupported claim/health-risk": "Exclude from copy; review claim compliance.",
    "Negative candidate": "Add to negatives.",
    "Semantic opportunity": "Candidate for visible copy / backend / PPC push.",
    "Ignore": "Already covered by mainstream copy terms.",
}


def _token_hit(kw_norm: str, tokens: list[str], match: str) -> bool:
    for tok in tokens:
        t = norm(tok)
        if not t:
            continue
        if match == "word":
            if re.search(r"(?<!\w)" + re.escape(t) + r"(?!\w)", kw_norm):
                return True
        else:
            if t in kw_norm:
                return True
    return False


def classify_keyword(kw: str, relev, sv, sheko_rank, best_comp, tcfg: dict) -> str:
    kw_norm = norm(kw)
    match = tcfg.get("match", "word")
    if _token_hit(kw_norm, tcfg["brand_tokens"], match):
        return "Competitor/brand term"
    if _token_hit(kw_norm, tcfg["form_tokens"], match):
        return "Wrong product form"
    if _token_hit(kw_norm, tcfg["claim_tokens"], match):
        return "Unsupported claim/health-risk"
    if _token_hit(kw_norm, tcfg["negative_tokens"], match):
        return "Negative candidate"
    so = tcfg["semantic_opportunity"]
    relev = relev or 0.0
    sv = sv or 0
    if relev >= so["min_relevancy"] and sv >= so["min_search_volume"]:
        if sheko_rank is None and so.get("include_if_sheko_not_ranking"):
            return "Semantic opportunity"
        if (
            sheko_rank is not None
            and best_comp is not None
            and (sheko_rank - best_comp) >= so["sheko_rank_worse_than_best_competitor_by"]
        ):
            return "Semantic opportunity"
    return "Ignore"


def build_outlier(
    ws: Worksheet,
    master_rows: list[list[str]],
    poe_terms: set[str],
    cfg: dict,
    result: dict,
    warnings: list[str],
) -> dict:
    tcfg = cfg["triage"]
    action_map = cfg["final_action"]["by_classification"]
    allowed_actions = set(cfg["final_action"]["allowed"])
    sheko_asin = cfg["product_anchor"]["asin"]

    header = master_rows[0]
    # Locate ASIN columns by ASIN-shaped header (B0XXXXXXXX). This skips
    # non-ASIN columns some DataDive exports add (e.g. "Sugg. bid & range").
    asin_cols = {
        header[i]: i
        for i in range(4, len(header))
        if re.match(r"^B0[A-Z0-9]{8}$", str(header[i]).strip())
    }
    if sheko_asin not in asin_cols:
        warnings.append(f"Outlier triage: Sheko ASIN {sheko_asin} not a master column.")
        sheko_idx = None
    else:
        sheko_idx = asin_cols[sheko_asin]
    comp_idxs = [i for a, i in asin_cols.items() if a != sheko_asin]

    out_header = [
        "Keyword", "Search Volume", "DataDive Relevancy", "Sheko Rank",
        "Best Competitor Rank", "Competitors Top 50", "POE Signal",
        "Classification", "Reason", "Recommended Use", "Source", "Final Action",
    ]

    snap = snapshot_styles(ws, upto_row=4)
    rows_out = [out_header]
    by_class: dict[str, int] = {}
    by_action: dict[str, int] = {}
    bad_actions: list[str] = []

    for row in master_rows[1:]:
        if len(row) < 4:
            continue
        kw = row[1]
        if not norm(kw):
            continue
        sv = to_int(row[2])
        relev = to_float(row[3])
        sheko_rank = to_int(row[sheko_idx]) if sheko_idx is not None and sheko_idx < len(row) else None
        comp_ranks = [to_int(row[i]) for i in comp_idxs if i < len(row)]
        comp_ranks = [r for r in comp_ranks if r is not None]
        best_comp = min(comp_ranks) if comp_ranks else None
        comps_top50 = sum(1 for r in comp_ranks if r <= 50)
        poe_signal = "In POE top 20" if norm(kw) in poe_terms else ""

        cls = classify_keyword(kw, relev, sv, sheko_rank, best_comp, tcfg)
        if cls == "Ignore":
            continue  # not an outlier; lives in the master list as normal copy

        action = action_map.get(cls, "Ignore")
        if action not in allowed_actions:
            bad_actions.append(f"{kw} -> {action}")

        rows_out.append(
            [
                kw, sv, relev, sheko_rank if sheko_rank is not None else "",
                best_comp if best_comp is not None else "", comps_top50, poe_signal,
                cls, REASONS.get(cls, ""), RECOMMENDED_USE.get(cls, ""),
                "DataDive Master List", action,
            ]
        )
        by_class[cls] = by_class.get(cls, 0) + 1
        by_action[action] = by_action.get(action, 0) + 1

    # Sort outliers by classification group then descending search volume.
    order = {c: i for i, c in enumerate(tcfg["categories_order"])}
    body = rows_out[1:]
    body.sort(key=lambda r: (order.get(r[7], 99), -(r[1] or 0)))
    rows_out = [out_header] + body

    clear_rows(ws)
    write_matrix(ws, rows_out, snap)
    if not ws.freeze_panes:
        ws.freeze_panes = "A2"

    result.update(
        {
            "rows_written": len(rows_out),
            "outlier_keywords": len(rows_out) - 1,
            "by_classification": by_class,
            "by_final_action": by_action,
            "invalid_actions": bad_actions,
        }
    )
    return result


def build_seo_text(ws: Worksheet, seo_path: str, result: dict, warnings: list[str]) -> dict:
    """Rebuild the '4.1 SEO Text' tab from the curated seo_content.json, adding
    a DataDive Ranking Juice column. 'Current Listing Text' is sourced from the
    existing tab (the live listing copy) unless overridden in the JSON."""
    if not os.path.exists(seo_path):
        warnings.append(f"SEO content file not found ({seo_path}); SEO tab left as-is.")
        result.update({"skipped": True})
        return result
    seo = json.load(open(seo_path, encoding="utf-8"))

    # Capture current listing text per section from the existing tab.
    current_by_section = {}
    for r in range(2, (ws.max_row or 1) + 1):
        sec = ws.cell(r, 1).value
        if sec:
            current_by_section[norm(sec)] = ws.cell(r, 2).value

    snap = snapshot_styles(ws, upto_row=12)
    rows: list[list[Any]] = [seo["header"]]
    for item in seo["rows"]:
        cur = item.get("current")
        if cur is None:
            cur = current_by_section.get(norm(item["section"]), "")
        rows.append(
            [item["section"], cur, item.get("new", ""), item.get("compliance", ""), item.get("rj", "")]
        )

    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    # Widen the new Ranking Juice column and ensure wrapping on body cells.
    ws.column_dimensions["E"].width = 46
    from openpyxl.styles import Alignment

    for r in range(2, len(rows) + 1):
        for c in range(2, 6):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    result.update({"rows_written": len(rows), "skipped": False})
    return result


def load_poe_terms(poe_search_terms_csv: str) -> set[str]:
    """Top search terms from the POE search-terms CSV (data under the row-5
    header). Used to flag 'In POE top 20' on the outlier tab."""
    rows = read_csv(poe_search_terms_csv)
    terms: set[str] = set()
    header_idx = None
    for i, r in enumerate(rows):
        if r and norm(r[0]) == "search term":
            header_idx = i
            break
    if header_idx is None:
        return terms
    for r in rows[header_idx + 1 :]:
        if r and norm(r[0]):
            terms.add(norm(r[0]))
    return terms


def load_json(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _capture_text(data: Any) -> str:
    if isinstance(data, str):
        return data
    if not isinstance(data, dict):
        return ""
    parts: list[str] = []
    for key in ("text", "rawText", "content"):
        if data.get(key):
            parts.append(str(data[key]))
    visible = data.get("visibleRoute")
    if isinstance(visible, dict):
        parts.append(_capture_text(visible))
    return "\n".join(parts)


def _first_table_rows(data: dict) -> tuple[list[str], list[list[Any]]]:
    tables = data.get("tables") or []
    if not tables or not isinstance(tables[0], dict):
        return [], []
    table = tables[0]
    headers = table.get("headers") or []
    rows = table.get("rows") or []
    return headers, rows


def _extract_review_topic_rows(data: dict, source: str) -> list[list[Any]]:
    headers, table_rows = _first_table_rows(data)
    if table_rows:
        out = []
        for row in table_rows:
            topic = row[0] if row else ""
            pct = row[2] if len(row) > 2 else row[1] if len(row) > 1 else ""
            snippets = " | ".join(str(v) for v in row[3:] if v)
            out.append([topic, pct, snippets, "POE Customer Review Insights", source])
        return out

    text = _capture_text(data)
    if not text:
        return []
    section = text
    start = section.find("Topic Subtopic %Mentions Review Snippets")
    if start >= 0:
        section = section[start + len("Topic Subtopic %Mentions Review Snippets") :]
    end_markers = ["Topic Impact on Star Rating", "Customer Review Topic Mentions Trends", "Disclaimer:"]
    end_positions = [section.find(m) for m in end_markers if section.find(m) >= 0]
    if end_positions:
        section = section[: min(end_positions)]
    section = re.sub(r"\s+", " ", section).strip()

    pattern = re.compile(
        r"(?P<topic>[A-ZÀ-Ýa-zà-ÿ0-9][^\"%]{2,90}?)\s+"
        r"(?P<pct>\d+(?:\.\d+)%)\s+"
        r"(?P<snips>\"[^\"]+\"(?:,\s*\"[^\"]+\")*)"
    )
    out: list[list[Any]] = []
    for m in pattern.finditer(section):
        topic = re.sub(r"\s+", " ", m.group("topic")).strip(" -:")
        topic = topic.replace("Topic Subtopic %Mentions Review Snippets", "").strip()
        snippets = m.group("snips").replace('", "', '" | "')
        out.append([topic, m.group("pct"), snippets, "POE Customer Review Insights", source])
    return out


def build_poe_reviews(ws: Worksheet, path: str, result: dict, warnings: list[str]) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    header = ["Topic", "% Mentions", "Review Snippets", "Evidence Type", "Source File"]
    if not path or not os.path.exists(path):
        rows = [header, ["Missing POE review evidence", "", "", "Warning", path]]
        warnings.append("POE Raw - Reviews: review JSON missing; wrote warning row.")
        status = "missing"
    else:
        data = load_json(path)
        topic_rows = _extract_review_topic_rows(data, path)
        if topic_rows:
            rows = [header] + topic_rows
            status = "parsed"
        else:
            rows = [header, ["POE review evidence captured but no topic rows parsed", "", "", "Warning", path]]
            warnings.append(f"POE Raw - Reviews: no topic rows parsed from {path}; wrote warning row.")
            status = "unparsed"
    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({"rows_written": len(rows), "topics": max(0, len(rows) - 1), "status": status, "path": path})
    return result


def _extract_returns_rows(data: dict, source: str) -> list[list[Any]]:
    headers, table_rows = _first_table_rows(data)
    if table_rows:
        return [[*(row[:4]), source] for row in table_rows]
    visible = data.get("visibleRoute") if isinstance(data, dict) else None
    text = _capture_text(visible if isinstance(visible, dict) else data)
    if not text or "Returns" not in text or "Customer Review Insights" in text and "Return" not in text.split("Customer Review Insights")[-1]:
        return []
    section = re.sub(r"\s+", " ", text).strip()
    pattern = re.compile(
        r"(?P<topic>[A-ZÀ-Ýa-zà-ÿ0-9][^\"%]{2,90}?)\s+"
        r"(?P<pct>\d+(?:\.\d+)%)\s+"
        r"(?P<detail>[^%]{0,180})(?=(?:[A-ZÀ-Ýa-zà-ÿ0-9][^\"%]{2,90}?\s+\d+(?:\.\d+)%|$))"
    )
    return [[m.group("topic").strip(), m.group("pct"), m.group("detail").strip(), "POE Returns", source] for m in pattern.finditer(section)]


def build_poe_returns(ws: Worksheet, path: str, result: dict, warnings: list[str]) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    header = ["Topic / Status", "% Mentions", "Detail", "Evidence Type", "Source File"]
    if not path or not os.path.exists(path):
        rows = [header, ["Missing POE returns evidence", "", "", "Warning", path]]
        warnings.append("POE Raw - Returns: returns JSON missing; wrote warning row.")
        status = "missing"
    else:
        data = load_json(path)
        return_rows = _extract_returns_rows(data, path)
        if return_rows:
            rows = [header] + return_rows
            status = "parsed"
        else:
            rows = [
                header,
                [
                    "Returns not exposed for this POE niche",
                    "",
                    "Seller Central returns route was captured, but no returns topics/table was visible for this niche.",
                    "Explicit not-exposed warning",
                    path,
                ],
            ]
            status = "not_exposed"
    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({"rows_written": len(rows), "returns_rows": max(0, len(rows) - 1), "status": status, "path": path})
    return result


def _poe_search_term_rows(path: str, limit: int = 12) -> list[list[str]]:
    rows = read_csv(path)
    header_idx = None
    for i, row in enumerate(rows):
        if row and norm(row[0]) == "search term":
            header_idx = i
            break
    if header_idx is None:
        return []
    header = rows[header_idx]
    body = rows[header_idx + 1 :]
    out = []
    for row in body:
        if not row or not norm(row[0]):
            continue
        row = row + [""] * max(0, len(header) - len(row))
        out.append(row)
        if len(out) >= limit:
            break
    return out


def _overview_metrics(data: dict) -> list[tuple[str, str]]:
    text = _capture_text(data.get("overview") if isinstance(data.get("overview"), dict) else data)
    metrics = []
    patterns = [
        ("Search Volume", r"Search Volume \(Past 360 days\)\s+([0-9,.]+)"),
        ("Search volume growth", r"Search volume growth \(Past 180 days\)\s+([+\-0-9,.%]+)"),
        ("Top clicked products", r"No\. of top clicked products\s+([0-9,.]+)"),
        ("Average price", r"Average price \(Past 360 days\)\s+([€$£0-9,.]+)"),
        ("Units sold range", r"Range of average units sold \(Past 360 days\)\s+([0-9,.]+\s+-\s+[0-9,.]+)"),
        ("Return rate", r"Return Rate \(Past 360 days\)\s+([0-9,.%]+)"),
    ]
    for label, pattern in patterns:
        m = re.search(pattern, text)
        if m:
            metrics.append((label, m.group(1)))
    return metrics


def build_semantic_insights(
    ws: Worksheet,
    args: dict,
    reviews_result: dict,
    returns_result: dict,
    related_result: dict,
    result: dict,
    warnings: list[str],
) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    header = ["Priority", "Insight Type", "Signal", "Value", "SEO / Listing Action", "Source"]
    rows: list[list[Any]] = [header]

    structured = {}
    if args.get("poe_structured_json") and os.path.exists(args["poe_structured_json"]):
        structured = load_json(args["poe_structured_json"])
        for label, value in _overview_metrics(structured):
            rows.append([1, "POE overview", label, value, "Use as niche context; do not treat as listing copy.", args["poe_structured_json"]])

    for i, row in enumerate(_poe_search_term_rows(args["poe_search_terms_csv"], limit=12), start=1):
        term = row[0]
        sv = row[1] if len(row) > 1 else ""
        conv = row[5] if len(row) > 5 else ""
        priority = 1 if i <= 6 else 2
        action = "Review for visible copy/backend coverage after relevance and claim check."
        rows.append([priority, "POE search term", term, f"SV {sv}; conversion {conv}", action, args["poe_search_terms_csv"]])

    if args.get("poe_reviews_json") and os.path.exists(args["poe_reviews_json"]):
        review_rows = _extract_review_topic_rows(load_json(args["poe_reviews_json"]), args["poe_reviews_json"])
        for topic, pct, snippets, _etype, source in review_rows[:10]:
            action = "Use as customer-language input; rewrite into compliant product/benefit wording."
            rows.append([1, "Review topic", topic, pct, action, source])
    else:
        rows.append([1, "Review topic", "Missing POE review evidence", "", "Do not finalize semantic insights until review capture exists.", args.get("poe_reviews_json", "")])

    if returns_result.get("status") == "not_exposed":
        rows.append([2, "Returns", "Returns not exposed", "", "Record as a source limitation; no return-topic SEO action.", returns_result.get("path", "")])
    elif returns_result.get("status") == "parsed":
        rows.append([1, "Returns", "Return topics captured", f"{returns_result.get('returns_rows', 0)} rows", "Review for product-experience objections.", returns_result.get("path", "")])

    for niche in related_result.get("kept", [])[:8]:
        rows.append([2, "Related niche", niche, "", "Use only if it matches the actual product form and ingredient.", args["related_niches_json"]])

    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({"rows_written": len(rows), "status": "rebuilt", "sources": {
        "poe_structured_json": args.get("poe_structured_json", ""),
        "poe_reviews_json": args.get("poe_reviews_json", ""),
        "poe_returns_json": args.get("poe_returns_json", ""),
    }})
    return result


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #
def excel_name_ok(name: str) -> bool:
    if not name or len(name) > 31:
        return False
    if any(ch in name for ch in "[]:*?/\\"):
        return False
    return True


def run_validations(wb, cfg, counts, related_result, paths, warnings) -> list[dict]:
    checks: list[dict] = []

    def add(name, ok, detail=""):
        checks.append({"check": name, "pass": bool(ok), "detail": detail})

    anchor = cfg["product_anchor"]["asin"]

    # 1. Product anchor is the configured ASIN and present as a master-list
    #    column (authoritative — it's a DataDive ASIN column). The ASINs tab is
    #    a curated carry-forward tab; if the anchor isn't in it, that tab is a
    #    template placeholder to curate for this market (warning, not failure).
    master_header = counts["_master_header"]
    in_master = anchor in master_header
    asins_tab_hit = False
    if "ASINs" in wb.sheetnames:
        ws = wb["ASINs"]
        for row in ws.iter_rows(values_only=True):
            if any(norm(v) == norm(anchor) for v in row):
                asins_tab_hit = True
                break
    add(
        f"Product anchor is configured ASIN {anchor}",
        bool(anchor) and in_master,
        f"config={anchor}, in_master={in_master}, in_ASINs_tab={asins_tab_hit}",
    )
    if not asins_tab_hit:
        warnings.append(
            f"Anchor {anchor} not in the carried-forward ASINs tab — curate the "
            f"ASINs / curated tabs for this market (currently template placeholders)."
        )

    # 2. Exact DataDive root/master row counts match source CSVs.
    add(
        "Root Keywords rows == roots CSV rows",
        counts["1. Root Keywords"] == counts["_roots_csv"],
        f"sheet={counts['1. Root Keywords']} csv={counts['_roots_csv']}",
    )
    add(
        "Master List rows == master CSV rows",
        counts["3.1. Master List DataDive"] == counts["_master_csv"],
        f"sheet={counts['3.1. Master List DataDive']} csv={counts['_master_csv']}",
    )

    # 3. Sheet names Excel/Google Sheets compatible (+ unique).
    names = wb.sheetnames
    bad = [n for n in names if not excel_name_ok(n)]
    dupes = len(names) != len(set(names))
    add(
        "Sheet names Excel/Sheets compatible",
        not bad and not dupes,
        f"bad={bad} duplicates={dupes}",
    )

    # 4. Related-niche filter excludes known irrelevant examples.
    excl = [norm(x) for x in cfg["related_niche_filter"]["exclude_examples"]]
    kept_norm = [norm(x) for x in related_result.get("kept", [])]
    leaked = [k for k in kept_norm if k in excl]
    keepset = [norm(x) for x in cfg["related_niche_filter"]["keep"]]
    non_keep = [k for k in kept_norm if k not in keepset]
    add(
        "Related-niche filter excludes drift",
        not leaked and not non_keep and not related_result.get("skipped"),
        f"leaked={leaked} non_keeplist={non_keep} "
        f"kept={len(kept_norm)} dropped={len(related_result.get('dropped', []))}",
    )

    # 5. POE products/search-terms sheet rows match source CSVs.
    add(
        "POE Products rows == products CSV rows",
        counts["POE Raw - Products"] == counts["_poe_products_csv"],
        f"sheet={counts['POE Raw - Products']} csv={counts['_poe_products_csv']}",
    )
    add(
        "POE Search Terms rows == search-terms CSV rows",
        counts["POE Raw - Search Terms"] == counts["_poe_search_terms_csv"],
        f"sheet={counts['POE Raw - Search Terms']} csv={counts['_poe_search_terms_csv']}",
    )

    # Bonus: every emitted Final Action is in the allowed enum.
    add(
        "Outlier Final Action values valid",
        not counts.get("_invalid_actions"),
        f"invalid={counts.get('_invalid_actions')}",
    )

    guard = cfg.get("stale_data_guard") or {}
    forbidden = [norm(x) for x in guard.get("forbidden_terms", []) if norm(x)]
    tabs = guard.get("tabs", [])
    leaks: list[str] = []
    for tab in tabs:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for row in ws.iter_rows(values_only=True):
            row_text = norm(" ".join(str(v) for v in row if v is not None))
            for term in forbidden:
                if term and term in row_text:
                    leaks.append(f"{tab}: {term}")
    if forbidden:
        add(
            "No stale forbidden terms in guarded tabs",
            not leaks,
            f"leaks={leaks[:12]} total={len(leaks)}",
        )

    required_tabs = cfg.get("required_current_tabs") or []
    missing_current = []
    for tab in required_tabs:
        if counts.get(tab, 0) <= 1:
            missing_current.append(tab)
    if required_tabs:
        add(
            "Required current-source tabs populated",
            not missing_current,
            f"missing_or_warning_only={missing_current}",
        )

    return checks


def build_curated_tabs(ws_book, path: str, result: dict, warnings: list[str]) -> dict:
    """Write curated content into named tabs from a JSON file:
    { "<Tab Name>": {"header": [...], "rows": [[...], ...]}, ... }.
    Used for POE Raw - Reviews/Returns/Semantic Insights on a new-market build so
    they carry real curated content (from POE captures) instead of a placeholder.
    Header style is preserved from the template tab."""
    written: list[str] = []
    if not path or not os.path.exists(path):
        result["tabs"] = written
        return result
    data = json.load(open(path, encoding="utf-8"))
    for tab, spec in data.items():
        if tab.startswith("_"):
            continue
        if tab not in ws_book.sheetnames:
            warnings.append(f"Curated tab '{tab}' not in template; skipped.")
            continue
        ws = ws_book[tab]
        snap = snapshot_styles(ws, upto_row=2)
        rows = [spec.get("header", [])] + spec.get("rows", [])
        clear_rows(ws)
        write_matrix(ws, rows, snap)
        if not ws.freeze_panes:
            ws.freeze_panes = "A2"
        written.append(tab)
    result["tabs"] = written
    return result


def clear_placeholder_tab(ws: Worksheet, note: str) -> None:
    """Blank a product-specific curated tab down to its header row + a note.
    Used when building a DIFFERENT product than the style template, so the new
    workbook never ships another product's curated content (e.g. collagen
    reviews in a fibre workbook). Keeps header styling and column widths."""
    snap = snapshot_styles(ws, upto_row=2)
    header = [ws.cell(1, c).value for c in range(1, (ws.max_column or 1) + 1)]
    clear_rows(ws)
    write_matrix(ws, [header], snap)
    cell = ws.cell(2, 1, value=note)
    apply_style(cell, snap, 2, 1)


# --------------------------------------------------------------------------- #
# Preflight (auto-generated cross-agent handoff)
# --------------------------------------------------------------------------- #
CODEX_INPUT_KEYS = [
    "roots_csv", "master_csv", "competitors_csv",
    "poe_products_csv", "poe_search_terms_csv", "related_niches_json",
    "poe_reviews_json", "poe_returns_json", "poe_structured_json",
]
SETUP_INPUT_KEYS = ["template", "seo_content"]


def _codex_handoff_block(cfg: dict, args: dict, missing: list[str]) -> str:
    pa = cfg["product_anchor"]
    lines = [
        "TASK: Gather the missing keyword-workbook inputs for the connected/internal browser + DataDive UI.",
        "",
        f"Objective: produce the inputs below for {pa.get('client')} {pa.get('product')} "
        f"({pa.get('marketplace')}), DataDive niche {pa.get('datadive_niche')}, anchor ASIN {pa.get('asin')}.",
        "",
        "Do NOT: run the builder, write SEO, edit listings, commit/push, or inspect cookies/session/credentials.",
        "",
        "Produce these files at exactly these paths:",
    ]
    label = {
        "roots_csv": "DataDive UI — niche analysis roots CSV",
        "master_csv": "DataDive UI — master keyword list CSV",
        "competitors_csv": "DataDive UI (or MCP fallback) — competitors CSV",
        "poe_products_csv": "POE — Niche Details > Products tab CSV",
        "poe_search_terms_csv": "POE — Niche Details > Search Terms tab CSV",
        "related_niches_json": "POE — related-niches capture JSON",
        "poe_reviews_json": "POE — Customer Review Insights capture JSON",
        "poe_returns_json": "POE — Returns route/table capture JSON",
        "poe_structured_json": "POE — structured overview capture JSON",
    }
    for k in missing:
        lines.append(f"  [ ] {label.get(k, k)}\n        -> {args[k]}")
    lines += [
        "",
        "Also capture (evidence): an online variation listing JSON (title/bullets/description/ingredients), "
        "review-insights + returns JSON, screenshots → evidence/{client}/opportunity-data/.",
        "Caveats to confirm: Seller Central account + marketplace; any ASIN that resolves to the wrong product "
        "family; ingredient/health-claim risks.",
        "",
        "Stop and report: the exact saved paths + the caveats above, then hand back to Claude.",
        "",
        "Protocol: /Users/victoruhl/Obsidian/Victors Second Brain/Context/codex-claude-handoff-protocol.md",
    ]
    return "\n".join(lines)


def _build_command(args: dict) -> str:
    return (
        ".venv/bin/python tools/sheko-keyword-workbook/build_keyword_workbook.py "
        f'--config "{args["config"]}"'
    )


def run_preflight(cfg: dict, args: dict) -> int:
    pa = cfg["product_anchor"]
    print("=" * 64)
    print(f"PREFLIGHT — {pa.get('client')} · {pa.get('product')}")
    print(f"  marketplace={pa.get('marketplace')}  niche={pa.get('datadive_niche')}  anchor={pa.get('asin')}")
    print("=" * 64)
    present = [k for k in CODEX_INPUT_KEYS + SETUP_INPUT_KEYS if os.path.exists(args[k])]
    missing_codex = [k for k in CODEX_INPUT_KEYS if not os.path.exists(args[k])]
    missing_setup = [k for k in SETUP_INPUT_KEYS if not os.path.exists(args[k])]

    print("PRESENT:")
    for k in present:
        print(f"  [x] {k}")
    if missing_codex or missing_setup:
        print("MISSING:")
        for k in missing_codex:
            print(f"  [ ] (CODEX)  {k}  -> {args[k]}")
        for k in missing_setup:
            print(f"  [ ] (setup)  {k}  -> {args[k]}")

    if missing_codex:
        print("\nSTATUS: WAITING ON CODEX. Paste the block below into Codex (do not hand-write it):\n")
        print(_codex_handoff_block(cfg, args, missing_codex))
    elif missing_setup:
        print(f"\nSTATUS: NEEDS SETUP — provide: {', '.join(missing_setup)}")
    else:
        print("\nSTATUS: READY — all inputs present. Hand to Claude to write SEO + build, or run:\n")
        print(_build_command(args))
    return 0


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> int:
    ap = argparse.ArgumentParser(description="Build the Sheko keyword workbook.")
    for k, v in DEFAULTS.items():
        ap.add_argument("--" + k.replace("_", "-"), default=v)
    ap.add_argument("--preflight", action="store_true",
                    help="Check inputs and print the cross-agent handoff / build-readiness status; do not build.")
    args = vars(ap.parse_args())
    args = {k.replace("-", "_"): v for k, v in args.items()}

    cfg = json.load(open(args["config"], encoding="utf-8"))
    warnings: list[str] = []

    # Input contract: paths may live in config `inputs{}` so they aren't recited
    # on the CLI. Precedence: explicit CLI flag > config inputs > built-in default.
    explicit = {
        a.lstrip("-").replace("-", "_").split("=")[0]
        for a in sys.argv[1:] if a.startswith("--")
    }

    def _resolve(p: str) -> str:
        p = os.path.expanduser(str(p))
        return p if os.path.isabs(p) else os.path.join(REPO, p)

    for k, v in (cfg.get("inputs") or {}).items():
        if k in DEFAULTS and k not in explicit:
            args[k] = _resolve(v)

    if args.get("preflight"):
        return run_preflight(cfg, args)

    # Verify sources exist.
    src_keys = [
        "template", "roots_csv", "master_csv", "competitors_csv",
        "poe_products_csv", "poe_search_terms_csv", "related_niches_json",
    ]
    rebuild_tabs = set(cfg["tabs"].get("rebuild", []))
    if "POE Raw - Reviews" in rebuild_tabs:
        src_keys.append("poe_reviews_json")
    if "POE Raw - Returns" in rebuild_tabs:
        src_keys.append("poe_returns_json")
    if "POE Semantic Insights" in rebuild_tabs and args.get("poe_structured_json"):
        src_keys.append("poe_structured_json")
    missing = [k for k in src_keys if not os.path.exists(args[k])]
    if missing:
        print("ERROR: missing source files:", file=sys.stderr)
        for k in missing:
            print(f"   {k}: {args[k]}", file=sys.stderr)
        print("\nRun with --preflight to generate the Codex handoff for the missing inputs.", file=sys.stderr)
        return 2

    print(f"Loading template: {args['template']}")
    wb = load_workbook(args["template"])

    counts: dict[str, Any] = {}

    # --- Exact-paste tabs --------------------------------------------------- #
    roots = read_csv(args["roots_csv"])
    master = read_csv(args["master_csv"])
    poe_products = read_csv(args["poe_products_csv"])
    poe_search = read_csv(args["poe_search_terms_csv"])

    counts["_roots_csv"] = len(roots)
    counts["_master_csv"] = len(master)
    counts["_poe_products_csv"] = len(poe_products)
    counts["_poe_search_terms_csv"] = len(poe_search)
    counts["_master_header"] = master[0] if master else []

    exact = cfg["tabs"]["exact_paste"]
    src_for = {
        "roots_csv": roots,
        "master_csv": master,
        "poe_products_csv": poe_products,
        "poe_search_terms_csv": poe_search,
    }
    for tab, src_key in exact.items():
        if tab not in wb.sheetnames:
            warnings.append(f"Exact-paste tab '{tab}' not in template; skipped.")
            continue
        n = build_exact_paste(wb[tab], src_for[src_key], warnings)
        counts[tab] = n
        print(f"  exact-paste {tab!r}: {n} rows")

    # --- Related niches (strict filter) ------------------------------------ #
    related_result: dict = {}
    if "POE Raw - Related Niches" in wb.sheetnames:
        build_related_niches(
            wb["POE Raw - Related Niches"], args["related_niches_json"], cfg, related_result, warnings
        )
        counts["POE Raw - Related Niches"] = related_result.get("rows_written", 0)
        print(
            f"  rebuilt 'POE Raw - Related Niches': kept {len(related_result.get('kept', []))} "
            f"niches, dropped {len(related_result.get('dropped', []))}"
        )
    else:
        warnings.append("'POE Raw - Related Niches' not in template; skipped.")

    # --- Outlier triage ----------------------------------------------------- #
    outlier_result: dict = {}
    if "Outlier - Opportunity KWs" in wb.sheetnames:
        poe_terms = load_poe_terms(args["poe_search_terms_csv"])
        build_outlier(
            wb["Outlier - Opportunity KWs"], master, poe_terms, cfg, outlier_result, warnings
        )
        counts["Outlier - Opportunity KWs"] = outlier_result.get("rows_written", 0)
        counts["_invalid_actions"] = outlier_result.get("invalid_actions", [])
        print(
            f"  rebuilt 'Outlier - Opportunity KWs': {outlier_result.get('outlier_keywords', 0)} "
            f"triaged keywords"
        )
    else:
        warnings.append("'Outlier - Opportunity KWs' not in template; skipped.")

    # --- SEO Text rewrite (curated, + Ranking Juice column) ----------------- #
    seo_result: dict = {}
    if "4.1 SEO Text" in wb.sheetnames:
        build_seo_text(wb["4.1 SEO Text"], args["seo_content"], seo_result, warnings)
        counts["4.1 SEO Text"] = seo_result.get("rows_written", 0)
        if not seo_result.get("skipped"):
            print(f"  rebuilt '4.1 SEO Text': {seo_result.get('rows_written', 0)} rows (+ Ranking Juice column)")

    # --- Curated tabs (real POE-sourced content, e.g. Reviews) -------------- #
    curated_result: dict = {}
    if args.get("curated_tabs"):
        build_curated_tabs(wb, args["curated_tabs"], curated_result, warnings)
        if curated_result.get("tabs"):
            print(f"  curated tabs: {', '.join(curated_result['tabs'])}")

    # --- Current POE evidence tabs ------------------------------------------- #
    current_poe_done: set[str] = set()
    reviews_result: dict = {}
    if "POE Raw - Reviews" in rebuild_tabs and "POE Raw - Reviews" in wb.sheetnames:
        build_poe_reviews(wb["POE Raw - Reviews"], args.get("poe_reviews_json", ""), reviews_result, warnings)
        counts["POE Raw - Reviews"] = reviews_result.get("rows_written", 0)
        current_poe_done.add("POE Raw - Reviews")
        print(
            f"  rebuilt 'POE Raw - Reviews': {reviews_result.get('topics', 0)} topics "
            f"({reviews_result.get('status')})"
        )

    returns_result: dict = {}
    if "POE Raw - Returns" in rebuild_tabs and "POE Raw - Returns" in wb.sheetnames:
        build_poe_returns(wb["POE Raw - Returns"], args.get("poe_returns_json", ""), returns_result, warnings)
        counts["POE Raw - Returns"] = returns_result.get("rows_written", 0)
        current_poe_done.add("POE Raw - Returns")
        print(
            f"  rebuilt 'POE Raw - Returns': {returns_result.get('returns_rows', 0)} rows "
            f"({returns_result.get('status')})"
        )

    semantic_result: dict = {}
    if "POE Semantic Insights" in rebuild_tabs and "POE Semantic Insights" in wb.sheetnames:
        build_semantic_insights(
            wb["POE Semantic Insights"],
            args,
            reviews_result,
            returns_result,
            related_result,
            semantic_result,
            warnings,
        )
        counts["POE Semantic Insights"] = semantic_result.get("rows_written", 0)
        current_poe_done.add("POE Semantic Insights")
        print(f"  rebuilt 'POE Semantic Insights': {semantic_result.get('rows_written', 0)} rows")

    # --- Clear product-specific curated placeholders (new-market builds) ---- #
    curated_done = set(curated_result.get("tabs", [])) | current_poe_done
    cleared_tabs = []
    note = (
        f"(Placeholder — curate for {cfg['product_anchor'].get('marketplace','this market')} / "
        f"{cfg['product_anchor'].get('product','this product')}. Carried-forward template content cleared.)"
    )
    for tab in cfg["tabs"].get("carry_forward_clear", []):
        if tab in curated_done:
            continue  # already populated with real content
        if tab in wb.sheetnames:
            clear_placeholder_tab(wb[tab], note)
            cleared_tabs.append(tab)
    if cleared_tabs:
        print(f"  cleared placeholder tabs: {', '.join(cleared_tabs)}")

    # --- Enforce tab order -------------------------------------------------- #
    desired = [t for t in cfg["tab_order"] if t in wb.sheetnames]
    wb._sheets.sort(key=lambda ws: desired.index(ws.title) if ws.title in desired else 999)

    # --- Validations -------------------------------------------------------- #
    checks = run_validations(wb, cfg, counts, related_result, args, warnings)

    # --- Save workbook ------------------------------------------------------ #
    os.makedirs(os.path.dirname(args["out"]), exist_ok=True)
    wb.save(args["out"])
    print(f"\nSaved workbook: {args['out']}")

    # Optional: copy the finished .xlsx into a delivery folder (e.g. synced Drive).
    delivered_to = ""
    if args.get("drive_dir"):
        if os.path.isdir(args["drive_dir"]):
            dest = os.path.join(args["drive_dir"], os.path.basename(args["out"]))
            shutil.copy2(args["out"], dest)
            delivered_to = dest
            print(f"Delivered copy to: {dest}")
        else:
            warnings.append(f"--drive-dir not a directory; skipped copy: {args['drive_dir']}")

    # --- Manifest ----------------------------------------------------------- #
    populated = {
        ws.title: ws.max_row for ws in wb.worksheets
    }
    manifest = {
        "generated_at": _dt.datetime.now().astimezone().isoformat(),
        "generator": "tools/sheko-keyword-workbook/build_keyword_workbook.py",
        "product_anchor": cfg["product_anchor"],
        "outputs": {
            "workbook_local": args["out"],
            "manifest": args["manifest"],
            "drive_copy": delivered_to,
            "drive_target_note": (
                "Use --drive-dir to copy the finished .xlsx into the synced Drive folder, "
                "or copy the reviewed local workbook there after confirmation."
            ),
        },
        "template": args["template"],
        "sources": {
            "roots_csv": {"path": args["roots_csv"], "rows": counts["_roots_csv"]},
            "master_csv": {"path": args["master_csv"], "rows": counts["_master_csv"]},
            "competitors_csv": {"path": args["competitors_csv"]},
            "poe_products_csv": {"path": args["poe_products_csv"], "rows": counts["_poe_products_csv"]},
            "poe_search_terms_csv": {
                "path": args["poe_search_terms_csv"], "rows": counts["_poe_search_terms_csv"]
            },
            "related_niches_json": {"path": args["related_niches_json"]},
            "poe_reviews_json": {
                "path": args.get("poe_reviews_json", ""),
                "rows": reviews_result.get("rows_written", 0),
                "status": reviews_result.get("status", ""),
            },
            "poe_returns_json": {
                "path": args.get("poe_returns_json", ""),
                "rows": returns_result.get("rows_written", 0),
                "status": returns_result.get("status", ""),
            },
            "poe_structured_json": {"path": args.get("poe_structured_json", "")},
        },
        "tab_production": cfg["tabs"],
        "tab_order": [ws.title for ws in wb.worksheets],
        "row_counts": populated,
        "related_niche_filter": {
            "keep_configured": cfg["related_niche_filter"]["keep"],
            "kept": related_result.get("kept", []),
            "dropped": related_result.get("dropped", []),
            "niche_rows_written": related_result.get("niche_rows_written", 0),
        },
        "outlier_triage": {
            "outlier_keywords": outlier_result.get("outlier_keywords", 0),
            "by_classification": outlier_result.get("by_classification", {}),
            "by_final_action": outlier_result.get("by_final_action", {}),
        },
        "seo_text": {
            "source": args["seo_content"],
            "rows_written": seo_result.get("rows_written", 0),
            "ranking_juice_snapshot": (
                json.load(open(args["seo_content"], encoding="utf-8")).get("ranking_juice_snapshot", {})
                if os.path.exists(args["seo_content"]) else {}
            ),
        },
        "poe_reviews": reviews_result,
        "poe_returns": returns_result,
        "poe_semantic_insights": semantic_result,
        "validations": checks,
        "validations_all_pass": all(c["pass"] for c in checks),
        "warnings": warnings,
    }
    os.makedirs(os.path.dirname(args["manifest"]), exist_ok=True)
    json.dump(manifest, open(args["manifest"], "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"Saved manifest: {args['manifest']}")

    # --- Console summary ---------------------------------------------------- #
    print("\n=== VALIDATIONS ===")
    for c in checks:
        print(f"  [{'PASS' if c['pass'] else 'FAIL'}] {c['check']}  ({c['detail']})")
    if warnings:
        print("\n=== WARNINGS ===")
        for w in warnings:
            print("  -", w)

    return 0 if manifest["validations_all_pass"] else 1


if __name__ == "__main__":
    sys.exit(main())
