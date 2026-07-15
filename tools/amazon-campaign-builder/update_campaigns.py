#!/usr/bin/env python3
"""
Amazon SP campaign UPDATE builder: real bulksheets IDs + a change-set -> Update/
Archive/Create bulk-upload .xlsx.

  # Preflight: check the change-set against the loaded export, list what's missing/invalid
  python3 tools/amazon-campaign-builder/update_campaigns.py --config <cfg> --preflight

  # Preview: print every planned change in plain English, write nothing
  python3 tools/amazon-campaign-builder/update_campaigns.py --config <cfg> --preview

  # Build: write the bulksheet .xlsx + _REVIEW.md, then run the QA gates
  python3 tools/amazon-campaign-builder/update_campaigns.py --config <cfg>

  # QA gates only (re-check an already-built file against the same export)
  python3 tools/amazon-campaign-builder/update_campaigns.py --config <cfg> --validate

Output is a FILE ONLY. This tool never uploads or touches live campaigns. Unlike
create mode, update-mode rows reference REAL, currently-live entities (via IDs read
from a bulksheets download), so uploading the resulting file WILL change/pause/archive
real campaigns. The _REVIEW.md this writes is mandatory reading before that upload;
nothing in this toolkit performs the upload itself.

Input: (a) `export_file`, a bulksheets download of the "Sponsored Products Campaigns"
tab (Bulk Operations > create/download a spreadsheet with the entities you want),
which is the ONLY valid source of real Campaign/Ad Group/Keyword/Product Targeting/
Portfolio IDs (never the console UI, never a prior session's remembered IDs); and
(b) a change-set config describing the intended changes (config.UPDATE.TEMPLATE.json).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path

HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent
sys.path.insert(0, str(HERE))

from campaign_model import COLUMNS, SHEET_NAMES  # noqa: E402
from update_model import build_change_set_rows, looks_like_real_id, read_export  # noqa: E402


def slugify(s):
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def _rel(p):
    try:
        return p.relative_to(REPO)
    except ValueError:
        return p


def load_config(path):
    cfg = json.loads(Path(path).read_text())
    cfg.setdefault("changes", {})
    cfg.setdefault("allow_end_date_clear", False)
    return cfg


def out_path(cfg, override=None):
    if override:
        return Path(override)
    stamp = cfg.get("date") or date.today().isoformat()
    brand = re.sub(r"\s+", "-", (cfg.get("brand") or cfg["client"]).strip())
    name = f"{stamp}_{brand}_{cfg.get('marketplace', 'US')}_SP_bulk_UPDATE.xlsx"
    return REPO / "output" / slugify(cfg["client"]) / "ads" / name


def _load_export_or_die(cfg):
    export_file = cfg.get("export_file")
    if not export_file or not Path(export_file).exists():
        print(f"export_file not found: {export_file!r}. Download a current bulksheets "
              f"export of the account first (Bulk Operations > create/download spreadsheet)")
        sys.exit(1)
    return read_export(export_file)


def _plan(cfg):
    export = _load_export_or_die(cfg)
    rows, review, errors = build_change_set_rows(
        cfg["changes"], export, allow_end_date_clear=bool(cfg.get("allow_end_date_clear")))
    return export, rows, review, errors


# ----------------------------------------------------------------- preflight
def preflight(cfg):
    print(f"Preflight: {cfg.get('client', '?')} ({cfg.get('marketplace', '?')}) [UPDATE mode]")
    for key in ("client", "marketplace", "export_file"):
        if not cfg.get(key):
            print(f"  [MISSING] config: `{key}` is required")
    if not cfg.get("client") or not cfg.get("marketplace") or not cfg.get("export_file"):
        print("\nNOT READY.")
        return 1

    export, rows, review, errors = _plan(cfg)
    for e in errors:
        print(f"  [MISSING] {e}")
    skipped = [r for r in review if r.startswith("SKIPPED")]
    for s in skipped:
        print(f"  [NOTE]    {s}")
    if errors:
        print(f"\nNOT READY. Fix the {len(errors)} item(s) above in the change-set.")
        return 1
    if not rows:
        print("  [NOTE]    change-set produces zero rows (everything no-op/cascade-skipped); "
              "nothing to build")
    n_campaigns = len({r["Campaign ID"] for r in rows if r.get("Campaign ID")})
    print(f"\nREADY: {len(rows)} row(s) planned across {n_campaigns} campaign(s), "
          f"{len(skipped)} skipped (no-op/cascade). Run without --preflight to build.")
    return 0


# ----------------------------------------------------------------- preview
def preview(cfg):
    export, rows, review, errors = _plan(cfg)
    print(f"Preview: {cfg.get('client','?')} ({cfg.get('marketplace','?')}) [UPDATE mode] · "
          f"{len(rows)} row(s), nothing written\n")
    for line in review:
        print(f"  {line}")
    if errors:
        print(f"\n{len(errors)} error(s) would block the build:")
        for e in errors:
            print(f"  [MISSING] {e}")
        return 1
    return 0


# ----------------------------------------------------------------- review file
def write_review(cfg, review, rows, xlsx):
    md = xlsx.with_name(xlsx.stem + "_REVIEW.md")
    applied = [r for r in review if not r.startswith("SKIPPED")]
    skipped = [r for r in review if r.startswith("SKIPPED")]
    n_archive = sum(1 for r in rows if r["Operation"] == "Archive")
    n_update = sum(1 for r in rows if r["Operation"] == "Update")
    n_create = sum(1 for r in rows if r["Operation"] == "Create")
    lines = [
        f"# Campaign UPDATE review: {cfg['client']} ({cfg.get('marketplace')})",
        "",
        f"- File: `{xlsx.name}` (sheet `{SHEET_NAMES['SP']}`, {len(rows)} rows)",
        f"- {n_update} update row(s) · {n_archive} archive row(s) · {n_create} create row(s) "
        f"(additions to existing ad groups only, no brand-new campaigns/ad groups)",
        f"- Source export: `{cfg.get('export_file')}`",
        "",
        "**This changes a LIVE account when uploaded.** Every line below is a plain-English "
        "description of exactly what will change. Read it before uploading.",
        "",
        "## What will change",
        "",
    ]
    lines += [f"- {r}" for r in applied] if applied else [
        "- (no effective changes: every requested change was a no-op or a cascade-skip)"]
    if skipped:
        lines += ["", "## Skipped (no-op / cascade; nothing uploaded for these)", ""]
        lines += [f"- {r}" for r in skipped]
    lines += [
        "",
        "## Upload (operator action, not automated)",
        "",
        "1. Campaign Manager > Bulk Operations > Upload your file.",
        "2. Amazon validates the sheet and reports updated/archived/created entities or errors.",
        "3. Re-download a fresh bulksheets export before running another update batch; IDs "
        "and current field values must always come from the latest export, never a cached one.",
        "",
        "Generated by `tools/amazon-campaign-builder/update_campaigns.py`. Do not hand-edit; "
        "change the change-set config and rebuild.",
    ]
    md.write_text("\n".join(lines))
    return md


# ----------------------------------------------------------------- build
def build(cfg, override_out=None):
    from openpyxl import Workbook

    export, rows, review, errors = _plan(cfg)
    if errors:
        print(f"Preflight failed with {len(errors)} error(s):")
        for e in errors:
            print(f"  [MISSING] {e}")
        return 1
    if not rows:
        print("Nothing to build: the change-set produced no effective rows (all no-op/cascade).")
        for r in review:
            print(f"  {r}")
        return 1

    xlsx = out_path(cfg, override_out)
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAMES["SP"]
    ws.append(COLUMNS["SP"])
    for row in rows:
        ws.append([row[c] for c in COLUMNS["SP"]])
    for i, col in enumerate(COLUMNS["SP"], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(len(col) + 2, 18)
    wb.save(xlsx)

    review_path = write_review(cfg, review, rows, xlsx)
    print(f"Built {_rel(xlsx)}")
    print(f"Review {_rel(review_path)}\n")
    for r in review:
        print(f"  {r}")
    print()
    return validate(cfg, override_out)


# ----------------------------------------------------------------- QA gates
PRIMARY_ID_COL = {
    "Campaign": "Campaign ID", "Ad Group": "Ad Group ID",
    "Keyword": "Keyword ID", "Negative Keyword": "Keyword ID",
    "Campaign Negative Keyword": "Keyword ID",
    "Product Targeting": "Product Targeting ID",
    "Negative Product Targeting": "Product Targeting ID",
    "Bidding Adjustment": "Campaign ID",
}
EXPORT_INDEX_FOR_ENTITY = {
    "Campaign": "campaigns", "Ad Group": "ad_groups", "Keyword": "keywords",
    "Negative Keyword": "negatives", "Campaign Negative Keyword": "negatives",
    "Product Targeting": "targets", "Negative Product Targeting": "neg_targets",
    "Bidding Adjustment": "campaigns",
}


def validate(cfg, override_out=None):
    from openpyxl import load_workbook

    xlsx = out_path(cfg, override_out)
    fails, warns = [], []
    if not xlsx.exists():
        print(f"VALIDATE: file not found at {xlsx}")
        return 1
    export = _load_export_or_die(cfg)

    wb = load_workbook(xlsx, data_only=True)
    if SHEET_NAMES["SP"] not in wb.sheetnames:
        fails.append(f"sheet '{SHEET_NAMES['SP']}' missing (found {wb.sheetnames})")
        return _report(fails, warns)
    ws = wb[SHEET_NAMES["SP"]]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if header != COLUMNS["SP"]:
        fails.append(f"header mismatch: {header} != expected {COLUMNS['SP']}")
        return _report(fails, warns)
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({h: (ws.cell(r, c + 1).value if ws.cell(r, c + 1).value is not None else "")
                     for c, h in enumerate(header)})

    archived_campaigns = {str(r["Campaign ID"]) for r in rows
                          if r["Entity"] == "Campaign" and r["Operation"] == "Archive"}
    archived_ad_groups = {str(r["Ad Group ID"]) for r in rows
                          if r["Entity"] == "Ad Group" and r["Operation"] == "Archive"}

    for i, r in enumerate(rows, 2):
        ent, op = r["Entity"], r["Operation"]
        if r["Product"] != "Sponsored Products":
            fails.append(f"row {i}: Product != 'Sponsored Products'")
        if op not in ("Update", "Archive", "Create"):
            fails.append(f"row {i}: Operation '{op}' invalid for an update-mode file")
            continue

        # gate: every Update/Archive row resolves to a real ID present in the export
        if op in ("Update", "Archive"):
            id_col = PRIMARY_ID_COL.get(ent)
            idx_name = EXPORT_INDEX_FOR_ENTITY.get(ent)
            if id_col is None or idx_name is None:
                fails.append(f"row {i}: unexpected Entity '{ent}' for Update/Archive")
            else:
                rid = str(r[id_col])
                if not looks_like_real_id(rid):
                    fails.append(f"row {i} ({ent} {op}): {id_col} '{rid}' is not a real "
                                 f"bulksheets ID (looks like a temp ID)")
                elif ent != "Bidding Adjustment" and rid not in getattr(export, idx_name):
                    fails.append(f"row {i} ({ent} {op}): {id_col} '{rid}' not found in the "
                                 f"loaded export")

        # gate: no stray Create rows for whole new Campaigns/Ad Groups in an update file
        if op == "Create" and ent in ("Campaign", "Ad Group"):
            fails.append(f"row {i}: stray Create row for a whole new {ent} in an update-mode "
                         f"file; brand-new campaigns/ad groups belong in create mode, not an "
                         f"update batch")
        if op == "Create" and ent in ("Keyword", "Negative Keyword", "Campaign Negative Keyword",
                                       "Product Targeting"):
            camp, agid = str(r.get("Campaign ID", "")), str(r.get("Ad Group ID", ""))
            if not looks_like_real_id(camp) or camp not in export.campaigns:
                fails.append(f"row {i} ({ent} Create): Campaign ID '{camp}' is not a real "
                             f"campaign from the export; new {ent} rows must attach to an "
                             f"EXISTING campaign/ad group, never a temp one")
            if ent != "Campaign Negative Keyword" and \
                    (not looks_like_real_id(agid) or agid not in export.ad_groups):
                fails.append(f"row {i} ({ent} Create): Ad Group ID '{agid}' is not a real ad "
                             f"group from the export")

        # gate: no-op Update rows dropped (defensive re-check on the built file)
        if op == "Update" and ent in ("Campaign", "Ad Group", "Keyword"):
            fields = {"Campaign": ("Campaign Name", "Daily Budget", "Bidding Strategy", "State"),
                      "Ad Group": ("Ad Group Name", "Ad Group Default Bid", "State"),
                      "Keyword": ("State",)}[ent]
            if not any(str(r.get(f, "")).strip() for f in fields):
                fails.append(f"row {i} ({ent} Update): no-op row; every editable field is "
                             f"blank; drop rows that change nothing")

        # gate: Portfolio ID present on campaign updates that belong to a portfolio
        if ent == "Campaign" and op == "Update":
            portfolio = export.campaigns.get(str(r["Campaign ID"]), {}).get("Portfolio ID", "")
            if str(portfolio).strip() and not str(r.get("Portfolio ID", "")).strip():
                fails.append(f"row {i}: Campaign {r['Campaign ID']} belongs to Portfolio "
                             f"'{portfolio}' but the Update row has no Portfolio ID; it would "
                             f"silently leave the portfolio")

        # gate: no parent+child double archive
        if ent == "Ad Group" and op == "Archive" and str(r["Campaign ID"]) in archived_campaigns:
            fails.append(f"row {i}: Ad Group {r['Ad Group ID']} archived while its parent "
                         f"Campaign {r['Campaign ID']} is also archived in this file; "
                         f"archiving the campaign already cascades to it")
        if ent in ("Keyword", "Negative Keyword", "Campaign Negative Keyword", "Product Targeting") \
                and op == "Archive":
            camp, agid = str(r.get("Campaign ID", "")), str(r.get("Ad Group ID", ""))
            if camp in archived_campaigns or (agid and agid in archived_ad_groups):
                fails.append(f"row {i}: {ent} archived while its parent Campaign/Ad Group is "
                             f"also archived in this file; redundant, drop the child row")

    return _report(fails, warns)


def _report(fails, warns):
    for w in warns:
        print(f"  [WARN] {w}")
    for f in fails:
        print(f"  [FAIL] {f}")
    if fails:
        print(f"VALIDATE: FAIL ({len(fails)} gate(s), {len(warns)} warning(s))")
        return 1
    print(f"VALIDATE: PASS ({len(warns)} warning(s))")
    return 0


# ----------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Amazon SP campaign UPDATE bulk-file builder")
    ap.add_argument("--config", required=True,
                    help="change-set config JSON (copy config.UPDATE.TEMPLATE.json)")
    ap.add_argument("--out", help="override the output .xlsx path")
    ap.add_argument("--preflight", action="store_true", help="check the change-set, list missing/invalid")
    ap.add_argument("--preview", action="store_true", help="print planned changes in plain English, write nothing")
    ap.add_argument("--validate", action="store_true", help="run QA gates on the built file only")
    args = ap.parse_args()

    cfg = load_config(args.config)
    if args.preflight:
        return preflight(cfg)
    if args.preview:
        return preview(cfg)
    if args.validate:
        return validate(cfg, args.out)
    if preflight(cfg) != 0:
        return 1
    print()
    return build(cfg, args.out)


if __name__ == "__main__":
    sys.exit(main())
