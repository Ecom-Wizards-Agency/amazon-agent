#!/usr/bin/env python3
"""
Amazon SP campaign builder: text brief -> config -> bulk-upload .xlsx.

  # Preflight: check the config contract, list what's missing or READY
  python3 tools/amazon-campaign-builder/build_campaigns.py --config <cfg> --preflight

  # Preview: print the planned campaigns without writing anything
  python3 tools/amazon-campaign-builder/build_campaigns.py --config <cfg> --preview

  # Build: write the bulksheet .xlsx + _REVIEW.md, then run the QA gates
  python3 tools/amazon-campaign-builder/build_campaigns.py --config <cfg>

  # QA gates only (re-check an already-built file)
  python3 tools/amazon-campaign-builder/build_campaigns.py --config <cfg> --validate

Output is a FILE ONLY. This tool never uploads or touches live campaigns.
Uploading via Campaign Manager > Bulk Operations is a separate, operator-
confirmed action (stop-before-risk). Everything client-specific lives in
the config (see config.TEMPLATE.json / NEW-CLIENT.md).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent
sys.path.insert(0, str(HERE))

from campaign_model import (  # noqa: E402
    AMAZON_BIDDING, AMAZON_MATCH, AMAZON_NEG_MATCH, AUTO_EXPRESSIONS,
    BIDDING_STRATEGIES, CAMPAIGN_PURPOSE_BIDDING, CAMPAIGN_PURPOSES,
    CAMPAIGN_TYPE_DEFAULT_PURPOSE, CAMPAIGN_TYPES, CAMPAIGN_TYPE_GOALS, CAMPAIGN_TYPE_MATCH,
    COLUMNS, DEFAULT_BID, DEFAULT_BUDGET, GOALS, MATCH_TYPES, MIN_BID,
    NAMING_PRESETS, NEGATIVE_LEVELS, NEGATIVE_MATCH_TYPES, PLACEMENT_LABELS, SHEET_NAMES, STATES,
    build_bulk_rows, generate_campaigns, parse_product_list,
    resolve_bidding_strategy, resolve_campaign_purpose,
)

# v2 default: the Ecom Wizards 8-slot naming convention (naming-convention.md).
# Every existing config that sets its own naming.variable_order is unaffected.
# See load_config(): an explicit variable_order always wins over this default.
# Opt back into the pre-v2 order with `"naming": {"preset": "LEGACY"}`.
NAMING_DEFAULTS = NAMING_PRESETS["EW"]

MAX_CAMPAIGN_NAME = 128
MAX_AD_GROUP_NAME = 255
MAX_BID = 1000.0
MIN_BUDGET = 1.0
ASIN_RE = re.compile(r"^[A-Z0-9]{10}$")


def slugify(s):
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def _rel(p):
    try:
        return p.relative_to(REPO)
    except ValueError:
        return p


def load_config(path):
    cfg = json.loads(Path(path).read_text())
    cfg.setdefault("defaults", {})
    naming_in = cfg.get("naming", {})
    preset = NAMING_PRESETS.get((naming_in.get("preset") or "").upper(), NAMING_DEFAULTS)
    cfg["naming"] = {**preset, **{k: v for k, v in naming_in.items() if k != "preset"}}
    return cfg


def out_path(cfg, override=None):
    if override:
        return Path(override)
    stamp = cfg.get("date") or date.today().isoformat()
    brand = re.sub(r"\s+", "-", (cfg.get("brand") or cfg["client"]).strip())
    name = f"{stamp}_{brand}_{cfg.get('marketplace', 'US')}_SP_bulk_campaigns.xlsx"
    return REPO / "output" / slugify(cfg["client"]) / "ads" / name


# ----------------------------------------------------------------- form assembly
def campaign_forms(cfg):
    """Merge config defaults into each campaigns[] entry -> app-form dicts."""
    d = cfg["defaults"]
    forms = []
    for spec in cfg.get("campaigns", []):
        def pick(key, fallback):
            v = spec.get(key)
            if v in (None, ""):
                v = d.get(key)
            return fallback if v in (None, "") else v

        ctype = spec.get("campaign_type", "")
        goals = CAMPAIGN_TYPE_GOALS.get(ctype, ["Discovery"])
        keywords = spec.get("target_asins") if ctype == "PAT" and spec.get("target_asins") \
            else spec.get("keywords", [])
        if isinstance(keywords, str):
            keywords = [k for k in re.split(r"\n", keywords)]
        form = {
            "campaign_type": ctype,
            "campaign_purpose": (spec.get("campaign_purpose") or "").strip().upper(),
            "goal": spec.get("goal") or goals[0],
            "product_name": spec.get("product_name", ""),
            "target_descriptor": spec.get("target_descriptor", ""),
            "sku": spec.get("sku", ""),
            "asin": spec.get("asin", ""),
            "keywords_raw": "\n".join(str(k) for k in keywords),
            "keywords_per_campaign": int(spec.get("keywords_per_campaign") or 0),
            "transpose_keywords": bool(spec.get("transpose_keywords")),
            "swap_name_order": bool(spec.get("swap_name_order")),
            "skw_include_keyword_in_name": bool(spec.get("skw_include_keyword_in_name", True)),
            "match_type": spec.get("match_type") or "",
            "bmm_modifier": bool(spec.get("bmm_modifier")),
            "daily_budget": float(pick("daily_budget", DEFAULT_BUDGET)),
            "keyword_bid": float(pick("keyword_bid", DEFAULT_BID)),
            "bidding_strategy": pick("bidding_strategy", ""),
            "portfolio_id": str(pick("portfolio_id", "")),
            "negative_keywords": spec.get("negative_keywords", []),
            "negative_match_type": spec.get("negative_match_type") or "NEGATIVE_EXACT",
            "negative_level": spec.get("negative_level") or "ad_group",
            "state": pick("state", "paused"),
            "start_date": str(pick("start_date", "")),
            "site_restriction": pick("site_restriction", "Amazon"),
        }
        for grp in ("close_match", "loose_match", "substitutes", "complements"):
            bid, state = spec.get(f"auto_{grp}_bid"), spec.get(f"auto_{grp}_state")
            form[f"auto_{grp}_bid"] = bid if bid not in ("", None) else None
            form[f"auto_{grp}_state"] = state if state in STATES else None
        forms.append(form)
    return forms


def generate_all(cfg):
    campaigns = []
    for form in campaign_forms(cfg):
        campaigns.extend(generate_campaigns(form, cfg["naming"]))
    return campaigns


# ----------------------------------------------------------------- preflight
def preflight(cfg):
    issues, notes = [], []
    for key in ("client", "marketplace"):
        if not cfg.get(key):
            issues.append(f"config: `{key}` is required")
    if not cfg.get("campaigns"):
        issues.append("config: `campaigns[]` is empty (nothing to build)")

    vendor = bool(cfg["defaults"].get("vendor_central_mode") or cfg.get("vendor_central_mode"))
    for i, (spec, form) in enumerate(zip(cfg.get("campaigns", []), campaign_forms(cfg)), 1):
        tag = f"campaign {i} ({form['campaign_type'] or '?'})"
        ctype = form["campaign_type"]
        if ctype not in CAMPAIGN_TYPES:
            issues.append(f"{tag}: campaign_type must be one of {'/'.join(CAMPAIGN_TYPES)}")
            continue
        if not form["product_name"]:
            issues.append(f"{tag}: product_name is required (used in the campaign name)")
        skus, asins = parse_product_list(form["sku"]), parse_product_list(form["asin"])
        if vendor and not asins:
            issues.append(f"{tag}: vendor mode needs asin(s) for the Product Ad rows")
        if not vendor and not skus:
            issues.append(f"{tag}: sku(s) required for the Product Ad rows (seller accounts advertise by SKU)")
        kws = [k for k in form["keywords_raw"].split("\n") if k.strip()]
        if ctype in ("SKW", "Halo", "BMM", "Phrase") and not kws:
            issues.append(f"{tag}: keywords[] is required for {ctype}")
        if ctype == "PAT" and not kws:
            issues.append(f"{tag}: target_asins[] is required for PAT")
        if ctype == "PAT":
            bad = [a for a in kws if not ASIN_RE.match(a.strip().upper())]
            if bad:
                issues.append(f"{tag}: target_asins entries not ASIN-shaped: {', '.join(bad[:5])}")
        if form["goal"] not in GOALS:
            issues.append(f"{tag}: goal must be one of {'/'.join(GOALS)}")
        elif form["goal"] not in CAMPAIGN_TYPE_GOALS[ctype]:
            notes.append(f"{tag}: goal '{form['goal']}' is unusual for {ctype} "
                         f"(app allows {'/'.join(CAMPAIGN_TYPE_GOALS[ctype])})")
        if form["match_type"] and form["match_type"] not in MATCH_TYPES:
            issues.append(f"{tag}: match_type must be one of {'/'.join(MATCH_TYPES)} (or empty for the "
                          f"{ctype} default {CAMPAIGN_TYPE_MATCH[ctype]})")
        if form["campaign_purpose"] and form["campaign_purpose"] not in CAMPAIGN_PURPOSES:
            issues.append(f"{tag}: campaign_purpose must be one of {'/'.join(CAMPAIGN_PURPOSES)} (or empty for "
                          f"the {ctype} default {CAMPAIGN_TYPE_DEFAULT_PURPOSE.get(ctype, '?')})")
        if form["bidding_strategy"] and form["bidding_strategy"] not in BIDDING_STRATEGIES:
            issues.append(f"{tag}: bidding_strategy must be one of {'/'.join(BIDDING_STRATEGIES)} (or empty)")
        elif form["bidding_strategy"] and ctype in CAMPAIGN_TYPE_DEFAULT_PURPOSE:
            purpose = form["campaign_purpose"] or CAMPAIGN_TYPE_DEFAULT_PURPOSE[ctype]
            expected = CAMPAIGN_PURPOSE_BIDDING.get(purpose)
            if expected and form["bidding_strategy"] != expected:
                notes.append(f"{tag}: bidding_strategy override '{form['bidding_strategy']}' differs from the "
                             f"naming-convention.md default '{expected}' for purpose {purpose} (QC-enforced table)")
        if form["state"] not in STATES:
            issues.append(f"{tag}: state must be enabled|paused")
        if not MIN_BID <= form["keyword_bid"] <= MAX_BID:
            issues.append(f"{tag}: keyword_bid {form['keyword_bid']} outside [{MIN_BID}, {MAX_BID}]")
        if form["daily_budget"] < MIN_BUDGET:
            issues.append(f"{tag}: daily_budget {form['daily_budget']} below Amazon's minimum {MIN_BUDGET}")
        if form["negative_match_type"] not in NEGATIVE_MATCH_TYPES:
            issues.append(f"{tag}: negative_match_type must be one of {'/'.join(NEGATIVE_MATCH_TYPES)}")
        if form["negative_level"] not in NEGATIVE_LEVELS:
            issues.append(f"{tag}: negative_level must be ad_group|campaign")
        if form["transpose_keywords"] and form["keywords_per_campaign"] < 1:
            issues.append(f"{tag}: transpose_keywords needs keywords_per_campaign >= 1")
        order = cfg["naming"]["variable_order"]
        fans_out = (form["transpose_keywords"] and form["keywords_per_campaign"] >= 1
                    and len(kws) > form["keywords_per_campaign"]) \
            or (ctype == "SKW" and len(kws) > 1 and not form["skw_include_keyword_in_name"])
        # A literal keyword slot disambiguates SKW fan-out on its own (each campaign gets a
        # different Keyword token); Counter/CampCounter both work for everything else, but
        # CampCounter only actually emits a value for Halo/Auto per naming-convention.md.
        disambiguated_by_keyword = "Keyword" in order and ctype == "SKW" and form["skw_include_keyword_in_name"]
        disambiguated_by_counter = "Counter" in order or ("CampCounter" in order and ctype in ("Halo", "Auto"))
        if fans_out and not disambiguated_by_keyword and not disambiguated_by_counter:
            issues.append(f"{tag}: fans out to several identically-named campaigns; add 'Counter' "
                          f"(or, for Halo/Auto, 'CampCounter') to naming.variable_order (Amazon rejects "
                          f"duplicate campaign names)")
        if ctype in ("BMM", "Phrase") and not form["negative_keywords"]:
            notes.append(f"{tag}: discovery campaign has no negative_keywords; naming-convention.md QC "
                         f"requires a Never-Ever/negative-phrase list at ad-group level from day one")
        purpose_for_qc = form["campaign_purpose"] or CAMPAIGN_TYPE_DEFAULT_PURPOSE.get(ctype, "")
        if ctype == "PAT" and form["match_type"] == "ASIN_EXPANDED" and purpose_for_qc == "SELF_TARGETING":
            notes.append(f"{tag}: Self-Targeting Expanded needs a negative product list of the targeted ASINs "
                         f"(naming-convention.md QC #7); this toolkit doesn't model Negative Product Targeting "
                         f"yet; add it via an update-mode change-set after the initial build")
        if form["start_date"]:
            try:
                sd = datetime.strptime(form["start_date"], "%Y-%m-%d").date()
                if sd < date.today():
                    issues.append(f"{tag}: start_date {form['start_date']} is in the past; Amazon "
                                  f"rejects past start dates on create")
            except ValueError:
                issues.append(f"{tag}: start_date must be YYYY-MM-DD")
        if form["state"] == "enabled":
            notes.append(f"{tag}: state=enabled means campaigns go LIVE on upload; the safety default is paused")

    campaigns = generate_all(cfg) if not issues else []
    for c in campaigns:
        if c["ad_group_name"] == c["campaign_name"]:
            notes.append(f"'{c['campaign_name']}': ad group name equals campaign name; naming-convention.md "
                         f"QC requires the ad group name to differ (drop prefix & suffix)")

    print(f"Preflight: {cfg.get('client', '?')} ({cfg.get('marketplace', '?')})")
    for msg in issues:
        print(f"  [MISSING] {msg}")
    for msg in notes:
        print(f"  [NOTE]    {msg}")
    if issues:
        print(f"\nNOT READY. Fix the {len(issues)} item(s) above in the config.")
        return 1
    print(f"\nREADY: {len(cfg['campaigns'])} spec(s) -> {len(campaigns)} campaign(s). "
          f"Run without --preflight to build.")
    return 0


# ----------------------------------------------------------------- preview / review
def summarize(cfg, campaigns):
    lines = []
    for c in campaigns:
        if c["targeting_type"] == "AUTO":
            targets = "4 auto groups"
        elif c["campaign_type"] == "PAT":
            targets = f"{len(c['asins'])} ASIN target(s)"
        else:
            targets = f"{len(c['keywords'])} keyword(s) [{AMAZON_MATCH[c['match_type']]}]"
        negs = f", {len(c['negative_keywords'])} neg ({c['negative_level']})" if c["negative_keywords"] else ""
        lines.append(f"{c['campaign_name']}\n"
                     f"    {c['campaign_type']} · {c['targeting_type']} · {c['state']} · "
                     f"budget {c['daily_budget']:.2f} · bid {c['keyword_bid']:.2f} · "
                     f"{AMAZON_BIDDING[c['bidding_strategy']]} · {targets}{negs}")
    return lines


def preview(cfg):
    campaigns = generate_all(cfg)
    print(f"Preview: {cfg['client']} ({cfg.get('marketplace')}) · "
          f"{len(campaigns)} campaign(s), nothing written\n")
    for line in summarize(cfg, campaigns):
        print(f"  {line}")
    total = sum(c["daily_budget"] for c in campaigns)
    print(f"\n  Combined daily budget: {total:.2f}")
    return 0


def write_review(cfg, campaigns, rows, xlsx):
    md = xlsx.with_name(xlsx.stem + "_REVIEW.md")
    n_kw = sum(1 for r in rows if r["Entity"] == "Keyword")
    n_pt = sum(1 for r in rows if r["Entity"] == "Product Targeting")
    n_neg = sum(1 for r in rows if "Negative" in r["Entity"])
    enabled = [c for c in campaigns if c["state"] == "enabled"]
    lines = [
        f"# Campaign build review: {cfg['client']} ({cfg.get('marketplace')})",
        "",
        f"- File: `{xlsx.name}` (sheet `{SHEET_NAMES['SP']}`, {len(rows)} rows)",
        f"- {len(campaigns)} campaign(s) · {n_kw} keyword target(s) · {n_pt} product target(s) · "
        f"{n_neg} negative(s)",
        f"- Combined daily budget: {sum(c['daily_budget'] for c in campaigns):.2f}",
        f"- State: {len(enabled)} enabled / {len(campaigns) - len(enabled)} paused"
        + (" (**enabled campaigns go live on upload**)" if enabled else ""),
        "",
        "| Campaign | Type | State | Budget | Bid | Strategy | Targets | Negatives |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for c in campaigns:
        targets = ("4 auto groups" if c["targeting_type"] == "AUTO"
                   else f"{len(c['asins'])} ASINs" if c["campaign_type"] == "PAT"
                   else f"{len(c['keywords'])} kw ({AMAZON_MATCH[c['match_type']]})")
        lines.append(f"| {c['campaign_name']} | {c['campaign_type']} | {c['state']} | "
                     f"{c['daily_budget']:.2f} | {c['keyword_bid']:.2f} | "
                     f"{AMAZON_BIDDING[c['bidding_strategy']]} | {targets} | "
                     f"{len(c['negative_keywords'])} ({c['negative_level']}) |")
    lines += [
        "",
        "## Upload (operator action, not automated)",
        "",
        "1. Campaign Manager > Bulk Operations > Upload your file.",
        "2. Amazon validates the sheet and reports created entities / errors.",
        "3. Campaigns arrive in the state above; flip paused campaigns on when ready.",
        "",
        "Generated by `tools/amazon-campaign-builder/`. Do not hand-edit the builder; "
        "change the config and rebuild.",
    ]
    md.write_text("\n".join(lines))
    return md


# ----------------------------------------------------------------- build
def build(cfg, override_out=None):
    from openpyxl import Workbook

    campaigns = generate_all(cfg)
    if not campaigns:
        print("Nothing to build: campaigns[] produced no campaigns.")
        return 1
    defaults = {**cfg["defaults"], "vendor_central_mode":
                bool(cfg["defaults"].get("vendor_central_mode") or cfg.get("vendor_central_mode"))}
    rows = build_bulk_rows(campaigns, defaults)

    xlsx = out_path(cfg, override_out)
    xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAMES["SP"]
    ws.append(COLUMNS["SP"])
    for row in rows:
        ws.append([row[c] for c in COLUMNS["SP"]])
    for i, col in enumerate(COLUMNS["SP"], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(len(col) + 2, 18)
    wb.save(xlsx)

    review = write_review(cfg, campaigns, rows, xlsx)
    print(f"Built {_rel(xlsx)}")
    print(f"Review {_rel(review)}\n")
    for line in summarize(cfg, campaigns):
        print(f"  {line}")
    print()
    return validate(cfg, override_out)


# ----------------------------------------------------------------- QA gates
def validate(cfg, override_out=None):
    from openpyxl import load_workbook

    xlsx = out_path(cfg, override_out)
    fails, warns = [], []
    if not xlsx.exists():
        print(f"VALIDATE: file not found at {xlsx}")
        return 1
    wb = load_workbook(xlsx, data_only=True)
    if SHEET_NAMES["SP"] not in wb.sheetnames:
        fails.append(f"sheet '{SHEET_NAMES['SP']}' missing (found {wb.sheetnames})")
        return _report(fails, warns)
    ws = wb[SHEET_NAMES["SP"]]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if header != COLUMNS["SP"]:
        fails.append(f"header mismatch: {header} != expected {COLUMNS['SP']}")
        return _report(fails, warns)

    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({h: (ws.cell(r, c + 1).value if ws.cell(r, c + 1).value is not None else "")
                     for c, h in enumerate(header)})

    camp_ids = {str(r["Campaign ID"]) for r in rows if r["Entity"] == "Campaign"}
    ag_ids = {(str(r["Campaign ID"]), str(r["Ad Group ID"])) for r in rows if r["Entity"] == "Ad Group"}
    today = date.today().isoformat().replace("-", "")
    camp_names, kw_seen = {}, {}

    for i, r in enumerate(rows, 2):
        ent = r["Entity"]
        if r["Product"] != "Sponsored Products":
            fails.append(f"row {i}: Product != 'Sponsored Products'")
        if r["Operation"] != "Create":
            fails.append(f"row {i}: Operation != 'Create'")
        if ent != "Campaign" and str(r["Campaign ID"]) not in camp_ids:
            fails.append(f"row {i} ({ent}): Campaign ID {r['Campaign ID']} has no Campaign row")
        if ent in ("Product Ad", "Keyword", "Negative Keyword", "Product Targeting") \
                and (str(r["Campaign ID"]), str(r["Ad Group ID"])) not in ag_ids:
            fails.append(f"row {i} ({ent}): Ad Group ID {r['Ad Group ID']} has no Ad Group row "
                         f"in campaign {r['Campaign ID']}")
        if r["State"] and r["State"] not in STATES:
            fails.append(f"row {i}: State '{r['State']}' invalid")

        if ent == "Campaign":
            name = r["Campaign Name"]
            if name in camp_names:
                fails.append(f"row {i}: duplicate campaign name '{name}' (also row {camp_names[name]}); "
                             f"Amazon rejects duplicate names on create")
            camp_names[name] = i
            if len(str(name)) > MAX_CAMPAIGN_NAME:
                fails.append(f"row {i}: campaign name > {MAX_CAMPAIGN_NAME} chars")
            if r["Targeting Type"] not in ("MANUAL", "AUTO"):
                fails.append(f"row {i}: Targeting Type '{r['Targeting Type']}' invalid")
            if r["Bidding Strategy"] not in AMAZON_BIDDING.values():
                fails.append(f"row {i}: Bidding Strategy '{r['Bidding Strategy']}' not an Amazon value")
            if float(r["Daily Budget"] or 0) < MIN_BUDGET:
                fails.append(f"row {i}: Daily Budget below {MIN_BUDGET}")
            sd = str(r["Start Date"])
            if not re.match(r"^\d{8}$", sd):
                fails.append(f"row {i}: Start Date '{sd}' not YYYYMMDD")
            elif sd < today:
                fails.append(f"row {i}: Start Date {sd} is in the past")
        elif ent == "Ad Group":
            if len(str(r["Ad Group Name"])) > MAX_AD_GROUP_NAME:
                fails.append(f"row {i}: ad group name > {MAX_AD_GROUP_NAME} chars")
            if not MIN_BID <= float(r["Ad Group Default Bid"] or 0) <= MAX_BID:
                fails.append(f"row {i}: Ad Group Default Bid outside [{MIN_BID}, {MAX_BID}]")
        elif ent == "Bidding Adjustment":
            if r["Placement"] not in PLACEMENT_LABELS.values():
                fails.append(f"row {i}: Placement '{r['Placement']}' not an Amazon value")
            pct = r["Percentage"]
            if not (isinstance(pct, int) or (isinstance(pct, float) and pct.is_integer())) \
                    or not 0 <= int(pct) <= 900:
                fails.append(f"row {i}: Percentage '{pct}' must be a whole number 0..900")
        elif ent == "Product Ad":
            if not r["SKU"] and not r["ASIN"]:
                fails.append(f"row {i}: Product Ad with neither SKU nor ASIN")
        elif ent == "Keyword":
            if r["Match Type"] not in AMAZON_MATCH.values():
                fails.append(f"row {i}: Match Type '{r['Match Type']}' not an Amazon value")
            if not MIN_BID <= float(r["Bid"] or 0) <= MAX_BID:
                fails.append(f"row {i}: Bid outside [{MIN_BID}, {MAX_BID}]")
            key = (str(r["Campaign ID"]), str(r["Ad Group ID"]),
                   str(r["Keyword Text"]).lower().strip(), r["Match Type"])
            if key in kw_seen:
                fails.append(f"row {i}: duplicate keyword+match in the same ad group "
                             f"('{r['Keyword Text']}', also row {kw_seen[key]})")
            kw_seen[key] = i
        elif ent in ("Negative Keyword", "Campaign Negative Keyword"):
            if r["Match Type"] not in AMAZON_NEG_MATCH.values():
                fails.append(f"row {i}: negative Match Type '{r['Match Type']}' not an Amazon value")
            words = str(r["Keyword Text"]).split()
            if len(words) > 10 or len(str(r["Keyword Text"])) > 80:
                warns.append(f"row {i}: negative '{r['Keyword Text']}' over Amazon's "
                             f"10-word/80-char cap")
        elif ent == "Product Targeting":
            expr = str(r["Product Targeting Expression"])
            ok = expr in AUTO_EXPRESSIONS or re.match(r'^(asin|asin-expanded)="[A-Z0-9]{10}"$', expr)
            if not ok:
                fails.append(f"row {i}: Product Targeting Expression '{expr}' invalid")
            if not MIN_BID <= float(r["Bid"] or 0) <= MAX_BID:
                fails.append(f"row {i}: Bid outside [{MIN_BID}, {MAX_BID}]")

    # cross-campaign self-competition + completeness
    across = {}
    for r in rows:
        if r["Entity"] == "Keyword":
            across.setdefault((str(r["Keyword Text"]).lower().strip(), r["Match Type"]), set()) \
                .add(str(r["Campaign ID"]))
    for (kw, mt), camps in across.items():
        if len(camps) > 1:
            warns.append(f"keyword '{kw}' ({mt}) appears in {len(camps)} campaigns (self-competition)")
    per_camp = {}
    for r in rows:
        per_camp.setdefault(str(r["Campaign ID"]), set()).add(r["Entity"])
    for cid, ents in per_camp.items():
        if "Campaign" not in ents:
            continue
        if "Product Ad" not in ents:
            fails.append(f"campaign {cid}: no Product Ad row, so nothing would be advertised")
        if not ents & {"Keyword", "Product Targeting"}:
            fails.append(f"campaign {cid}: no Keyword or Product Targeting row, so no targets")

    return _report(fails, warns)


def _report(fails, warns):
    for w in warns:
        print(f"  [WARN] {w}")
    for f in fails:
        print(f"  [FAIL] {f}")
    if fails:
        print(f"VALIDATE: FAIL ({len(fails)} gate(s), {len(warns)} warning(s))")
        return 1
    print(f"VALIDATE: PASS ({len(warns)} warning(s))")
    return 0


# ----------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Amazon SP campaign bulk-file builder")
    ap.add_argument("--config", "--brief", dest="config", required=True,
                    help="per-client campaign config JSON (copy config.TEMPLATE.json)")
    ap.add_argument("--out", help="override the output .xlsx path")
    ap.add_argument("--keyword-file", help="keyword-research workbook (.xlsx) to source "
                    "campaigns[] from (see keyword_workbook.py); merges ahead of any "
                    "campaigns[] already in the config")
    ap.add_argument("--keyword-sheet", help="sheet name inside --keyword-file to parse "
                    "(default: auto-detect '5. Campaign Structure' or fall back to the first sheet)")
    ap.add_argument("--preflight", action="store_true", help="check the config, list missing fields")
    ap.add_argument("--preview", action="store_true", help="print planned campaigns, write nothing")
    ap.add_argument("--validate", action="store_true", help="run QA gates on the built file only")
    args = ap.parse_args()

    cfg = load_config(args.config)
    if args.keyword_file:
        from keyword_workbook import parse_keyword_workbook
        kwd = cfg.get("keyword_file_defaults", {})
        specs = parse_keyword_workbook(args.keyword_file, sheet=args.keyword_sheet,
                                        product_name=kwd.get("product_name", ""),
                                        sku=kwd.get("sku"), asin=kwd.get("asin"))
        cfg["campaigns"] = specs + cfg.get("campaigns", [])
        print(f"keyword-file: {len(specs)} campaign spec(s) parsed from {args.keyword_file}\n")
    if args.preflight:
        return preflight(cfg)
    if args.preview:
        return preview(cfg)
    if args.validate:
        return validate(cfg, args.out)
    if preflight(cfg) != 0:
        return 1
    print()
    return build(cfg, args.out)


if __name__ == "__main__":
    sys.exit(main())
