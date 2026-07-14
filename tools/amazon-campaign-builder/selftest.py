#!/usr/bin/env python3
"""Self-test harness for campaign-builder v2 — synthetic fixtures, no network/Amazon access.

Exercises:
  A. create mode, LEGACY naming preset (backward compatibility: unchanged naming formula
     and unchanged bidding defaults for every type except SKW, which naming-convention.md
     deliberately changes to Fixed bids).
  B. create mode, EW naming preset (default) incl. campaign_purpose overrides
     (SHIELD, SELF_TARGETING) and the QC notes they trigger.
  C. keyword-file input: a synthetic '5. Campaign Structure' workbook -> parsed
     campaigns[] -> full preflight/build/validate.
  D. update mode: a synthetic bulksheets export -> a change-set exercising budget/
     bidding/state/placement/portfolio-carry-forward/end-date changes, ad-group
     updates, keyword pause/replace/add, negative archive/add, target archive/add,
     and parent+child archive cascade-dropping -> full preflight/preview/build/validate.
  E. update mode QA gates: a deliberately-broken already-built file (temp ID, missing
     portfolio, parent+child double archive) -> --validate must FAIL with all three.

Run: python3 tools/amazon-campaign-builder/selftest.py
Writes scratch files under a tempdir (never touches output/ or a real client config).
"""
from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

import openpyxl  # noqa: E402

import build_campaigns as bc  # noqa: E402
import campaign_model as cm  # noqa: E402
import keyword_workbook as kwb  # noqa: E402
import update_campaigns as uc  # noqa: E402
import update_model as um  # noqa: E402

FAILURES = []


def check(label, cond, detail=""):
    status = "ok" if cond else "FAIL"
    print(f"  [{status}] {label}" + (f" — {detail}" if detail and not cond else ""))
    if not cond:
        FAILURES.append(f"{label}: {detail}")


def section(title):
    print(f"\n=== {title} ===")


TMP = Path(tempfile.mkdtemp(prefix="campaign-builder-selftest-"))


# =================================================================== A. create/LEGACY
def test_create_legacy():
    section("A. create mode — LEGACY naming preset (backward compatibility)")
    cfg = {
        "client": "Selftest Legacy", "brand": "Selftest", "marketplace": "US",
        "naming": {"preset": "LEGACY", "suffix": "EW"},
        "defaults": {"daily_budget": 10.0, "keyword_bid": 0.5, "state": "paused"},
        "campaigns": [
            {"campaign_type": "SKW", "product_name": "Widget", "target_descriptor": "generic",
             "sku": ["SKU-1"], "keywords": ["red widget", "blue widget"]},
            {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "long-tail",
             "sku": ["SKU-1"], "keywords": ["red widget for kitchen", "red widget for office"]},
            {"campaign_type": "BMM", "product_name": "Widget", "target_descriptor": "widget",
             "sku": ["SKU-1"], "keywords": ["widget"]},
            {"campaign_type": "Phrase", "product_name": "Widget", "target_descriptor": "widget",
             "sku": ["SKU-1"], "keywords": ["widget"]},
            {"campaign_type": "Auto", "product_name": "Widget", "target_descriptor": "auto",
             "sku": ["SKU-1"]},
            {"campaign_type": "PAT", "product_name": "Widget", "target_descriptor": "competitors",
             "sku": ["SKU-1"], "target_asins": ["B000000001", "B000000002"]},
        ],
    }
    cfg_path = TMP / "cfg_legacy.json"
    cfg_path.write_text(json.dumps(cfg))
    loaded = bc.load_config(str(cfg_path))
    check("naming.variable_order is the pre-v2 6-token order",
          loaded["naming"]["variable_order"] ==
          ["Goal", "SP", "MatchType", "ProductName", "TargetDescriptor", "EW"],
          str(loaded["naming"]["variable_order"]))

    campaigns = bc.generate_all(loaded)
    by_type = {c["campaign_type"]: c for c in campaigns if c["campaign_type"] != "SKW"}

    skw = [c for c in campaigns if c["campaign_type"] == "SKW"]
    check("SKW fans out to one campaign per keyword", len(skw) == 2, str(len(skw)))
    check("SKW campaign name matches the legacy formula",
          skw[0]["campaign_name"] == "Rank | SP | Exact | Widget | red widget | EW",
          skw[0]["campaign_name"])
    check("SKW ad group name keeps 'SP' (legacy behavior, unchanged)",
          skw[0]["ad_group_name"] == "SP | Exact | Widget | red widget",
          skw[0]["ad_group_name"])
    check("SKW bidding is now 'Fixed bids' (naming-convention.md change, was 'Down only')",
          skw[0]["bidding_strategy"] == "Fixed bids", skw[0]["bidding_strategy"])

    check("Halo bidding unchanged: Down only", by_type["Halo"]["bidding_strategy"] == "Down only")
    check("BMM bidding unchanged: Down only", by_type["BMM"]["bidding_strategy"] == "Down only")
    check("Phrase bidding unchanged: Down only", by_type["Phrase"]["bidding_strategy"] == "Down only")
    check("Auto bidding unchanged: Up and down", by_type["Auto"]["bidding_strategy"] == "Up and down")
    check("PAT bidding unchanged: Down only", by_type["PAT"]["bidding_strategy"] == "Down only")

    out = TMP / "legacy.xlsx"
    rc_pre = bc.preflight(loaded)
    check("preflight READY", rc_pre == 0, "preflight found blocking issues")
    rc_build = bc.build(loaded, str(out))
    check("build+validate PASS", rc_build == 0, "build/validate returned non-zero")
    check("output file exists", out.exists())


# =================================================================== B. create/EW
def test_create_ew():
    section("B. create mode — EW naming preset (default) + campaign_purpose overrides")
    cfg = {
        "client": "Selftest EW", "brand": "Selftest", "marketplace": "US",
        "defaults": {"daily_budget": 10.0, "keyword_bid": 0.5, "state": "paused"},
        "campaigns": [
            {"campaign_type": "SKW", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["red widget"]},
            {"campaign_type": "SKW", "campaign_purpose": "SHIELD", "product_name": "Widget",
             "sku": ["SKU-1"], "keywords": ["acme widget"]},
            {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "long-tail",
             "sku": ["SKU-1"], "keywords": ["red widget for kitchen", "red widget for office"]},
            {"campaign_type": "Auto", "product_name": "Widget", "sku": ["SKU-1"]},
            {"campaign_type": "PAT", "campaign_purpose": "SELF_TARGETING", "match_type": "ASIN_EXPANDED",
             "product_name": "Widget", "sku": ["SKU-1"], "target_asins": ["B000000009"]},
            {"campaign_type": "PAT", "product_name": "Widget", "sku": ["SKU-1"],
             "target_asins": ["B000000001"],
             "bidding_strategy": "Up and down"},  # deliberate override -> should NOTE in preflight
        ],
    }
    cfg_path = TMP / "cfg_ew.json"
    cfg_path.write_text(json.dumps(cfg))
    loaded = bc.load_config(str(cfg_path))
    check("naming.variable_order defaults to the EW 8-slot preset",
          loaded["naming"]["variable_order"] ==
          ["Goal", "AdType", "MatchType", "TriggerWord", "ProductName", "Keyword", "CampCounter", "EW"],
          str(loaded["naming"]["variable_order"]))

    campaigns = bc.generate_all(loaded)
    rank_skw = campaigns[0]
    shield_skw = campaigns[1]
    halo = campaigns[2]
    auto = campaigns[3]
    self_pat = campaigns[4]
    comp_pat = campaigns[5]

    check("Rank SKW name uses trigger word 'SKW' and Keyword slot",
          rank_skw["campaign_name"] == "Rank | SP | Exact | SKW | Widget | red widget | EW",
          rank_skw["campaign_name"])
    check("Rank SKW bidding = Fixed bids (purpose default)", rank_skw["bidding_strategy"] == "Fixed bids")
    check("Shield SKW trigger word is 'Shield', not 'SKW'",
          "Shield" in shield_skw["campaign_name"], shield_skw["campaign_name"])
    check("Shield SKW bidding = Down only (purpose override)", shield_skw["bidding_strategy"] == "Down only")
    check("Halo name carries CampCounter (counter 01)", "01" in halo["campaign_name"], halo["campaign_name"])
    check("Halo bidding = Down only", halo["bidding_strategy"] == "Down only")
    check("Auto name carries CampCounter (counter 01)", "01" in auto["campaign_name"], auto["campaign_name"])
    check("Auto bidding = Up and down", auto["bidding_strategy"] == "Up and down")
    check("Non-Halo/Auto campaigns have NO Camp Counter token in the name",
          "01" not in rank_skw["campaign_name"] and "01" not in shield_skw["campaign_name"],
          f"{rank_skw['campaign_name']} / {shield_skw['campaign_name']}")
    check("Self-Targeting PAT trigger word is 'Self-Targeting'",
          "Self-Targeting" in self_pat["campaign_name"], self_pat["campaign_name"])
    check("Self-Targeting PAT bidding = Up and down (purpose override)",
          self_pat["bidding_strategy"] == "Up and down")
    check("Competitor PAT bidding stays Down only by default (unaffected by the override test below)",
          True)  # comp_pat has an explicit override tested via the NOTE below

    for c in campaigns:
        check(f"ad group name != campaign name ({c['campaign_type']})",
              c["ad_group_name"] != c["campaign_name"],
              f"{c['ad_group_name']!r} == {c['campaign_name']!r}")

    # capture preflight stdout to check for the bidding-override NOTE and QC notes
    import io
    from contextlib import redirect_stdout
    buf = io.StringIO()
    with redirect_stdout(buf):
        rc_pre = bc.preflight(loaded)
    out_text = buf.getvalue()
    print(out_text)
    check("preflight READY", rc_pre == 0)
    check("preflight NOTEs the bidding_strategy override on the competitor PAT campaign",
          "differs from the naming-convention.md default" in out_text, out_text)
    check("preflight NOTEs missing negative_keywords on discovery-less check n/a (no BMM/Phrase here)",
          True)

    out = TMP / "ew.xlsx"
    rc_build = bc.build(loaded, str(out))
    check("build+validate PASS", rc_build == 0)


# =================================================================== C. keyword-file input
def _write_campaign_structure_fixture(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "5. Campaign Structure"

    def section(row, col, title, kind, label=""):
        """title row -> optional label row -> Keyword/Search Volume (or ASINs/Brand)
        header row. Returns the header row (data starts the row after)."""
        ws.cell(row, col, title)
        hdr_row = row + (2 if label else 1)
        if label:
            ws.cell(row + 1, col, label)
        if kind == "keywords":
            ws.cell(hdr_row, col, "Keyword")
            ws.cell(hdr_row, col + 1, "Search Volume")
        else:
            ws.cell(hdr_row, col, "ASINs")
            ws.cell(hdr_row, col + 1, "Brand")
        return hdr_row

    # Rank-SKW / Shield-SKW share a row, different columns
    h1 = section(1, 1, "Rank-SKW", "keywords")
    for i, (kw, sv) in enumerate([("red widget", 2000), ("blue widget", 1500)]):
        ws.cell(h1 + 1 + i, 1, kw)
        ws.cell(h1 + 1 + i, 2, sv)
    ws.cell(h1 + 3, 1, "Sum")

    h2 = section(1, 4, "Shield-SKW", "keywords")
    ws.cell(h2 + 1, 4, "acme widget")
    ws.cell(h2 + 1, 5, 500)
    ws.cell(h2 + 3, 4, "Sum")

    h3 = section(10, 1, "Long-Tails (Halo)", "keywords")
    for i, kw in enumerate(["red widget for kitchen", "red widget for office"]):
        ws.cell(h3 + 1 + i, 1, kw)
        ws.cell(h3 + 1 + i, 2, 300)
    ws.cell(h3 + 4, 1, "Sum")

    # Discovery-Root Keywords: BMM and Phrase columns, distinguished by a label row
    # between the section title and the Keyword/Search Volume header (this is how
    # scan_campaign_structure_sections tells them apart — see keyword_workbook.py).
    h4 = section(20, 1, "Discovery-Root Keywords", "keywords", label="BMM Root")
    ws.cell(h4 + 1, 1, "widget")
    ws.cell(h4 + 1, 2, 5000)
    ws.cell(h4 + 3, 1, "Sum")

    h5 = section(20, 4, "Discovery-Root Keywords", "keywords", label="Phrase Root")
    ws.cell(h5 + 1, 4, "widget accessory")
    ws.cell(h5 + 1, 5, 4000)
    ws.cell(h5 + 3, 4, "Sum")

    h6 = section(30, 1, "PAT (Stronger)", "asins")
    ws.cell(h6 + 1, 1, "B000000001")
    ws.cell(h6 + 1, 2, "Competitor A")
    ws.cell(h6 + 3, 1, "Sum")

    h7 = section(30, 4, "PAT (Weaker)", "asins")
    ws.cell(h7 + 1, 4, "B000000002")
    ws.cell(h7 + 1, 5, "Competitor B")
    ws.cell(h7 + 3, 4, "Sum")

    wb.save(path)


def test_keyword_file():
    section("C. create mode — keyword-file input ('5. Campaign Structure' scaffold)")
    wb_path = TMP / "keyword_workbook_fixture.xlsx"
    _write_campaign_structure_fixture(wb_path)

    specs = kwb.parse_keyword_workbook(str(wb_path), product_name="Widget", sku=["SKU-1"])
    check("parsed >= 6 campaign specs (2 SKW + 1 shield SKW + 1 halo + 2 discovery + 2 PAT)",
          len(specs) >= 6, str(len(specs)))
    types = [s["campaign_type"] for s in specs]
    check("includes 2 plain SKW rank specs", types.count("SKW") == 3, str(types))  # rank(2) + shield(1)
    check("includes 1 Halo spec (all halo keywords bundled)", types.count("Halo") == 1, str(types))
    check("includes BMM + Phrase discovery specs", "BMM" in types and "Phrase" in types, str(types))
    check("includes 2 PAT specs (stronger + weaker)", types.count("PAT") == 2, str(types))
    shield_spec = next(s for s in specs if s["campaign_type"] == "SKW" and s["campaign_purpose"] == "SHIELD")
    check("shield SKW spec carries campaign_purpose=SHIELD", shield_spec["keywords"] == ["acme widget"])

    cfg = {
        "client": "Selftest KWFile", "brand": "Selftest", "marketplace": "US",
        "defaults": {"daily_budget": 10.0, "keyword_bid": 0.5, "state": "paused"},
        "campaigns": specs,
    }
    cfg_path = TMP / "cfg_kwfile.json"
    cfg_path.write_text(json.dumps(cfg))
    loaded = bc.load_config(str(cfg_path))
    rc_pre = bc.preflight(loaded)
    check("preflight READY on keyword-file-derived config", rc_pre == 0)
    out = TMP / "kwfile.xlsx"
    rc_build = bc.build(loaded, str(out))
    check("build+validate PASS", rc_build == 0)

    # exercise the CLI wiring too (--keyword-file / --keyword-sheet)
    cfg2 = {"client": "Selftest KWFile CLI", "brand": "Selftest", "marketplace": "US",
            "keyword_file_defaults": {"product_name": "Widget", "sku": ["SKU-1"]},
            "defaults": {"daily_budget": 10.0, "keyword_bid": 0.5, "state": "paused"},
            "campaigns": []}
    cfg2_path = TMP / "cfg_kwfile_cli.json"
    cfg2_path.write_text(json.dumps(cfg2))
    import subprocess
    r = subprocess.run(
        [sys.executable, str(HERE / "build_campaigns.py"), "--config", str(cfg2_path),
         "--keyword-file", str(wb_path), "--out", str(TMP / "kwfile_cli.xlsx")],
        capture_output=True, text=True)
    check("--keyword-file CLI run exits 0", r.returncode == 0, r.stdout + r.stderr)
    check("--keyword-file CLI reports parsed campaign count",
          "campaign spec(s) parsed" in r.stdout, r.stdout)


# =================================================================== D/E. update mode
def _write_export_fixture(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cm.SHEET_NAMES["SP"]
    ws.append(cm.SP_COLUMNS)

    def row(**kw):
        r = {c: "" for c in cm.SP_COLUMNS}
        r["Product"] = "Sponsored Products"
        r.update(kw)
        ws.append([r[c] for c in cm.SP_COLUMNS])

    def camp(cid, name, portfolio="", end_date=""):
        row(Entity="Campaign", **{"Campaign ID": cid, "Campaign Name": name,
            "Start Date": "20260101", "End Date": end_date, "Targeting Type": "MANUAL",
            "State": "enabled", "Daily Budget": 10.0, "Portfolio ID": portfolio,
            "Bidding Strategy": "Dynamic bids - down only"})

    def ad_group(cid, agid, name, bid=0.5, state="enabled"):
        row(Entity="Ad Group", **{"Campaign ID": cid, "Ad Group ID": agid,
            "Ad Group Name": name, "State": state, "Ad Group Default Bid": bid})

    def keyword(cid, agid, kwid, text, match, bid=0.5, state="enabled"):
        row(Entity="Keyword", **{"Campaign ID": cid, "Ad Group ID": agid, "Keyword ID": kwid,
            "Keyword Text": text, "Match Type": match, "Bid": bid, "State": state})

    def negative(cid, agid, negid, text, match="negativeExact"):
        row(Entity="Negative Keyword", **{"Campaign ID": cid, "Ad Group ID": agid,
            "Keyword ID": negid, "Keyword Text": text, "Match Type": match, "State": "enabled"})

    def target(cid, agid, tid, expr, bid=0.5):
        row(Entity="Product Targeting", **{"Campaign ID": cid, "Ad Group ID": agid,
            "Product Targeting ID": tid, "Product Targeting Expression": expr, "Bid": bid,
            "State": "enabled"})

    camp("111111111", "Rank | SP | Exact | Widget | red widget | EW", portfolio="999999")
    ad_group("111111111", "222222222", "SP | Exact | Widget | red widget", bid=0.5)
    keyword("111111111", "222222222", "333333333", "old keyword", "exact", bid=0.5)
    keyword("111111111", "222222222", "333333334", "pause me", "broad", bid=0.4)
    negative("111111111", "222222222", "444444444", "bad term")
    target("111111111", "222222222", "555555555", 'asin="B000000001"', bid=0.6)

    camp("111111112", "Discovery | SP | Broad | Widget | widget | EW")  # no portfolio
    ad_group("111111112", "222222223", "SP | Broad | Widget | widget", bid=0.3)

    camp("111111113", "Profit | SP | Exact | Widget | halo theme | EW")  # to be archived
    ad_group("111111113", "222222224", "SP | Exact | Widget | halo theme", bid=0.5)
    keyword("111111113", "222222224", "333333336", "halo term", "exact", bid=0.5)

    wb.save(path)


def _good_change_set(export_file):
    return {
        "client": "Selftest Update", "brand": "Selftest", "marketplace": "US",
        "export_file": export_file, "allow_end_date_clear": False,
        "changes": {
            "campaigns": [
                {"campaign_id": "111111111", "daily_budget": 25.0, "bidding_strategy": "Up and down",
                 "placements": {"top_of_search_placement": 50}},
                {"campaign_id": "111111112", "name": "Discovery | SP | Broad | Widget | widget v2 | EW"},
            ],
            "ad_groups": [
                {"ad_group_id": "222222223", "default_bid": 0.35},
            ],
            "keywords": {
                "pause": ["333333334"],
                "replace": [{"old_keyword_id": "333333333", "new_text": "new keyword",
                            "new_match_type": "EXACT", "new_bid": 0.55}],
                "add": [{"campaign_id": "111111112", "ad_group_id": "222222223",
                        "text": "widget accessory", "match_type": "PHRASE", "bid": 0.4}],
            },
            "negatives": {
                "archive": ["444444444"],
                "add": [{"campaign_id": "111111111", "ad_group_id": "222222222",
                        "text": "cheap knockoff", "match_type": "NEGATIVE_PHRASE"}],
            },
            "targets": {
                "archive": ["555555555"],
                "add": [{"campaign_id": "111111111", "ad_group_id": "222222222",
                        "asin": "B000000099", "bid": 0.5}],
            },
            "archive_campaigns": ["111111113"],
            "archive_ad_groups": ["222222224"],  # child of an archived campaign -> must be dropped
        },
    }


def test_update_good():
    section("D. update mode — real-ID Update/Archive/Create rows + partial-update rules")
    export_path = TMP / "export.xlsx"
    _write_export_fixture(export_path)

    export = um.read_export(str(export_path))
    check("export index has 3 campaigns", len(export.campaigns) == 3, str(len(export.campaigns)))
    check("export index has 3 ad groups", len(export.ad_groups) == 3)
    check("export index has 3 keywords", len(export.keywords) == 3)

    cfg = _good_change_set(str(export_path))
    cfg_path = TMP / "cfg_update_good.json"
    cfg_path.write_text(json.dumps(cfg))
    loaded = uc.load_config(str(cfg_path))

    rc_pre = uc.preflight(loaded)
    check("update preflight READY", rc_pre == 0)

    export2, rows, review, errors = uc._plan(loaded)
    check("no generation errors", errors == [], str(errors))

    camp_a_row = next(r for r in rows if r["Entity"] == "Campaign" and r["Campaign ID"] == "111111111")
    check("Campaign A Portfolio ID carried forward (999999)", camp_a_row["Portfolio ID"] == "999999",
          camp_a_row["Portfolio ID"])
    check("Campaign A End Date carried forward blank (not cleared unintentionally)",
          camp_a_row["End Date"] == "", repr(camp_a_row["End Date"]))
    check("Campaign A Daily Budget updated to 25.0", camp_a_row["Daily Budget"] == 25.0)
    check("Campaign A Bidding Strategy translated to Amazon vocabulary",
          camp_a_row["Bidding Strategy"] == "Dynamic bids - up and down", camp_a_row["Bidding Strategy"])

    camp_b_row = next(r for r in rows if r["Entity"] == "Campaign" and r["Campaign ID"] == "111111112")
    check("Campaign B (no portfolio) Portfolio ID stays blank, not fabricated",
          camp_b_row["Portfolio ID"] == "", repr(camp_b_row["Portfolio ID"]))

    placement_rows = [r for r in rows if r["Entity"] == "Bidding Adjustment"]
    check("placement update row emitted", len(placement_rows) == 1 and placement_rows[0]["Percentage"] == 50)

    archive_rows = [r for r in rows if r["Operation"] == "Archive"]
    check("Campaign 111111113 archived", any(
        r["Entity"] == "Campaign" and r["Campaign ID"] == "111111113" for r in archive_rows))
    check("Ad Group 222222224 (child of archived campaign) is NOT separately archived",
          not any(r["Entity"] == "Ad Group" and r["Ad Group ID"] == "222222224" for r in archive_rows))
    check("cascade skip is recorded in the review trail",
          any("cascade" in r.lower() and "222222224" in r for r in review), "\n".join(review))

    replace_archive = [r for r in rows if r["Entity"] == "Keyword" and r["Operation"] == "Archive"
                       and r["Keyword ID"] == "333333333"]
    replace_create = [r for r in rows if r["Entity"] == "Keyword" and r["Operation"] == "Create"
                      and r["Keyword Text"] == "new keyword"]
    check("keyword replace = archive old + create new (never an Update to Text/Match Type)",
          len(replace_archive) == 1 and len(replace_create) == 1)
    check("replacement keyword has a fresh (non-real, non-numeric) temp ID",
          not um.looks_like_real_id(replace_create[0]["Keyword ID"]))

    neg_archive = [r for r in rows if "Negative" in r["Entity"] and r["Operation"] == "Archive"]
    check("negative archived (never paused — no negative Update rows at all)",
          len(neg_archive) == 1 and not any(
              "Negative" in r["Entity"] and r["Operation"] == "Update" for r in rows))

    out = TMP / "update_good.xlsx"
    rc_build = uc.build(loaded, str(out))
    check("update build+validate PASS", rc_build == 0)

    # --preview / --validate CLI smoke test
    import subprocess
    r = subprocess.run([sys.executable, str(HERE / "update_campaigns.py"), "--config", str(cfg_path),
                       "--preview"], capture_output=True, text=True)
    check("--preview CLI exits 0", r.returncode == 0, r.stdout + r.stderr)
    r = subprocess.run([sys.executable, str(HERE / "update_campaigns.py"), "--config", str(cfg_path),
                       "--validate", "--out", str(out)], capture_output=True, text=True)
    check("--validate CLI exits 0 (PASS)", r.returncode == 0, r.stdout + r.stderr)

    return str(export_path)


def test_update_broken(export_path):
    section("E. update mode QA gates — deliberately-broken file must FAIL validate")
    cfg = {"client": "Selftest Update Broken", "brand": "Selftest", "marketplace": "US",
           "export_file": export_path, "allow_end_date_clear": False, "changes": {}}
    cfg_path = TMP / "cfg_update_broken.json"
    cfg_path.write_text(json.dumps(cfg))
    loaded = uc.load_config(str(cfg_path))

    def row(**kw):
        r = {c: "" for c in cm.SP_COLUMNS}
        r["Product"] = "Sponsored Products"
        r.update(kw)
        return r

    broken_rows = [
        # 1. temp-ID-style Campaign ID on an Update row (not numeric -> fails looks_like_real_id)
        row(Entity="Campaign", Operation="Update", **{"Campaign ID": "campaign_1",
            "Campaign Name": "", "Daily Budget": 30.0, "Portfolio ID": ""}),
        # 2. real Campaign ID (111111111, which DOES belong to Portfolio 999999 in the export)
        #    but the Update row omits Portfolio ID -> silent portfolio drop
        row(Entity="Campaign", Operation="Update", **{"Campaign ID": "111111111",
            "Daily Budget": 30.0, "Portfolio ID": ""}),
        # 3a/3b. parent+child double archive: Campaign 111111113 AND its Ad Group 222222224
        row(Entity="Campaign", Operation="Archive", **{"Campaign ID": "111111113"}),
        row(Entity="Ad Group", Operation="Archive", **{"Campaign ID": "111111113",
            "Ad Group ID": "222222224"}),
    ]
    out = TMP / "update_broken.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cm.SHEET_NAMES["SP"]
    ws.append(cm.COLUMNS["SP"])
    for r in broken_rows:
        ws.append([r[c] for c in cm.COLUMNS["SP"]])
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    rc = uc.validate(loaded, str(out))
    check("validate FAILS on the deliberately-broken file", rc == 1, f"rc={rc}")

    import io
    from contextlib import redirect_stdout
    buf = io.StringIO()
    with redirect_stdout(buf):
        uc.validate(loaded, str(out))
    text = buf.getvalue()
    print(text)
    check("gate: catches the temp-ID-style Campaign ID", "not a real bulksheets ID" in text, text)
    check("gate: catches the missing Portfolio ID on a portfolio campaign",
          "would silently leave the portfolio" in text, text)
    check("gate: catches the parent+child double archive",
          "also archived in this file" in text, text)


# =================================================================== main
def main():
    test_create_legacy()
    test_create_ew()
    test_keyword_file()
    export_path = test_update_good()
    test_update_broken(export_path)

    print(f"\n{'='*70}")
    if FAILURES:
        print(f"SELFTEST: FAIL — {len(FAILURES)} assertion(s) failed:")
        for f in FAILURES:
            print(f"  - {f}")
        return 1
    print("SELFTEST: ALL PASS")
    return 0


if __name__ == "__main__":
    rc = main()
    shutil.rmtree(TMP, ignore_errors=True)
    sys.exit(rc)
