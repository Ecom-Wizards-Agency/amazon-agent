#!/usr/bin/env python3
"""Update-mode pure logic: read a bulksheets export (current live account state) + a
change-set -> real-ID Update/Archive/Create rows.

See references/bulksheets-2.0-reference.md section 4 ("Update semantics... the
critical rules") for the rules enforced here — in short:

  - Update/Archive rows need the REAL, bulksheets-download ID, never a temp ID.
  - Blank fields on an Update row = "leave unchanged" — EXCEPT End Date, where blank
    CLEARS an existing end date. This module carries the export's existing End Date
    forward by default on every Campaign Update row, and only touches it when the
    change-set explicitly asks for a new one or opts into clearing it.
  - Portfolio ID must be re-included on every Campaign Update row (never left blank),
    or the campaign silently drops out of its portfolio.
  - Keyword Text / Match Type / Product Targeting Expression are immutable — a
    "change" is always Archive-old + Create-new (fresh temp ID for the new entity;
    the parent Campaign/Ad Group ID is the existing real one, not a temp one).
  - Archiving a Campaign/Ad Group cascades to its children — this module drops any
    explicit child Archive request whose parent is archived in the same change-set,
    logging the skip to the review trail instead of emitting a redundant row.
  - Negatives can only be archived, never paused.

This module only reads a .xlsx (never writes) and returns row dicts + a plain-English
review trail; update_campaigns.py does the file I/O and CLI. Nothing here touches a
live account — only the operator's later manual bulk-upload does.
"""
from __future__ import annotations

import re

from campaign_model import AMAZON_BIDDING, AMAZON_MATCH, AMAZON_NEG_MATCH, PLACEMENT_LABELS, SHEET_NAMES, SP_COLUMNS, STATES

REAL_ID_RE = re.compile(r"^\d+$")


def looks_like_real_id(value) -> bool:
    """Amazon bulksheets IDs are purely numeric. Temp IDs (e.g. 'campaign_1', this
    toolkit's own create-mode counters reused by mistake) and console-displayed IDs
    with any non-digit character fail this check."""
    return bool(REAL_ID_RE.match(str(value or "").strip()))


def _s(v):
    return "" if v is None else str(v).strip()


def _money(v):
    return round(float(v), 2)


def _yyyymmdd(v):
    v = _s(v)
    if not v:
        return ""
    if re.match(r"^\d{8}$", v):
        return v
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", v)
    return "".join(m.groups()) if m else v


# --------------------------------------------------------------- reading the export
class ExportIndex:
    """Indexed view of a bulksheets download's 'Sponsored Products Campaigns' tab."""

    def __init__(self):
        self.campaigns = {}    # Campaign ID -> row dict
        self.ad_groups = {}    # Ad Group ID -> row dict
        self.keywords = {}     # Keyword ID -> row dict (Entity == Keyword)
        self.negatives = {}    # Keyword ID -> row dict (Negative Keyword / Campaign Negative Keyword)
        self.targets = {}      # Product Targeting ID -> row dict (Entity == Product Targeting)
        self.neg_targets = {}  # Product Targeting ID -> row dict (Negative Product Targeting)
        self.raw_rows = []

    def campaign_portfolio(self, campaign_id):
        c = self.campaigns.get(_s(campaign_id))
        return _s(c.get("Portfolio ID")) if c else ""

    def campaign_end_date(self, campaign_id):
        c = self.campaigns.get(_s(campaign_id))
        return _s(c.get("End Date")) if c else ""

    def ad_group_campaign(self, ad_group_id):
        ag = self.ad_groups.get(_s(ad_group_id))
        return _s(ag.get("Campaign ID")) if ag else ""

    def keyword_parents(self, keyword_id):
        k = self.keywords.get(_s(keyword_id))
        return (_s(k.get("Campaign ID")), _s(k.get("Ad Group ID"))) if k else ("", "")

    def negative_parents(self, negative_id):
        n = self.negatives.get(_s(negative_id))
        return (_s(n.get("Campaign ID")), _s(n.get("Ad Group ID"))) if n else ("", "")

    def target_parents(self, target_id):
        t = self.targets.get(_s(target_id))
        return (_s(t.get("Campaign ID")), _s(t.get("Ad Group ID"))) if t else ("", "")


def read_export(path, sheet_name=None):
    """Read a downloaded bulksheets .xlsx -> ExportIndex. Tolerant of the tab name
    (falls back to any sheet containing 'sponsored products')."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    name = sheet_name or SHEET_NAMES["SP"]
    if name not in wb.sheetnames:
        name = next((s for s in wb.sheetnames if "sponsored products" in s.lower()), None)
    if not name:
        raise ValueError(f"no '{SHEET_NAMES['SP']}' tab found in {path} (sheets: {wb.sheetnames})")
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"'{name}' tab is empty in {path}")
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    idx = ExportIndex()
    for raw in rows[1:]:
        row = {header[i]: (raw[i] if i < len(raw) and raw[i] is not None else "")
               for i in range(len(header))}
        if not any(str(v).strip() for v in row.values()):
            continue
        idx.raw_rows.append(row)
        ent = _s(row.get("Entity"))
        if ent == "Campaign":
            idx.campaigns[_s(row.get("Campaign ID"))] = row
        elif ent == "Ad Group":
            idx.ad_groups[_s(row.get("Ad Group ID"))] = row
        elif ent == "Keyword":
            idx.keywords[_s(row.get("Keyword ID"))] = row
        elif ent in ("Negative Keyword", "Campaign Negative Keyword"):
            idx.negatives[_s(row.get("Keyword ID"))] = row
        elif ent == "Product Targeting":
            idx.targets[_s(row.get("Product Targeting ID"))] = row
        elif ent == "Negative Product Targeting":
            idx.neg_targets[_s(row.get("Product Targeting ID"))] = row
    return idx


# --------------------------------------------------------------- row building
def _empty_row(operation):
    row = {c: "" for c in SP_COLUMNS}
    row["Product"] = "Sponsored Products"
    row["Operation"] = operation
    return row


def build_change_set_rows(changes, export, *, allow_end_date_clear=False):
    """changes: the 'changes' dict from a config.UPDATE.TEMPLATE.json-shaped config.
    Returns (rows, review, errors). `review` is a flat list of plain-English lines
    (both applied changes and SKIPPED no-op/cascade lines) suitable for _REVIEW.md.
    `errors` are hard failures (unresolved ID, disallowed end-date clear, bad enum) —
    if non-empty the caller should refuse to build."""
    rows, review, errors = [], [], []
    counter = iter(range(1, 10 ** 9))

    def next_temp_id():
        return f"new_{next(counter)}"

    archived_campaigns = {_s(c) for c in changes.get("archive_campaigns", [])}
    for cid in archived_campaigns:
        if cid not in export.campaigns:
            errors.append(f"archive_campaigns: Campaign ID {cid!r} not found in the export")

    archived_ad_groups = set()
    for ag_id in {_s(a) for a in changes.get("archive_ad_groups", [])}:
        if ag_id not in export.ad_groups:
            errors.append(f"archive_ad_groups: Ad Group ID {ag_id!r} not found in the export")
            continue
        camp = export.ad_group_campaign(ag_id)
        if camp in archived_campaigns:
            review.append(f"SKIPPED (parent+child archive): Ad Group {ag_id} — Campaign {camp} "
                          f"is already archived in this file; archiving a campaign cascades to "
                          f"its ad groups, so a separate Archive row for the child is redundant")
            continue
        archived_ad_groups.add(ag_id)

    def parent_archived(campaign_id, ad_group_id=""):
        if campaign_id in archived_campaigns:
            return campaign_id, "Campaign"
        if ad_group_id and ad_group_id in archived_ad_groups:
            return ad_group_id, "Ad Group"
        return None, None

    # ---------------------------------------------------- campaign updates
    for ch in changes.get("campaigns", []):
        cid = _s(ch.get("campaign_id"))
        if not looks_like_real_id(cid):
            errors.append(f"changes.campaigns: campaign_id {ch.get('campaign_id')!r} is not a "
                          f"real bulksheets ID — Update/Archive rows must use IDs sourced from "
                          f"a bulksheets download, never a temp ID")
            continue
        if cid not in export.campaigns:
            errors.append(f"changes.campaigns: campaign_id {cid!r} not found in the loaded export")
            continue
        if cid in archived_campaigns:
            review.append(f"SKIPPED: Campaign {cid} update — also archived in this file; the "
                          f"archive supersedes the update")
            continue
        export_row = export.campaigns[cid]
        row = _empty_row("Update")
        row["Entity"] = "Campaign"
        row["Campaign ID"] = cid
        row["Campaign Name"] = export_row.get("Campaign Name", "")

        changed = {}
        if ch.get("name"):
            row["Campaign Name"] = ch["name"]
            changed["Campaign Name"] = (export_row.get("Campaign Name", ""), ch["name"])
        if ch.get("daily_budget") is not None:
            new_budget = _money(ch["daily_budget"])
            if _s(export_row.get("Daily Budget")) != _s(new_budget):
                row["Daily Budget"] = new_budget
                changed["Daily Budget"] = (export_row.get("Daily Budget", ""), new_budget)
        if ch.get("bidding_strategy"):
            new_bidding = AMAZON_BIDDING.get(ch["bidding_strategy"], ch["bidding_strategy"])
            if _s(export_row.get("Bidding Strategy")) != _s(new_bidding):
                row["Bidding Strategy"] = new_bidding
                changed["Bidding Strategy"] = (export_row.get("Bidding Strategy", ""), new_bidding)
        if ch.get("state"):
            if ch["state"] not in STATES:
                errors.append(f"changes.campaigns/{cid}: state must be one of {'/'.join(STATES)}")
            elif _s(export_row.get("State")) != ch["state"]:
                row["State"] = ch["state"]
                changed["State"] = (export_row.get("State", ""), ch["state"])

        # Portfolio ID — ALWAYS re-included (reference 4.3): omitting it on an Update
        # row silently drops the campaign from its portfolio, regardless of the
        # general blank-means-unchanged rule.
        row["Portfolio ID"] = export.campaign_portfolio(cid)

        # End Date — the one field where blank means CLEAR, not "unchanged" (4.2).
        # Default: carry the export's existing End Date forward untouched. Only
        # change it if the change-set gives a new end_date, or explicitly opts into
        # clearing it (clear_end_date AND the file-level allow_end_date_clear both true).
        existing_end_date = export.campaign_end_date(cid)
        if ch.get("clear_end_date"):
            if not allow_end_date_clear:
                errors.append(f"changes.campaigns/{cid}: clear_end_date is set but the "
                              f"change-set's top-level allow_end_date_clear is false — blank "
                              f"End Date clears an existing end date, so clearing must be "
                              f"opted into explicitly")
            else:
                row["End Date"] = ""
                changed["End Date"] = (existing_end_date or "(none)", "(cleared — runs indefinitely)")
        elif ch.get("end_date"):
            new_end = _yyyymmdd(ch["end_date"])
            row["End Date"] = new_end
            if new_end != existing_end_date:
                changed["End Date"] = (existing_end_date or "(none)", new_end)
        else:
            row["End Date"] = existing_end_date  # carry forward — never accidentally clear

        placement_rows = []
        for key, label in PLACEMENT_LABELS.items():
            pct = (ch.get("placements") or {}).get(key)
            if pct is not None:
                prow = _empty_row("Update")
                prow["Entity"] = "Bidding Adjustment"
                prow["Campaign ID"] = cid
                prow["Campaign Name"] = export_row.get("Campaign Name", "")
                prow["Placement"] = label
                prow["Percentage"] = int(pct)
                placement_rows.append(prow)
                changed[f"Placement {label}"] = ("?", int(pct))

        if not changed and not placement_rows:
            review.append(f"SKIPPED (no-op): Campaign {cid} ({export_row.get('Campaign Name','')}) "
                          f"— no fields differ from the export")
            continue

        rows.append(row)
        rows.extend(placement_rows)
        for field, (old, new) in changed.items():
            review.append(f"UPDATE Campaign {cid} ({export_row.get('Campaign Name','')}): "
                          f"{field} '{old}' -> '{new}'")

    # ---------------------------------------------------- campaign archives
    for cid in archived_campaigns:
        if cid not in export.campaigns:
            continue
        row = _empty_row("Archive")
        row["Entity"] = "Campaign"
        row["Campaign ID"] = cid
        row["Campaign Name"] = export.campaigns[cid].get("Campaign Name", "")
        rows.append(row)
        review.append(f"ARCHIVE Campaign {cid} ({row['Campaign Name']}) — cascades to all of "
                      f"its Ad Groups, Keywords, Product Targeting, and Negatives")

    # ---------------------------------------------------- ad group updates
    for ch in changes.get("ad_groups", []):
        agid = _s(ch.get("ad_group_id"))
        if not looks_like_real_id(agid):
            errors.append(f"changes.ad_groups: ad_group_id {ch.get('ad_group_id')!r} is not a "
                          f"real ID")
            continue
        if agid not in export.ad_groups:
            errors.append(f"changes.ad_groups: ad_group_id {agid!r} not found in the loaded export")
            continue
        campaign_id = export.ad_group_campaign(agid)
        parent, parent_kind = parent_archived(campaign_id, agid)
        if parent:
            review.append(f"SKIPPED (parent archived): Ad Group {agid} update — its "
                          f"{parent_kind} {parent} is archived in this file")
            continue
        export_row = export.ad_groups[agid]
        row = _empty_row("Update")
        row["Entity"] = "Ad Group"
        row["Campaign ID"] = campaign_id
        row["Ad Group ID"] = agid
        row["Ad Group Name"] = export_row.get("Ad Group Name", "")

        changed = {}
        if ch.get("name"):
            row["Ad Group Name"] = ch["name"]
            changed["Ad Group Name"] = (export_row.get("Ad Group Name", ""), ch["name"])
        if ch.get("default_bid") is not None:
            new_bid = _money(ch["default_bid"])
            if _s(export_row.get("Ad Group Default Bid")) != _s(new_bid):
                row["Ad Group Default Bid"] = new_bid
                changed["Ad Group Default Bid"] = (export_row.get("Ad Group Default Bid", ""), new_bid)
        if ch.get("state"):
            if ch["state"] not in STATES:
                errors.append(f"changes.ad_groups/{agid}: state must be one of {'/'.join(STATES)}")
            elif _s(export_row.get("State")) != ch["state"]:
                row["State"] = ch["state"]
                changed["State"] = (export_row.get("State", ""), ch["state"])

        if not changed:
            review.append(f"SKIPPED (no-op): Ad Group {agid} ({export_row.get('Ad Group Name','')}) "
                          f"— no fields differ from the export")
            continue
        rows.append(row)
        for field, (old, new) in changed.items():
            review.append(f"UPDATE Ad Group {agid} ({export_row.get('Ad Group Name','')}): "
                          f"{field} '{old}' -> '{new}'")

    # ---------------------------------------------------- ad group archives
    for agid in archived_ad_groups:
        row = _empty_row("Archive")
        row["Entity"] = "Ad Group"
        row["Campaign ID"] = export.ad_group_campaign(agid)
        row["Ad Group ID"] = agid
        row["Ad Group Name"] = export.ad_groups[agid].get("Ad Group Name", "")
        rows.append(row)
        review.append(f"ARCHIVE Ad Group {agid} ({row['Ad Group Name']}) — cascades to its "
                      f"Keywords, Product Targeting, and Negatives")

    # ---------------------------------------------------- keywords
    kw = changes.get("keywords", {})
    for action, target_state in (("pause", "paused"), ("enable", "enabled")):
        for kwid in kw.get(action, []):
            kwid = _s(kwid)
            if kwid not in export.keywords:
                errors.append(f"keywords.{action}: Keyword ID {kwid!r} not found in the export")
                continue
            camp, ag = export.keyword_parents(kwid)
            parent, parent_kind = parent_archived(camp, ag)
            if parent:
                review.append(f"SKIPPED (parent archived): Keyword {kwid} {action} — its "
                              f"{parent_kind} {parent} is archived in this file")
                continue
            exrow = export.keywords[kwid]
            if _s(exrow.get("State")) == target_state:
                review.append(f"SKIPPED (no-op): Keyword {kwid} ({exrow.get('Keyword Text','')}) "
                              f"— already {target_state}")
                continue
            row = _empty_row("Update")
            row["Entity"] = "Keyword"
            row["Campaign ID"], row["Ad Group ID"], row["Keyword ID"] = camp, ag, kwid
            row["State"] = target_state
            rows.append(row)
            review.append(f"UPDATE Keyword {kwid} ({exrow.get('Keyword Text','')}): State "
                          f"'{exrow.get('State','')}' -> '{target_state}'")

    for kwid in {_s(k) for k in kw.get("archive", [])}:
        if kwid not in export.keywords:
            errors.append(f"keywords.archive: Keyword ID {kwid!r} not found in the export")
            continue
        camp, ag = export.keyword_parents(kwid)
        parent, parent_kind = parent_archived(camp, ag)
        if parent:
            review.append(f"SKIPPED (parent archived): Keyword {kwid} archive — its "
                          f"{parent_kind} {parent} is already archived in this file")
            continue
        exrow = export.keywords[kwid]
        row = _empty_row("Archive")
        row["Entity"] = "Keyword"
        row["Campaign ID"], row["Ad Group ID"], row["Keyword ID"] = camp, ag, kwid
        rows.append(row)
        review.append(f"ARCHIVE Keyword {kwid} ({exrow.get('Keyword Text','')}, "
                      f"{exrow.get('Match Type','')})")

    for rep in kw.get("replace", []):
        old_id = _s(rep.get("old_keyword_id"))
        if old_id not in export.keywords:
            errors.append(f"keywords.replace: old_keyword_id {rep.get('old_keyword_id')!r} not "
                          f"found in the export")
            continue
        camp, ag = export.keyword_parents(old_id)
        exrow = export.keywords[old_id]
        parent, parent_kind = parent_archived(camp, ag)
        if not parent:
            arow = _empty_row("Archive")
            arow["Entity"] = "Keyword"
            arow["Campaign ID"], arow["Ad Group ID"], arow["Keyword ID"] = camp, ag, old_id
            rows.append(arow)
        new_text = rep.get("new_text") or exrow.get("Keyword Text", "")
        new_match = AMAZON_MATCH.get(rep.get("new_match_type"), rep.get("new_match_type")) \
            or exrow.get("Match Type", "")
        new_bid = _money(rep["new_bid"]) if rep.get("new_bid") is not None else exrow.get("Bid", "")
        crow = _empty_row("Create")
        crow["Entity"] = "Keyword"
        crow["Campaign ID"], crow["Ad Group ID"] = camp, ag
        crow["Keyword ID"] = next_temp_id()
        crow["State"] = rep.get("state") or exrow.get("State") or "enabled"
        crow["Bid"] = new_bid
        crow["Keyword Text"] = new_text
        crow["Match Type"] = new_match
        rows.append(crow)
        note = f" [archive of old ID skipped: parent {parent_kind} {parent} already archived]" if parent else ""
        review.append(f"REPLACE Keyword {old_id} ({exrow.get('Keyword Text','')}, "
                      f"{exrow.get('Match Type','')}) -> new Keyword '{new_text}' ({new_match}) "
                      f"— Keyword Text/Match Type are immutable, so this is Archive-old + "
                      f"Create-new, never an Update{note}")

    for add in kw.get("add", []):
        camp, ag = _s(add.get("campaign_id")), _s(add.get("ad_group_id"))
        if not looks_like_real_id(camp) or camp not in export.campaigns:
            errors.append(f"keywords.add: campaign_id {add.get('campaign_id')!r} not found in "
                          f"the export — new keywords attach to an EXISTING (real-ID) campaign")
            continue
        if not looks_like_real_id(ag) or ag not in export.ad_groups:
            errors.append(f"keywords.add: ad_group_id {add.get('ad_group_id')!r} not found in "
                          f"the export")
            continue
        parent, parent_kind = parent_archived(camp, ag)
        if parent:
            review.append(f"SKIPPED (parent archived): new keyword {add.get('text')!r} — "
                          f"{parent_kind} {parent} is archived in this file")
            continue
        row = _empty_row("Create")
        row["Entity"] = "Keyword"
        row["Campaign ID"], row["Ad Group ID"] = camp, ag
        row["Keyword ID"] = next_temp_id()
        row["State"] = add.get("state") or "enabled"
        if add.get("bid") is not None:
            row["Bid"] = _money(add["bid"])
        row["Keyword Text"] = add.get("text", "")
        row["Match Type"] = AMAZON_MATCH.get(add.get("match_type"), add.get("match_type", ""))
        rows.append(row)
        review.append(f"ADD Keyword '{add.get('text','')}' ({row['Match Type']}) to Ad Group "
                      f"{ag} (Campaign {camp})")

    # ---------------------------------------------------- negatives (archive-only, 4.7)
    neg = changes.get("negatives", {})
    for negid in {_s(n) for n in neg.get("archive", [])}:
        if negid not in export.negatives:
            errors.append(f"negatives.archive: Negative Keyword ID {negid!r} not found in the export")
            continue
        camp, ag = export.negative_parents(negid)
        parent, parent_kind = parent_archived(camp, ag)
        if parent:
            review.append(f"SKIPPED (parent archived): Negative {negid} archive — its "
                          f"{parent_kind} {parent} is already archived in this file")
            continue
        exrow = export.negatives[negid]
        row = _empty_row("Archive")
        row["Entity"] = exrow.get("Entity", "Negative Keyword")
        row["Campaign ID"] = camp
        if ag:
            row["Ad Group ID"] = ag
        row["Keyword ID"] = negid
        rows.append(row)
        review.append(f"ARCHIVE Negative {negid} ({exrow.get('Keyword Text','')}) — negatives "
                      f"can only be archived, never paused (reference 4.7)")

    for add in neg.get("add", []):
        camp, ag = _s(add.get("campaign_id")), _s(add.get("ad_group_id"))
        level = add.get("level") or ("campaign" if not ag else "ad_group")
        if not looks_like_real_id(camp) or camp not in export.campaigns:
            errors.append(f"negatives.add: campaign_id {add.get('campaign_id')!r} not found in the export")
            continue
        if level == "ad_group" and (not looks_like_real_id(ag) or ag not in export.ad_groups):
            errors.append(f"negatives.add: ad_group_id {add.get('ad_group_id')!r} not found in the export")
            continue
        parent, parent_kind = parent_archived(camp, ag if level == "ad_group" else "")
        if parent:
            review.append(f"SKIPPED (parent archived): new negative {add.get('text')!r} — "
                          f"{parent_kind} {parent} is archived in this file")
            continue
        row = _empty_row("Create")
        row["Entity"] = "Campaign Negative Keyword" if level == "campaign" else "Negative Keyword"
        row["Campaign ID"] = camp
        if level == "ad_group":
            row["Ad Group ID"] = ag
        row["Keyword ID"] = next_temp_id()
        row["State"] = "enabled"
        row["Keyword Text"] = add.get("text", "")
        row["Match Type"] = AMAZON_NEG_MATCH.get(add.get("match_type"), add.get("match_type", ""))
        rows.append(row)
        review.append(f"ADD Negative '{add.get('text','')}' ({row['Match Type']}) at {level} "
                      f"level (Campaign {camp}" + (f", Ad Group {ag}" if level == "ad_group" else "") + ")")

    # ---------------------------------------------------- product targeting (PAT)
    tgt = changes.get("targets", {})
    for tid in {_s(t) for t in tgt.get("archive", [])}:
        if tid not in export.targets:
            errors.append(f"targets.archive: Product Targeting ID {tid!r} not found in the export")
            continue
        camp, ag = export.target_parents(tid)
        parent, parent_kind = parent_archived(camp, ag)
        if parent:
            review.append(f"SKIPPED (parent archived): Target {tid} archive — its {parent_kind} "
                          f"{parent} is already archived in this file")
            continue
        exrow = export.targets[tid]
        row = _empty_row("Archive")
        row["Entity"] = "Product Targeting"
        row["Campaign ID"], row["Ad Group ID"] = camp, ag
        row["Product Targeting ID"] = tid
        rows.append(row)
        review.append(f"ARCHIVE Product Targeting {tid} ({exrow.get('Product Targeting Expression','')})")

    for add in tgt.get("add", []):
        camp, ag = _s(add.get("campaign_id")), _s(add.get("ad_group_id"))
        if not looks_like_real_id(camp) or camp not in export.campaigns:
            errors.append(f"targets.add: campaign_id {add.get('campaign_id')!r} not found in the export")
            continue
        if not looks_like_real_id(ag) or ag not in export.ad_groups:
            errors.append(f"targets.add: ad_group_id {add.get('ad_group_id')!r} not found in the export")
            continue
        parent, parent_kind = parent_archived(camp, ag)
        if parent:
            review.append(f"SKIPPED (parent archived): new target {add.get('asin')!r} — "
                          f"{parent_kind} {parent} is archived in this file")
            continue
        expr = (f'asin-expanded="{_s(add.get("asin")).upper()}"' if add.get("expanded")
                else f'asin="{_s(add.get("asin")).upper()}"')
        row = _empty_row("Create")
        row["Entity"] = "Product Targeting"
        row["Campaign ID"], row["Ad Group ID"] = camp, ag
        row["Product Targeting ID"] = next_temp_id()
        row["State"] = add.get("state") or "enabled"
        if add.get("bid") is not None:
            row["Bid"] = _money(add["bid"])
        row["Product Targeting Expression"] = expr
        rows.append(row)
        review.append(f"ADD Product Targeting {expr} to Ad Group {ag} (Campaign {camp})")

    return rows, review, errors
