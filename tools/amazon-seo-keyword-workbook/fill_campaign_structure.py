#!/usr/bin/env python3
"""Fill the keyword workbook's '5. Campaign Structure' tab — VISUAL PLAN ONLY.

Hybrid pipeline (no bulk files are ever emitted by this tool):

  1. --extract candidates.json   mechanical extraction from the built workbook
  2. (agent step, no script)     the AI agent classifies candidates into buckets
                                 using _local/ads-strategy/strategy.md
  3. --apply classification.json validate + write keywords/ASINs into the
                                 scaffold and append a Proposed Campaign Names
                                 block (copy-paste material for the bulk-creator
                                 webapp; uploading anything stays a separate
                                 operator action)

Strategy thresholds and campaign naming live LOCAL-ONLY in
_local/ads-strategy/strategy.json (copy of ads-strategy.TEMPLATE.json); this
script refuses to run while the file is missing or still has <placeholders>.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import time
from typing import Any

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "amazon-campaign-builder"))

import openpyxl  # noqa: E402
import build_keyword_workbook as bkw  # noqa: E402
import campaign_model  # noqa: E402

norm, to_int, to_float = bkw.norm, bkw.to_int, bkw.to_float

SCHEMA_CANDIDATES = "amazon-agent.campaign-candidates.v1"
SCHEMA_CLASSIFICATION = "amazon-agent.campaign-classification.v1"
SCHEMA_STRATEGY = "amazon-agent.ads-strategy.v1"

CAMPAIGN_TAB = "5. Campaign Structure"

# scaffold section anchors -> (bucket, kind). Titles normalized via _norm_title.
SECTION_ANCHORS = {
    "rank-skw": ("rank_skw", "keywords"),
    "shield-skw": ("shield_skw", "keywords"),
    "long-tails (halo)": ("halo", "keywords"),
    "discovery-root keywords": ("discovery", "keywords"),
    "shield discovery-brand keywords": ("shield_discovery", "keywords"),
    "pat (stronger)": ("pat_stronger", "asins"),
    "pat (weaker)": ("pat_weaker", "asins"),
}
EXPECTED_BUCKETS = {"rank_skw", "shield_skw", "halo", "discovery_bmm", "discovery_phrase",
                    "shield_discovery_bmm", "shield_discovery_phrase", "pat_stronger", "pat_weaker"}
KEYWORD_BUCKETS = {"rank_skw", "shield_skw", "halo", "discovery_bmm", "discovery_phrase",
                   "shield_discovery_bmm", "shield_discovery_phrase"}
DISCOVERY_BUCKETS = {"discovery_bmm", "discovery_phrase", "shield_discovery_bmm", "shield_discovery_phrase"}
ASIN_RE = re.compile(r"^B0[A-Z0-9]{8}$")


def _norm_title(s: Any) -> str:
    """Normalize section titles: lowercase, collapse whitespace, and strip
    spaces around hyphens so 'Shield Discovery- Brand Keywords' and
    'Long-Tails (Halo)' both match their canonical anchors."""
    return re.sub(r"\s*-\s*", "-", norm(s))


def die(msg: str, code: int = 2) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(code)


# --------------------------------------------------------------- strategy
def load_strategy(path: str) -> dict:
    if not os.path.exists(path):
        die(f"strategy file missing: {path}\n"
            "  Copy tools/amazon-seo-keyword-workbook/ads-strategy.TEMPLATE.json to "
            "_local/ads-strategy/strategy.json and fill it in.\n"
            "  Claude: refresh values from the Notion strategy playbooks. "
            "Codex: ask the operator — do not guess thresholds.")
    raw = open(path, encoding="utf-8").read()
    if "<" in re.sub(r'"_comment[^"]*":\s*"[^"]*"', "", raw):
        die(f"strategy file still contains <placeholders>: {path} — fill in real values first.")
    strat = json.loads(raw)
    if strat.get("schema") != SCHEMA_STRATEGY:
        die(f"unexpected strategy schema in {path}: {strat.get('schema')!r}")
    return strat


# --------------------------------------------------------------- scaffold scan
def scan_scaffold(ws) -> dict:
    """Locate every campaign slot on the Campaign Structure sheet.

    Returns {"slots": [slot...], "last_sum_row": int}. Each slot:
      {bucket, label, kw_col, sv_col, header_row, data_start, data_end, capacity}
    Aborts if the sheet doesn't contain exactly the expected section kinds.
    """
    grid: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid[(cell.row, cell.column)] = cell.value

    anchors = []  # (row, col, bucket, kind)
    for (r, c), v in grid.items():
        key = _norm_title(v)
        if key in SECTION_ANCHORS:
            bucket, kind = SECTION_ANCHORS[key]
            anchors.append((r, c, bucket, kind))
    if not anchors:
        die(f"no section anchors found on '{ws.title}' — template scaffold changed?")

    pair_headers = {"keywords": ("keyword", "search volume"), "asins": ("asins", "brand")}
    slots: list[dict] = []
    last_sum_row = 0

    # group anchors by row so sections sharing a row (Rank/Shield) split by column
    by_row: dict[int, list] = {}
    for a in anchors:
        by_row.setdefault(a[0], []).append(a)

    for row_anchors in by_row.values():
        row_anchors.sort(key=lambda a: a[1])
        for i, (ar, ac, bucket, kind) in enumerate(row_anchors):
            col_lo = ac
            col_hi = row_anchors[i + 1][1] - 1 if i + 1 < len(row_anchors) else ws.max_column
            first, second = pair_headers[kind]
            # find the pair-header row within the next 3 rows
            header_row = None
            for hr in range(ar + 1, ar + 4):
                if any(norm(grid.get((hr, c))) == first for c in range(col_lo, col_hi + 1)):
                    header_row = hr
                    break
            if header_row is None:
                die(f"section '{bucket}' at row {ar}: no '{first}' header row found within 3 rows")
            for c in range(col_lo, col_hi + 1):
                if norm(grid.get((header_row, c))) != first:
                    continue
                sv_col = next((c2 for c2 in range(c + 1, min(c + 3, col_hi + 1))
                               if norm(grid.get((header_row, c2))) == second), None)
                if sv_col is None:
                    continue
                # label: nearest non-empty cell above the header in this column
                label = ""
                for lr in range(header_row - 1, ar, -1):
                    if grid.get((lr, c)) is not None:
                        label = str(grid[(lr, c)]).strip()
                        break
                # sum row: scan down this kw column
                sum_row = next((sr for sr in range(header_row + 1, header_row + 40)
                                if norm(grid.get((sr, c))) == "sum"), None)
                if sum_row is None:
                    die(f"section '{bucket}' column {c}: no Sum row found below header row {header_row}")
                slot_bucket = bucket
                if bucket in ("discovery", "shield_discovery"):
                    slot_bucket = f"{bucket}_bmm" if "bmm" in norm(label) else f"{bucket}_phrase"
                slots.append({
                    "bucket": slot_bucket, "label": label, "kw_col": c, "sv_col": sv_col,
                    "header_row": header_row, "data_start": header_row + 1,
                    "data_end": sum_row - 1, "capacity": sum_row - 1 - header_row,
                })
                last_sum_row = max(last_sum_row, sum_row)

    found = {s["bucket"] for s in slots}
    missing = EXPECTED_BUCKETS - found
    if missing:
        die(f"scaffold scan incomplete — missing buckets {sorted(missing)}. "
            "Template scaffold changed? Update fill_campaign_structure.py.")
    return {"slots": slots, "last_sum_row": last_sum_row}


# --------------------------------------------------------------- workbook readers
def _sheet_rows(wb, name: str) -> list[list[Any]]:
    return [list(r) for r in wb[name].iter_rows(values_only=True)] if name in wb.sheetnames else []


def _resolve_tab(wb, *patterns: str) -> str | None:
    for s in wb.sheetnames:
        sl = s.lower()
        if any(all(tok in sl for tok in p.split()) for p in patterns):
            return s
    return None


def read_roots(wb) -> list[dict]:
    tab = _resolve_tab(wb, "root keywords")
    rows = _sheet_rows(wb, tab) if tab else []
    if not rows:
        return []
    hdr = [norm(h) for h in rows[0]]

    def col(*names):
        return next((i for i, h in enumerate(hdr) if h in names), None)

    ci = {"star": col("important"), "root": col("root keyword", "root"),
          "freq": col("frequency"), "sv": col("broad search volume", "broad sv"),
          "score": col("root score", "score"), "cat": col("category")}
    out = []
    for row in rows[1:]:
        root = row[ci["root"]] if ci["root"] is not None and ci["root"] < len(row) else None
        if not norm(root):
            continue
        out.append({
            "root": str(root).strip(),
            "star": str(row[ci["star"]] or "").strip() if ci["star"] is not None else "",
            "frequency": to_int(row[ci["freq"]]) if ci["freq"] is not None else None,
            "broad_sv": to_int(row[ci["sv"]]) if ci["sv"] is not None else None,
            "score": to_float(row[ci["score"]]) if ci["score"] is not None else None,
            "category": str(row[ci["cat"]] or "").strip() if ci["cat"] is not None else "",
            "word_count": len(bkw._word_tokens(str(root))),
        })
    return out


def read_outlier_flags(wb) -> dict[str, dict]:
    tab = _resolve_tab(wb, "outlier")
    rows = _sheet_rows(wb, tab) if tab else []
    if not rows:
        return {}
    hdr = [norm(h) for h in rows[0]]
    ki = next((i for i, h in enumerate(hdr) if h in ("keyword", "search terms", "search term")), 0)
    ci = next((i for i, h in enumerate(hdr) if h == "classification"), None)
    if ci is None:
        return {}
    out = {}
    for row in rows[1:]:
        if ki >= len(row) or not norm(row[ki]) or ci >= len(row):
            continue
        cls = norm(row[ci])
        out[norm(row[ki])] = {
            "competitor_brand": "competitor" in cls or "brand term" in cls,
            "claim": "claim" in cls or "health-risk" in cls,
            "form": "wrong product form" in cls,
            "semantic": "semantic" in cls,
        }
    return out


def read_never_terms(wb) -> set[str]:
    tab = _resolve_tab(wb, "never")
    rows = _sheet_rows(wb, tab) if tab else []
    if not rows:
        return set()
    hdr = [norm(h) for h in rows[0]]
    pi = next((i for i, h in enumerate(hdr) if "phrase" in h or "word" in h or "keyword" in h), None)
    terms = set()
    for row in rows[1:]:
        if pi is not None and pi < len(row) and norm(row[pi]):
            terms.add(norm(row[pi]))
    return terms


def read_asins(wb) -> tuple[list[dict], str]:
    rows = _sheet_rows(wb, "ASINs")
    if not rows:
        return [], ""
    hdr = [norm(h) for h in rows[0]]

    def col(name):
        return next((i for i, h in enumerate(hdr) if h == name), None)

    ai, ri, bi = col("asin"), col("role"), col("brand")
    comps, anchor_brand = [], ""
    for row in rows[1:]:
        if ai is None or ai >= len(row) or not norm(row[ai]):
            continue
        role = norm(row[ri]) if ri is not None and ri < len(row) else ""
        brand = str(row[bi] or "").strip() if bi is not None and bi < len(row) else ""
        if role == "anchor":
            anchor_brand = brand
        elif role == "competitor":
            comps.append({"asin": str(row[ai]).strip().upper(), "brand": brand})
    return comps, anchor_brand


def read_mkl(wb) -> tuple[dict[str, dict], dict[str, dict]]:
    """(core_terms, expanded_terms) keyed by normalized keyword.
    Values: {keyword, sv, relevancy, sugg_bid}."""
    def parse(tab):
        rows = _sheet_rows(wb, tab) if tab else []
        if not rows:
            return {}
        out = bkw._keyword_rows_by_term(rows)
        hdr = [norm(h) for h in rows[0]]
        bid_i = next((i for i, h in enumerate(hdr) if "bid" in h), None)
        if bid_i is not None:
            kw_i = next((i for i, h in enumerate(hdr)
                         if h in {"search terms", "search term", "keyword", "keywords"}), 1)
            for row in rows[1:]:
                if kw_i < len(row) and norm(row[kw_i]) in out and bid_i < len(row):
                    m = re.search(r"[\d.]+", str(row[bid_i] or ""))
                    out[norm(row[kw_i])]["sugg_bid"] = float(m.group()) if m else None
        for v in out.values():
            v.setdefault("sugg_bid", None)
        return out

    core = parse(bkw._resolve_master_sheet(wb, ""))
    expanded = parse(_resolve_tab(wb, "mkl 1%", "expanded mkl", "expanded 1%"))
    return core, expanded


def _stem(tok: str) -> str:
    return re.sub(r"(?:'s|’s|s)$", "", tok) if len(tok) > 3 else tok


def brand_matcher(cfg_tokens: list[str], anchor_brand: str):
    """Own-brand test. Config entries match as single stemmed tokens or phrases;
    the anchor brand contributes only multi-token consecutive phrases (its
    individual words are often generic product words, e.g. 'manuka')."""
    phrases: list[list[str]] = []
    singles: set[str] = set()
    for t in cfg_tokens:
        toks = [_stem(x) for x in bkw._word_tokens(t)]
        if len(toks) == 1:
            singles.add(toks[0])
        elif toks:
            phrases.append(toks)
    a_toks = [_stem(x) for x in bkw._word_tokens(anchor_brand)]
    if len(a_toks) >= 2:
        phrases.append(a_toks)
        phrases.append(a_toks[:2])

    def match(keyword: str) -> bool:
        kt = [_stem(x) for x in bkw._word_tokens(keyword)]
        if singles.intersection(kt):
            return True
        for p in phrases:
            for i in range(len(kt) - len(p) + 1):
                if kt[i:i + len(p)] == p:
                    return True
        return False

    return match


def assign_root(keyword: str, roots: list[dict]) -> dict | None:
    kw_tokens = set(bkw._word_tokens(keyword))
    best = None
    for r in roots:
        r_tokens = set(bkw._word_tokens(r["root"]))
        if r_tokens and r_tokens <= kw_tokens:
            if best is None or (r["word_count"], r["score"] or 0) > (best["word_count"], best["score"] or 0):
                best = r
    return best


# --------------------------------------------------------------- extract
def cmd_extract(cfg: dict, cfg_path: str, args: argparse.Namespace) -> int:
    cs = cfg.get("campaign_structure", {})
    strat_json_path = args.strategy or cs.get("strategy_json", "_local/ads-strategy/strategy.json")
    strat = load_strategy(strat_json_path)
    wb_path = args.workbook or cs.get("workbook") or cfg.get("inputs", {}).get("out", "")
    if not wb_path or not os.path.exists(wb_path):
        die(f"workbook not found: {wb_path!r} (set campaign_structure.workbook, inputs.out, or --workbook)")
    wb = openpyxl.load_workbook(wb_path)
    if CAMPAIGN_TAB not in wb.sheetnames:
        die(f"'{CAMPAIGN_TAB}' tab not found in {wb_path}")
    scaffold = scan_scaffold(wb[CAMPAIGN_TAB])

    roots = read_roots(wb)
    outlier = read_outlier_flags(wb)
    never = read_never_terms(wb)
    comps, anchor_brand = read_asins(wb)
    core, expanded = read_mkl(wb)

    own_cfg = [t for t in cs.get("own_brand_tokens", []) if norm(t)]
    is_own_brand = brand_matcher(own_cfg, anchor_brand)
    launched = {norm(k) for k in cs.get("already_launched", [])}
    comp_brand_terms = sorted({k for k, f in outlier.items() if f["competitor_brand"]})

    bands = strat["sv_bands"]
    rank_lo, rank_hi = bands["rank_skw"]["min"], bands["rank_skw"]["max"]
    halo_max = bands["halo"]["max"]
    xf = strat.get("expanded_candidate_filter", {})

    # candidate pool: all core terms + bounded expanded terms
    pool: dict[str, dict] = dict(core)
    for term, v in expanded.items():
        if term in pool:
            continue
        if (v["sv"] or 0) <= xf.get("max_sv", halo_max) and (v["relevancy"] or 0) >= xf.get("min_relevancy", 0):
            pool[term] = {**v, "_expanded_only": True}

    kws = []
    for term, v in sorted(pool.items(), key=lambda kv: -(kv[1]["sv"] or 0)):
        fl = outlier.get(term, {})
        flags = {
            "own_brand": is_own_brand(term),
            "competitor_brand": bool(fl.get("competitor_brand")),
            "never": term in never,
            "claim": bool(fl.get("claim")),
            "form": bool(fl.get("form")),
            "already_launched": term in launched,
        }
        r = assign_root(v["keyword"], roots)
        sv = v["sv"] or 0
        if flags["never"] or flags["claim"] or flags["form"]:
            bucket, why = "exclude", "never/claim/form flag"
        elif flags["competitor_brand"]:
            bucket, why = "pat_context", "competitor brand term — PAT target context, not a keyword"
        elif flags["already_launched"]:
            bucket, why = "skip", "already launched"
        elif flags["own_brand"]:
            bucket, why = "shield", "contains own-brand token"
        elif rank_lo <= sv <= rank_hi:
            bucket, why = "rank_skw", f"SV {sv} in rank band"
        elif 0 < sv < halo_max:
            bucket, why = "halo", f"SV {sv} below halo max"
        else:
            bucket, why = "review", f"SV {sv} outside bands — agent decides"
        kws.append({
            "keyword": v["keyword"], "sv": v["sv"], "relevancy": v["relevancy"],
            "sugg_bid": v.get("sugg_bid"),
            "root": r["root"] if r else None, "root_star": r["star"] if r else "",
            "flags": flags, "expanded_only": bool(v.get("_expanded_only")),
            "suggested_bucket": bucket, "suggested_reason": why,
        })

    # PAT suggestion by revenue (competitors csv / legacy tab), else null
    rev = {}
    comp_csv = args.competitors_csv or cfg.get("inputs", {}).get("competitors_csv", "")
    comp_csv = os.path.expanduser(comp_csv) if comp_csv else ""
    if comp_csv and os.path.exists(comp_csv):
        import csv as _csv
        with open(comp_csv, encoding="utf-8-sig") as f:
            rev_rows = list(_csv.reader(f))
        cmap = bkw.competitor_map(rev_rows)
        for asin, d in cmap.items():
            m = re.search(r"[\d,.]+", str(d.get("30d revenue") or d.get("revenue") or ""))
            if m:
                rev[asin.upper()] = float(m.group().replace(",", ""))
    revs = sorted(rev.values())
    median = revs[len(revs) // 2] if revs else None
    asins_out = []
    for c in comps:
        r = rev.get(c["asin"])
        sugg = None
        if r is not None and median is not None:
            sugg = "stronger" if r >= median else "weaker"
        asins_out.append({**c, "revenue_30d": r,
                          "revenue_source": "competitors_csv" if r is not None else "none",
                          "suggested_pat": sugg})

    out = {
        "schema": SCHEMA_CANDIDATES,
        "workbook": wb_path, "config": cfg_path,
        "strategy_json": strat_json_path,
        "strategy_md": cs.get("strategy_md", "_local/ads-strategy/strategy.md"),
        "product_anchor": cfg.get("product_anchor", {}),
        "scaffold": {
            "slots": [{k: s[k] for k in ("bucket", "label", "capacity")} for s in scaffold["slots"]],
        },
        "thresholds": {"sv_bands": bands, "caps": strat["caps"], "discovery": strat.get("discovery", {})},
        "own_brand_tokens": own_cfg,
        "anchor_brand": anchor_brand,
        "competitor_brand_terms": comp_brand_terms,
        "keywords": kws,
        "roots": roots,
        "asins": asins_out,
        "next_step": ("Agent: read strategy_md, assign keywords/ASINs to scaffold slots, write "
                      "classification.json (schema amazon-agent.campaign-classification.v1), then run --apply. "
                      "VISUAL FILL ONLY — do not create campaign bulk files from this data."),
    }
    dest = args.extract
    os.makedirs(os.path.dirname(dest) or ".", exist_ok=True)
    json.dump(out, open(dest, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"candidates: {len(kws)} keywords ({sum(1 for k in kws if k['suggested_bucket']=='rank_skw')} rank-band, "
          f"{sum(1 for k in kws if k['suggested_bucket']=='halo')} halo-band, "
          f"{sum(1 for k in kws if k['suggested_bucket']=='exclude')} excluded), "
          f"{len(asins_out)} competitor ASINs ({sum(1 for a in asins_out if a['revenue_30d'] is not None)} with revenue)")
    print(f"wrote {dest}")
    return 0


# --------------------------------------------------------------- apply
def _fmt_descriptor(tpl: str, *, keyword: str = "", root: str = "", counter: int | None = None) -> str:
    return (tpl.replace("{keyword}", keyword).replace("{root}", root)
               .replace("{counter}", f"{counter:02d}" if counter is not None else "")).strip()


def cmd_apply(cfg: dict, cfg_path: str, args: argparse.Namespace) -> int:
    cs = cfg.get("campaign_structure", {})
    strat = load_strategy(args.strategy or cs.get("strategy_json", "_local/ads-strategy/strategy.json"))
    cls = json.load(open(args.apply, encoding="utf-8"))
    if cls.get("schema") != SCHEMA_CLASSIFICATION:
        die(f"unexpected classification schema: {cls.get('schema')!r}")

    wb_path = args.workbook or cs.get("workbook") or cfg.get("inputs", {}).get("out", "")
    if not wb_path or not os.path.exists(wb_path):
        die(f"workbook not found: {wb_path!r}")
    wb = openpyxl.load_workbook(wb_path)
    ws = wb[CAMPAIGN_TAB]
    scaffold = scan_scaffold(ws)
    slots = {(s["bucket"], norm(s["label"])): s for s in scaffold["slots"]}

    # re-read source-of-truth data for validation
    core, expanded = read_mkl(wb)
    all_terms = {**expanded, **core}
    outlier = read_outlier_flags(wb)
    never = read_never_terms(wb)
    comps, anchor_brand = read_asins(wb)
    comp_asins = {c["asin"] for c in comps}
    own_cfg = [t for t in cs.get("own_brand_tokens", []) if norm(t)]
    is_own_brand = brand_matcher(own_cfg, anchor_brand)
    launched = {norm(k) for k in cs.get("already_launched", [])}
    roots = read_roots(wb)
    root_names = {norm(r["root"]) for r in roots}

    bands, caps = strat["sv_bands"], strat["caps"]
    fails: list[str] = []
    warns: list[str] = []
    skips: list[str] = []

    def slot_for(a) -> dict | None:
        label = str(a.get("campaign", "")).strip()
        if isinstance(a.get("campaign"), int):  # allow bare index for discovery columns
            matches = [s for s in scaffold["slots"] if s["bucket"] == a["bucket"]]
            idx = a["campaign"] - 1
            return matches[idx] if 0 <= idx < len(matches) else None
        return slots.get((a["bucket"], norm(label)))

    # group assignments per slot
    per_slot: dict[int, list[dict]] = {}
    seen_kw: dict[str, str] = {}
    seen_asin: dict[str, str] = {}
    for a in cls.get("assignments", []):
        s = slot_for(a)
        if s is None:
            fails.append(f"unknown slot: bucket={a.get('bucket')!r} campaign={a.get('campaign')!r}")
            continue
        if s["bucket"] in KEYWORD_BUCKETS:
            term = norm(a.get("keyword", ""))
            if not term:
                fails.append(f"{s['bucket']}/{s['label']}: assignment missing 'keyword'")
                continue
            if term in launched:
                skips.append(f"already launched, skipped: {a['keyword']}")
                continue
            row = all_terms.get(term)
            is_discovery = s["bucket"] in DISCOVERY_BUCKETS
            if row is None and not (is_discovery and term in root_names):
                fails.append(f"{s['bucket']}/{s['label']}: keyword not in workbook MKL tabs: {a['keyword']!r}")
                continue
            fl = outlier.get(term, {})
            if term in never or fl.get("claim") or fl.get("form"):
                fails.append(f"{s['bucket']}/{s['label']}: never/claim/form-flagged keyword rejected: {a['keyword']!r}")
                continue
            if fl.get("competitor_brand"):
                fails.append(f"{s['bucket']}/{s['label']}: competitor brand term rejected (TARGET it in PAT instead): {a['keyword']!r}")
                continue
            if s["bucket"].startswith("shield") and not is_own_brand(term):
                fails.append(f"{s['bucket']}/{s['label']}: shield keyword lacks an own-brand token: {a['keyword']!r}"
                             + ("" if own_cfg else " (set campaign_structure.own_brand_tokens in the config)"))
                continue
            # dedupe: strict within/across exact-match buckets; discovery may reuse
            # the same root across BMM and Phrase (intended duplication per theory),
            # but not twice within the same bucket; exact-vs-discovery overlap = WARN.
            prev = seen_kw.get(term)
            if prev:
                same_family = (prev in DISCOVERY_BUCKETS) == (is_discovery)
                if same_family and (not is_discovery or prev == s["bucket"]):
                    fails.append(f"duplicate keyword: {a['keyword']!r} ({prev} and {s['bucket']})")
                    continue
                if not same_family:
                    warns.append(f"{a['keyword']!r} used as both exact keyword ({prev if prev not in DISCOVERY_BUCKETS else s['bucket']}) "
                                 f"and discovery root — legitimate but check intent")
            seen_kw[term] = s["bucket"]
            sv = (row or {}).get("sv")
            if s["bucket"] == "rank_skw" and sv is not None and not (bands["rank_skw"]["min"] <= sv <= bands["rank_skw"]["max"]):
                msg = f"rank_skw/{s['label']}: SV {sv} outside rank band: {a['keyword']!r}"
                (fails if bands["rank_skw"].get("severity_outside") == "fail" else warns).append(msg)
            if s["bucket"] == "halo" and sv is not None and sv >= bands["halo"]["max"]:
                msg = f"halo/{s['label']}: SV {sv} at/above halo max: {a['keyword']!r}"
                (fails if bands["halo"].get("severity_above", "fail") == "fail" else warns).append(msg)
            per_slot.setdefault(id(s), []).append({**a, "_slot": s, "_sv": sv,
                                                   "_sugg_bid": (row or {}).get("sugg_bid")})
        else:  # PAT
            asin = str(a.get("asin", "")).strip().upper()
            if not ASIN_RE.match(asin):
                fails.append(f"{s['bucket']}/{s['label']}: invalid ASIN {a.get('asin')!r}")
                continue
            if asin not in comp_asins:
                fails.append(f"{s['bucket']}/{s['label']}: ASIN not a competitor in the ASINs tab: {asin}")
                continue
            if asin in seen_asin:
                fails.append(f"duplicate ASIN across PAT campaigns: {asin}")
                continue
            seen_asin[asin] = s["bucket"]
            brand = a.get("brand") or next((c["brand"] for c in comps if c["asin"] == asin), "")
            per_slot.setdefault(id(s), []).append({**a, "_slot": s, "_brand": brand})

    # per-slot rules
    for entries in per_slot.values():
        s = entries[0]["_slot"]
        if len(entries) > s["capacity"]:
            fails.append(f"{s['bucket']}/{s['label']}: {len(entries)} entries exceed capacity {s['capacity']}")
        if s["bucket"] == "halo":
            if len(entries) > caps["halo_keywords_per_campaign"]:
                fails.append(f"halo/{s['label']}: {len(entries)} keywords exceed cap {caps['halo_keywords_per_campaign']}")
            kw_roots = {norm(assign_root(e['keyword'], roots)["root"]) if assign_root(e['keyword'], roots) else "?" for e in entries}
            if len(kw_roots) > 1:
                fails.append(f"halo/{s['label']}: mixed roots {sorted(kw_roots)} — one root theme per halo campaign")
        if s["bucket"] in DISCOVERY_BUCKETS:
            if len(entries) != 1:
                fails.append(f"{s['bucket']}/{s['label']}: exactly one root keyword per discovery campaign (got {len(entries)})")
            for e in entries:
                wc = len(bkw._word_tokens(e["keyword"]))
                if wc < strat.get("discovery", {}).get("min_root_words", 1):
                    warns.append(f"{s['bucket']}/{s['label']}: root {e['keyword']!r} has {wc} word(s) — specificity check")
        if s["bucket"].startswith("pat") and len(entries) > caps["pat_asins_per_campaign"]:
            fails.append(f"{s['bucket']}/{s['label']}: exceeds PAT cap {caps['pat_asins_per_campaign']}")

    # ---- report
    for w in warns:
        print(f"WARN: {w}")
    for k in skips:
        print(f"SKIP: {k}")
    if fails:
        for f_ in fails:
            print(f"FAIL: {f_}", file=sys.stderr)
        print(f"\n{len(fails)} validation failure(s) — nothing written.", file=sys.stderr)
        return 1

    # ---- build write plan
    plan: list[tuple[int, int, Any]] = []  # (row, col, value)
    for entries in per_slot.values():
        s = entries[0]["_slot"]
        entries.sort(key=lambda e: (e.get("order") is None, e.get("order", 0), -(e.get("_sv") or 0)))
        for i, e in enumerate(entries):
            r = s["data_start"] + i
            if s["bucket"].startswith("pat"):
                plan.append((r, s["kw_col"], e["asin"].strip().upper()))
                plan.append((r, s["sv_col"], e.get("_brand", "")))
            else:
                plan.append((r, s["kw_col"], e["keyword"]))
                if e.get("_sv") is not None:
                    plan.append((r, s["sv_col"], int(e["_sv"])))

    # names + bids block
    naming = strat["naming"]
    settings = {k: naming.get(k, "") for k in ("variable_order", "delimiter", "suffix", "custom1_value", "custom2_value")}
    product_name = cs.get("product_name_for_naming") or cfg.get("product_anchor", {}).get("product", "")
    if len(product_name) > 40:
        warns.append(f"product name for campaign names is {len(product_name)} chars (full listing title?) — "
                     "set campaign_structure.product_name_for_naming to a short name")
        print(f"WARN: {warns[-1]}")
    bids = strat.get("bids", {})
    pct = bids.get("start_bid_pct_of_recommended")
    names_rows: list[list[Any]] = [
        [f"PROPOSED CAMPAIGN NAMES — generated by fill_campaign_structure.py from the local naming template "
         f"(_local/ads-strategy/strategy.json). Visual plan only; edit freely."],
        ["Bucket", "Campaign/Wave", "Target", "Proposed Campaign Name", "Proposed Ad Group Name",
         "Daily budget", "Start bid", "ToS %", "RoS %", "PP %"],
    ]
    counters: dict[str, int] = {}

    def name_row(bucket, label, target, descriptor, sugg_bid=None):
        by = naming.get("by_bucket", {}).get(bucket, {})
        counters[bucket] = counters.get(bucket, 0) + 1
        ctx = {"goal": by.get("goal", ""), "match_type": by.get("match_type", ""),
               "campaign_type": by.get("campaign_type", ""),
               "product_name": product_name,
               "target_descriptor": descriptor, "counter": None}
        cname = campaign_model.generate_campaign_name(settings, ctx)
        agname = campaign_model.generate_ad_group_name(settings, ctx)
        b = bids.get("by_bucket", {}).get(bucket, {})
        if sugg_bid is not None and pct is not None:
            start_bid = round(sugg_bid * (1 + pct / 100.0), 2)
        elif pct is not None:
            start_bid = f"{pct:+d}% of recommended"
        else:
            start_bid = ""
        names_rows.append([bucket, label, target, cname, agname,
                           b.get("daily_budget", ""), start_bid,
                           b.get("top_of_search_placement", ""), b.get("rest_of_search_placement", ""),
                           b.get("product_pages_placement", "")])

    for entries in sorted(per_slot.values(), key=lambda es: (es[0]["_slot"]["header_row"], es[0]["_slot"]["kw_col"])):
        s = entries[0]["_slot"]
        by = naming.get("by_bucket", {}).get(s["bucket"], {})
        dtpl = by.get("descriptor", "{keyword}")
        if s["bucket"] in ("rank_skw", "shield_skw"):
            for e in entries:  # one campaign per keyword (SKW)
                name_row(s["bucket"], s["label"], e["keyword"],
                         _fmt_descriptor(dtpl, keyword=e["keyword"]), e.get("_sugg_bid"))
        elif s["bucket"] in DISCOVERY_BUCKETS:
            e = entries[0]
            name_row(s["bucket"], s["label"], e["keyword"], _fmt_descriptor(dtpl, root=e["keyword"]))
        elif s["bucket"] == "halo":
            root = (assign_root(entries[0]["keyword"], roots) or {}).get("root", "")
            name_row(s["bucket"], s["label"], f"{len(entries)} keywords (root: {root})",
                     _fmt_descriptor(dtpl, root=root))
        else:  # PAT
            cnt = counters.get(s["bucket"], 0) + 1
            name_row(s["bucket"], s["label"], f"{len(entries)} ASINs",
                     _fmt_descriptor(dtpl, counter=cnt))

    names_start = scaffold["last_sum_row"] + 2

    if args.dry_run:
        print(f"\nDRY RUN — workbook untouched: {wb_path}")
        print(f"scaffold: {len(scaffold['slots'])} slots, last Sum row {scaffold['last_sum_row']}")
        print(f"planned cell writes: {len(plan)}; names block: {len(names_rows)} rows at row {names_start}")
        for entries in sorted(per_slot.values(), key=lambda es: (es[0]["_slot"]["header_row"], es[0]["_slot"]["kw_col"])):
            s = entries[0]["_slot"]
            print(f"  {s['bucket']:<24} {s['label']:<12} {len(entries)}/{s['capacity']}")
        for row in names_rows[2:]:
            print("  name:", row[3])
        return 0

    # ---- clear + write
    forbidden = set()
    for s in scaffold["slots"]:
        for r in range(s["data_start"], s["data_end"] + 1):
            ws.cell(r, s["kw_col"]).value = None
            ws.cell(r, s["sv_col"]).value = None
    for r in range(names_start, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
    for r, c, v in plan:
        ok = any(s["data_start"] <= r <= s["data_end"] and c in (s["kw_col"], s["sv_col"])
                 for s in scaffold["slots"])
        if not ok:
            die(f"internal error: planned write outside scaffold data ranges at r{r} c{c}")
        ws.cell(r, c, value=v)
    from openpyxl.styles import Font
    for i, row in enumerate(names_rows):
        r = names_start + i
        for j, v in enumerate(row, start=1):
            cell = ws.cell(r, j, value=v)
            if i in (0, 1):
                cell.font = Font(bold=True)

    out_path = args.out or wb_path
    if out_path == wb_path:
        bak = f"{wb_path}.bak-{time.strftime('%Y%m%d-%H%M%S')}.xlsx"
        shutil.copy2(wb_path, bak)
        print(f"backup: {bak}")
    wb.save(out_path)
    print(f"applied: {len(plan)} cell writes + {len(names_rows)} name-block rows -> {out_path}")
    if warns:
        print(f"{len(warns)} warning(s) above.")
    return 0


# --------------------------------------------------------------- main
def main() -> int:
    ap = argparse.ArgumentParser(description="Fill the workbook '5. Campaign Structure' tab (visual plan only).")
    ap.add_argument("--config", required=True, help="per-client config.<client>.json")
    ap.add_argument("--workbook", help="built workbook .xlsx (default: campaign_structure.workbook or inputs.out)")
    ap.add_argument("--strategy", help="strategy.json override (default: campaign_structure.strategy_json)")
    ap.add_argument("--competitors-csv", help="competitors CSV for PAT revenue enrichment")
    ap.add_argument("--extract", metavar="CANDIDATES_JSON", help="write the candidates file for agent classification")
    ap.add_argument("--apply", metavar="CLASSIFICATION_JSON", help="validate + write a classification into the tab")
    ap.add_argument("--dry-run", action="store_true", help="with --apply: print the plan, write nothing")
    ap.add_argument("--out", help="with --apply: write to this copy instead of in-place")
    args = ap.parse_args()

    if bool(args.extract) == bool(args.apply):
        ap.error("use exactly one of --extract or --apply")
    cfg = json.load(open(args.config, encoding="utf-8"))
    if args.extract:
        return cmd_extract(cfg, args.config, args)
    return cmd_apply(cfg, args.config, args)


if __name__ == "__main__":
    sys.exit(main())
