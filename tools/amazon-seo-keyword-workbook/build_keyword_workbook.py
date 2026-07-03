#!/usr/bin/env python3
"""
Amazon SEO keyword-research workbook builder.

Deterministically rebuilds a keyword workbook from raw DataDive +
Amazon Product Opportunity Explorer (POE) exports, while preserving the style
and structure of an existing template workbook.

Design
------
* Template-base: an existing .xlsx is loaded as the styled base, so column
  widths, freeze panes, tab colors, and formulas can be reused.
* Exact-paste tabs (DataDive roots/core MKL/expanded MKL, POE products/search terms) are
  cleared and re-pasted 1:1 from their source CSVs. They are NEVER transformed.
* Rebuilt tabs:
    - "ASINs", "2.2 Never KWs", "3.2 Outlier KWs", "POE Raw - Reviews",
      "POE Raw - Returns", "POE Semantic Insights", and blank/skipped tabs are
      generated from the current run. ("4.2 Competitors" is legacy/optional —
      built only if a config lists it; the current template carries competitor
      copy in the ASINs tab instead.)
    - "POE Raw - Related Niches": rebuilt from the related-niches JSON with a
      configurable keep-list relevance filter.
    - "3.2 Outlier KWs": rule-based triage over the master keyword
      list, plus a Final Action column.
* The template is a style source only. No product-specific tab is carried forward.
* All rules live in config JSON so the workflow is auditable and reusable.
* Validation + an evidence manifest (JSON) are emitted every run.

Usage
-----
    # Client-agnostic: pick a client/product/market config; paths come from its
    # inputs{} block. Copy config.TEMPLATE.json to start a new client.
    .venv/bin/python tools/amazon-seo-keyword-workbook/build_keyword_workbook.py \
        --config tools/amazon-seo-keyword-workbook/config.<client>-<product>-<market>.json

    # --preflight first to see which inputs are still missing.
    # Any path can be overridden with a flag (--template/--out/--roots-csv/...).
"""
from __future__ import annotations

import argparse
import copy
import csv
import datetime as _dt
import json
import os
import re
import shutil
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------- #
# Repo + default paths
# --------------------------------------------------------------------------- #
REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
HOME = os.path.expanduser("~")

# Client-agnostic by design: there are NO client/product defaults. Every path
# comes from the chosen config's `inputs{}` block (or an explicit --flag). The
# only default is `config`, which points at the commented TEMPLATE so a bare run
# / --help is self-documenting and fails preflight with a clear "fill the
# template" message rather than silently building someone else's client.
DEFAULTS = {
    "config": os.path.join(os.path.dirname(__file__), "config.TEMPLATE.json"),
    "template": "",
    "out": "",
    "manifest": "",
    "roots_csv": "",
    "master_csv": "",
    "expanded_mkl_csv": "",
    "competitors_csv": "",
    "poe_products_csv": "",
    "poe_search_terms_csv": "",
    "related_niches_json": "",
    "poe_reviews_json": "",
    "poe_returns_json": "",
    "poe_structured_json": "",
    "listing_reference_json": "",
    "seo_content": "",
    "curated_tabs": "",  # optional: JSON of curated tab content to write deterministically.
    "drive_dir": "",  # optional: folder to copy the finished .xlsx into (e.g. the synced Drive folder). Off by default.
    "handoff_note": "",
}


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def read_csv(path: str) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def drop_leading_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    """Strip leading column(s) that are blank in the header AND every data row.

    DataDive's Core/Expanded MKL CSV exports prepend a blank column (empty
    header + empty cells, the UI's select/checkbox gutter). Pasted 1:1 it shows
    up as an empty column A in the Master List / Expanded MKL tabs. This removes
    only fully-blank leading columns, so CSVs without the gutter (roots,
    competitors, POE) are untouched, and the last column is never stripped.
    """
    if not rows:
        return rows

    def col0_blank(r: list[str]) -> bool:
        return len(r) == 0 or r[0] is None or str(r[0]).strip() == ""

    while len(rows[0]) > 1 and all(col0_blank(r) for r in rows):
        rows = [r[1:] for r in rows]
    return rows


def norm(s: Any) -> str:
    """Normalize text for matching: lowercase, collapse whitespace."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def to_int(x: Any):
    try:
        return int(float(str(x).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def to_float(x: Any):
    try:
        return float(str(x).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# --------------------------------------------------------------------------- #
# Style preservation
# --------------------------------------------------------------------------- #
def snapshot_styles(ws: Worksheet, upto_row: int = 8) -> list[list[dict]]:
    """Capture per-cell style for the first `upto_row` rows so they can be
    re-applied after the tab is cleared and re-pasted."""
    snap: list[list[dict]] = []
    rows = min(ws.max_row, upto_row) if ws.max_row else 0
    cols = ws.max_column or 1
    for r in range(1, rows + 1):
        row_styles = []
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            row_styles.append(
                {
                    "font": copy.copy(cell.font),
                    "fill": copy.copy(cell.fill),
                    "border": copy.copy(cell.border),
                    "alignment": copy.copy(cell.alignment),
                    "number_format": cell.number_format,
                }
            )
        snap.append(row_styles)
    return snap


def apply_style(cell, snap: list[list[dict]], row_idx: int, col_idx: int) -> None:
    """Apply a captured style. Rows past the snapshot reuse the last captured
    row (uniform data styling); columns past it reuse the last captured col."""
    if not snap:
        return
    r = min(row_idx, len(snap)) - 1
    row_styles = snap[r]
    if not row_styles:
        return
    c = min(col_idx, len(row_styles)) - 1
    st = row_styles[c]
    cell.font = copy.copy(st["font"])
    cell.fill = copy.copy(st["fill"])
    cell.border = copy.copy(st["border"])
    cell.alignment = copy.copy(st["alignment"])
    cell.number_format = st["number_format"]


def clear_rows(ws: Worksheet) -> None:
    """Remove all row content but keep sheet-level props (widths via
    column_dimensions, freeze_panes, tabColor all persist)."""
    if ws.max_row and ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)


def write_matrix(ws: Worksheet, rows: list[list[Any]], snap: list[list[dict]]) -> int:
    """Write a 2D matrix starting at A1, applying captured styles. Returns the
    number of rows written."""
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(i, j, value=val)
            apply_style(cell, snap, i, j)
    return len(rows)


def ensure_sheet(wb, title: str, after: Any = None) -> Worksheet:
    if title in wb.sheetnames:
        return wb[title]
    # Reuse old long-tail tab as the style slot for the expanded MKL when present.
    ws = wb.create_sheet(title)
    if after and after in wb.sheetnames:
        wb._sheets.remove(ws)
        idx = wb.sheetnames.index(after) + 1
        wb._sheets.insert(idx, ws)
    return ws


# --------------------------------------------------------------------------- #
# Tab builders
# --------------------------------------------------------------------------- #
def _triage_vocab(core_rows: list[list[str]], poe_terms: set[str], args: dict, cfg: dict) -> dict:
    """The shared vocabulary/token sets used to classify both roots and
    Never-Ever words — one consistent logic across the two tabs."""
    tcfg = cfg.get("triage", {})
    ncfg = cfg.get("never_ever", {})
    core = _keyword_rows_by_term(core_rows)
    return {
        "core_words": {tok for kw in core for tok in _word_tokens(kw)},
        "poe_words": {tok for term in poe_terms for tok in _word_tokens(term)},
        "listing_words": set(_word_tokens(_listing_language(args, cfg))),
        "brand": {norm(x) for x in tcfg.get("brand_tokens", [])},
        "form": {norm(x) for x in tcfg.get("form_tokens", [])},
        "claim": {norm(x) for x in tcfg.get("claim_tokens", [])},
        "negative": {norm(x) for x in tcfg.get("negative_tokens", [])},
        "explicit_never": {norm(x) for x in ncfg.get("explicit_never_words", [])},
        "relevant": {norm(x) for x in ncfg.get("relevant_words", [])},
        "misspell": {norm(x) for x in ncfg.get("misspell_or_variant_words", [])},
        "stop": {norm(x) for x in ncfg.get("stop_words", [])},
        "related_excludes": {
            tok
            for phrase in cfg.get("related_niche_filter", {}).get("exclude_examples", [])
            for tok in _word_tokens(phrase)
        },
    }


def _root_category(root: str, vocab: dict) -> str:
    """Semantic bucket for a root. Guard buckets (Brand/Claim/Form/Off-niche)
    win over Core/Product so an ad-root marker can never land on them."""
    words = _word_tokens(root)
    if any(w in vocab["brand"] for w in words):
        return "Brand"
    if any(w in vocab["claim"] for w in words):
        return "Claim risk"
    if any(w in vocab["form"] for w in words):
        return "Wrong form"
    if any(w in vocab["negative"] or w in vocab["explicit_never"] or w in vocab["related_excludes"] for w in words):
        return "Off-niche"
    if any(
        w in vocab["core_words"] or w in vocab["listing_words"] or w in vocab["poe_words"] or w in vocab["relevant"]
        for w in words
    ):
        return "Core/Product"
    return "Review"


def transform_roots(rows: list[list[str]], cfg: dict, vocab: dict | None = None) -> list[list[Any]]:
    """Clean up the DataDive roots CSV for the '1. Root Keywords' tab.

    The raw export is `(empty) | Normalized Root | Frequency | Broad Search
    Volume | (unnamed 0-1 relevance score)`. Output columns: `Important | Root
    Keyword | Frequency | Broad Search Volume | Root Score | Category`.

    Importance marking (config block `root_importance`) — the operator's ad
    targets, so this is the load-bearing column:
    - Tiered mode (when `ad_min_sv`/`ad_min_score` is configured): `ad_marker`
      (default '⭐⭐') when score >= ad_min_score AND Broad SV >= ad_min_sv AND
      the root's Category is Core/Product or Review — the roots that should
      seed SKW/rank campaigns. `marker` ('⭐') for score >= min_score without
      the SV floor. Brand/Claim/Form/Off-niche roots never get the ad marker.
    - Legacy mode (only min_score/marker configured): unchanged binary marking.

    Category comes from `_root_category()` (same token sets as the Never-Ever
    ladder) when vocab is provided. Keeps exactly one output row per input row
    so the roots row-count QA gate is unaffected.
    """
    if not rows:
        return rows
    imp = cfg.get("root_importance") or {}
    try:
        min_score = float(imp.get("min_score", 0.10))
    except (TypeError, ValueError):
        min_score = 0.10
    marker = imp.get("marker", "⭐")
    tiered = "ad_min_sv" in imp or "ad_min_score" in imp
    try:
        ad_min_score = float(imp.get("ad_min_score", min_score))
    except (TypeError, ValueError):
        ad_min_score = min_score
    try:
        ad_min_sv = int(imp.get("ad_min_sv", 500))
    except (TypeError, ValueError):
        ad_min_sv = 500
    ad_marker = imp.get("ad_marker", "⭐⭐")

    out: list[list[Any]] = [["Important", "Root Keyword", "Frequency", "Broad Search Volume", "Root Score", "Category"]]
    for r in rows[1:]:
        root = r[1] if len(r) > 1 else ""
        freq = r[2] if len(r) > 2 else ""
        sv = r[3] if len(r) > 3 else ""
        raw_score = r[4] if len(r) > 4 else ""
        score = to_float(raw_score)
        category = _root_category(root, vocab) if vocab else ""
        if score is None:
            out.append(["", root, freq, sv, raw_score, category])
            continue
        mark = marker if score >= min_score else ""
        if tiered and score >= ad_min_score and category in ("Core/Product", "Review", ""):
            sv_val = to_int(sv)
            if sv_val is not None and sv_val >= ad_min_sv:
                mark = ad_marker
        out.append([mark, root, freq, sv, round(score, 2), category])
    return out


def build_exact_paste(ws: Worksheet, csv_rows: list[list[str]], warnings: list[str]) -> int:
    """Clear a tab and paste CSV rows 1:1 with preserved styling."""
    snap = snapshot_styles(ws, upto_row=8)
    clear_rows(ws)
    n = write_matrix(ws, csv_rows, snap)
    return n


def competitor_map(competitors_rows: list[list[str]]) -> dict[str, dict[str, Any]]:
    if not competitors_rows:
        return {}
    raw_header = [str(h).strip() for h in competitors_rows[0]]
    # The genuine DataDive UI export (Niche Tracker > Export Competitors) is
    # TRANSPOSED: attribute names in rows ('Brand', 'ASIN', 'Strength', ...)
    # and one column per competitor ASIN. Detect ASIN column headers and pivot.
    asin_cols = {i: h.upper() for i, h in enumerate(raw_header) if re.match(r"^B0[A-Z0-9]{8}$", h, re.I)}
    if asin_cols:
        label_idx = next((i for i, h in enumerate(raw_header) if norm(h) == "competitors"), 1)
        out_t: dict[str, dict[str, Any]] = {a: {"asin": a} for a in asin_cols.values()}
        for row in competitors_rows[1:]:
            label = norm(row[label_idx]) if label_idx < len(row) else ""
            if not label:
                continue
            for i, a in asin_cols.items():
                out_t[a].setdefault(label, row[i] if i < len(row) else "")
        return out_t
    # Row-per-competitor shape (MCP-derived fallback exports).
    header = [norm(h) for h in competitors_rows[0]]
    out: dict[str, dict[str, Any]] = {}
    for row in competitors_rows[1:]:
        item = {}
        for i, key in enumerate(header):
            item[key] = row[i] if i < len(row) else ""
        asin = item.get("asin") or item.get("asins")
        if asin:
            out[str(asin).strip()] = item
    return out


def build_asins(
    ws: Worksheet,
    cfg: dict,
    competitors_rows: list[list[str]],
    listing_path: str,
    result: dict,
    warnings: list[str],
) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    comp_by_asin = competitor_map(competitors_rows)
    listing = load_json(listing_path) if listing_path and os.path.exists(listing_path) else {}
    child_asins = listing.get("child_asins") or listing.get("children") or []
    if isinstance(child_asins, dict):
        child_asins = [{"asin": k, **(v if isinstance(v, dict) else {})} for k, v in child_asins.items()]

    anchor = cfg["product_anchor"].get("asin", "")
    mkt = cfg["product_anchor"].get("marketplace", "")
    client = cfg["product_anchor"].get("client", "")

    # Index any captured listing copy (anchor and/or competitors) by ASIN so the
    # template's title / bullet_points columns fill when the data exists and stay
    # blank — never fabricated — until the listing capture provides them.
    copy_items = []
    if isinstance(listing, dict):
        copy_items = listing.get("listings") or ([listing] if listing.get("asin") else [])
    copy_idx: dict[str, dict[str, str]] = {}
    for it in copy_items:
        if isinstance(it, dict) and it.get("asin"):
            bl = it.get("bullets") or it.get("bullet_points") or []
            if isinstance(bl, list):
                bl = "\n".join(str(b) for b in bl)
            link = it.get("link") or it.get("canonical") or it.get("url") or ""
            copy_idx[str(it["asin"]).strip()] = {
                "title": it.get("title", "") or "", "bullets": bl or "", "link": link or "",
            }

    def title_for(a: str, fallback: str = "") -> str:
        return copy_idx.get(a, {}).get("title") or fallback

    def bullets_for(a: str) -> str:
        return copy_idx.get(a, {}).get("bullets", "")

    def link_for(a: str) -> str:
        # Clean canonical 'amazon.<tld>/dp/<ASIN>' — this column doubles as the
        # input link list for the listing-capture scrape (matches the ZeroWork
        # 'Format Link' normalization). Falls back to any captured link.
        if a and mkt:
            tld = {"italy": "it", "germany": "de", "uk": "co.uk", "spain": "es", "france": "fr"}.get(mkt.lower(), "com")
            return f"https://www.amazon.{tld}/dp/{a}"
        return copy_idx.get(a, {}).get("link", "")

    # Header matches the Google template (+ 'link' column): the captured listing
    # scrape (link | title | bullet_points | ASIN) lands here by default.
    rows = [["ASIN", "role", "marketplace", "brand", "title", "bullet_points", "link", "source"]]
    rows.append([
        anchor, "Anchor", mkt, client,
        title_for(anchor, cfg["product_anchor"].get("product", "")),
        bullets_for(anchor), link_for(anchor),
        "listing_reference_json" if anchor in copy_idx else "config.product_anchor",
    ])
    for child in child_asins:
        ca = child if isinstance(child, str) else str(child.get("asin", "")).strip()
        brand = "" if isinstance(child, str) else child.get("brand", "")
        fb_title = "" if isinstance(child, str) else child.get("title", "")
        rows.append([ca, "Child", mkt, brand, title_for(ca, fb_title), bullets_for(ca), link_for(ca), "listing_reference_json"])

    # Same-brand non-anchor ASINs are siblings, not competitors. Detect by an
    # explicit config list (asin_roles.siblings) or by brand matching the anchor.
    sibling_set = {str(a).strip().upper() for a in (cfg.get("asin_roles", {}).get("siblings") or [])}
    client_n = norm(client)

    def comp_role(asin: str, brand: str) -> str:
        if str(asin).strip().upper() in sibling_set:
            return "Sibling"
        if client_n and norm(brand) == client_n:
            return "Sibling"
        return "Competitor"

    anchor_u = str(anchor).strip().upper()
    for asin, comp in comp_by_asin.items():
        if str(asin).strip().upper() == anchor_u:
            continue  # anchor already written as the Anchor row; don't re-add it
        rows.append([
            asin, comp_role(asin, comp.get("brand", "")), mkt, comp.get("brand", ""),
            title_for(asin), bullets_for(asin), link_for(asin), "competitors_csv",
        ])

    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({"rows_written": len(rows), "competitors": len(comp_by_asin), "children": max(0, len(rows) - len(comp_by_asin) - 2)})
    if not listing_path or not os.path.exists(listing_path):
        warnings.append("ASINs: listing_reference_json missing; ASIN tab built from config + competitors only.")
    return result


def build_competitors(ws: Worksheet, competitors_rows: list[list[str]], result: dict, warnings: list[str]) -> dict:
    n = build_exact_paste(ws, competitors_rows, warnings)
    result.update({"rows_written": n})
    return result


def _keyword_rows_by_term(rows: list[list[str]]) -> dict[str, dict[str, Any]]:
    if not rows:
        return {}
    header = [norm(h) for h in rows[0]]
    kw_idx = next((i for i, h in enumerate(header) if h in {"search terms", "search term", "keyword", "keywords"}), 1)
    sv_idx = next((i for i, h in enumerate(header) if h in {"sv", "search volume"}), 2)
    rel_idx = next((i for i, h in enumerate(header) if h.startswith("relev")), 3)
    out: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        if kw_idx >= len(row) or not norm(row[kw_idx]):
            continue
        term = norm(row[kw_idx])
        out[term] = {
            "keyword": row[kw_idx],
            "sv": to_int(row[sv_idx]) if sv_idx < len(row) else None,
            "relevancy": to_float(row[rel_idx]) if rel_idx < len(row) else None,
        }
    return out


def _word_tokens(text: str) -> list[str]:
    return re.findall(r"[0-9A-Za-zÀ-ÖØ-öø-ÿ]+(?:['’][0-9A-Za-zÀ-ÖØ-öø-ÿ]+)?", norm(text))


def _listing_language(args: dict, cfg: dict) -> str:
    parts = [json.dumps(cfg.get("product_anchor", {}), ensure_ascii=False)]
    for key in ("listing_reference_json", "seo_content"):
        path = args.get(key, "")
        if path and os.path.exists(path):
            try:
                parts.append(json.dumps(load_json(path), ensure_ascii=False))
            except Exception:
                pass
    return " ".join(parts)


def build_never_keywords(
    ws: Worksheet,
    expanded_rows: list[list[str]],
    core_rows: list[list[str]],
    poe_terms: set[str],
    args: dict,
    cfg: dict,
    result: dict,
    warnings: list[str],
) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    expanded = _keyword_rows_by_term(expanded_rows)
    vocab = _triage_vocab(core_rows, poe_terms, args, cfg)

    ncfg = cfg.get("never_ever", {})
    min_freq = int(ncfg.get("min_frequency", 2))
    max_rel = float(ncfg.get("max_relevancy_for_auto_negative", 0.3))
    review_band = float(ncfg.get("review_band_max_relevancy", 0.5))
    phrase_candidates = bool(ncfg.get("phrase_candidates", True))

    bucket: dict[str, dict[str, Any]] = {}
    for kw, meta in expanded.items():
        seen_in_kw = set(_word_tokens(kw))
        for word in seen_in_kw:
            if len(word) < 2 or word in vocab["stop"]:
                continue
            b = bucket.setdefault(word, {"frequency": 0, "sv": [], "rel": [], "examples": []})
            b["frequency"] += 1
            if meta.get("sv") is not None:
                b["sv"].append(meta["sv"])
            if meta.get("relevancy") is not None:
                b["rel"].append(meta["relevancy"])
            if len(b["examples"]) < 5:
                b["examples"].append(meta["keyword"])

    # Sectioned audit view: the classification ladder is unchanged (protection
    # first), but everything it computes is now WRITTEN instead of discarded —
    # category, why, SV/relevancy evidence, example phrases — and the
    # non-negative categories (brands, claim risk, review band) get their own
    # sections so a human can audit and act per campaign type.
    header = ["Negative Marker", "Frequency", "Word / Phrase", "Category", "Why", "Max SV", "Max Relevancy", "Example Phrases"]
    never_rows: list[list[Any]] = []
    brand_rows: list[list[Any]] = []
    claim_rows: list[list[Any]] = []
    review_rows: list[list[Any]] = []
    reviewed = 0
    never_words: set[str] = set()

    for word, data in bucket.items():
        reviewed += 1
        freq = data["frequency"]
        rels = data["rel"] or [0.0]
        max_rel_seen = max(rels)
        in_core = word in vocab["core_words"]
        in_poe = word in vocab["poe_words"]
        in_product = word in vocab["listing_words"]
        max_sv = max(data["sv"]) if data["sv"] else ""
        examples = "; ".join(data["examples"])
        evidence = [freq, word, "", "", max_sv, round(max_rel_seen, 3), examples]

        if word in vocab["relevant"] or in_product or in_core or in_poe:
            continue  # protected — appears in core/product/POE language or configured relevant words
        if word in vocab["misspell"]:
            continue  # variant still represents relevant product intent
        if word in vocab["brand"]:
            evidence[2:4] = ["Competitor brand", "Negative in rank/SKW campaigns, TARGET in PAT/conquest — decide per campaign. Never a blanket negative."]
            brand_rows.append(["Campaign-dependent"] + evidence)
            continue
        if word in vocab["claim"]:
            evidence[2:4] = ["Claim risk", "Health/compliance risk — do not use in copy automatically; see health-claims-compliance.md / compliance.claims_audit."]
            claim_rows.append(["Do not auto-use"] + evidence)
            continue
        if word in vocab["explicit_never"] or word in vocab["negative"] or word in vocab["form"] or word in vocab["related_excludes"]:
            category = (
                "Wrong form" if word in vocab["form"]
                else "Configured junk" if word in vocab["explicit_never"]
                else "Off-niche/irrelevant"
            )
            evidence[2:4] = [category, "Configured irrelevant/wrong-form/off-niche token."]
            never_rows.append(["Negative phrase"] + evidence)
            never_words.add(word)
            continue
        if freq >= min_freq and max_rel_seen <= max_rel:
            evidence[2:4] = [
                "Auto (low-relevancy discovery)",
                f"Frequent in 1% discovery but absent from core/product/POE context; max relevancy <= {max_rel}.",
            ]
            never_rows.append(["Negative phrase"] + evidence)
            never_words.add(word)
            continue
        if freq >= min_freq and max_rel < max_rel_seen <= review_band:
            evidence[2:4] = [
                "Near-miss band",
                f"Absent from core/product/POE but relevancy {round(max_rel_seen, 3)} is above the {max_rel} auto cut — human call.",
            ]
            review_rows.append(["Review manually"] + evidence)

    # Phrase-level negative candidates: full 1% keywords whose combination is
    # low-relevancy although no single word earned a word-level negative — the
    # combination-irrelevant edge case. Kept small (top 30 by SV).
    phrase_rows: list[list[Any]] = []
    if phrase_candidates:
        core_kws = set(_keyword_rows_by_term(core_rows))
        for kw, meta in expanded.items():
            rel = meta.get("relevancy")
            if rel is None or rel > max_rel:
                continue
            words = _word_tokens(kw)
            if len(words) < 2 or any(w in never_words for w in words):
                continue
            # Brand/claim words have their own sections with their own handling;
            # don't duplicate those phrases here.
            if any(w in vocab["brand"] or w in vocab["claim"] for w in words):
                continue
            if norm(kw) in poe_terms or norm(kw) in core_kws:
                continue
            phrase_rows.append([
                "Review (phrase)", "", meta["keyword"], "Phrase-level candidate",
                f"Combination relevancy {round(rel, 3)} <= {max_rel} but no single word qualified as a word-level negative — negate the exact phrase if truly off-product.",
                meta.get("sv") or "", round(rel, 3), "",
            ])
        phrase_rows.sort(key=lambda r: -(to_int(r[5]) or 0))
        phrase_rows = phrase_rows[:30]

    def _freq_sorted(rows_: list[list[Any]]) -> list[list[Any]]:
        return sorted(rows_, key=lambda r: (-(to_int(r[1]) or 0), str(r[2])))

    rows: list[list[Any]] = [header]
    sections = [
        ("SECTION: NEVER EVER — apply as negative phrase", _freq_sorted(never_rows)),
        ("SECTION: COMPETITOR BRANDS — campaign-dependent (not blanket negatives)", _freq_sorted(brand_rows)),
        ("SECTION: CLAIM RISK — compliance review, do not auto-use", _freq_sorted(claim_rows)),
        ("SECTION: REVIEW MANUALLY — near-miss band", _freq_sorted(review_rows)),
        ("SECTION: PHRASE-LEVEL NEGATIVE CANDIDATES — review", phrase_rows),
    ]
    for title, srows in sections:
        if not srows:
            continue
        rows.append([title, "", "", "", "", "", "", ""])
        rows.extend(srows)

    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({
        "rows_written": len(rows),
        "never_ever_words": len(never_rows),
        "brand_words": len(brand_rows),
        "claim_words": len(claim_rows),
        "review_words": len(review_rows),
        "phrase_candidates": len(phrase_rows),
        "words_reviewed": reviewed,
    })
    return result


def build_blank_or_skipped(ws: Worksheet, title: str, reason: str, result: dict) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    rows = [["Status", "Reason", "Generated At"], ["Skipped", reason, _dt.datetime.now().astimezone().isoformat()]]
    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result[title] = {"rows_written": len(rows), "status": "skipped", "reason": reason}
    return result


def build_related_niches(
    ws: Worksheet, related_json_path: str, cfg: dict, result: dict, warnings: list[str]
) -> dict:
    """Rebuild the related-niches tab from the POE related-niches JSON with the
    strict keep-list filter. Preserves the template's metadata + column-label
    header rows (everything above the niche data block)."""
    keep = [norm(x) for x in cfg["related_niche_filter"]["keep"]]
    exclude = [norm(x) for x in cfg["related_niche_filter"]["exclude_examples"]]

    data = json.load(open(related_json_path, encoding="utf-8"))
    # Two capture shapes are supported:
    #  (a) tables[0] = {headers, rows} (chrome table scrape; the scraper
    #      mis-labels the first niche row as `headers`, so real rows = headers + rows)
    #  (b) relatedNiches = [{cells, niche, href}, ...] (visible-structured capture)
    #  (c) relatedNiches = [{niche, href}, ...] (amazon-agent.poe-related-niches.v1)
    niche_rows: list = []
    tables = data.get("tables") or []
    if tables and isinstance(tables[0], dict) and (tables[0].get("rows") or tables[0].get("headers")):
        table = tables[0]
        if isinstance(table.get("headers"), list):
            niche_rows.append(table["headers"])
        niche_rows.extend(table.get("rows", []))
    elif isinstance(data.get("relatedNiches"), list):
        for item in data["relatedNiches"]:
            if not isinstance(item, dict):
                continue
            cells = item.get("cells")
            if isinstance(cells, list):
                niche_rows.append(cells)
            elif item.get("niche"):
                # {niche, href} shape: synthesize a row keyed on the niche name.
                niche_rows.append([item.get("niche"), item.get("href", "")])

    if not niche_rows:
        warnings.append(f"Related niches: no usable rows in {related_json_path}; tab left as-is.")
        result.update({"kept": [], "dropped": [], "rows_written": 0, "skipped": True})
        return result

    # The template's tab carries the column-label header in the rows above the
    # niche data block (freeze_panes marks the data start). Snapshot the whole
    # tab style, capture the header block, then rewrite niches under it.
    snap = snapshot_styles(ws, upto_row=8)
    freeze = ws.freeze_panes  # e.g. "A4" -> niche data starts at row 4
    header_rows_count = 3
    if freeze:
        m = re.match(r"[A-Z]+(\d+)", freeze)
        if m:
            header_rows_count = int(m.group(1)) - 1

    header_block = []
    for r in range(1, header_rows_count + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, (ws.max_column or 1) + 1)]
        header_block.append(row_vals)

    # Label rows that can repeat inside a scraped table; never niches.
    label_skip = {"niche details", "niche name", ""}
    kept, dropped = [], []
    kept_rows = []
    for nrow in niche_rows:
        name = norm(nrow[0]) if nrow else ""
        if name in label_skip:
            continue
        if name in keep:
            kept.append(nrow[0])
            kept_rows.append(nrow)
        else:
            dropped.append(nrow[0])

    # Validation safety: ensure no known drift example survived.
    survived_excluded = [k for k in kept if norm(k) in exclude]

    clear_rows(ws)
    out_rows = header_block + kept_rows
    write_matrix(ws, out_rows, snap)
    if freeze:
        ws.freeze_panes = freeze

    result.update(
        {
            "kept": kept,
            "dropped": dropped,
            "rows_written": len(out_rows),
            "niche_rows_written": len(kept_rows),
            "survived_excluded": survived_excluded,
            "skipped": False,
        }
    )
    return result


# --- Outlier triage ------------------------------------------------------- #
REASONS = {
    "Competitor/brand term": "Competitor or third-party brand token in the term.",
    "Wrong product form": "Term implies a non-powder form (capsule/drink/serum/etc.).",
    "Unsupported claim/health-risk": "Term implies a health/disease claim not supported for copy.",
    "Negative candidate": "Term is off-product/irrelevant to the product.",
    "Semantic opportunity": "Relevant term where the anchor product under-ranks vs competitors.",
    "Ignore": "No special signal; normal mainstream term in the master list.",
}
RECOMMENDED_USE = {
    "Competitor/brand term": "Keep for competitor/PPC research; do not use in copy.",
    "Wrong product form": "Add as negative; not relevant to powder form.",
    "Unsupported claim/health-risk": "Exclude from copy; review claim compliance.",
    "Negative candidate": "Add to negatives.",
    "Semantic opportunity": "Candidate for visible copy / backend / PPC push.",
    "Ignore": "Already covered by mainstream copy terms.",
}


def _token_hit(kw_norm: str, tokens: list[str], match: str) -> bool:
    for tok in tokens:
        t = norm(tok)
        if not t:
            continue
        if match == "word":
            if re.search(r"(?<!\w)" + re.escape(t) + r"(?!\w)", kw_norm):
                return True
        else:
            if t in kw_norm:
                return True
    return False


def classify_keyword(kw: str, relev, sv, anchor_rank, best_comp, tcfg: dict) -> str:
    kw_norm = norm(kw)
    match = tcfg.get("match", "word")
    if _token_hit(kw_norm, tcfg["brand_tokens"], match):
        return "Competitor/brand term"
    if _token_hit(kw_norm, tcfg["form_tokens"], match):
        return "Wrong product form"
    if _token_hit(kw_norm, tcfg["claim_tokens"], match):
        return "Unsupported claim/health-risk"
    if _token_hit(kw_norm, tcfg["negative_tokens"], match):
        return "Negative candidate"
    so = tcfg["semantic_opportunity"]
    include_if_not_ranking = so.get("include_if_anchor_not_ranking")
    rank_gap = so.get("anchor_rank_worse_than_best_competitor_by")
    relev = relev or 0.0
    sv = sv or 0
    if relev >= so["min_relevancy"] and sv >= so["min_search_volume"]:
        if anchor_rank is None and include_if_not_ranking:
            return "Semantic opportunity"
        if (
            anchor_rank is not None
            and best_comp is not None
            and rank_gap is not None
            and (anchor_rank - best_comp) >= rank_gap
        ):
            return "Semantic opportunity"
    return "Ignore"


def build_outlier(
    ws: Worksheet,
    master_rows: list[list[str]],
    expanded_rows: list[list[str]],
    poe_terms: set[str],
    cfg: dict,
    result: dict,
    warnings: list[str],
) -> dict:
    tcfg = cfg["triage"]
    action_map = cfg["final_action"]["by_classification"]
    allowed_actions = set(cfg["final_action"]["allowed"])
    anchor_asin = cfg["product_anchor"]["asin"]
    # The anchor rank column is labelled with the client/brand for readability,
    # generically falling back to "Anchor".
    anchor_label = (
        cfg["product_anchor"].get("client")
        or cfg["product_anchor"].get("brand")
        or "Anchor"
    )

    header = master_rows[0]
    # Locate key columns by header NAME (robust to a leading blank column or to
    # it being stripped — see drop_leading_empty_columns). ASIN rank columns are
    # found by their ASIN-shaped header (B0XXXXXXXX), skipping label columns.
    def _col(names: set[str], default: int, prefix: bool = False) -> int:
        for i, h in enumerate(header):
            hn = norm(h)
            if (prefix and hn.startswith(tuple(names))) or (not prefix and hn in names):
                return i
        return default

    kw_idx = _col({"search terms", "search term", "keyword", "keywords"}, 1)
    sv_idx = _col({"sv", "search volume"}, 2)
    rel_idx = _col({"relev"}, 3, prefix=True)
    asin_cols = {
        header[i]: i
        for i in range(len(header))
        if re.match(r"^B0[A-Z0-9]{8}$", str(header[i]).strip())
    }
    if anchor_asin not in asin_cols:
        warnings.append(f"Outlier triage: anchor ASIN {anchor_asin} not a master column.")
        anchor_idx = None
    else:
        anchor_idx = asin_cols[anchor_asin]
    # Same-brand non-anchor ASINs are siblings, not competitors. Excluding them
    # from the competitor set keeps "Best Competitor Rank" honest — a sibling
    # ranking well must not suppress a genuine opportunity (see config
    # asin_roles.siblings).
    sibling_set = {str(a).strip().upper() for a in (cfg.get("asin_roles", {}).get("siblings") or [])}
    comp_idxs = [
        i for a, i in asin_cols.items()
        if a != anchor_asin and str(a).strip().upper() not in sibling_set
    ]

    out_header = [
        "Keyword", "Search Volume", "DataDive Relevancy", f"{anchor_label} Rank",
        "Best Competitor Rank", "Competitors Top 50", "POE Signal",
        "Classification", "Reason", "Recommended Use", "Source", "Final Action",
    ]

    snap = snapshot_styles(ws, upto_row=4)
    rows_out = [out_header]
    by_class: dict[str, int] = {}
    by_action: dict[str, int] = {}
    bad_actions: list[str] = []

    core_terms_seen: set[str] = set()
    for row in master_rows[1:]:
        if not row or kw_idx >= len(row):
            continue
        kw = row[kw_idx]
        if not norm(kw):
            continue
        core_terms_seen.add(norm(kw))
        sv = to_int(row[sv_idx]) if sv_idx < len(row) else None
        relev = to_float(row[rel_idx]) if rel_idx < len(row) else None
        anchor_rank = to_int(row[anchor_idx]) if anchor_idx is not None and anchor_idx < len(row) else None
        comp_ranks = [to_int(row[i]) for i in comp_idxs if i < len(row)]
        comp_ranks = [r for r in comp_ranks if r is not None]
        best_comp = min(comp_ranks) if comp_ranks else None
        comps_top50 = sum(1 for r in comp_ranks if r <= 50)
        poe_signal = "In POE top 20" if norm(kw) in poe_terms else ""

        cls = classify_keyword(kw, relev, sv, anchor_rank, best_comp, tcfg)
        if cls == "Ignore":
            continue  # not an outlier; lives in the master list as normal copy

        action = action_map.get(cls, "Ignore")
        if action not in allowed_actions:
            bad_actions.append(f"{kw} -> {action}")

        rows_out.append(
            [
                kw, sv, relev, anchor_rank if anchor_rank is not None else "",
                best_comp if best_comp is not None else "", comps_top50, poe_signal,
                cls, REASONS.get(cls, ""), RECOMMENDED_USE.get(cls, ""),
                "DataDive Master List", action,
            ]
        )
        by_class[cls] = by_class.get(cls, 0) + 1
        by_action[action] = by_action.get(action, 0) + 1

    expanded = _keyword_rows_by_term(expanded_rows)
    for kw_norm, meta in expanded.items():
        if kw_norm in core_terms_seen:
            continue
        kw = meta["keyword"]
        sv = meta.get("sv")
        relev = meta.get("relevancy")
        cls = classify_keyword(kw, relev, sv, None, None, tcfg)
        if cls == "Ignore":
            continue
        action = action_map.get(cls, "Ignore")
        if action not in allowed_actions:
            bad_actions.append(f"{kw} -> {action}")
        poe_signal = "In POE top 20" if norm(kw) in poe_terms else "Expanded 1% only"
        rows_out.append(
            [
                kw, sv, relev, "", "", "", poe_signal,
                cls, REASONS.get(cls, ""), RECOMMENDED_USE.get(cls, ""),
                "DataDive Expanded MKL 1%", action,
            ]
        )
        by_class[cls] = by_class.get(cls, 0) + 1
        by_action[action] = by_action.get(action, 0) + 1

    # Sort outliers by classification group then descending search volume.
    order = {c: i for i, c in enumerate(tcfg["categories_order"])}
    body = rows_out[1:]
    body.sort(key=lambda r: (order.get(r[7], 99), -(r[1] or 0)))
    rows_out = [out_header] + body

    clear_rows(ws)
    write_matrix(ws, rows_out, snap)
    if not ws.freeze_panes:
        ws.freeze_panes = "A2"

    result.update(
        {
            "rows_written": len(rows_out),
            "outlier_keywords": len(rows_out) - 1,
            "by_classification": by_class,
            "by_final_action": by_action,
            "invalid_actions": bad_actions,
        }
    )
    return result


def build_seo_text(ws: Worksheet, seo_path: str, result: dict, warnings: list[str]) -> dict:
    """Rebuild the '4. SEO Text' tab from the curated seo_content.json, adding
    a DataDive Ranking Juice column. 'Current Listing Text' is sourced from the
    existing tab (the live listing copy) unless overridden in the JSON."""
    if not os.path.exists(seo_path):
        warnings.append(f"SEO content file not found ({seo_path}); SEO tab left as-is.")
        result.update({"skipped": True})
        return result
    seo = json.load(open(seo_path, encoding="utf-8"))

    snap = snapshot_styles(ws, upto_row=12)
    # Columns: Section | Old Listing | New Listing | Notes / Compliance.
    # Col A states what the row is (the section label) in its own column; Col B is
    # the current/old copy; Col C holds ONLY the publishable new copy (no notes);
    # Col D carries the compliance + Ranking Juice notes, kept out of the copy
    # column so the brand-token QA gate scans pure copy only.
    rows: list[list[Any]] = [["Section", "Old Listing", "New Listing", "Notes / Compliance"]]
    for item in seo["rows"]:
        section = item.get("section", "")
        cur = item.get("current") or ""
        new_cell = item.get("new", "")
        notes = []
        if item.get("compliance"):
            notes.append(f"Compliance: {item['compliance']}")
        if item.get("rj"):
            notes.append(f"Ranking Juice: {item['rj']}")
        rows.append([section, cur, new_cell, "\n".join(notes)])

    clear_rows(ws)
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, value=val)
    ws.freeze_panes = "A2"

    from openpyxl.styles import Alignment, Border, Side, Font

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 50

    # Deterministic, UNIFORM styling. Do NOT bleed the template's heterogeneous
    # per-row styles onto the content — that is what put stray bold + random
    # horizontal separator lines on arbitrary rows (e.g. Bullet 5, Backend).
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_align = Alignment(wrap_text=True, vertical="top")

    # Header row 1: reuse the template's captured header font+fill (brand accent),
    # forced bold + a clean box border.
    hdr_snap = snap[0] if snap else []
    for c in range(1, 5):
        cell = ws.cell(1, c)
        st = hdr_snap[min(c, len(hdr_snap)) - 1] if hdr_snap else None
        if st:
            f = copy.copy(st["font"]); f.bold = True; cell.font = f
            cell.fill = copy.copy(st["fill"])
        else:
            cell.font = Font(bold=True)
        cell.border = box
        cell.alignment = top_align

    # Data rows 2..N: ONE consistent style. Section label (col A) bold for
    # scannability; the copy columns regular. Uniform box border everywhere —
    # no random separators, no lone bold bullet.
    data_font = None
    for probe in (4, 3, 2):  # a clean template data row → font family/size/color
        if len(snap) >= probe and snap[probe - 1]:
            data_font = snap[probe - 1][min(2, len(snap[probe - 1])) - 1]["font"]
            break
    for r in range(2, len(rows) + 1):
        for c in range(1, 5):
            cell = ws.cell(r, c)
            f = copy.copy(data_font) if data_font is not None else copy.copy(cell.font)
            f.bold = (c == 1)
            cell.font = f
            cell.border = box
            cell.alignment = top_align

    result.update({"rows_written": len(rows), "skipped": False})
    return result


def build_seo_variations(wb, seo_path: str, result: dict, warnings: list[str]) -> dict:
    """Optional 'SEO Variations' tab — one column per supply/pack size for the
    fields that differ by variation (e.g. Item Highlights, Bullet 5). Driven by
    seo_content['supply_variations']; if that block is absent the tab is skipped
    entirely (other clients/products are unaffected)."""
    if not os.path.exists(seo_path):
        result.update({"skipped": True})
        return result
    seo = json.load(open(seo_path, encoding="utf-8"))
    sv = seo.get("supply_variations") or {}
    labels = sv.get("labels") or []
    var_rows = sv.get("rows") or []
    if not labels or not var_rows:
        result.update({"skipped": True})
        return result

    ws = ensure_sheet(wb, "SEO Variations", after="4. SEO Text")
    clear_rows(ws)
    matrix: list[list[Any]] = [["Field"] + list(labels)]
    for row in var_rows:
        vals = row.get("values", {})
        matrix.append([row.get("field", "")] + [vals.get(l, "") for l in labels])
    if sv.get("shared_note"):
        matrix.append(["Shared fields"] + [sv["shared_note"]] + [""] * (len(labels) - 1))
    for i, r in enumerate(matrix, start=1):
        for j, v in enumerate(r, start=1):
            ws.cell(i, j, value=v)

    from openpyxl.styles import Alignment, Border, Side, Font
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    ta = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 22
    for c in range(2, len(labels) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 62
    for r in range(1, len(matrix) + 1):
        for c in range(1, len(labels) + 2):
            cell = ws.cell(r, c)
            cell.font = Font(bold=(r == 1 or c == 1))
            cell.border = box
            cell.alignment = ta
    result.update({"rows_written": len(matrix), "skipped": False})
    return result


def load_poe_terms(poe_search_terms_csv: str) -> set[str]:
    """Top search terms from the POE search-terms CSV (data under the row-5
    header). Used to flag 'In POE top 20' on the outlier tab."""
    rows = read_csv(poe_search_terms_csv)
    terms: set[str] = set()
    header_idx = None
    for i, r in enumerate(rows):
        if r and norm(r[0]) == "search term":
            header_idx = i
            break
    if header_idx is None:
        return terms
    for r in rows[header_idx + 1 :]:
        if r and norm(r[0]):
            terms.add(norm(r[0]))
    return terms


def load_json(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _capture_text(data: Any) -> str:
    if isinstance(data, str):
        return data
    if not isinstance(data, dict):
        return ""
    parts: list[str] = []
    for key in ("text", "rawText", "content"):
        if data.get(key):
            parts.append(str(data[key]))
    visible = data.get("visibleRoute")
    if isinstance(visible, dict):
        parts.append(_capture_text(visible))
    return "\n".join(parts)


def _first_table_rows(data: dict) -> tuple[list[str], list[list[Any]]]:
    tables = data.get("tables") or []
    if not tables or not isinstance(tables[0], dict):
        return [], []
    table = tables[0]
    headers = table.get("headers") or []
    rows = table.get("rows") or []
    return headers, rows


def _extract_review_topic_rows(data: dict, source: str) -> list[list[Any]]:
    headers, table_rows = _first_table_rows(data)
    if table_rows:
        out = []
        for row in table_rows:
            topic = row[0] if row else ""
            pct = row[2] if len(row) > 2 else row[1] if len(row) > 1 else ""
            snippets = " | ".join(str(v) for v in row[3:] if v)
            out.append([topic, pct, snippets, "POE Customer Review Insights", source])
        return out

    text = _capture_text(data)
    if not text:
        return []
    section = text
    # Section start markers — English UI and the German Seller Central UI
    # (EU accounts whose display language is Deutsch render the review-insights
    # table header as "Thema Unterthema %Erwähnungen Snippets überprüfen").
    start_markers = [
        "Topic Subtopic %Mentions Review Snippets",
        "Thema Unterthema %Erwähnungen Snippets überprüfen",
    ]
    for sm in start_markers:
        idx = section.find(sm)
        if idx >= 0:
            section = section[idx + len(sm) :]
            break
    end_markers = [
        "Topic Impact on Star Rating", "Customer Review Topic Mentions Trends", "Disclaimer:",
        "Auswirkungen des Themas auf die Sternebewertung",
        "Kundenrezensionen, Themen, Erwähnungen", "Haftungsausschluss:",
    ]
    end_positions = [section.find(m) for m in end_markers if section.find(m) >= 0]
    if end_positions:
        section = section[: min(end_positions)]
    section = re.sub(r"\s+", " ", section).strip()

    # Percentage accepts dot OR comma decimal and an optional space before '%'
    # (EU number format renders "33,32 %"); topic text excludes digits so the
    # leading topic token can't swallow the previous row's percentage.
    pattern = re.compile(
        r"(?P<topic>[A-ZÀ-Ýa-zà-ÿ][^\"%0-9]{2,90}?)\s+"
        r"(?P<pct>\d+(?:[.,]\d+)?\s*%)\s+"
        r"(?P<snips>\"[^\"]+\"(?:,\s*\"[^\"]+\")*)"
    )
    out: list[list[Any]] = []
    for m in pattern.finditer(section):
        topic = re.sub(r"\s+", " ", m.group("topic")).strip(" -:")
        topic = topic.replace("Topic Subtopic %Mentions Review Snippets", "").strip()
        snippets = m.group("snips").replace('", "', '" | "')
        out.append([topic, m.group("pct"), snippets, "POE Customer Review Insights", source])
    return out


def build_poe_reviews(ws: Worksheet, path: str, result: dict, warnings: list[str]) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    header = ["Topic", "% Mentions", "Review Snippets", "Evidence Type", "Source File"]
    if not path or not os.path.exists(path):
        rows = [header, ["Missing POE review evidence", "", "", "Warning", path]]
        warnings.append("POE Raw - Reviews: review JSON missing; wrote warning row.")
        status = "missing"
    else:
        data = load_json(path)
        topic_rows = _extract_review_topic_rows(data, path)
        if topic_rows:
            rows = [header] + topic_rows
            status = "parsed"
        else:
            rows = [header, ["POE review evidence captured but no topic rows parsed", "", "", "Warning", path]]
            warnings.append(f"POE Raw - Reviews: no topic rows parsed from {path}; wrote warning row.")
            status = "unparsed"
    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({"rows_written": len(rows), "topics": max(0, len(rows) - 1), "status": status, "path": path})
    return result


def _extract_returns_rows(data: dict, source: str) -> list[list[Any]]:
    headers, table_rows = _first_table_rows(data)
    if table_rows:
        return [[*(row[:4]), source] for row in table_rows]
    visible = data.get("visibleRoute") if isinstance(data, dict) else None
    text = _capture_text(visible if isinstance(visible, dict) else data)
    if not text or "Returns" not in text or "Customer Review Insights" in text and "Return" not in text.split("Customer Review Insights")[-1]:
        return []
    section = re.sub(r"\s+", " ", text).strip()
    # Bound to the topic table: starts after the 'Topic %Mentions' header inside
    # 'Product Returns Insights', ends before the 'Topic Returns Trend' chart
    # (whose axis ticks like '0% 20% 40%' would otherwise parse as topics).
    if "Product Returns Insights" in section:
        section = section.split("Product Returns Insights", 1)[1]
    if "Topic Returns Trend" in section:
        section = section.split("Topic Returns Trend", 1)[0]
    if "%Mentions" in section:
        section = section.split("%Mentions", 1)[1]
    pattern = re.compile(r"(?P<topic>[^%]+?)\s+(?P<pct>\d{1,3}(?:[.,]\d+)?%)")
    return [
        [m.group("topic").strip(), m.group("pct"), "", "POE Returns", source]
        for m in pattern.finditer(section)
        if m.group("topic").strip()
    ]


def build_poe_returns(ws: Worksheet, path: str, result: dict, warnings: list[str]) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    header = ["Topic / Status", "% Mentions", "Detail", "Evidence Type", "Source File"]
    if not path or not os.path.exists(path):
        rows = [header, ["Missing POE returns evidence", "", "", "Warning", path]]
        warnings.append("POE Raw - Returns: returns JSON missing; wrote warning row.")
        status = "missing"
    else:
        data = load_json(path)
        return_rows = _extract_returns_rows(data, path)
        if return_rows:
            rows = [header] + return_rows
            status = "parsed"
        else:
            rows = [
                header,
                [
                    "Returns not exposed for this POE niche",
                    "",
                    "Seller Central returns route was captured, but no returns topics/table was visible for this niche.",
                    "Explicit not-exposed warning",
                    path,
                ],
            ]
            status = "not_exposed"
    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({"rows_written": len(rows), "returns_rows": max(0, len(rows) - 1), "status": status, "path": path})
    return result


def _poe_search_term_rows(path: str, limit: int = 12) -> list[list[str]]:
    rows = read_csv(path)
    header_idx = None
    for i, row in enumerate(rows):
        if row and norm(row[0]) == "search term":
            header_idx = i
            break
    if header_idx is None:
        return []
    header = rows[header_idx]
    body = rows[header_idx + 1 :]
    out = []
    for row in body:
        if not row or not norm(row[0]):
            continue
        row = row + [""] * max(0, len(header) - len(row))
        out.append(row)
        if len(out) >= limit:
            break
    return out


def _overview_metrics(data: dict) -> list[tuple[str, str]]:
    text = _capture_text(data.get("overview") if isinstance(data.get("overview"), dict) else data)
    metrics = []
    patterns = [
        ("Search Volume", r"Search Volume \(Past 360 days\)\s+([0-9,.]+)"),
        ("Search volume growth", r"Search volume growth \(Past 180 days\)\s+([+\-0-9,.%]+)"),
        ("Top clicked products", r"No\. of top clicked products\s+([0-9,.]+)"),
        ("Average price", r"Average price \(Past 360 days\)\s+([€$£0-9,.]+)"),
        ("Units sold range", r"Range of average units sold \(Past 360 days\)\s+([0-9,.]+\s+-\s+[0-9,.]+)"),
        ("Return rate", r"Return Rate \(Past 360 days\)\s+([0-9,.%]+)"),
    ]
    for label, pattern in patterns:
        m = re.search(pattern, text)
        if m:
            metrics.append((label, m.group(1)))
    return metrics


def build_semantic_insights(
    ws: Worksheet,
    args: dict,
    reviews_result: dict,
    returns_result: dict,
    related_result: dict,
    result: dict,
    warnings: list[str],
) -> dict:
    snap = snapshot_styles(ws, upto_row=4)
    header = ["Priority", "Insight Type", "Signal", "Value", "SEO / Listing Action", "Source"]
    rows: list[list[Any]] = [header]

    structured = {}
    if args.get("poe_structured_json") and os.path.exists(args["poe_structured_json"]):
        structured = load_json(args["poe_structured_json"])
        for label, value in _overview_metrics(structured):
            rows.append([1, "POE overview", label, value, "Use as niche context; do not treat as listing copy.", args["poe_structured_json"]])

    for i, row in enumerate(_poe_search_term_rows(args["poe_search_terms_csv"], limit=12), start=1):
        term = row[0]
        sv = row[1] if len(row) > 1 else ""
        conv = row[5] if len(row) > 5 else ""
        priority = 1 if i <= 6 else 2
        action = "Review for visible copy/backend coverage after relevance and claim check."
        rows.append([priority, "POE search term", term, f"SV {sv}; conversion {conv}", action, args["poe_search_terms_csv"]])

    if args.get("poe_reviews_json") and os.path.exists(args["poe_reviews_json"]):
        review_rows = _extract_review_topic_rows(load_json(args["poe_reviews_json"]), args["poe_reviews_json"])
        for topic, pct, snippets, _etype, source in review_rows[:10]:
            action = "Use as customer-language input; rewrite into compliant product/benefit wording."
            rows.append([1, "Review topic", topic, pct, action, source])
    else:
        rows.append([1, "Review topic", "Missing POE review evidence", "", "Do not finalize semantic insights until review capture exists.", args.get("poe_reviews_json", "")])

    if returns_result.get("status") == "not_exposed":
        rows.append([2, "Returns", "Returns not exposed", "", "Record as a source limitation; no return-topic SEO action.", returns_result.get("path", "")])
    elif returns_result.get("status") == "parsed":
        rows.append([1, "Returns", "Return topics captured", f"{returns_result.get('returns_rows', 0)} rows", "Review for product-experience objections.", returns_result.get("path", "")])

    for niche in related_result.get("kept", [])[:8]:
        rows.append([2, "Related niche", niche, "", "Use only if it matches the actual product form and ingredient.", args["related_niches_json"]])

    clear_rows(ws)
    write_matrix(ws, rows, snap)
    ws.freeze_panes = "A2"
    result.update({"rows_written": len(rows), "status": "rebuilt", "sources": {
        "poe_structured_json": args.get("poe_structured_json", ""),
        "poe_reviews_json": args.get("poe_reviews_json", ""),
        "poe_returns_json": args.get("poe_returns_json", ""),
    }})
    return result


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #
def excel_name_ok(name: str) -> bool:
    if not name or len(name) > 31:
        return False
    if any(ch in name for ch in "[]:*?/\\"):
        return False
    return True


_RJ_TOKEN_RE = re.compile(r"[a-z0-9äöüß%]+")


def _seo_copy_by_section(wb) -> tuple[dict, str | None]:
    """{section_norm: new_copy} from the SEO Text tab (col A label, col C copy)."""
    seo_tab = next((s for s in ("4.1 SEO Text", "4. SEO Text") if s in wb.sheetnames), None)
    out: dict[str, str] = {}
    if not seo_tab:
        return out, None
    for row in wb[seo_tab].iter_rows(min_row=2, values_only=True):
        sec = norm(row[0] if row and row[0] is not None else "")
        new = str(row[2]).strip() if (len(row) > 2 and row[2] is not None) else ""
        if sec and sec not in out:
            out[sec] = new
    return out, seo_tab


def _resolve_master_sheet(wb, preferred: str) -> str | None:
    """The Core 30% MKL tab — config-named (preferred) or, across template schemes,
    a tab matching MKL+30% or 'Master List' (e.g. '3.1 MKL DataDive 30%' vs
    '3.1. Master List DataDive')."""
    if preferred and preferred in wb.sheetnames:
        return preferred
    for s in wb.sheetnames:
        sl = s.lower()
        if ("mkl" in sl and "30" in sl) or "master list" in sl:
            return s
    return None


def _master_sv_map(wb, master_sheet: str) -> dict:
    """{keyword_norm: sv_int} read from the built master MKL tab (keyword + SV cols
    located by header name, mirroring the outlier-tab logic)."""
    out: dict[str, int] = {}
    sheet = _resolve_master_sheet(wb, master_sheet)
    if not sheet:
        return out
    rows = list(wb[sheet].iter_rows(values_only=True))
    if not rows:
        return out
    header = [norm(h) for h in rows[0]]

    def col(names: set, default: int) -> int:
        for i, h in enumerate(header):
            if h in names:
                return i
        return default

    kw_i = col({"search terms", "search term", "keyword", "keywords"}, 1)
    sv_i = col({"sv", "search volume"}, 2)
    for r in rows[1:]:
        if not r or kw_i >= len(r):
            continue
        kw = norm(r[kw_i])
        if not kw:
            continue
        out[kw] = (to_int(r[sv_i]) if sv_i < len(r) else None) or 0
    return out


def _rj_tokens(text: str) -> set:
    return set(_RJ_TOKEN_RE.findall(norm(text)))


def _kw_is_covered(kw_norm: str, copy_tokens: set) -> bool:
    """A keyword is covered if every token appears in the copy (exact, or a >=4-char
    stem prefix match — a light German-stem approximation, e.g. ballaststoff(e))."""
    for t in _RJ_TOKEN_RE.findall(kw_norm):
        if t in copy_tokens:
            continue
        if len(t) >= 4 and any(ct.startswith(t) or t.startswith(ct) for ct in copy_tokens if len(ct) >= 4):
            continue
        return False
    return True


def compute_rj_coverage(master_sv: dict, copy_text: str, exclude_tokens: list) -> dict:
    """Ranking-Juice coverage of the publishable copy against the MKL.
    'addressable' excludes keywords matching the triage brand/form/negative tokens
    (competitor brands + off-product forms routed to PPC/Negative, not copy)."""
    copy_tokens = _rj_tokens(copy_text)
    excl = [norm(t) for t in exclude_tokens if norm(t)]
    total = covered = addr = covered_addr = 0
    for kw, sv in master_sv.items():
        is_excl = _token_hit(kw, excl, "word") if excl else False
        is_cov = _kw_is_covered(kw, copy_tokens)
        total += sv
        if not is_excl:
            addr += sv
        if is_cov:
            covered += sv
            if not is_excl:
                covered_addr += sv

    def pct(a, b):
        return round(100.0 * a / b, 1) if b else 0.0

    return {
        "total": total, "covered": covered, "covered_pct": pct(covered, total),
        "addressable": addr, "covered_addressable": covered_addr,
        "covered_addressable_pct": pct(covered_addr, addr),
    }


def run_validations(wb, cfg, counts, related_result, paths, warnings) -> list[dict]:
    checks: list[dict] = []

    def add(name, ok, detail=""):
        checks.append({"check": name, "pass": bool(ok), "detail": detail})

    anchor = cfg["product_anchor"]["asin"]

    # Sheet titles for the exact-paste DataDive MKL tabs come from the config
    # (source of truth via tabs.exact_paste), not hardcoded — so validation stays
    # aligned with whatever the config names them (e.g. '3.1. Master List DataDive').
    exact_paste = cfg["tabs"]["exact_paste"]
    sheet_for = {src: sheet for sheet, src in exact_paste.items()}
    master_sheet = sheet_for.get("master_csv", "")
    expanded_sheet = sheet_for.get("expanded_mkl_csv", "")

    # 1. Product anchor is the configured ASIN and present as a master-list
    #    column (authoritative — it's a DataDive ASIN column). The ASINs tab is
    #    a curated carry-forward tab; if the anchor isn't in it, that tab is a
    #    template placeholder to curate for this market (warning, not failure).
    master_header = counts["_master_header"]
    in_master = anchor in master_header
    asins_tab_hit = False
    if "ASINs" in wb.sheetnames:
        ws = wb["ASINs"]
        for row in ws.iter_rows(values_only=True):
            if any(norm(v) == norm(anchor) for v in row):
                asins_tab_hit = True
                break
    add(
        f"Product anchor is configured ASIN {anchor}",
        bool(anchor) and in_master,
        f"config={anchor}, in_master={in_master}, in_ASINs_tab={asins_tab_hit}",
    )
    if not asins_tab_hit:
        warnings.append(
            f"Anchor {anchor} not in the carried-forward ASINs tab — curate the "
            f"ASINs / curated tabs for this market (currently template placeholders)."
        )

    # 2. Exact DataDive root/master row counts match source CSVs.
    add(
        "Root Keywords rows == roots CSV rows",
        counts["1. Root Keywords"] == counts["_roots_csv"],
        f"sheet={counts['1. Root Keywords']} csv={counts['_roots_csv']}",
    )
    add(
        "Master List rows == master CSV rows",
        counts.get(master_sheet) == counts["_master_csv"],
        f"sheet={counts.get(master_sheet)} csv={counts['_master_csv']}",
    )
    if expanded_sheet in wb.sheetnames:
        add(
            "Expanded MKL rows == 1% CSV rows",
            counts.get(expanded_sheet) == counts.get("_expanded_mkl_csv"),
            f"sheet={counts.get(expanded_sheet)} csv={counts.get('_expanded_mkl_csv')}",
        )
    add(
        "Core and Expanded MKL source paths are distinct",
        os.path.abspath(paths.get("master_csv", "")) != os.path.abspath(paths.get("expanded_mkl_csv", "")),
        f"core={paths.get('master_csv', '')} expanded={paths.get('expanded_mkl_csv', '')}",
    )

    metadata = cfg.get("datadive_exports") or {}
    missing_meta = []
    for group in ("core_mkl", "expanded_mkl"):
        item = metadata.get(group) or {}
        for field in ("min_relevancy", "visible_keyword_count", "visible_search_volume", "export_timestamp", "niche_id", "marketplace", "hero_keyword"):
            value = str(item.get(field, "")).strip()
            if not value or value.startswith("TO_RECORD"):
                missing_meta.append(f"{group}.{field}")
    add(
        "DataDive export metadata present for core and expanded MKL",
        not missing_meta,
        f"missing={missing_meta}",
    )

    # 3. Sheet names Excel/Google Sheets compatible (+ unique).
    names = wb.sheetnames
    bad = [n for n in names if not excel_name_ok(n)]
    dupes = len(names) != len(set(names))
    add(
        "Sheet names Excel/Sheets compatible",
        not bad and not dupes,
        f"bad={bad} duplicates={dupes}",
    )

    # 4. Related-niche filter excludes known irrelevant examples.
    excl = [norm(x) for x in cfg["related_niche_filter"]["exclude_examples"]]
    kept_norm = [norm(x) for x in related_result.get("kept", [])]
    leaked = [k for k in kept_norm if k in excl]
    keepset = [norm(x) for x in cfg["related_niche_filter"]["keep"]]
    non_keep = [k for k in kept_norm if k not in keepset]
    # A skipped related-niche build (POE exposed no related-niches grid → tab
    # handled via generated_blank) has zero kept rows and therefore no drift to
    # catch; the drift gate is satisfied. The gate only fails on real drift:
    # an excluded example or a non-keeplist niche that survived the filter.
    add(
        "Related-niche filter excludes drift",
        not leaked and not non_keep,
        f"leaked={leaked} non_keeplist={non_keep} "
        f"kept={len(kept_norm)} dropped={len(related_result.get('dropped', []))} "
        f"skipped={bool(related_result.get('skipped'))}",
    )

    # 5. POE products/search-terms sheet rows match source CSVs.
    add(
        "POE Products rows == products CSV rows",
        counts["POE Raw - Products"] == counts["_poe_products_csv"],
        f"sheet={counts['POE Raw - Products']} csv={counts['_poe_products_csv']}",
    )
    add(
        "POE Search Terms rows == search-terms CSV rows",
        counts["POE Raw - Search Terms"] == counts["_poe_search_terms_csv"],
        f"sheet={counts['POE Raw - Search Terms']} csv={counts['_poe_search_terms_csv']}",
    )

    # Bonus: every emitted Final Action is in the allowed enum.
    add(
        "Outlier Final Action values valid",
        not counts.get("_invalid_actions"),
        f"invalid={counts.get('_invalid_actions')}",
    )
    # Template's '2.2 Never KWs' columns: Negative Marker | Frequency | Negative Phrase.
    # Sectioned Never-KWs tab: word rows (Never Ever / brand / claim / review
    # band) must be single words; multi-word entries are allowed only in the
    # phrase-candidates section. Every data row must carry Category + Why.
    never_bad = []
    never_sheet = next((s for s in ("2 Never KWs", "2.2 Never KWs") if s in wb.sheetnames), None)
    never_rows = 0
    never_sections = 0
    in_phrase_section = False
    if never_sheet:
        ws = wb[never_sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            col_a = norm(row[0] if row else "")
            if col_a.startswith("section:"):
                never_sections += 1
                in_phrase_section = "phrase" in col_a
                continue
            phrase = norm(row[2] if row and len(row) > 2 else "")
            if not phrase:
                continue
            never_rows += 1
            if len(phrase.split()) != 1 and not in_phrase_section:
                never_bad.append(f"multiword:{phrase}")
            if len(row) < 5 or not norm(row[3]) or not norm(row[4]):
                never_bad.append(f"missing-category/why:{phrase}")
    add(
        "Never KWs tab is sectioned; word rows are single words with Category+Why",
        bool(never_sheet) and (never_sections >= 1 or never_rows == 0) and not never_bad,
        f"sheet={never_sheet!r} sections={never_sections} rows={never_rows} bad={never_bad[:12]} total={len(never_bad)}",
    )

    # Defensive: every tab declared in tabs.rebuild must have actually been rebuilt
    # this run (i.e. have a counts entry). Catches sheet-name mismatches that would
    # otherwise silently leave the style-template product's content in place
    # (the collagen-SEO-in-a-fibre-workbook class of bug).
    rebuild_tabs = cfg["tabs"].get("rebuild", [])
    not_rebuilt = [t for t in rebuild_tabs if t not in counts]
    add(
        "All configured rebuild tabs were rebuilt",
        not not_rebuilt,
        f"not_rebuilt={not_rebuilt}",
    )

    guard = cfg.get("stale_data_guard") or {}
    forbidden = [norm(x) for x in guard.get("forbidden_terms", []) if norm(x)]
    tabs = guard.get("tabs", [])
    leaks: list[str] = []
    for tab in tabs:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for row in ws.iter_rows(values_only=True):
            row_text = norm(" ".join(str(v) for v in row if v is not None))
            for term in forbidden:
                if term and term in row_text:
                    leaks.append(f"{tab}: {term}")
    if forbidden:
        add(
            "No stale forbidden terms in guarded tabs",
            not leaks,
            f"leaks={leaks[:12]} total={len(leaks)}",
        )

    required_tabs = cfg.get("required_current_tabs") or []
    missing_current = []
    for tab in required_tabs:
        if counts.get(tab, 0) <= 1:
            missing_current.append(tab)
    if required_tabs:
        add(
            "Required current-source tabs populated",
            not missing_current,
            f"missing_or_warning_only={missing_current}",
        )

    # Competitor brand names must never reach publishable SEO copy (title,
    # item highlights, bullets, description, backend). Scans the New Listing
    # column (col C) of the SEO Text tab for triage.brand_tokens; the own brand
    # and the Notes/annotation rows (Guardrail, Ranking Juice, Semantic
    # direction) are exempt.
    brand_tokens = [
        t for t in (cfg.get("triage", {}).get("brand_tokens") or [])
        if norm(t) and norm(t) != norm(cfg["product_anchor"].get("client", ""))
    ]
    seo_tab = next((s for s in ("4.1 SEO Text", "4. SEO Text") if s in wb.sheetnames), None)
    if brand_tokens and seo_tab:
        copy_hits: list[str] = []
        for row in wb[seo_tab].iter_rows(min_row=2, values_only=True):
            # Col A = section label (own column); Col C = publishable new copy.
            # Match by prefix so transition variants ("Title (≤75 char …)") and
            # the new "Item Highlights" field are still scanned.
            section = norm(row[0] if row and row[0] is not None else "")
            is_copy = (
                section.startswith("title")
                or section.startswith("item highlight")
                or section.startswith("bullet")
                or section.startswith("description")
                or "backend search terms" in section
            )
            if not is_copy:
                continue
            new_copy = norm(row[2] if len(row) > 2 and row[2] is not None else "")
            for tok in brand_tokens:
                if _token_hit(new_copy, [tok], "word"):
                    copy_hits.append(f"{section}: {tok}")
        add(
            "No competitor brand tokens in SEO copy",
            not copy_hits,
            f"sheet={seo_tab!r} hits={copy_hits[:8]} tokens_checked={len(brand_tokens)}",
        )

    # --- Product-facts intake + title (root-vs-tracked) + RJ + semantic gates ---
    pf = cfg.get("product_facts")
    if pf is None:
        warnings.append(
            "No product_facts block in config — add it (see config.TEMPLATE.json) so the "
            "title framing (blend vs single) and compliance are gated."
        )
    else:
        ings = [i for i in (pf.get("ingredients") or []) if norm(i.get("name"))]
        blend = norm(pf.get("blend_or_single"))
        certs = pf.get("certifications") or {}
        add(
            "Product facts present (ingredients + blend flag + certifications)",
            bool(ings) and blend in {"blend", "single"} and isinstance(certs, dict) and bool(certs),
            f"ingredients={len(ings)} blend_or_single={blend!r} certifications={list(certs)}",
        )

    seo_by_section, seo_tab2 = _seo_copy_by_section(wb)
    master_sv = _master_sv_map(wb, master_sheet)

    if seo_tab2 and master_sv:
        title_text = next(
            (v for k, v in seo_by_section.items() if k.startswith(("title (≤75", "title (<=75"))),
            next((v for k, v in seo_by_section.items() if k.startswith("title")), ""),
        )
        client_tokens = _rj_tokens(cfg["product_anchor"].get("client", ""))
        title_tokens = _rj_tokens(title_text)
        covered_kws = sorted(
            (kw for kw in master_sv if _kw_is_covered(kw, title_tokens)),
            key=lambda k: master_sv[k], reverse=True,
        )
        mkl_vocab: set = set()
        for kw in master_sv:
            mkl_vocab |= _rj_tokens(kw)
        lead_tokens = [t for t in _RJ_TOKEN_RE.findall(norm(title_text)) if t not in client_tokens]
        lead = lead_tokens[0] if lead_tokens else ""
        lead_tracked = bool(lead) and (
            lead in mkl_vocab
            or any(len(lead) >= 4 and (v.startswith(lead) or lead.startswith(v)) for v in mkl_vocab if len(v) >= 4)
        )
        add(
            "Title leads with a tracked MKL keyword (not a root)",
            bool(covered_kws) and lead_tracked,
            f"lead={lead!r} lead_in_MKL={lead_tracked} title_covers_kws={len(covered_kws)} top={covered_kws[:3]}",
        )

        # RJ coverage of the full publishable copy — informational (covered SV > 0);
        # the numbers print in the validation/manifest output, NOT a new workbook tab.
        publishable = " ".join(
            v for k, v in seo_by_section.items()
            if k.startswith(("title", "item highlight", "bullet", "description")) or "backend search terms" in k
        )
        tri = cfg.get("triage", {})
        excl_tokens = (tri.get("brand_tokens") or []) + (tri.get("form_tokens") or []) + (tri.get("negative_tokens") or [])
        rj = compute_rj_coverage(master_sv, publishable, excl_tokens)
        add(
            "Ranking-Juice coverage computed (covered SV > 0)",
            rj["covered"] > 0,
            f"covered={rj['covered']}/{rj['total']} ({rj['covered_pct']}% total) | "
            f"addressable {rj['covered_addressable']}/{rj['addressable']} ({rj['covered_addressable_pct']}%)",
        )
        if rj["covered_addressable_pct"] < 60:
            warnings.append(
                f"Ranking-Juice addressable coverage is {rj['covered_addressable_pct']}% (<60%) — "
                f"check the title/Item Highlights front-load the head cluster."
            )

        # Compliance tax: SV of addressable keywords gated by claim tokens —
        # visible copy can't cover them compliantly, so they deflate the
        # addressable % silently unless reported here. Informational row;
        # recovery is via the rewrite ladder (health-claims-compliance.md).
        claim_toks = tri.get("claim_tokens") or []
        if claim_toks:
            rj_ex_claims = compute_rj_coverage(master_sv, publishable, excl_tokens + claim_toks)
            claim_gated = rj["addressable"] - rj_ex_claims["addressable"]
            pct_gated = round(100.0 * claim_gated / rj["addressable"], 1) if rj["addressable"] else 0.0
            add(
                "Compliance tax computed (claim-gated SV reported)",
                True,
                f"claim-gated {claim_gated} of addressable {rj['addressable']} SV ({pct_gated}%) | "
                f"addressable excl. claims: {rj_ex_claims['covered_addressable']}/{rj_ex_claims['addressable']} "
                f"({rj_ex_claims['covered_addressable_pct']}%)",
            )

        # Blend title guard: a blend title shouldn't LEAD with one ingredient name.
        if pf and norm(pf.get("blend_or_single")) == "blend" and lead:
            led_by_ing = next(
                (norm(i.get("name")) for i in (pf.get("ingredients") or [])
                 if norm(i.get("name")) and lead in _RJ_TOKEN_RE.findall(norm(i.get("name")))),
                None,
            )
            if led_by_ing:
                warnings.append(
                    f"Blend title appears to lead with a single ingredient ({led_by_ing!r}) — use a generic "
                    f"blend signal (-Komplex/-Mix) and move ingredient names to Item Highlights."
                )

    if seo_tab2:
        sem = next((v for k, v in seo_by_section.items() if k.startswith("semantic")), None)
        add(
            "Semantic / Alexa AI direction present",
            bool(sem and sem.strip()),
            f"present={sem is not None} non_empty={bool(sem and sem.strip())}",
        )

    # Regulated-category nudge: the health-claims self-check must have run
    # before delivery (skills/amazon-seo/references/health-claims-compliance.md).
    comp = cfg.get("compliance") or {}
    if norm(comp.get("category_tier")) == "regulated" and not comp.get("checked"):
        warnings.append(
            "Regulated category: health-claims self-check not recorded — run "
            "/health-claims-check for this listing and set compliance.checked=true."
        )

    return checks


def build_curated_tabs(ws_book, path: str, result: dict, warnings: list[str]) -> dict:
    """Write curated content into named tabs from a JSON file:
    { "<Tab Name>": {"header": [...], "rows": [[...], ...]}, ... }.
    Used for POE Raw - Reviews/Returns/Semantic Insights on a new-market build so
    they carry real curated content (from POE captures) instead of a placeholder.
    Header style is preserved from the template tab."""
    written: list[str] = []
    if not path or not os.path.exists(path):
        result["tabs"] = written
        return result
    data = json.load(open(path, encoding="utf-8"))
    for tab, spec in data.items():
        if tab.startswith("_"):
            continue
        if tab not in ws_book.sheetnames:
            warnings.append(f"Curated tab '{tab}' not in template; skipped.")
            continue
        ws = ws_book[tab]
        snap = snapshot_styles(ws, upto_row=2)
        rows = [spec.get("header", [])] + spec.get("rows", [])
        clear_rows(ws)
        write_matrix(ws, rows, snap)
        if not ws.freeze_panes:
            ws.freeze_panes = "A2"
        written.append(tab)
    result["tabs"] = written
    return result


def clear_placeholder_tab(ws: Worksheet, note: str) -> None:
    """Blank a product-specific curated tab down to its header row + a note.
    Used when building a DIFFERENT product than the style template, so the new
    workbook never ships another product's curated content (e.g. collagen
    reviews in a fibre workbook). Keeps header styling and column widths."""
    snap = snapshot_styles(ws, upto_row=2)
    header = [ws.cell(1, c).value for c in range(1, (ws.max_column or 1) + 1)]
    clear_rows(ws)
    write_matrix(ws, [header], snap)
    cell = ws.cell(2, 1, value=note)
    apply_style(cell, snap, 2, 1)


def write_handoff_note(path: str, cfg: dict, args: dict, manifest: dict) -> str:
    if not path:
        return ""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    pa = cfg.get("product_anchor", {})
    lines = [
        "---",
        f"created: {_dt.date.today().isoformat()}",
        f"client: {pa.get('client', '')}",
        "project: Amazon SEO Keyword Workbook",
        f"marketplace: {pa.get('marketplace', '')}",
        f"amazon_account: {pa.get('amazon_account', '')}",
        f"product: {pa.get('product', '')}",
        f"asin: {pa.get('asin', '')}",
        "status: generated",
        "---",
        "",
        f"# {pa.get('client', 'Client')} {pa.get('product', 'Product')} Keyword Workbook Handoff",
        "",
        "## Outputs",
        "",
        f"- Workbook: `{manifest['outputs']['workbook_local']}`",
        f"- Manifest: `{manifest['outputs']['manifest']}`",
        f"- Drive copy: `{manifest['outputs'].get('drive_copy') or 'not copied in this run'}`",
        "",
        "## DataDive Exports",
        "",
    ]
    for key, value in (cfg.get("datadive_exports") or {}).items():
        if not isinstance(value, dict):
            continue
        lines.append(f"- `{key}`: min relevancy `{value.get('min_relevancy','')}`, visible keywords `{value.get('visible_keyword_count','')}`, visible SV `{value.get('visible_search_volume','')}`, exported `{value.get('export_timestamp','')}`")
    lines += [
        "",
        "## Source Files",
        "",
    ]
    for key, value in manifest.get("sources", {}).items():
        if isinstance(value, dict):
            lines.append(f"- `{key}`: `{value.get('path','')}`")
    lines += [
        "",
        "## QA",
        "",
        f"- Validations all pass: `{manifest.get('validations_all_pass')}`",
    ]
    for check in manifest.get("validations", []):
        lines.append(f"- {'PASS' if check.get('pass') else 'FAIL'}: {check.get('check')} ({check.get('detail','')})")
    lines += [
        "",
        "## Next-Agent Prompt",
        "",
        "```text",
        f"You are working in `<repo-root>`.",
        "Use the project-local skill `skills/amazon-seo-keyword-workflow/SKILL.md`.",
        f"Continue from workbook `{manifest['outputs']['workbook_local']}` and manifest `{manifest['outputs']['manifest']}`.",
        "Do not edit live listings or upload anything to Seller Central without the operator's explicit approval.",
        "Verify health-claim-sensitive language against the actual product label and Amazon/EU compliance before final copy.",
        "```",
        "",
    ]
    if manifest.get("warnings"):
        lines += ["## Warnings", ""]
        for warning in manifest["warnings"]:
            lines.append(f"- {warning}")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


# --------------------------------------------------------------------------- #
# Preflight (auto-generated cross-agent handoff)
# --------------------------------------------------------------------------- #
# DataDive inputs the MCP can supply — Claude generates these from saved MCP JSON
# via datadive_mcp_to_csv.py (get_niche_roots / get_niche_keywords / get_niche_competitors),
# so they are NOT part of the Codex browser task. Validated byte-for-data-identical to the
# UI exports on a validation run (roots 222/222, Core 257/257, 0 mismatches). The Expanded
# 1% MKL stays in CODEX because the MCP returns only the ~visible/tracked set, not the 1% tail.
MCP_INPUT_KEYS = ["roots_csv", "master_csv", "competitors_csv"]
CODEX_INPUT_KEYS = [
    "expanded_mkl_csv",
    "poe_products_csv", "poe_search_terms_csv", "related_niches_json",
    "poe_reviews_json", "poe_returns_json", "poe_structured_json", "listing_reference_json",
]
SETUP_INPUT_KEYS = ["template", "seo_content"]


def _codex_handoff_block(cfg: dict, args: dict, missing: list[str]) -> str:
    pa = cfg["product_anchor"]
    client = (pa.get("client") or "").strip()
    product = (pa.get("product") or "").strip()
    # Avoid "Acme Acme Collagen" when the product name already starts with the client.
    subject = product if product.lower().startswith(client.lower()) and client else f"{client} {product}".strip()
    lines = [
        "TASK: Gather the missing keyword-workbook inputs for the connected/internal browser + DataDive UI.",
        "",
        f"Objective: produce the inputs below for {subject} "
        f"({pa.get('marketplace')}), DataDive niche {pa.get('datadive_niche')}, anchor ASIN {pa.get('asin')}.",
        "",
        "Do NOT: run the builder, write SEO, edit listings, commit/push, or inspect cookies/session/credentials.",
        "",
        "Produce these files at exactly these paths:",
    ]
    label = {
        "roots_csv": "DataDive UI — niche analysis roots CSV",
        "master_csv": "DataDive UI — Core MKL CSV at 30% Min Rel.",
        "expanded_mkl_csv": "DataDive UI — Expanded MKL CSV at 1% Min Rel.",
        "competitors_csv": "DataDive UI (or MCP fallback) — competitors CSV",
        "poe_products_csv": "POE — Niche Details > Products tab CSV",
        "poe_search_terms_csv": "POE — Niche Details > Search Terms tab CSV",
        "related_niches_json": "POE — related-niches capture JSON",
        "poe_reviews_json": "POE — Customer Review Insights capture JSON",
        "poe_returns_json": "POE — Returns route/table capture JSON",
        "poe_structured_json": "POE — structured overview capture JSON",
        "listing_reference_json": (
            "Listing copy for the ANCHOR + COMPETITORS — use the `amazon-listing-capture` skill "
            "/ `tools/listing-capture/extract-amazon-listing-copy.js` (local language; title #productTitle; "
            "bullets #feature-bullets ul, fallback #productFactsDesktopExpander > div:first-child ul; "
            "deterministic ASIN). Output ONE file per `tools/listing-capture/listing-reference.schema.v1.json` "
            "(listings:[{asin,title,bullets[],link,status}])"
        ),
    }
    for k in missing:
        lines.append(f"  [ ] {label.get(k, k)}\n        -> {args[k]}")
    lines += [
        "",
        "Listing capture covers the anchor + the niche's competitor ASINs (input = the ASINs tab `link` "
        "column once built, or the DataDive competitor set). Keep a row even when title/bullets fail (set status).",
        "Caveats to confirm: Seller Central account + marketplace; any ASIN that resolves to the wrong product "
        "family; ingredient/health-claim risks (flag joint/skin/hair/nail/anti-age claims in the live copy).",
        "",
        "Known capture quirks (from prior runs):",
        "- DataDive export buttons may emit NO detectable download event. If exports don't land, ask the operator to "
        "click them manually, then map the files in ~/Downloads by filename/timestamp/rows/headers "
        "(Core 30% includes a 'Sugg. bid & range' column; the Expanded 1% file has many more rows). Report the "
        "row counts + headers per file so Claude can cross-check them against the DataDive MCP niche statistics.",
        "- POE: direct tab URLs may render only the tab header — click the in-page tab to load the actual "
        "Products/Search Terms/Reviews/Returns content before capturing.",
        "- POE Download clicks DO work even when the browser download event times out — check ~/Downloads for "
        "the new file and rename it to the contract path.",
        "- Amazon listing pages may render in English despite locale URL params — switch the Amazon site "
        "language preference to the marketplace language (e.g. 'italiano - IT'), then re-run the capture.",
        "- After Claude confirms the canonical files, delete duplicate/raw intermediate downloads "
        "(NEVER the canonical contract paths).",
        "",
        "Stop and report: the exact saved paths + the caveats above, then hand back to Claude.",
        "",
        # Per-run protocol lives in the client's own Obsidian folder (config.inputs.handoff_note),
        # so each client/run is self-contained rather than appended to one shared file. Falls back
        # to the reusable Context protocol only when a run-specific note isn't configured.
        f"Protocol: {(args.get('handoff_note') or '').strip() or '<your-vault>/Context/codex-claude-handoff-protocol.md'}",
    ]
    return "\n".join(lines)


def _build_command(args: dict) -> str:
    return (
        ".venv/bin/python tools/amazon-seo-keyword-workbook/build_keyword_workbook.py "
        f'--config "{args["config"]}"'
    )


def run_preflight(cfg: dict, args: dict) -> int:
    pa = cfg["product_anchor"]
    print("=" * 64)
    print(f"PREFLIGHT — {pa.get('client')} · {pa.get('product')}")
    print(f"  marketplace={pa.get('marketplace')}  niche={pa.get('datadive_niche')}  anchor={pa.get('asin')}")
    print("=" * 64)
    present = [k for k in MCP_INPUT_KEYS + CODEX_INPUT_KEYS + SETUP_INPUT_KEYS if os.path.exists(args[k])]
    missing_mcp = [k for k in MCP_INPUT_KEYS if not os.path.exists(args[k])]
    missing_codex = [k for k in CODEX_INPUT_KEYS if not os.path.exists(args[k])]
    missing_setup = [k for k in SETUP_INPUT_KEYS if not os.path.exists(args[k])]

    print("PRESENT:")
    for k in present:
        print(f"  [x] {k}")
    if missing_mcp or missing_codex or missing_setup:
        print("MISSING:")
        for k in missing_mcp:
            print(f"  [ ] (MCP)    {k}  -> {args[k]}")
        for k in missing_codex:
            print(f"  [ ] (CODEX)  {k}  -> {args[k]}")
        for k in missing_setup:
            print(f"  [ ] (setup)  {k}  -> {args[k]}")

    if missing_mcp:
        niche = cfg["product_anchor"].get("datadive_niche", "<NICHE>")
        anchor = cfg["product_anchor"].get("asin", "<ANCHOR>")
        print(
            "\nSTATUS: NEEDS MCP GEN — Claude generates these from the DataDive MCP (no browser download):\n"
            "  1) Call get_niche_roots / get_niche_keywords / get_niche_competitors for niche "
            f"{niche}; save each raw JSON response to a file.\n"
            "  2) Cross-check len(keywords) == get_niche_competitors numVisibleKeywords before trusting Core.\n"
            "  3) Run datadive_mcp_to_csv.py to write roots_csv/master_csv/competitors_csv:\n"
            f"     .venv/bin/python tools/amazon-seo-keyword-workbook/datadive_mcp_to_csv.py --anchor {anchor} \\\n"
            "       --roots-json <roots.json> --keywords-json <keywords.json> --competitors-json <comps.json> \\\n"
            f"       --out-roots \"{args['roots_csv']}\" --out-core \"{args['master_csv']}\" --out-competitors \"{args['competitors_csv']}\"\n"
        )
    if missing_codex:
        print("\nSTATUS: WAITING ON CODEX. Paste the block below into Codex (do not hand-write it):\n")
        print(_codex_handoff_block(cfg, args, missing_codex))
    elif missing_mcp:
        pass
    elif missing_setup:
        print(f"\nSTATUS: NEEDS SETUP — provide: {', '.join(missing_setup)}")
    else:
        print("\nSTATUS: READY — all inputs present. Hand to Claude to write SEO + build, or run:\n")
        print(_build_command(args))
    return 0


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> int:
    ap = argparse.ArgumentParser(description="Build an Amazon SEO keyword workbook.")
    for k, v in DEFAULTS.items():
        ap.add_argument("--" + k.replace("_", "-"), default=v)
    ap.add_argument("--preflight", action="store_true",
                    help="Check inputs and print the cross-agent handoff / build-readiness status; do not build.")
    args = vars(ap.parse_args())
    args = {k.replace("-", "_"): v for k, v in args.items()}

    cfg = json.load(open(args["config"], encoding="utf-8"))
    warnings: list[str] = []

    # Compliance layer: audit-derived prohibited terms (from a health-claims
    # self-check or an Amazon SAS audit) extend the hand-authored claim_tokens.
    comp = cfg.get("compliance") or {}
    audited_prohibited = [
        e.get("term") for e in (comp.get("claims_audit") or [])
        if norm(e.get("verdict")) == "prohibited" and norm(e.get("term")) and not e.get("term", "").startswith("<")
    ]
    if audited_prohibited:
        tri = cfg.setdefault("triage", {})
        existing = {norm(t) for t in tri.get("claim_tokens", [])}
        tri.setdefault("claim_tokens", []).extend(
            t for t in audited_prohibited if norm(t) not in existing
        )

    # Input contract: paths may live in config `inputs{}` so they aren't recited
    # on the CLI. Precedence: explicit CLI flag > config inputs > built-in default.
    explicit = {
        a.lstrip("-").replace("-", "_").split("=")[0]
        for a in sys.argv[1:] if a.startswith("--")
    }

    def _resolve(p: str) -> str:
        p = os.path.expanduser(str(p))
        return p if os.path.isabs(p) else os.path.join(REPO, p)

    for k, v in (cfg.get("inputs") or {}).items():
        if k in DEFAULTS and k not in explicit:
            args[k] = _resolve(v)

    if args.get("preflight"):
        return run_preflight(cfg, args)

    # Verify sources exist.
    src_keys = [
        "template", "roots_csv", "master_csv", "expanded_mkl_csv", "competitors_csv",
        "poe_products_csv", "poe_search_terms_csv", "related_niches_json",
    ]
    rebuild_tabs = set(cfg["tabs"].get("rebuild", []))
    if "POE Raw - Reviews" in rebuild_tabs:
        src_keys.append("poe_reviews_json")
    if "POE Raw - Returns" in rebuild_tabs:
        src_keys.append("poe_returns_json")
    if "POE Semantic Insights" in rebuild_tabs and args.get("poe_structured_json"):
        src_keys.append("poe_structured_json")
    if "ASINs" in rebuild_tabs:
        src_keys.append("listing_reference_json")
    missing = [k for k in src_keys if not os.path.exists(args[k])]
    if missing:
        print("ERROR: missing source files:", file=sys.stderr)
        for k in missing:
            print(f"   {k}: {args[k]}", file=sys.stderr)
        print("\nRun with --preflight to generate the Codex handoff for the missing inputs.", file=sys.stderr)
        return 2

    print(f"Loading template: {args['template']}")
    wb = load_workbook(args["template"])

    counts: dict[str, Any] = {}

    # --- Exact-paste tabs --------------------------------------------------- #
    roots = read_csv(args["roots_csv"])
    master = drop_leading_empty_columns(read_csv(args["master_csv"]))
    expanded_mkl = drop_leading_empty_columns(read_csv(args["expanded_mkl_csv"]))
    competitors = read_csv(args["competitors_csv"])
    poe_products = read_csv(args["poe_products_csv"])
    poe_search = read_csv(args["poe_search_terms_csv"])

    counts["_roots_csv"] = len(roots)
    counts["_master_csv"] = len(master)
    counts["_expanded_mkl_csv"] = len(expanded_mkl)
    counts["_competitors_csv"] = len(competitors)
    counts["_poe_products_csv"] = len(poe_products)
    counts["_poe_search_terms_csv"] = len(poe_search)
    counts["_master_header"] = master[0] if master else []

    # The DataDive roots CSV ships a leading empty column and an unnamed
    # trailing 0–1 relevance score; transform_roots() renames the columns,
    # marks important/ad roots (config: root_importance — tiered ⭐⭐ ad roots
    # when ad_min_sv/ad_min_score are set), and buckets each root with the same
    # vocabulary the Never-Ever ladder uses.
    poe_terms = load_poe_terms(args["poe_search_terms_csv"])
    roots_vocab = _triage_vocab(master, poe_terms, args, cfg)
    roots_disp = transform_roots(roots, cfg, roots_vocab)

    exact = cfg["tabs"]["exact_paste"]
    src_for = {
        "roots_csv": roots_disp,
        "master_csv": master,
        "expanded_mkl_csv": expanded_mkl,
        "poe_products_csv": poe_products,
        "poe_search_terms_csv": poe_search,
    }
    for tab, src_key in exact.items():
        ensure_sheet(wb, tab)
        if tab not in wb.sheetnames:
            warnings.append(f"Exact-paste tab '{tab}' not in template; skipped.")
            continue
        n = build_exact_paste(wb[tab], src_for[src_key], warnings)
        counts[tab] = n
        print(f"  exact-paste {tab!r}: {n} rows")

    # --- ASINs / competitors / Never Ever from current sources ---------------- #
    asins_result: dict = {}
    if "ASINs" in wb.sheetnames or "ASINs" in cfg["tabs"].get("rebuild", []):
        ws = ensure_sheet(wb, "ASINs")
        build_asins(ws, cfg, competitors, args.get("listing_reference_json", ""), asins_result, warnings)
        counts["ASINs"] = asins_result.get("rows_written", 0)
        print(f"  rebuilt 'ASINs': {asins_result.get('rows_written', 0)} rows")

    competitors_result: dict = {}
    # Legacy/optional: the current template drops "4.2 Competitors" (competitor
    # copy lives in the ASINs tab). This stays dormant unless a config/template
    # reintroduces the tab — then it rebuilds it from competitors_csv.
    if "4.2 Competitors" in wb.sheetnames or "4.2 Competitors" in cfg["tabs"].get("rebuild", []):
        ws = ensure_sheet(wb, "4.2 Competitors")
        build_competitors(ws, competitors, competitors_result, warnings)
        counts["4.2 Competitors"] = competitors_result.get("rows_written", 0)
        print(f"  rebuilt '4.2 Competitors': {competitors_result.get('rows_written', 0)} rows")

    never_result: dict = {}
    # Tab is '2 Never KWs' in the current template (legacy name '2.2 Never KWs');
    # resolve whichever exists so the negatives actually rebuild instead of leaving
    # stale carried-forward template rows.
    never_sheet = next(
        (s for s in ("2 Never KWs", "2.2 Never KWs") if s in wb.sheetnames),
        None,
    )
    if never_sheet:
        ws = ensure_sheet(wb, never_sheet)
        build_never_keywords(ws, expanded_mkl, master, poe_terms, args, cfg, never_result, warnings)
        counts[never_sheet] = never_result.get("rows_written", 0)
        print(
            f"  rebuilt {never_sheet!r}: {never_result.get('never_ever_words', 0)} one-word negatives, "
            f"{never_result.get('brand_words', 0)} brand, {never_result.get('claim_words', 0)} claim-risk, "
            f"{never_result.get('review_words', 0)} review-band, {never_result.get('phrase_candidates', 0)} phrase candidates"
        )

    # --- Related niches (strict filter) ------------------------------------ #
    related_result: dict = {}
    if "POE Raw - Related Niches" in wb.sheetnames:
        build_related_niches(
            wb["POE Raw - Related Niches"], args["related_niches_json"], cfg, related_result, warnings
        )
        counts["POE Raw - Related Niches"] = related_result.get("rows_written", 0)
        print(
            f"  rebuilt 'POE Raw - Related Niches': kept {len(related_result.get('kept', []))} "
            f"niches, dropped {len(related_result.get('dropped', []))}"
        )
    else:
        warnings.append("'POE Raw - Related Niches' not in template; skipped.")

    # --- Outlier triage ----------------------------------------------------- #
    # The opportunity-KW tab is named 'Outlier - Opportunity KWs' in the current
    # template (legacy name was '3.2 Outlier KWs'); resolve whichever exists so the
    # triage actually rebuilds instead of silently leaving stale carried-forward rows.
    outlier_result: dict = {}
    outlier_sheet = next(
        (s for s in ("Outlier - Opportunity KWs", "3.2 Outlier KWs") if s in wb.sheetnames),
        None,
    )
    if outlier_sheet:
        poe_terms = load_poe_terms(args["poe_search_terms_csv"])
        build_outlier(
            wb[outlier_sheet], master, expanded_mkl, poe_terms, cfg, outlier_result, warnings
        )
        counts[outlier_sheet] = outlier_result.get("rows_written", 0)
        counts["_invalid_actions"] = outlier_result.get("invalid_actions", [])
        print(
            f"  rebuilt {outlier_sheet!r}: {outlier_result.get('outlier_keywords', 0)} "
            f"triaged keywords"
        )
    else:
        warnings.append("Outlier opportunity-KW tab not in template; skipped.")

    # --- SEO Text rewrite (curated, + Ranking Juice column) ----------------- #
    # Tab is '4.1 SEO Text' in the current template (legacy name '4. SEO Text');
    # resolve whichever exists so the SEO copy actually rebuilds instead of leaving
    # the style-template product's (e.g. collagen) SEO text in place.
    seo_result: dict = {}
    seo_sheet = next(
        (s for s in ("4.1 SEO Text", "4. SEO Text") if s in wb.sheetnames),
        None,
    )
    if seo_sheet:
        build_seo_text(wb[seo_sheet], args["seo_content"], seo_result, warnings)
        counts[seo_sheet] = seo_result.get("rows_written", 0)
        if not seo_result.get("skipped"):
            print(f"  rebuilt {seo_sheet!r}: {seo_result.get('rows_written', 0)} rows (+ Ranking Juice column)")
    else:
        warnings.append("SEO Text tab not in template; skipped.")

    # Optional per-supply 'SEO Variations' tab (only if seo_content defines it).
    seo_var_result: dict = {}
    build_seo_variations(wb, args["seo_content"], seo_var_result, warnings)
    if not seo_var_result.get("skipped"):
        counts["SEO Variations"] = seo_var_result.get("rows_written", 0)
        print(f"  built 'SEO Variations': {seo_var_result.get('rows_written', 0)} rows")

    # --- Curated tabs (real POE-sourced content, e.g. Reviews) -------------- #
    curated_result: dict = {}
    if args.get("curated_tabs"):
        build_curated_tabs(wb, args["curated_tabs"], curated_result, warnings)
        if curated_result.get("tabs"):
            print(f"  curated tabs: {', '.join(curated_result['tabs'])}")

    # --- Current POE evidence tabs ------------------------------------------- #
    current_poe_done: set[str] = set()
    reviews_result: dict = {}
    if "POE Raw - Reviews" in rebuild_tabs and "POE Raw - Reviews" in wb.sheetnames:
        build_poe_reviews(wb["POE Raw - Reviews"], args.get("poe_reviews_json", ""), reviews_result, warnings)
        counts["POE Raw - Reviews"] = reviews_result.get("rows_written", 0)
        current_poe_done.add("POE Raw - Reviews")
        print(
            f"  rebuilt 'POE Raw - Reviews': {reviews_result.get('topics', 0)} topics "
            f"({reviews_result.get('status')})"
        )

    returns_result: dict = {}
    if "POE Raw - Returns" in rebuild_tabs and "POE Raw - Returns" in wb.sheetnames:
        build_poe_returns(wb["POE Raw - Returns"], args.get("poe_returns_json", ""), returns_result, warnings)
        counts["POE Raw - Returns"] = returns_result.get("rows_written", 0)
        current_poe_done.add("POE Raw - Returns")
        print(
            f"  rebuilt 'POE Raw - Returns': {returns_result.get('returns_rows', 0)} rows "
            f"({returns_result.get('status')})"
        )

    semantic_result: dict = {}
    if "POE Semantic Insights" in rebuild_tabs and "POE Semantic Insights" in wb.sheetnames:
        build_semantic_insights(
            wb["POE Semantic Insights"],
            args,
            reviews_result,
            returns_result,
            related_result,
            semantic_result,
            warnings,
        )
        counts["POE Semantic Insights"] = semantic_result.get("rows_written", 0)
        current_poe_done.add("POE Semantic Insights")
        print(f"  rebuilt 'POE Semantic Insights': {semantic_result.get('rows_written', 0)} rows")

    # --- Clear product-specific curated placeholders (new-market builds) ---- #
    curated_done = set(curated_result.get("tabs", [])) | current_poe_done
    cleared_tabs = []
    note = (
        f"(Placeholder — curate for {cfg['product_anchor'].get('marketplace','this market')} / "
        f"{cfg['product_anchor'].get('product','this product')}. Carried-forward template content cleared.)"
    )
    for tab in cfg["tabs"].get("carry_forward_clear", []):
        if tab in curated_done:
            continue  # already populated with real content
        if tab in wb.sheetnames:
            clear_placeholder_tab(wb[tab], note)
            cleared_tabs.append(tab)
    if cleared_tabs:
        print(f"  cleared placeholder tabs: {', '.join(cleared_tabs)}")

    skipped_result: dict = {}
    for tab, reason in (cfg.get("tabs", {}).get("generated_blank") or {}).items():
        ws = ensure_sheet(wb, tab)
        build_blank_or_skipped(ws, tab, reason, skipped_result)
        counts[tab] = skipped_result[tab]["rows_written"]
        print(f"  generated blank/skipped {tab!r}: {reason}")

    # --- Enforce tab order -------------------------------------------------- #
    desired = [t for t in cfg["tab_order"] if t in wb.sheetnames]
    wb._sheets.sort(key=lambda ws: desired.index(ws.title) if ws.title in desired else 999)

    # --- Validations -------------------------------------------------------- #
    checks = run_validations(wb, cfg, counts, related_result, args, warnings)

    # --- Save workbook ------------------------------------------------------ #
    os.makedirs(os.path.dirname(args["out"]), exist_ok=True)
    wb.save(args["out"])
    print(f"\nSaved workbook: {args['out']}")

    # Optional: copy the finished .xlsx into a delivery folder (e.g. synced Drive).
    delivered_to = ""
    if args.get("drive_dir"):
        if os.path.isdir(args["drive_dir"]):
            dest = os.path.join(args["drive_dir"], os.path.basename(args["out"]))
            shutil.copy2(args["out"], dest)
            delivered_to = dest
            print(f"Delivered copy to: {dest}")
        else:
            warnings.append(f"--drive-dir not a directory; skipped copy: {args['drive_dir']}")

    # --- Manifest ----------------------------------------------------------- #
    populated = {
        ws.title: ws.max_row for ws in wb.worksheets
    }
    manifest = {
        "generated_at": _dt.datetime.now().astimezone().isoformat(),
            "generator": "tools/amazon-seo-keyword-workbook/build_keyword_workbook.py",
        "product_anchor": cfg["product_anchor"],
        "outputs": {
            "workbook_local": args["out"],
            "manifest": args["manifest"],
            "drive_copy": delivered_to,
            "drive_target_note": (
                "Use --drive-dir to copy the finished .xlsx into the synced Drive folder, "
                "or copy the reviewed local workbook there after confirmation."
            ),
        },
        "template": args["template"],
        "sources": {
            "roots_csv": {"path": args["roots_csv"], "rows": counts["_roots_csv"]},
            "master_csv": {"path": args["master_csv"], "rows": counts["_master_csv"]},
            "expanded_mkl_csv": {"path": args["expanded_mkl_csv"], "rows": counts["_expanded_mkl_csv"]},
            "competitors_csv": {"path": args["competitors_csv"], "rows": counts["_competitors_csv"]},
            "poe_products_csv": {"path": args["poe_products_csv"], "rows": counts["_poe_products_csv"]},
            "poe_search_terms_csv": {
                "path": args["poe_search_terms_csv"], "rows": counts["_poe_search_terms_csv"]
            },
            "related_niches_json": {"path": args["related_niches_json"]},
            "poe_reviews_json": {
                "path": args.get("poe_reviews_json", ""),
                "rows": reviews_result.get("rows_written", 0),
                "status": reviews_result.get("status", ""),
            },
            "poe_returns_json": {
                "path": args.get("poe_returns_json", ""),
                "rows": returns_result.get("rows_written", 0),
                "status": returns_result.get("status", ""),
            },
            "poe_structured_json": {"path": args.get("poe_structured_json", "")},
        },
        "tab_production": cfg["tabs"],
        "datadive_exports": cfg.get("datadive_exports", {}),
        "tab_order": [ws.title for ws in wb.worksheets],
        "row_counts": populated,
        "related_niche_filter": {
            "keep_configured": cfg["related_niche_filter"]["keep"],
            "kept": related_result.get("kept", []),
            "dropped": related_result.get("dropped", []),
            "niche_rows_written": related_result.get("niche_rows_written", 0),
        },
        "outlier_triage": {
            "outlier_keywords": outlier_result.get("outlier_keywords", 0),
            "by_classification": outlier_result.get("by_classification", {}),
            "by_final_action": outlier_result.get("by_final_action", {}),
        },
        "asins": asins_result,
        "competitors": competitors_result,
        "never_ever": never_result,
        "generated_blank": skipped_result,
        "seo_text": {
            "source": args["seo_content"],
            "rows_written": seo_result.get("rows_written", 0),
            "ranking_juice_snapshot": (
                json.load(open(args["seo_content"], encoding="utf-8")).get("ranking_juice_snapshot", {})
                if os.path.exists(args["seo_content"]) else {}
            ),
        },
        "poe_reviews": reviews_result,
        "poe_returns": returns_result,
        "poe_semantic_insights": semantic_result,
        "validations": checks,
        "validations_all_pass": all(c["pass"] for c in checks),
        "warnings": warnings,
    }
    handoff_path = write_handoff_note(args.get("handoff_note", ""), cfg, args, manifest)
    if handoff_path:
        manifest["outputs"]["handoff_note"] = handoff_path
    os.makedirs(os.path.dirname(args["manifest"]), exist_ok=True)
    json.dump(manifest, open(args["manifest"], "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"Saved manifest: {args['manifest']}")
    if handoff_path:
        print(f"Saved handoff note: {handoff_path}")

    # --- Console summary ---------------------------------------------------- #
    print("\n=== VALIDATIONS ===")
    for c in checks:
        print(f"  [{'PASS' if c['pass'] else 'FAIL'}] {c['check']}  ({c['detail']})")
    if warnings:
        print("\n=== WARNINGS ===")
        for w in warnings:
            print("  -", w)

    return 0 if manifest["validations_all_pass"] else 1


if __name__ == "__main__":
    sys.exit(main())
