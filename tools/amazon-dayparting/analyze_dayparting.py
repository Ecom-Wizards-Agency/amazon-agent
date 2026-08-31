#!/usr/bin/env python3
"""Build a deterministic hourly Sponsored Products dayparting analysis."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError


DAYS = ("MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN")
REQUIRED = ("start_date", "start_time", "campaign", "currency", "impressions", "clicks", "spend", "orders", "sales")
ALIASES = {
    "start_date": ("start date", "date"),
    "start_time": ("start time", "hour", "time"),
    "campaign": ("campaign name", "campaign"),
    "currency": ("currency", "currency code"),
    "impressions": ("impressions",),
    "clicks": ("clicks",),
    "spend": ("spend", "cost"),
    "orders": ("orders", "7 day total orders", "14 day total orders", "purchases"),
    "sales": ("7 day total sales", "14 day total sales", "30 day total sales", "sales", "attributed sales"),
    "portfolio": ("portfolio name", "portfolio"),
}


class DaypartingError(ValueError):
    pass


def normalized_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def resolve_headers(headers: list[object]) -> dict[str, str]:
    normalized = {normalized_header(header): str(header) for header in headers if header is not None}
    resolved = {}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                resolved[canonical] = normalized[alias]
                break
    missing = [field for field in REQUIRED if field not in resolved]
    if missing:
        raise DaypartingError("hourly report fingerprint missing: " + ", ".join(missing))
    return resolved


def parse_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace("\u00a0", " ")
    negative = raw.startswith("(") and raw.endswith(")")
    raw = re.sub(r"[^0-9,\.\-+]", "", raw)
    if not raw:
        return 0.0
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        if re.search(r",\d{1,2}$", raw):
            raw = raw.replace(",", ".")
        else:
            raw = raw.replace(",", "")
    number = float(raw)
    return -number if negative else number


def parse_date(value: object, date_format: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    formats = {
        "mdy": ("%m/%d/%Y", "%m-%d-%Y"),
        "dmy": ("%d/%m/%Y", "%d-%m-%Y"),
        "ymd": ("%Y-%m-%d", "%Y/%m/%d"),
    }
    candidates = formats[date_format] if date_format != "auto" else formats["ymd"] + formats["mdy"]
    for fmt in candidates:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    raise DaypartingError(f"unrecognized date {raw!r}; set --date-format mdy, dmy, or ymd")


def parse_hour(value: object) -> int:
    if isinstance(value, datetime):
        return value.hour
    if hasattr(value, "hour") and not isinstance(value, str):
        return int(value.hour)
    raw = str(value).strip()
    match = re.match(r"^(\d{1,2})(?::\d{2})?", raw)
    if not match or not 0 <= int(match.group(1)) <= 23:
        raise DaypartingError(f"unrecognized hour {raw!r}; expected 0 through 23 or HH:MM")
    return int(match.group(1))


def read_source(path: Path) -> tuple[list[dict[str, object]], list[object]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(handle, dialect=dialect)
            return list(reader), list(reader.fieldnames or [])
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(values))
        except StopIteration as exc:
            raise DaypartingError("hourly workbook is empty") from exc
        rows = [dict(zip(headers, row)) for row in values]
        return rows, headers
    raise DaypartingError("input must be .csv, .xlsx, or .xlsm")


def empty_totals() -> dict[str, float]:
    return {key: 0.0 for key in ("impressions", "clicks", "orders", "spend", "sales")}


def add_totals(target: dict[str, float], row: dict[str, object]) -> None:
    for key in target:
        target[key] += float(row[key])


def metrics(totals: dict[str, float]) -> dict[str, float | None]:
    impressions, clicks = totals["impressions"], totals["clicks"]
    orders, spend, sales = totals["orders"], totals["spend"], totals["sales"]
    return {
        **totals,
        "ctr": clicks / impressions if impressions else None,
        "cvr": orders / clicks if clicks else None,
        "cpc": spend / clicks if clicks else None,
        "acos": spend / sales if sales else None,
        "roas": sales / spend if spend else None,
        "rpc": sales / clicks if clicks else None,
        "actc": clicks / orders if orders else None,
    }


def close_enough(left: float, right: float) -> bool:
    return math.isclose(left, right, rel_tol=1e-9, abs_tol=1e-6)


def analyze_file(
    input_path: Path,
    out_dir: Path,
    timezone: str,
    date_format: str = "auto",
    currency_filter: str | None = None,
    confidence_multiplier: float = 5.0,
    min_clicks_override: int | None = None,
    exclude_recent_days: int = 2,
    allow_thin_data: bool = False,
    max_increase: int = 300,
    max_decrease: int = 99,
    campaign_pattern: str | None = None,
    portfolio_pattern: str | None = None,
) -> dict[str, object]:
    try:
        ZoneInfo(timezone)
    except ZoneInfoNotFoundError as exc:
        raise DaypartingError(f"unknown IANA timezone {timezone!r}") from exc
    raw_rows, headers = read_source(input_path)
    mapping = resolve_headers(headers)
    if portfolio_pattern and "portfolio" not in mapping:
        raise DaypartingError("--portfolio-regex was supplied but the source has no portfolio column")
    try:
        campaign_re = re.compile(campaign_pattern, re.I) if campaign_pattern else None
        portfolio_re = re.compile(portfolio_pattern, re.I) if portfolio_pattern else None
    except re.error as exc:
        raise DaypartingError(f"invalid scope regex: {exc}") from exc
    parsed = []
    for index, source in enumerate(raw_rows, 2):
        if not any(value not in (None, "") for value in source.values()):
            continue
        try:
            currency = str(source[mapping["currency"]] or "").strip().upper()
            if currency_filter and currency != currency_filter.upper():
                continue
            row = {
                "date": parse_date(source[mapping["start_date"]], date_format),
                "hour": parse_hour(source[mapping["start_time"]]),
                "campaign": str(source[mapping["campaign"]] or "").strip(),
                "currency": currency,
                "portfolio": str(source.get(mapping.get("portfolio", ""), "") or "").strip(),
            }
            if campaign_re and not campaign_re.search(row["campaign"]):
                continue
            if portfolio_re and not portfolio_re.search(row["portfolio"]):
                continue
            for key in ("impressions", "clicks", "spend", "orders", "sales"):
                row[key] = parse_number(source[mapping[key]])
            if row["impressions"] < 0 or row["clicks"] < 0 or row["orders"] < 0 \
                    or row["spend"] < 0 or row["sales"] < 0:
                raise DaypartingError("additive metrics cannot be negative")
            parsed.append(row)
        except (ValueError, TypeError, DaypartingError) as exc:
            raise DaypartingError(f"row {index}: {exc}") from exc
    if not parsed:
        raise DaypartingError("no hourly rows remained after parsing and filters")

    currencies = sorted({str(row["currency"]) for row in parsed})
    if len(currencies) != 1:
        raise DaypartingError(
            "mixed currencies cannot be aggregated: " + ", ".join(currencies) + "; rerun with --currency"
        )
    max_date = max(row["date"] for row in parsed)
    cutoff = max_date - timedelta(days=exclude_recent_days)
    settled = [row for row in parsed if row["date"] <= cutoff] if exclude_recent_days else parsed
    if not settled:
        raise DaypartingError("excluding recent attribution days left no settled rows")

    total = empty_totals()
    day_totals = {day: empty_totals() for day in DAYS}
    hour_totals = {hour: empty_totals() for hour in range(24)}
    block_totals = {start: empty_totals() for start in range(0, 24, 4)}
    grid_totals = {(day, hour): empty_totals() for day in DAYS for hour in range(24)}
    dates_by_day: dict[str, set[date]] = defaultdict(set)
    for row in settled:
        day = DAYS[row["date"].weekday()]
        add_totals(total, row)
        add_totals(day_totals[day], row)
        add_totals(hour_totals[row["hour"]], row)
        add_totals(block_totals[(row["hour"] // 4) * 4], row)
        add_totals(grid_totals[(day, row["hour"])], row)
        dates_by_day[day].add(row["date"])

    for field in total:
        if not close_enough(total[field], sum(day_totals[day][field] for day in DAYS)):
            raise DaypartingError(f"day reconciliation failed for {field}")
        if not close_enough(total[field], sum(hour_totals[hour][field] for hour in range(24))):
            raise DaypartingError(f"hour reconciliation failed for {field}")
        if not close_enough(total[field], sum(block_totals[start][field] for start in range(0, 24, 4))):
            raise DaypartingError(f"four-hour reconciliation failed for {field}")
        if not close_enough(total[field], sum(grid_totals[key][field] for key in grid_totals)):
            raise DaypartingError(f"grid reconciliation failed for {field}")

    overall = metrics(total)
    actc = overall["actc"]
    min_clicks = min_clicks_override
    if min_clicks is None and actc is not None:
        min_clicks = math.ceil(float(actc) * confidence_multiplier)
    distinct_dates = sorted({row["date"] for row in settled})
    enough_dates = len(distinct_dates) >= 14 or allow_thin_data
    recommendations_available = bool(min_clicks is not None and overall["rpc"] and enough_dates)

    def enrich(bucket: dict[str, float]) -> dict[str, object]:
        result = metrics(bucket)
        trusted = bool(min_clicks is not None and bucket["clicks"] >= min_clicks)
        recommendation = None
        if recommendations_available and trusted and result["rpc"] is not None:
            raw_change = round((float(result["rpc"]) / float(overall["rpc"]) - 1.0) * 100)
            recommendation = max(-max_decrease, min(max_increase, raw_change))
        result["trusted"] = trusted
        result["suggested_change_pct"] = recommendation
        return result

    by_day = [{"day": day, **enrich(day_totals[day])} for day in DAYS]
    by_hour = [{"hour": hour, **enrich(hour_totals[hour])} for hour in range(24)]
    by_block = [
        {"start_hour": start, "end_hour": start + 3, "window": f"{start:02d}-{start + 3:02d}",
         **enrich(block_totals[start])}
        for start in range(0, 24, 4)
    ]
    grid = [
        {"day": day, "hour": hour, **enrich(grid_totals[(day, hour)])}
        for day in DAYS for hour in range(24)
    ]

    warnings = []
    if len(distinct_dates) < 14:
        warnings.append(f"thin data: {len(distinct_dates)} settled dates; recommendations require 14 by default")
    if min_clicks is None:
        warnings.append("no orders in scope: aCTC and the default confidence threshold are unavailable")
    day_counts = {day: len(dates_by_day[day]) for day in DAYS}
    if max(day_counts.values()) - min(day_counts.values()) > 1:
        warnings.append("weekday coverage is uneven; partial weeks may overweight some days")
    if exclude_recent_days:
        warnings.append(f"excluded report dates after {cutoff.isoformat()} for attribution settlement")

    out_dir.mkdir(parents=True, exist_ok=True)
    write_metric_csv(out_dir / "by_day.csv", by_day, "day")
    write_metric_csv(out_dir / "by_hour.csv", by_hour, "hour")
    write_metric_csv(out_dir / "by_4hour.csv", by_block, "start_hour", "end_hour", "window")
    write_metric_csv(out_dir / "grid.csv", grid, "day", "hour")

    grid_lookup = {(row["day"], row["hour"]): row for row in grid}
    with (out_dir / "adlabs_grid.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["day"] + list(range(24)))
        for day in DAYS:
            writer.writerow([day] + [
                grid_lookup[(day, hour)]["suggested_change_pct"]
                if grid_lookup[(day, hour)]["suggested_change_pct"] is not None else 0
                for hour in range(24)
            ])

    rules = [
        {"days": row["day"], "hours": [row["hour"]], "change_pct": row["suggested_change_pct"]}
        for row in grid if row["suggested_change_pct"] not in (None, 0)
    ]
    (out_dir / "adlabs_rules.json").write_text(json.dumps(rules, indent=2) + "\n", encoding="utf-8")

    trusted = [row for row in grid if row["suggested_change_pct"] is not None]
    trusted_blocks = [row for row in by_block if row["suggested_change_pct"] is not None]
    best_block = max(trusted_blocks, key=lambda row: float(row["rpc"])) if trusted_blocks else None
    worst_block = min(trusted_blocks, key=lambda row: float(row["rpc"])) if trusted_blocks else None
    summary = {
        "input_file": input_path.name,
        "timezone": timezone,
        "currency": currencies[0],
        "source_rows": len(raw_rows),
        "parsed_rows": len(parsed),
        "settled_rows": len(settled),
        "excluded_recent_rows": len(parsed) - len(settled),
        "date_start": min(distinct_dates).isoformat(),
        "date_end": max(distinct_dates).isoformat(),
        "distinct_dates": len(distinct_dates),
        "dates_per_weekday": day_counts,
        "campaign_count": len({row["campaign"] for row in settled}),
        "campaigns": sorted({row["campaign"] for row in settled}),
        "portfolio_count": len({row["portfolio"] for row in settled if row["portfolio"]}),
        "portfolios": sorted({row["portfolio"] for row in settled if row["portfolio"]}),
        "campaign_regex": campaign_pattern,
        "portfolio_regex": portfolio_pattern,
        "confidence_multiplier": confidence_multiplier,
        "minimum_clicks": min_clicks,
        "recommendations_available": recommendations_available,
        "trusted_grid_cells": len(trusted),
        "best_trusted_4hour_window": best_block,
        "worst_trusted_4hour_window": worst_block,
        "overall": overall,
        "warnings": warnings,
    }
    (out_dir / "summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    return summary


def write_metric_csv(path: Path, rows: list[dict[str, object]], *leading: str) -> None:
    metric_fields = (
        "impressions", "clicks", "orders", "spend", "sales", "ctr", "cvr", "cpc", "acos", "roas",
        "rpc", "actc", "trusted", "suggested_change_pct",
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(leading) + list(metric_fields))
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field) for field in writer.fieldnames})


def main() -> int:
    parser = argparse.ArgumentParser(description="Analyze an hourly Sponsored Products campaign report")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--out-dir", required=True, type=Path)
    parser.add_argument("--timezone", required=True, help="advertising profile timezone, e.g. America/New_York")
    parser.add_argument("--date-format", choices=("auto", "mdy", "dmy", "ymd"), default="auto")
    parser.add_argument("--currency", help="keep one currency when the source contains more than one")
    parser.add_argument("--confidence-multiplier", type=float, default=5.0)
    parser.add_argument("--min-clicks", type=int, help="override the aCTC-based confidence threshold")
    parser.add_argument("--exclude-recent-days", type=int, default=2)
    parser.add_argument("--allow-thin-data", action="store_true")
    parser.add_argument("--max-increase", type=int, default=300)
    parser.add_argument("--max-decrease", type=int, default=99)
    parser.add_argument("--campaign-regex", help="case-insensitive regex for campaigns to include")
    parser.add_argument("--portfolio-regex", help="case-insensitive regex for portfolios to include")
    args = parser.parse_args()
    if args.confidence_multiplier < 0 or args.exclude_recent_days < 0:
        parser.error("confidence multiplier and excluded recent days cannot be negative")
    if args.min_clicks is not None and args.min_clicks < 1:
        parser.error("--min-clicks must be at least 1")
    if not 0 <= args.max_increase <= 300 or not 0 <= args.max_decrease <= 99:
        parser.error("AdLabs bounds require max increase 0..300 and max decrease 0..99")
    try:
        summary = analyze_file(
            args.input, args.out_dir, args.timezone, args.date_format, args.currency,
            args.confidence_multiplier, args.min_clicks, args.exclude_recent_days,
            args.allow_thin_data, args.max_increase, args.max_decrease,
            args.campaign_regex, args.portfolio_regex,
        )
    except DaypartingError as exc:
        print(f"DAYPARTING: FAIL: {exc}", file=sys.stderr)
        return 1
    print(f"DAYPARTING: PASS: {summary['settled_rows']} settled rows, "
          f"{summary['distinct_dates']} dates, {summary['trusted_grid_cells']} trusted cells")
    for warning in summary["warnings"]:
        print(f"[WARN] {warning}")
    print(f"Wrote {args.out_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
