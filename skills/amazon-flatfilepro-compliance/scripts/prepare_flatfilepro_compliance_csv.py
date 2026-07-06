#!/usr/bin/env python3
"""Create a narrow FlatFilePro upload CSV from reviewed field changes.

Inputs:
  --source: FlatFilePro export as .csv/.tsv/.xlsx/.xlsm
  --changes: CSV with columns sku, attribute, value
  --output: destination CSV

The script validates that every requested attribute exists in the source
headers and writes only sku plus changed attributes. It does not decide which
values are compliant; that review must happen before this script is used.
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path


def read_source_headers(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SystemExit("openpyxl is required to read Excel sources.") from exc
        wb = load_workbook(path, read_only=True, data_only=False)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True))
        for row in rows:
            values = [str(cell).strip() if cell is not None else "" for cell in row]
            if "sku" in values:
                return values
        raise SystemExit("Could not find a header row containing 'sku' in the first 10 rows.")

    delimiter = "\t" if suffix == ".tsv" else ","
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            values = [cell.strip() for cell in row]
            if "sku" in values:
                return values
    raise SystemExit("Could not find a header row containing 'sku'.")


def read_changes(path: Path) -> OrderedDict[str, dict[str, str]]:
    by_sku: OrderedDict[str, dict[str, str]] = OrderedDict()
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        required = {"sku", "attribute", "value"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise SystemExit(f"Changes file missing columns: {', '.join(sorted(missing))}")
        for line_no, row in enumerate(reader, start=2):
            sku = (row.get("sku") or "").strip()
            attr = (row.get("attribute") or "").strip()
            value = row.get("value")
            if value is None:
                value = ""
            if not sku or not attr:
                raise SystemExit(f"Line {line_no}: sku and attribute are required.")
            by_sku.setdefault(sku, {})[attr] = value
    return by_sku


def read_source_rows(path: Path, headers: list[str]) -> dict[str, dict[str, str]]:
    """Read full source data rows keyed by sku (csv/tsv/xlsx), using the located header row."""
    suffix = path.suffix.lower()
    raw_rows: list[list[str]] = []
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=False)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            raw_rows.append([str(cell).strip() if cell is not None else "" for cell in row])
    else:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.reader(f, delimiter=delimiter):
                raw_rows.append([cell.strip() for cell in row])

    header_idx = None
    for idx, row in enumerate(raw_rows):
        if "sku" in row:
            header_idx = idx
            break
    if header_idx is None:
        raise SystemExit("Could not locate the header row while reading source data rows.")

    sku_pos = raw_rows[header_idx].index("sku")
    out: dict[str, dict[str, str]] = {}
    for row in raw_rows[header_idx + 1 :]:
        if sku_pos >= len(row) or not row[sku_pos]:
            continue
        rec = {}
        for i, header in enumerate(raw_rows[header_idx]):
            if header and i < len(row):
                rec[header] = row[i]
        out[row[sku_pos]] = rec
    return out


def write_output(
    path: Path,
    rows_by_sku: OrderedDict[str, dict[str, str]],
    headers: list[str],
    source_rows: dict[str, dict[str, str]] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    used_attrs: list[str] = []
    seen = set()
    for changes in rows_by_sku.values():
        for attr in changes:
            if attr not in seen:
                seen.add(attr)
                used_attrs.append(attr)

    source_order = {header: idx for idx, header in enumerate(headers)}
    ordered_attrs = sorted(used_attrs, key=lambda h: source_order.get(h, 10**9))
    output_headers = ["sku"] + [attr for attr in ordered_attrs if attr != "sku"]

    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=output_headers)
        writer.writeheader()
        for sku, changes in rows_by_sku.items():
            out = {header: "" for header in output_headers}
            out["sku"] = sku
            if source_rows is not None:
                current = source_rows.get(sku, {})
                for attr in output_headers:
                    if attr != "sku" and current.get(attr):
                        out[attr] = current[attr]
            for attr, value in changes.items():
                if attr != "sku":
                    out[attr] = value
            writer.writerow(out)


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--changes", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument(
        "--fill-unchanged",
        action="store_true",
        help="Fill every output cell that has no change value with the SKU's current "
        "source value (full-grid output). Prevents a mapped-but-empty cell from "
        "clearing a live value at upload time.",
    )
    args = parser.parse_args(argv)

    headers = read_source_headers(args.source)
    rows_by_sku = read_changes(args.changes)
    valid_headers = set(headers)

    missing_attrs = defaultdict(list)
    for sku, changes in rows_by_sku.items():
        for attr in changes:
            if attr != "sku" and attr not in valid_headers:
                missing_attrs[attr].append(sku)
    if missing_attrs:
        for attr, skus in sorted(missing_attrs.items()):
            print(f"Missing source header: {attr} (SKUs: {', '.join(skus)})", file=sys.stderr)
        return 2

    source_rows = read_source_rows(args.source, headers) if args.fill_unchanged else None
    write_output(args.output, rows_by_sku, headers, source_rows)
    mode = " (full-grid: unchanged cells carried from source)" if args.fill_unchanged else ""
    print(f"Wrote {len(rows_by_sku)} rows to {args.output}{mode}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
